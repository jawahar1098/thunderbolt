from app.ticketing import tickect_bp
from app.ticketing.database import *
import json
from app import app
from bson.objectid import ObjectId
from bson import ObjectId
from random import randint
from datetime import datetime
from mongoengine import Q
import pandas as pd
import numpy as np
import io
import math
import re
import datetime
import openpyxl
from app.auth.routes import token_required
import flask_cors
from dateutil import parser as date_parser
from dateutil.parser import ParserError
from werkzeug.utils import secure_filename
from flask import  jsonify, request, Response, session,send_file
from pymongo import MongoClient
from os import environ
import os
import bcrypt
import zipfile
import time
from functools import wraps
from bson.objectid import ObjectId
import json
from datetime import datetime, timedelta
from MongoClinet import thunderbolt as mongoticket
db = mongoticket().db
from MongoClinet import CDAT as mongocdat
import threading
sus_db = mongocdat().db

cors_allowed_ip = "*"

class MyEncoder(json.JSONEncoder):
    def default(self, obj):
        if isinstance(obj, ObjectId):
            return str(obj)
        return super(MyEncoder, self).default(obj)

tickect_bp.json_encoder = MyEncoder

admin_email = 'super@gmail.com'
admin_password = '123'
userDetails = {

  "name": "Super",
  "email": "super@gmail.com",
  "password1": "123",
  "password2": "123",
  "mobilenumber": "9637645941",
  "designation": "Administrator",
  "officername": "Super Admin",
  "question": "What's the name of the first school you attended?",
  "answer": "123",
  "status": "Active",
  "logged": "0"
}

@tickect_bp.route('/allinONE', methods=["POST", "GET"])
def allinONE():
    email = request.json.get('email')
    team = request.json.get("team")
    cat = request.json.get("cat")

    modules = request.json.get("modules")
    reqtype = request.json.get("type")
    officername = request.json.get("officername")
    if cat == 'Data Request':
        category = ['Data Request']
    elif cat == 'Analysis Request':
        category = ['Analysis Request','Note']
    #print(reqtype,'//////////////////////////////////////////////////////')
    application = ['NEXUS','CDAT','TICKETING','DATABASE','ANALYSIS','VIGOR']
    if email == admin_email and userDetails['password1'] == admin_password:
            collection = db['users']
            alldata = collection.find()
            all_data = []
            for i in alldata:
                i['_id'] = str(i['_id'])
                print(i['Application'])
                s = set(i['Application'])
                i['noaccess'] = [x for x in application if x not in s]
                all_data.append(i)
            return jsonify(all_data), 200
    user = Users.objects_safe(email=email).first().to_mongo().to_dict()
    designation = user['designation']
    role = user["role"]
    collection = db['users']
    coll = db['tickets']
    query = {}
    projection = {
            '_id' : 1,
            'priority' : 1,
            'useremail':1,
            'date' : 1,
            'token' : 1,
            'approval' : 1,
            'pending' : 1,
            'requestcategory' : 1,
            'requesttype' : 1,
            'requesttypes' : 1,
            'subtypes' : 1,
            'status_count': 1,
            'result': 1,
            'newnumber' : 1,
            'proforma_data' : 1,

        }
    
    projection1 = {
            '_id' : 1,
            "team": 1,
            'priority' : 1,
            'date' : 1,
            'useremail' : 1,
            'token' : 1,
            'approval' : 1,
            'officername' : 1,
            'pending' : 1,
            'requestcategory' : 1,
            'requesttype' : 1,
            'subtypes' : 1,
            'requesttypes' : 1,
            'edit':1,
            'status_count': 1,
            'result': 1,
            'newnumber' : 1,
            'proforma_data' : 1,

    }
    filter_conditions = []
    if team == "ADMIN":
        if reqtype == 'My_tickets':
            # Assuming page and limit are passed from the frontend
            page = request.json.get('page')
            limit = request.json.get('limit')
            # #print(type(page),type(limit))
            # Calculate skip based on page number and limit
            if page is None or limit is None:
                page = 1
                limit = 6
            skip = (page - 1) * limit
            # Fetch tickets with pagination
            #print(skip)
            My_tickets = coll.find({'useremail': email,'requestcategory': {'$in' :  category}}, projection).sort([('_id', -1)]).skip(skip).limit(limit)
            total_tickets = coll.count_documents({'useremail': email,'requestcategory': {'$in' :  category}})  # Total count of tickets
            # Create a list to store filtered tickets
            filtered_tickets = []
            res_len = 0
            newno_len = 0
            total_len = 0
            for res in My_tickets:
                if res['pending'] == 'Mail Under Process' and (res['subtypes'] == 'Phone' or res['requesttype'] == 'CDR'):
                    res_len = len(res['result'])
                    total_len = res_len - int(res['status_count'])
                    
                elif res['pending'] == 'Ticket_Closed' and (res['subtypes'] == 'Phone' or res['requesttype'] == 'CDR'):
                    res_len = len(res['result'])
                    total_len = res_len
                    
                elif res['pending'] == 'Mail Under Process' and (res['subtypes'] == 'IMEI' or res['requesttype'] == 'IMEI CDR'):
                    newno_len = len(res['newnumber'])
                    total_len = newno_len - int(res['status_count'])
                    
                elif res['pending'] == 'Ticket_Closed' and (res['subtypes'] == 'IMEI' or res['requesttype'] == 'IMEI CDR'):
                    newno_len = len(res['newnumber'])
                    total_len = newno_len
                    

                filtered_tickets.append({
                    '_id': res['_id'],
                    'priority': res['priority'],
                    'useremail':res['useremail'],
                    'date': res['date'],
                    'token': res['token'],
                    'approval': res['approval'],
                    'pending': res['pending'],
                    'requestcategory': res['requestcategory'],
                    'requesttype': res['requesttype'],
                    'requesttypes': res['requesttypes'],
                    'subtypes': res['subtypes'],
                    'status_count': res['status_count'],
                    'total_len': total_len,
                    "res_len": res_len,
                    "newno_len": newno_len
                })
            # Calculate pagination information
            total_pages = math.ceil(total_tickets / limit) # limit
            pagination_info = {
                'total_pages': int(total_pages),
                'current_page': page,
                'total_tickets': total_tickets
            }
            #print(pagination_info)
            # Send filtered_tickets and pagination_info to frontend
            return jsonify({'mytickets': filtered_tickets, 'pagination': pagination_info})
        elif reqtype == 'Others':
            # Similarly, add pagination for 'Others' request type
            page = request.json.get('page')
            limit = request.json.get('limit')
            #print(type(page),type(limit))
            # Calculate skip based on page number and limit
            if page is None or limit is None:
                page = 1
                limit = 6
            #print(page,limit)
            skip = (page - 1) * limit
            My_tickets = coll.find({'useremail': {'$nin' : [email]}, 'requestcategory': {'$in' :  category},}, projection1).sort([('_id', -1)]).skip(skip).limit(limit)
            total_tickets1 = coll.count_documents({'useremail': {'$nin' : [email]},'requestcategory': {'$in' :  category}})  # Total count of tickets
            filtered_tickets1 = []
            res_len = 0
            newno_len = 0
            total_len = 0
            for res in My_tickets:
                if res['pending'] == 'Mail Under Process' and (res['subtypes'] == 'Phone' or res['requesttype'] == 'CDR'):
                    res_len = len(res['result'])
                    total_len = res_len - int(res['status_count'])                    
                elif res['pending'] == 'Mail Under Process' and (res['subtypes'] == 'IMEI' or res['requesttype'] == 'IMEI CDR'):
                    newno_len = len(res['newnumber'])
                    total_len = newno_len - int(res['status_count'])
                    
                #print(res['_id'])
                filtered_tickets1.append({
                    '_id': res['_id'],
                    'priority': res['priority'],
                    'useremail':res['useremail'],
                    'date': res['date'],
                    'team': res['team'],
                    'token': res['token'],
                    "officername": res['officername'],
                    'approval': res['approval'],
                    'pending': res['pending'],
                    'requestcategory': res['requestcategory'],
                    'requesttype': res['requesttype'],
                    'subtypes': res['subtypes'],
                    'proforma_data': res['proforma_data'],
                    'requesttypes': res['requesttypes'],
                    'edit':res['edit'],
                    'status_count': res['status_count'],
                    'total_len': total_len,
                    "res_len": res_len,
                    "newno_len": newno_len
                    })
            total_pages1 = math.ceil(total_tickets1 / limit) # limit
            pagination_info1 = {
                'total_pages1': int(total_pages1),
                'current_page1': page,
                'total_tickets1': total_tickets1
            }
            return jsonify({'others': filtered_tickets1, 'pagination1': pagination_info1})

    elif designation == "SP" and team != 'ADMIN':
        if reqtype == 'My_tickets':
            # Assuming page and limit are passed from the frontend
            page = request.json.get('page')
            limit = request.json.get('limit')
            #print(type(page),type(limit))
            # Calculate skip based on page number and limit
            if page is None or limit is None:
                page = 1
                limit = 6
            skip = (page - 1) * limit
            # Fetch tickets with pagination
            My_tickets = coll.find({'useremail': email,'requestcategory': {'$in' :  category}}, projection).sort([('_id', -1)]).skip(skip).limit(limit)
            total_tickets = coll.count_documents({'useremail': email,'requestcategory': {'$in' :  category}})  # Total count of tickets
            # Create a list to store filtered tickets
            filtered_tickets = []
            res_len = 0
            newno_len = 0
            total_len = 0
            for res in My_tickets:
                if res['pending'] == 'Mail Under Process' and (res['subtypes'] == 'Phone' or res['requesttype'] == 'CDR'):
                    res_len = len(res['result'])
                    total_len = res_len - int(res['status_count'])
        
                elif res['pending'] == 'Mail Under Process' and (res['subtypes'] == 'IMEI' or res['requesttype'] == 'IMEI CDR'):
                    newno_len = len(res['newnumber'])
                    total_len = newno_len - int(res['status_count'])   
                filtered_tickets.append({
                    '_id': res['_id'],
                    'priority': res['priority'],
                    'useremail':res['useremail'],
                    'date': res['date'],
                    'token': res['token'],
                    'approval': res['approval'],
                    'pending': res['pending'],
                    'requestcategory': res['requestcategory'],
                    'requesttype': res['requesttype'],
                    'requesttypes': res['requesttypes'],
                    'subtypes': res['subtypes'],
                    'status_count': res['status_count'],
                    'proforma_data': res['proforma_data'],
                    'total_len': total_len,
                    "res_len": res_len,
                    "newno_len": newno_len
                })
            # Calculate pagination information
            total_pages = math.ceil(total_tickets / limit) # limit
            pagination_info = {
                'total_pages': int(total_pages),
                'current_page': page,
                'total_tickets': total_tickets
            }
            #print(pagination_info)
            # Send filtered_tickets and pagination_info to frontend
            return jsonify({'mytickets': filtered_tickets, 'pagination': pagination_info})
        elif reqtype == 'Others':
            # Similarly, add pagination for 'Others' request type
            page = request.json.get('page')
            limit = request.json.get('limit')
            #print(type(page),type(limit))
            # Calculate skip based on page number and limit
            if page is None or limit is None:
                page = 1
                limit = 6
            #print(page,limit)
            skip = (page - 1) * limit
            if team == 'CAT':
                query = {'$and' : [
                    {'useremail': {'$nin': [email]}},
                    {'team': team},
                    {'requestcategory': {'$in' :  category}},
                    {"$or": [
                        {"pending": "ADDL-SP/DSP Approval Pending", "team": team},
                        {"$and": [{"pending": {"$regex": officername}}, {"team": team}]},
                        {"pending": "Mail Under Process", "token": {"$regex": "_CO"}},
                        {"approval": "Approved by CAT DSP", "token": {"$regex": "_CO"}},
                        {"pending": "--", "requestcategory": "Analysis Request"},
                        {"approval": {"$regex": "^Assigned"}, "requestcategory": "Analysis Request"},
                        {"pending": "SP Approval Pending", "requestcategory": "Analysis Request"},
                        {"pending": "SP Approval Pending", "team": team},
                        {"pending": "Mail Under Process", "team": team},
                        {"pending": "Ticket_Closed", "team": team},
                        {"pending": {"$regex": "^Rejected"}, "team": team},
                        {"approval": {"$regex": "^Report"}, "requestcategory": "Analysis Request", "team": team},

                    ]}]}
            else:
                query = {'$and' : [
                    {'useremail': {'$nin': [email]}},
                    {'team': team},
                    {'requestcategory': {'$in' :  category}},
                    {"$or": [
                        {"pending": "SP Approval Pending", "team": team,},
                        {"pending": "SP Approval Pending", "team": team,},
                        {"pending": "SP Approval Pending", "team": team},
                        {"approval": "Approved by SP", "team": team,},
                        {"approval": "Approved by SP", "team": team,},
                        {"token": {"$regex": "_CO"}, "team": team},
                        {"pending": {"$regex": officername}, "team": team},
                        {"pending": "Assign to CAT Ins", "requestcategory": "Analysis Request", "team": team,},
                        {"pending": "SP Approval Pending", "requestcategory": "Analysis Request", "team": team,},
                        {"pending": "Assign to CAT Ins", "requestcategory": "Analysis Request", "team": team,},
                        {"pending": "Assign to CAT Ins", "requestcategory": "Analysis Request", "team": team,},
                        {"pending": "SP Approval Pending", "requestcategory": "Analysis Request", "team": team,},
                        {"pending": "SP Approval Pending", "requestcategory": "Analysis Request", "team": team,},
                        {"approval": {"$regex": "^Assigned"}, "requestcategory": "Analysis Request", "team": team},
                        {"approval": {"$regex": "^Report"}, "requestcategory": "Analysis Request", "team": team},
                        {"pending": "Mail Under Process", "team": team},
                        {"pending": {"$regex": "^Rejected"}, "team": team}
                    ]}]}

            My_tickets = coll.find(query, projection1).sort([('_id', -1)]).skip(skip).limit(limit)
            total_tickets1 = coll.count_documents(query)  # Total count of tickets
            filtered_tickets1 = []
            res_len = 0
            newno_len = 0
            total_len = 0
            for res in My_tickets:
                if res['pending'] == 'Mail Under Process' and (res['subtypes'] == 'Phone' or res['requesttype'] == 'CDR'):
                    res_len = len(res['result'])
                    total_len = res_len - int(res['status_count'])
                elif res['pending'] == 'Mail Under Process' and (res['subtypes'] == 'IMEI' or res['requesttype'] == 'IMEI CDR'):
                    newno_len = len(res['newnumber'])
                    total_len = newno_len - int(res['status_count'])

                #print(res['_id'])
                filtered_tickets1.append({
                    '_id': res['_id'],
                    'priority': res['priority'],
                    'useremail':res['useremail'],
                    'date': res['date'],
                    'token': res['token'],
                     'team': res['team'],
                    "officername": res['officername'],
                    'approval': res['approval'],
                    'pending': res['pending'],
                    'requestcategory': res['requestcategory'],
                    'requesttype': res['requesttype'],
                    'subtypes': res['subtypes'],
                    'requesttypes': res['requesttypes'],
                    'edit':res['edit'],
                    'status_count': res['status_count'],
                    'proforma_data': res['proforma_data'],
                    'total_len': total_len,
                    "res_len": res_len,
                    "newno_len": newno_len
                    })
            total_pages1 = math.ceil(total_tickets1 / limit)
            pagination_info1 = {
                'total_pages1': int(total_pages1),
                'current_page1': page,
                'total_tickets1': total_tickets1
            }
            return jsonify({'others': filtered_tickets1, 'pagination1': pagination_info1})
    elif designation == "ADDL-SP/DSP":
        if reqtype == 'My_tickets':
            # Assuming page and limit are passed from the frontend
            page = request.json.get('page')
            limit = request.json.get('limit')
            #print(type(page),type(limit))
            # Calculate skip based on page number and limit
            if page is None or limit is None:
                page = 1
                limit = 6
            skip = (page - 1) * limit
            # Fetch tickets with pagination
            My_tickets = coll.find({'useremail': email,'requestcategory': {'$in' :  category}}, projection).sort([('_id', -1)]).skip(skip).limit(limit)
            total_tickets = coll.count_documents({'useremail': email,'requestcategory': {'$in' :  category}})  # Total count of tickets
            # Create a list to store filtered tickets
            filtered_tickets = []
            res_len = 0
            newno_len = 0
            total_len = 0
            for res in My_tickets:
                if res['pending'] == 'Mail Under Process' and (res['subtypes'] == 'Phone' or res['requesttype'] == 'CDR'):
                    res_len = len(res['result'])
                    total_len = res_len - int(res['status_count'])                    
                elif res['pending'] == 'Mail Under Process' and (res['subtypes'] == 'IMEI' or res['requesttype'] == 'IMEI CDR'):
                    newno_len = len(res['newnumber'])
                    total_len = newno_len - int(res['status_count'])
                    
                filtered_tickets.append({
                    '_id': res['_id'],
                    'priority': res['priority'],
                    'useremail':res['useremail'],
                    'date': res['date'],
                    'token': res['token'],
                    'approval': res['approval'],
                    'pending': res['pending'],
                    'requestcategory': res['requestcategory'],
                    'requesttype': res['requesttype'],
                    'requesttypes': res['requesttypes'],
                    'subtypes': res['subtypes'],
                    'status_count': res['status_count'],
                    'total_len': total_len,
                    "res_len": res_len,
                    "newno_len": newno_len
                    
                })
            # Calculate pagination information
            total_pages = math.ceil(total_tickets / limit) # limit
            pagination_info = {
                'total_pages': int(total_pages),
                'current_page': page,
                'total_tickets': total_tickets
            }
            #print(pagination_info)
            # Send filtered_tickets and pagination_info to frontend
            return jsonify({'mytickets': filtered_tickets, 'pagination': pagination_info})
        elif reqtype == 'Others':
            # Similarly, add pagination for 'Others' request type
            page = request.json.get('page')
            limit = request.json.get('limit')
            #print(type(page),type(limit))
            # Calculate skip based on page number and limit
            if page is None or limit is None:
                page = 1
                limit = 6
            #print(page,limit)
            skip = (page - 1) * limit
            if team == "CAT":
                query = {'$and' : [
                    {'useremail': {'$nin': [email]}},
                    {'requestcategory': {'$in' :  category}},
                    {"$or": [
                        {"pending": "Mail Under Process", "token": {"$regex": "_CO"}, "designation": {"$ne": "SP"}},
                        {"$and": [{"pending": {"$regex": officername}}, {"team": team}]},
                        {"pending": "Ticket_Closed", "team": team, "designation": {"$ne": "SP"}},
                        {"approval": "Approved by CAT Ins", "token": {"$regex": "_CO"}, "designation": {"$ne": "SP"}},
                        {"pending": "ADDL-SP/DSP Approval Pending", "requestcategory": "Analysis Request"},
                        {"$and": [{"pending": "--"}, {"requestcategory": "Analysis Request"}, {"designation": {"$ne": "SP"}}]},
                        {"$and": [{"pending": "--"}, {"requestcategory": "Analysis Request"}, {"designation": {"$ne": "SP"}}]},
                        {"$and": [{"pending": "ADDL-SP/DSP Approval Pending"}, {"team": team}]},
                        {"approval": {"$regex": "^Assigned"}, "requestcategory": "Analysis Request", "designation": {"$ne": "SP"}},
                        {"approval": {"$regex": "^Assigned"}, "requestcategory": "Analysis Request", "designation": {"$ne": "SP"}},
                        {"pending": "SP Approval Pending", "team": team, "designation": {"$ne": "ADDL-SP/DSP"}},
                        {"pending": "SP Approval Pending", "requestcategory": "Analysis Request","designation": {"$ne": "ADDL-SP/DSP"}},
                        {"pending": "Mail Under Process", "team": team, "designation": {"$ne": "SP"}},
                        {"pending": {"$regex": "^Rejected"}, "team": team}
                        ]}
                ]
                    }
            else:
                query = {'$and' : [
                    {'useremail': {'$nin': [email]}},
                    {'requestcategory': {'$in' :  category}},
                    {"$or": [
                        {"$and": [{"pending": {"$regex": officername}}, {"team": team}]},
                        {"$and": [{"token": {"$regex": "_CO"}}, {"team": team}, {"others": {"$in": [officername]}}]},
                        {"pending": "ADDL-SP/DSP Approval Pending", "team": team, "requestcategory": "Data Request", "superior": officername},
                        {"pending": "ADDL-SP/DSP Approval Pending", "team": team, "requestcategory": "Data Request", "others": officername},
                        {"pending": "SP Approval Pending", "team": team, "others": {"$in": [officername]}},
                        {"pending": "SP Approval Pending", "team": team, "superior": officername},
                        {"pending": "SP Approval Pending", "team": team, "others": officername},
                        {"pending": {"$regex": "Rejected"}, "team": team, "superior": officername},
                        {"pending": {"$regex": "Rejected"}, "team": team, "others": officername},
                        {"pending": "ADDL-SP/DSP Approval Pending", "requestcategory": "Analysis Request", "team": team, "others": {"$in": [officername]}},
                        {"pending": "ADDL-SP/DSP Approval Pending", "requestcategory": "Analysis Request", "team": team, "superior": officername},
                        {"pending": "ADDL-SP/DSP Approval Pending", "requestcategory": "Analysis Request", "team": team, "others": officername},
                        {"pending": "Assign to CAT Ins", "requestcategory": "Analysis Request", "team": team, "designation": {"$nin": ["SP"]}, "others": {"$in": [officername]}},
                        {"pending": "Assign to CAT Ins", "requestcategory": "Analysis Request", "team": team, "designation": {"$nin": ["SP"]}, "superior": officername},
                        {"pending": "Assign to CAT Ins", "requestcategory": "Analysis Request", "team": team, "designation": {"$nin": ["SP"]}, "others": officername},
                        {"pending": "Mail Under Process", "team": team, "designation": {"$ne": "SP"}, "others": officername},
                        {"pending": "Mail Under Process", "team": team, "designation": {"$ne": "SP"}, "superior": officername},
                        {"approval": {"$regex": "Assigned"}, "requestcategory": "Analysis Request", "team": team, "designation": {"$nin": ["SP"]}, "others": {"$in": [officername]}},
                        {"approval": {"$regex": "Assigned"}, "requestcategory": "Analysis Request", "team": team, "designation": {"$nin": ["SP"]}, "superior": officername},
                        {"approval": {"$regex": "Assigned"}, "requestcategory": "Analysis Request", "team": team, "designation": {"$nin": ["SP"]}, "others": officername}
                        ]}
                ]}

            My_tickets = coll.find(query, projection1).sort([('_id', -1)]).skip(skip).limit(limit)
            total_tickets1 = coll.count_documents(query)  # Total count of tickets
            filtered_tickets1 = []
            res_len = 0
            newno_len = 0
            total_len = 0
            for res in My_tickets:
                if res['pending'] == 'Mail Under Process' and (res['subtypes'] == 'Phone' or res['requesttype'] == 'CDR'):
                    res_len = len(res['result'])
                    total_len = res_len - int(res['status_count'])                    
                elif res['pending'] == 'Mail Under Process' and (res['subtypes'] == 'IMEI' or res['requesttype'] == 'IMEI CDR'):
                    newno_len = len(res['newnumber'])
                    total_len = newno_len - int(res['status_count'])
                 
                #print(res['_id'])
                filtered_tickets1.append({
                    '_id': res['_id'],
                    'priority': res['priority'],
                    'useremail':res['useremail'],
                    'date': res['date'],
                    'token': res['token'],
                     'team': res['team'],
                    "officername": res['officername'],
                    'approval': res['approval'],
                    'pending': res['pending'],
                    'requestcategory': res['requestcategory'],
                    'requesttype': res['requesttype'],
                    'subtypes': res['subtypes'],
                    'requesttypes': res['requesttypes'],
                    'edit':res['edit'],
                    'status_count': res['status_count'],
                    'total_len': total_len,
                    "res_len": res_len,
                    "newno_len": newno_len
                    })
            total_pages1 = math.ceil(total_tickets1 / limit)
            pagination_info1 = {
                'total_pages1': int(total_pages1),
                'current_page1': page,
                'total_tickets1': total_tickets1
            }
            return jsonify({'others': filtered_tickets1, 'pagination1': pagination_info1})

    elif designation == "INSPR" and role != "Analyst":
        if reqtype == 'My_tickets':
            # Assuming page and limit are passed from the frontend
            page = request.json.get('page')
            limit = request.json.get('limit')
            #print(type(page),type(limit))
            # Calculate skip based on page number and limit
            if page is None or limit is None:
                page = 1
                limit = 6
            skip = (page - 1) * limit
            # Fetch tickets with pagination
            My_tickets = coll.find({'useremail': email,'requestcategory': {'$in' :  category}}, projection).sort([('_id', -1)]).skip(skip).limit(limit)
            total_tickets = coll.count_documents({'useremail': email,'requestcategory': {'$in' :  category}})  # Total count of tickets
            # Create a list to store filtered tickets
            filtered_tickets = []
            res_len = 0
            newno_len = 0
            total_len = 0
            for res in My_tickets:
                if res['pending'] == 'Mail Under Process' and (res['subtypes'] == 'Phone' or res['requesttype'] == 'CDR'):
                    res_len = len(res['result'])
                    total_len = res_len - int(res['status_count'])
                
                elif res['pending'] == 'Mail Under Process' and (res['subtypes'] == 'IMEI' or res['requesttype'] == 'IMEI CDR'):
                    newno_len = len(res['newnumber'])
                    total_len = newno_len - int(res['status_count'])
                    
                filtered_tickets.append({
                    '_id': res['_id'],
                    'priority': res['priority'],
                    'useremail':res['useremail'],
                    'date': res['date'],
                    'token': res['token'],
                    'approval': res['approval'],
                    'pending': res['pending'],
                    'requestcategory': res['requestcategory'],
                    'requesttype': res['requesttype'],
                    'requesttypes': res['requesttypes'],
                    'subtypes': res['subtypes'],
                    'status_count': res['status_count'],
                    'total_len': total_len,
                    "res_len": res_len,
                    "newno_len": newno_len
                })
            # Calculate pagination information
            total_pages = math.ceil(total_tickets / limit) # limit
            pagination_info = {
                'total_pages': int(total_pages),
                'current_page': page,
                'total_tickets': total_tickets
            }
            #print(pagination_info)
            # Send filtered_tickets and pagination_info to frontend
            return jsonify({'mytickets': filtered_tickets, 'pagination': pagination_info})
        elif reqtype == 'Others':
            # Similarly, add pagination for 'Others' request type
            page = request.json.get('page')
            limit = request.json.get('limit')
            #print(type(page),type(limit))
            # Calculate skip based on page number and limit
            if page is None or limit is None:
                page = 1
                limit = 6
            #print(page,limit)
            skip = (page - 1) * limit
            if team == 'CAT':
                #print(team)
                query = {
                            '$and': [
                                {'useremail': {'$nin': [email]}},
                                {'requestcategory': {'$in' :  category}},
                                {"$or": [
                                    {"pending": "Ins Approval Pending", "team": team, 'modules': modules},
                                    {"pending": "ADDL-SP/DSP Approval Pending", "team": team, 'modules': modules, "designation": {"$nin": ["SP", "DSP", "ADDL-SP/DSP"]}},
                                    {"pending": "ADDL-SP/DSP Approval Pending", 'modules': modules, "token": {"$regex": "_CO"}, "designation": {"$nin": ["SP", "DSP", "ADDL-SP/DSP"]}},
                                    {"pending": {"$regex": officername}},
                                    {"pending": "Ticket_Closed", "team": team, 'modules': modules, "designation": {"$nin": ["SP", "ADDL-SP/DSP"]}},
                                    {"pending": "Mail Under Process", "token": {"$regex": "_CO"}, 'modules': modules, "designation": {"$nin": ["SP", "DSP", "ADDL-SP/DSP"]}},
                                    {"pending": "Annexe Send to CAT INS", "token": {"$regex": "_CO"}, 'modules': modules, "designation": {"$nin": ["SP", "DSP", "ADDL-SP/DSP"]}},
                                    {"requestcategory": "Analysis Request", 'team': {'$nin': ['ADMIN']}},
                                    {"pending": "Mail Under Process", 'modules': modules, "team": team, "designation": {"$nin": ["SP", "DSP", "ADDL-SP/DSP"]}},
                                    {"pending": {"$regex": "^Rejected"}, "team": team, 'modules': modules, "designation": {"$nin": ["SP", "DSP", "ADDL-SP/DSP"]}}
                                ]
                                },
                        ]}

            else:
                query = {'$and' : [
                {'useremail': {'$nin': [email]}},
                {'requestcategory': {'$in' :  category}},
                {"$or": [
                        # DATA REQ
                        {"pending": "Ins Approval Pending", "team": team,'modules':modules, "others": {"$in": [officername, '']},"designation": {"$nin": ["SP", "DSP", "ADDL-SP/DSP"]}},
                        {"pending": "ADDL-SP/DSP Approval Pending",'modules':modules, "team": team, "others": {"$in": [officername, '']},"designation": {"$nin": ["SP", "DSP", "ADDL-SP/DSP"]}},
                        {"pending": {"$regex": officername}, "team": team,'modules':modules,"designation": {"$nin": ["SP", "DSP", "ADDL-SP/DSP"]}},
                        {"pending": "Annexe Send to CAT INS", "team": team, 'modules':modules,"others": {"$in": [officername, '']},"designation": {"$nin": ["SP", "DSP", "ADDL-SP/DSP"]}},
                        {"pending": {"$regex": "^Rejected"}, "team": team, 'modules':modules,"others": {"$in": [officername, '']},"designation": {"$nin": ["SP", "DSP", "ADDL-SP/DSP"]}},
                        {"pending": "Mail Under Process", "team": team,'modules':modules, "others": {"$in": [officername, '']},"designation": {"$nin": ["SP", "DSP", "ADDL-SP/DSP"]}},
                        {"pending": "SP Approval Pending", "team": team, 'modules':modules,"others": {"$in": [officername, '']},"designation": {"$nin": ["SP", "DSP", "ADDL-SP/DSP"]}},
                        # ANALYSIS REQ
                        {"pending": "Ins Approval Pending", 'modules':modules,"requestcategory": "Analysis Request", "team": team,"designation": {"$nin": ["SP", "DSP", "ADDL-SP/DSP"]}},
                        {"approval": {"$regex": "^Assigned"}, 'modules':modules,"requestcategory": "Analysis Request", "team": team,"designation": {"$nin": ["SP", "DSP", "ADDL-SP/DSP"]}},
                        {"pending": "ADDL-SP/DSP Approval Pending", 'modules':modules,"requestcategory": "Analysis Request", "team": team,"designation": {"$nin": ["SP", "DSP", "ADDL-SP/DSP"]}},
                        {"pending": "SP Approval Pending", 'modules':modules,"requestcategory": "Analysis Request", "team": team,"designation": {"$nin": ["SP", "DSP", "ADDL-SP/DSP"]}},
                        {"pending": "Assign to CAT Ins", 'modules':modules,"requestcategory": "Analysis Request", "team": team,"designation": {"$nin": ["SP", "DSP", "ADDL-SP/DSP"]}}
                    ]
                }]}


            My_tickets = coll.find(query, projection1).sort([('_id', -1)]).skip(skip).limit(limit)
            total_tickets1 = coll.count_documents(query)  # Total count of tickets
            filtered_tickets1 = []
            res_len = 0
            newno_len = 0
            total_len = 0
            for res in My_tickets:
                if res['pending'] == 'Mail Under Process' and (res['subtypes'] == 'Phone' or res['requesttype'] == 'CDR'):
                    res_len = len(res['result'])
                    total_len = res_len - int(res['status_count'])
                
                elif res['pending'] == 'Mail Under Process' and (res['subtypes'] == 'IMEI' or res['requesttype'] == 'IMEI CDR'):
                    newno_len = len(res['newnumber'])
                    total_len = newno_len - int(res['status_count'])
                    
                #print(res['_id'])

                filtered_tickets1.append({
                    '_id': res['_id'],
                    'priority': res['priority'],
                    'useremail':res['useremail'],
                    'date': res['date'],
                    'token': res['token'],
                     'team': res['team'],
                    "officername": res['officername'],
                    'approval': res['approval'],
                    'pending': res['pending'],
                    'requestcategory': res['requestcategory'],
                    'requesttype': res['requesttype'],
                    'subtypes': res['subtypes'],
                    'requesttypes': res['requesttypes'],
                    'edit':res['edit'],
                    'status_count': res['status_count'],
                    'total_len': total_len,
                    "res_len": res_len,
                    "newno_len": newno_len
                    })
            total_pages1 = math.ceil(total_tickets1 / limit)
            pagination_info1 = {
                'total_pages1': int(total_pages1),
                'current_page1': page,
                'total_tickets1': total_tickets1
            }
            return jsonify({'others': filtered_tickets1, 'pagination1': pagination_info1})
    elif role == "Analyst":
        
        if reqtype == 'My_tickets':
            # Assuming page and limit are passed from the frontend
            page = request.json.get('page')
            limit = request.json.get('limit')
            #print(type(page),type(limit))
            # Calculate skip based on page number and limit
            if page is None or limit is None:
                page = 1
                limit = 6
            skip = (page - 1) * limit
            # Fetch tickets with pagination
            project = {
                '_id' : 1,
                'priority' : 1,
                'date' : 1,
                'useremail' : 1,
                'token' : 1,
                'approval' : 1,
                'officername' : 1,
                'pending' : 1,
                'requestcategory' : 1,
                'requesttype' : 1,
                'subtypes' : 1,
                'requesttypes': 1,
                'edit': 1,
                'status_count': 1,
                'result': 1,
                'newnumber' : 1
            }
            My_tickets = coll.find({'useremail': email,'requestcategory': {'$in' :  category}}, project).sort([('_id', -1)]).skip(skip).limit(limit)
            total_tickets = coll.count_documents({'useremail': email,'requestcategory': {'$in' :  category}})  # Total count of tickets

            # Create a list to store filtered tickets
            filtered_tickets = []
            res_len = 0
            newno_len = 0
            total_len = 0
            for res in My_tickets:
                if res['pending'] == 'Mail Under Process' and (res['subtypes'] == 'Phone' or res['requesttype'] == 'CDR'):
                    res_len = len(res['result'])
                    total_len = res_len - int(res['status_count'])
                
                elif res['pending'] == 'Mail Under Process' and (res['subtypes'] == 'IMEI' or res['requesttype'] == 'IMEI CDR'):
                    newno_len = len(res['newnumber'])
                    total_len = newno_len - int(res['status_count'])
                
                filtered_tickets.append({
                    '_id': res['_id'],
                    'priority': res['priority'],
                    'useremail':res['useremail'],
                    'date': res['date'],
                    'token': res['token'],
                    'approval': res['approval'],
                    'pending': res['pending'],
                    'requestcategory': res['requestcategory'],
                    'requesttype': res['requesttype'],
                    'requesttypes': res['requesttypes'],
                    'subtypes': res['subtypes'],
                    'status_count': res['status_count'],
                    'total_len': total_len,
                    "res_len": res_len,
                    "newno_len": newno_len
                })
            # Calculate pagination information
            total_pages = math.ceil(total_tickets / limit) # limit
            pagination_info = {
                'total_pages': int(total_pages),
                'current_page': page,
                'total_tickets': total_tickets
            }
            #print(pagination_info)
            # Send filtered_tickets and pagination_info to frontend
            return jsonify({'mytickets': filtered_tickets, 'pagination': pagination_info})
        elif reqtype == 'Others':
            # Similarly, add pagination for 'Others' request type
            page = request.json.get('page')
            limit = request.json.get('limit')
            #print(type(page),type(limit))
            # Calculate skip based on page number and limit
            if page is None or limit is None:
                page = 1
                limit = 6
            #print(page,limit)
            skip = (page - 1) * limit
            query = {'$and' : [
                {'useremail': {'$nin': [email]}}, 
                {"$or": [
                    {'assign_Officer': {'$in':[officername]}}
                ]
            }]}
            project1 = {
                '_id' : 1,
                'priority' : 1,
                'date' : 1,
                'team' : 1,
                'useremail' : 1,
                'token' : 1,
                'approval' : 1,
                'officername' : 1,
                'pending' : 1,
                'requestcategory' : 1,
                'requesttype' : 1,
                'subtypes' : 1,
                'requesttypes': 1,
                'edit': 1,
                'assign_Officer': 1,
            }
            My_tickets = coll.find(query, project1).sort([('_id', -1)]).skip(skip).limit(limit)
            total_tickets1 = coll.count_documents(query)  # Total count of tickets
            filtered_tickets1 = []
            for res in My_tickets:
                #print(res['_id'])

                filtered_tickets1.append({
                    '_id': res['_id'],
                    'priority': res['priority'],
                    'useremail':res['useremail'],
                    'date': res['date'],
                    'token': res['token'],
                     'team': res['team'],
                    "officername": res['officername'],
                    'approval': res['approval'],
                    'pending': res['pending'],
                    'requestcategory': res['requestcategory'],
                    'requesttype': res['requesttype'],
                    'subtypes': res['subtypes'],
                    'requesttypes': res['requesttypes'],
                    'edit':res['edit'],
                    'assign_Officer' : res['assign_Officer']
                    })
            total_pages1 = math.ceil(total_tickets1 / limit)
            pagination_info1 = {
                'total_pages1': int(total_pages1),
                'current_page1': page,
                'total_tickets1': total_tickets1
            }
            return jsonify({'others': filtered_tickets1, 'pagination1': pagination_info1})
    elif role == "Mailer":
        if reqtype == 'My_tickets':
            # Assuming page and limit are passed from the frontend
            page = request.json.get('page')
            limit = request.json.get('limit')
            #print(type(page),type(limit))
            # Calculate skip based on page number and limit
            if page is None or limit is None:
                page = 1
                limit = 6
            skip = (page - 1) * limit
            # Fetch tickets with pagination
            My_tickets = coll.find({'useremail': email,'requestcategory': {'$in' :  category}}, projection).sort([('_id', -1)]).skip(skip).limit(limit)
            total_tickets = coll.count_documents({'useremail': email,'requestcategory': {'$in' :  category}})  # Total count of tickets
            # Create a list to store filtered tickets
            filtered_tickets = []
            res_len = 0
            newno_len = 0
            total_len = 0
            for res in My_tickets:
                if res['pending'] == 'Mail Under Process' and (res['subtypes'] == 'Phone' or res['requesttype'] == 'CDR'):
                    res_len = len(res['result'])
                    total_len = res_len - int(res['status_count'])
                    
                elif res['pending'] == 'Ticket_Closed' and (res['subtypes'] == 'Phone' or res['requesttype'] == 'CDR'):
                    res_len = len(res['result'])
                    total_len = res_len
                    
                elif res['pending'] == 'Mail Under Process' and (res['subtypes'] == 'IMEI' or res['requesttype'] == 'IMEI CDR'):
                    newno_len = len(res['newnumber'])
                    total_len = newno_len - int(res['status_count'])
                    
                elif res['pending'] == 'Ticket_Closed' and (res['subtypes'] == 'IMEI' or res['requesttype'] == 'IMEI CDR'):
                    newno_len = len(res['newnumber'])
                    total_len = newno_len
                    
                filtered_tickets.append({
                    '_id': res['_id'],
                    'priority': res['priority'],
                    'useremail':res['useremail'],
                    'date': res['date'],
                    'token': res['token'],
                    'approval': res['approval'],
                    'pending': res['pending'],
                    'requestcategory': res['requestcategory'],
                    'requesttype': res['requesttype'],
                    'subtypes': res['subtypes'],
                    'status_count': res['status_count'],
                    'total_len': total_len,
                    "res_len": res_len,
                    "newno_len": newno_len
                })
            # Calculate pagination information
            total_pages = math.ceil(total_tickets / limit) # limit
            pagination_info = {
                'total_pages': int(total_pages),
                'current_page': page,
                'total_tickets': total_tickets
            }
            #print(pagination_info)
            # Send filtered_tickets and pagination_info to frontend
            return jsonify({'mytickets': filtered_tickets, 'pagination': pagination_info})
        elif reqtype == 'Others':
            # Similarly, add pagination for 'Others' request type
            page = request.json.get('page')
            limit = request.json.get('limit')
            #print(type(page),type(limit))
            # Calculate skip based on page number and limit
            if page is None or limit is None:
                page = 1
                limit = 6
            #print(page,limit)
            skip = (page - 1) * limit
            query = {'$and' : [
                {'useremail': {'$nin': [email]}},
                {"$or": [
                    {"requestcategory": "Data Request", 'pending' : {'$in': ['Mail Under Process',
                    'Ticket_Closed']}},
                ]
            }]}

            My_tickets = coll.find(query, projection1).sort([('_id', -1)]).skip(skip).limit(limit)
            total_tickets1 = coll.count_documents(query)  # Total count of tickets
            filtered_tickets1 = []
            res_len = 0
            newno_len = 0
            total_len = 0
            for res in My_tickets:
                if res['pending'] == 'Mail Under Process' and (res['subtypes'] == 'Phone' or res['requesttype'] == 'CDR'):
                    res_len = len(res['result'])
                    total_len = res_len - int(res['status_count'])
                    
                elif res['pending'] == 'Ticket_Closed' and (res['subtypes'] == 'Phone' or res['requesttype'] == 'CDR'):
                    res_len = len(res['result'])
                    total_len = res_len
                    
                elif res['pending'] == 'Mail Under Process' and (res['subtypes'] == 'IMEI' or res['requesttype'] == 'IMEI CDR'):
                    newno_len = len(res['newnumber'])
                    total_len = newno_len - int(res['status_count'])
                    
                elif res['pending'] == 'Ticket_Closed' and (res['subtypes'] == 'IMEI' or res['requesttype'] == 'IMEI CDR'):
                    newno_len = len(res['newnumber'])
                    total_len = newno_len
                    
                #print(res['_id'])
                filtered_tickets1.append({
                    '_id': res['_id'],
                    'priority': res['priority'],
                    'useremail':res['useremail'],
                    'date': res['date'],
                    'token': res['token'],
                    'team': res['team'],
                    "officername": res['officername'],
                    'approval': res['approval'],
                    'pending': res['pending'],
                    'requestcategory': res['requestcategory'],
                    'requesttype': res['requesttype'],
                    'subtypes': res['subtypes'],
                    'requesttypes': res['requesttypes'],
                    'edit':res['edit'],
                    'status_count': res['status_count'],
                    'total_len': total_len,
                    "res_len": res_len,
                    "newno_len": newno_len
                    })
            total_pages1 = math.ceil(total_tickets1 / limit)
            pagination_info1 = {
                'total_pages1': int(total_pages1),
                'current_page1': page,
                'total_tickets1': total_tickets1
            }
            return jsonify({'others': filtered_tickets1, 'pagination1': pagination_info1})
    else:
        if reqtype == 'My_tickets':
            # Assuming page and limit are passed from the frontend
            page = request.json.get('page')
            limit = request.json.get('limit')
            #print(type(page),type(limit))
            # Calculate skip based on page number and limit
            if page is None or limit is None:
                page = 1
                limit = 6
            skip = (page - 1) * limit
            # Fetch tickets with pagination
            My_tickets = coll.find({'useremail': email,'requestcategory': {'$in' :  category}}, projection).sort([('_id', -1)]).skip(skip).limit(limit)
            total_tickets = coll.count_documents({'useremail': email,'requestcategory': {'$in' :  category}})  # Total count of tickets
            # Create a list to store filtered tickets
            filtered_tickets = []
            res_len = 0
            newno_len = 0
            total_len = 0
            for res in My_tickets:
                if res['pending'] == 'Mail Under Process' and (res['subtypes'] == 'Phone' or res['requesttype'] == 'CDR'):
                    res_len = len(res['result'])
                    total_len = res_len - int(res['status_count'])
                    
                elif res['pending'] == 'Ticket_Closed' and (res['subtypes'] == 'Phone' or res['requesttype'] == 'CDR'):
                    res_len = len(res['result'])
                    total_len = res_len
                    
                elif res['pending'] == 'Mail Under Process' and (res['subtypes'] == 'IMEI' or res['requesttype'] == 'IMEI CDR'):
                    newno_len = len(res['newnumber'])
                    total_len = newno_len - int(res['status_count'])
                    
                elif res['pending'] == 'Ticket_Closed' and (res['subtypes'] == 'IMEI' or res['requesttype'] == 'IMEI CDR'):
                    newno_len = len(res['newnumber'])
                    total_len = newno_len
                    
                filtered_tickets.append({
                    '_id': res['_id'],
                    'priority': res['priority'],
                    'useremail':res['useremail'],
                    'date': res['date'],
                    'token': res['token'],
                    'approval': res['approval'],
                    'pending': res['pending'],
                    'requestcategory': res['requestcategory'],
                    'requesttype': res['requesttype'],
                    'subtypes': res['subtypes'],
                     'status_count': res['status_count'],
                    'total_len': total_len,
                    "res_len": res_len,
                    "newno_len": newno_len
                })
            # Calculate pagination information
            total_pages = math.ceil(total_tickets / limit) # limit
            pagination_info = {
                'total_pages': int(total_pages),
                'current_page': page,
                'total_tickets': total_tickets
            }
            #print(pagination_info)
            # Send filtered_tickets and pagination_info to frontend
            return jsonify({'mytickets': filtered_tickets, 'pagination': pagination_info})
        elif reqtype == 'Others':
            # Similarly, add pagination for 'Others' request type
            page = request.json.get('page')
            limit = request.json.get('limit')
            #print(type(page),type(limit))
            # Calculate skip based on page number and limit
            if page is None or limit is None:
                page = 1
                limit = 6
            #print(page,limit)
            skip = (page - 1) * limit
            My_tickets = 0
            total_tickets1 = 0  # Total count of tickets
            filtered_tickets1 = []
            total_pages1 = math.ceil(total_tickets1 / limit)
            pagination_info1 = {
                'total_pages1': 0,
                'current_page1': 0,
                'total_tickets1': 0
            }
            return jsonify({'others': filtered_tickets1, 'pagination1': pagination_info1})


# -------------------------------------------------------------------------------------------------------------------------------------------
# -------------------------------------------------------------------------------------------------------------------------------------------
@tickect_bp.route('/updateuser', methods=['POST'])
def update_user():
    data = request.json
    id = data.get('id')
    # #print(id)
    coll = db['tickets']

    # Retrieve the user with the provided email from the database
    user = Users.objects_safe(id = id).first()
    # #print(user)
    if user:
        # Update the user fields with the edited data
        user.name = data.get('name')
        user.email = data.get('email')
        user.mobilenumber = data.get('mobilenumber')
        user.password1 = data.get('password1')
        user.password2 = data.get('password2')
        user.designation = data.get('designation')
        user.officername = data.get('officername')
        user.team = data.get('team')
        user.modules = data.get('modules')
        user.status = data.get('status')
        user.superior = data.get('superior')
        user.role = data.get('role')
        user.Application = data.get('Application')
        # Save the changes to the database
        user.save()

        coll.update_many({"useremail": user.email}, {'$set': {"officername": user.officername}})

        return jsonify({'message': 'User details updated successfully'})
    else:
        return jsonify({'message': 'User not found'})
    
@tickect_bp.route('/deleteuser', methods=['POST'])
def deleteuser():
    id = request.json.get('id')
    email = request.json.get('email')
    id1 = ObjectId(id)
    user = Users.objects_safe(id=id1, email=email).first()
    print(user,"..................")
    if user:
        user.delete()
        return jsonify({'message': 'User deleted successfully'})
    else:
        return jsonify({'message': 'User not found'})
# Endpoint for New Ticket Generate------------------------------------------------------------------------------>



# @tickect_bp.route('/newform', methods=['POST'])
# def newform():
#     now = datetime.now()
#     day_with_millisec = now.strftime("%d%m%Y%H%M%S")+f"{now.microsecond // 1000:03d}"
#     coll2 = db['tickets']
#     last_token = 1001
#     db_len = coll2.find_one({"status": {"$ne": "Converted"}}, sort=[('_id', -1)])
#     if db_len:
#         last_token = db_len.get('token','').split("_")
#         if last_token[0] != day_with_millisec:
#             last_token=int(last_token[1])+1
#             tokens = f"{day_with_millisec}_" + str(last_token)
#     else:
#         tokens=f"{day_with_millisec}_"+ str(last_token)
#     res = {}
#     newnumber = request.json.get('newnumber')
#     requesttype = request.json.get('requesttype')
#     subtypes = request.json.get('subtype')
#     print("!!!!!!!!!!!!!!!!!!!!!")
#     if subtypes == "Phone" or requesttype == "CDR" or requesttype == "CAF" or requesttype == "SDR" or requesttype == "CAF/CDR" or requesttype == "ISD" or requesttype == "TOWER CDR":
#         nickname =  request.json.get('nickname')
#         pending = "Approval Pending"
#         print("@@@@@@@@@@@@@@@@@@@@@")
#     else:
#         nickname = ''
#         pending = "Approval Pending"
#         print("#####################")
#     requesttype = request.json.get('requesttype')    
#     demo = []
#     if (requesttype == "GPRS" and subtypes == "IMEI") or (requesttype == "IPDR" and subtypes == "IMEI") or (request == 'IMEI CDR'):
#         for phoneNumber in newnumber:
#             if "IMEI" in phoneNumber and (len(phoneNumber["IMEI"]) == 15 or len(phoneNumber["IMEI"]) == 14):
#                 phoneNumber["IMEI"] = phoneNumber["IMEI"][:14] + '0'
#             # #print(phoneNumber)
#             demo.append(phoneNumber)
#         newnumber=demo
    
#     result = []

#     try:
#         for res in newnumber:
#             if res['result']['SDR'] != '':
#                 activation_date = res['result']['ActivationDate']
#                 res['result']['FetchedNickName'] = res['Nickname']
#                 from_date = res['From_Date']
#                 to_date = res['To_Date']
#                 mnp = res['result']['MNP']
#                 print(".................................")
#                 if activation_date == '' or activation_date == 'None' or activation_date != '':
#                     res['result']['From_Date'] = from_date
#                     res['result']['To_Date'] = to_date
#                 else:
#                     if 'JO' in mnp or 'Jio' in mnp or 'jio' in mnp or 'JIO' in mnp:
#                         from_date1=res['From_Date']
#                         from_date=datetime.strptime(from_date1,"%d/%m/%Y")
#                         if len(activation_date)>10:
#                             jo=datetime.strptime(activation_date,"%Y-%m-%d %H:%M:%S")
#                             jo_date1=jo.strftime("%d/%m/%Y")
#                             jo_date=datetime.strptime(jo_date1,'%d/%m/%Y')
                            
#                             if from_date >= jo_date:
#                                 res['result']['From_Date'] = from_date1
#                                 res['result']['To_Date'] = to_date
#                             else:
#                                 res['result']['From_Date'] = jo_date1
#                                 res['result']['To_Date'] = to_date                        
#                         else:
#                             jo=datetime.strptime(activation_date,"%Y-%m-%d")
#                             jo_date1=jo.strftime("%d/%m/%Y")
#                             jo_date=datetime.strptime(jo_date1,'%d/%m/%Y')
#                             if from_date >= jo_date:
#                                 res['result']['From_Date'] = from_date1
#                                 res['result']['To_Date'] = to_date
#                             else:
#                                 res['result']['From_Date'] = jo_date1
#                                 res['result']['To_Date'] = to_date

#                     elif 'CO' in mnp:
#                         from_date1=res['From_Date']
#                         from_date=datetime.strptime(from_date1,"%d/%m/%Y")
#                         bsnl = datetime.strptime(activation_date, "%d-%b-%y")
#                         bsnl_date1=bsnl.strftime("%d/%m/%Y")
#                         bsnl_date=datetime.strptime(bsnl_date1,'%d/%m/%Y')
#                         if from_date >= bsnl_date:
#                             res['result']['From_Date'] = from_date1
#                             res['result']['To_Date'] = to_date
#                         else:
#                             res['result']['From_Date'] = bsnl_date1
#                             res['result']['To_Date'] = to_date

#                     elif 'VI' in mnp:
#                         from_date1=res['From_Date']
#                         from_date=datetime.strptime(from_date1,"%d/%m/%Y")
#                         vi=datetime.strptime(activation_date, "%d-%b-%Y")
#                         vi_date1=vi.strftime("%d/%m/%Y")
#                         vi_date = datetime.strptime(vi_date1,'%d/%m/%Y')
#                         if from_date >= vi_date :
#                             res['result']['From_Date'] = from_date1
#                             res['result']['To_Date'] = to_date
#                         else:
#                             res['result']['From_Date'] = vi_date1
#                             res['result']['To_Date'] = to_date

#                     elif 'AT' in mnp:
#                         from_date1=res['From_Date']
#                         from_date=datetime.strptime(from_date1,"%d/%m/%Y")
#                         at=datetime.strptime(activation_date, "%a, %d %b %Y %H:%M:%S %Z")
#                         at_date1=at.strftime("%d/%m/%Y")
#                         at_date = datetime.strptime(at_date1,'%d/%m/%Y')
#                         if from_date >= at_date:
#                             res['result']['From_Date'] = from_date1
#                             res['result']['To_Date'] = to_date
#                         else:
#                             res['result']['From_Date'] = at_date1
#                             res['result']['To_Date'] = to_date

                
#                 pending = "Approval Pending"
#                 result.append(res['result'])

#     except Exception as e:
#         print(e)
    
#     Tickets(
#             useremail =request.json.get('useremail'),
#             superior =request.json.get('superior'),
#             officername =request.json.get('officername'),
#             role =request.json.get('role'),
#             designation = request.json.get("designation"),
#             type = request.json.get('update'),
#             requestcategory = "Data Request",
#             requesttype = request.json.get('requesttype'),
#             subtypes = subtypes,
#             team = request.json.get('team'),
#             modules = request.json.get('modules'),
#             relation =  request.json.get('relation'),
#             nickname =  nickname,
#             suspect =  request.json.get('suspect'),
#             name =  request.json.get('name'),
#             location =  request.json.get('location'),
#             source = request.json.get('source'),
#             reason = request.json.get('reason'),
#             refno = request.json.get('refno'),
#             refdate=str(request.json.get('refdate')),
#             priority = request.json.get('priority'),
#             token = tokens,
#             pending = pending,
#             approval = "-",
#             others='',
#             newnumber = newnumber,
#             remarks = request.json.get('remarks'),
#             date = request.json.get("date"),
#             result = result,
#             raise_time = request.json.get("raise_time")
        
#     ).save()
#     coll = db['tickets']
#     user1= coll.find_one({'token':tokens})
#     try:
#         if any(("CO" in res2["result"]["MNP"]) for res2 in user1["newnumber"]) or ('MNP' in res2 and"CO" in res2["MNP"] for res2 in user1["newnumber"]):
#             split_documents(tokens)
#     except:
#         pass
#     return jsonify({'message': 'Data Insert successful'})
@tickect_bp.route('/newform', methods=['POST'])
def newform():
    now = datetime.now()
    day_with_millisec = now.strftime("%d%m%Y%H%M%S")+f"{now.microsecond // 1000:03d}"
    coll2 = db['tickets']
    last_token = 1001
    db_len = coll2.find_one({"status": {"$ne": "Converted"}}, sort=[('_id', -1)])
    if db_len:
        last_token = db_len.get('token','').split("_")
        if last_token[0] != day_with_millisec:
            last_token=int(last_token[1])+1
            tokens = f"{day_with_millisec}_" + str(last_token)
    else:
        tokens=f"{day_with_millisec}_"+ str(last_token)
    res = {}
    newnumber = request.json.get('newnumber')
    requesttype = request.json.get('requesttype')
    subtypes = request.json.get('subtype')
    if subtypes == "Phone" or requesttype == "CDR" or requesttype == "CAF" or requesttype == "SDR" or requesttype == "CAF/CDR" or requesttype == "ISD":
        nickname =  request.json.get('nickname')
        pending = "Approval Pending"
    else:
        nickname = ''
        pending = "Approval Pending"
    requesttype = request.json.get('requesttype')    
    demo = []
    if (requesttype == "GPRS" and subtypes == "IMEI") or (requesttype == "IPDR" and subtypes == "IMEI") or (request == 'IMEI CDR'):
        for phoneNumber in newnumber:
            if "IMEI" in phoneNumber and (len(phoneNumber["IMEI"]) == 15 or len(phoneNumber["IMEI"]) == 14):
                phoneNumber["IMEI"] = phoneNumber["IMEI"][:14] + '0'
            # #print(phoneNumber)
            demo.append(phoneNumber)
        newnumber=demo
    
    result = []

    try:
        for res in newnumber:
            if res['result']['SDR'] != '':
                activation_date = res['result']['ActivationDate']
                res['result']['FetchedNickName'] = res['Nickname']
                from_date = res['From_Date']
                to_date = res['To_Date']
                mnp = res['result']['MNP']
                if activation_date == '' or activation_date == 'None' or activation_date != '':
                    res['result']['From_Date'] = from_date
                    res['result']['To_Date'] = to_date
                else:
                    if 'JO' in mnp or 'Jio' in mnp or 'jio' in mnp or 'JIO' in mnp:
                        from_date1=res['From_Date']
                        from_date=datetime.strptime(from_date1,"%d/%m/%Y")
                        if len(activation_date)>10:
                            jo=datetime.strptime(activation_date,"%Y-%m-%d %H:%M:%S")
                            jo_date1=jo.strftime("%d/%m/%Y")
                            jo_date=datetime.strptime(jo_date1,'%d/%m/%Y')
                            
                            if from_date >= jo_date:
                                res['result']['From_Date'] = from_date1
                                res['result']['To_Date'] = to_date
                            else:
                                res['result']['From_Date'] = jo_date1
                                res['result']['To_Date'] = to_date                        
                        else:
                            jo=datetime.strptime(activation_date,"%Y-%m-%d")
                            jo_date1=jo.strftime("%d/%m/%Y")
                            jo_date=datetime.strptime(jo_date1,'%d/%m/%Y')
                            if from_date >= jo_date:
                                res['result']['From_Date'] = from_date1
                                res['result']['To_Date'] = to_date
                            else:
                                res['result']['From_Date'] = jo_date1
                                res['result']['To_Date'] = to_date

                    elif 'CO' in mnp:
                        from_date1=res['From_Date']
                        from_date=datetime.strptime(from_date1,"%d/%m/%Y")
                        bsnl = datetime.strptime(activation_date, "%d-%b-%y")
                        bsnl_date1=bsnl.strftime("%d/%m/%Y")
                        bsnl_date=datetime.strptime(bsnl_date1,'%d/%m/%Y')
                        if from_date >= bsnl_date:
                            res['result']['From_Date'] = from_date1
                            res['result']['To_Date'] = to_date
                        else:
                            res['result']['From_Date'] = bsnl_date1
                            res['result']['To_Date'] = to_date

                    elif 'VI' in mnp:
                        from_date1=res['From_Date']
                        from_date=datetime.strptime(from_date1,"%d/%m/%Y")
                        vi=datetime.strptime(activation_date, "%d-%b-%Y")
                        vi_date1=vi.strftime("%d/%m/%Y")
                        vi_date = datetime.strptime(vi_date1,'%d/%m/%Y')
                        if from_date >= vi_date :
                            res['result']['From_Date'] = from_date1
                            res['result']['To_Date'] = to_date
                        else:
                            res['result']['From_Date'] = vi_date1
                            res['result']['To_Date'] = to_date

                    elif 'AT' in mnp:
                        from_date1=res['From_Date']
                        from_date=datetime.strptime(from_date1,"%d/%m/%Y")
                        at=datetime.strptime(activation_date, "%a, %d %b %Y %H:%M:%S %Z")
                        at_date1=at.strftime("%d/%m/%Y")
                        at_date = datetime.strptime(at_date1,'%d/%m/%Y')
                        if from_date >= at_date:
                            res['result']['From_Date'] = from_date1
                            res['result']['To_Date'] = to_date
                        else:
                            res['result']['From_Date'] = at_date1
                            res['result']['To_Date'] = to_date

                
                pending = "Approval Pending"
                result.append(res['result'])

    except Exception as e:
        print(e)
    
    Tickets(
            useremail =request.json.get('useremail'),
            superior =request.json.get('superior'),
            officername =request.json.get('officername'),
            role =request.json.get('role'),
            designation = request.json.get("designation"),
            type = request.json.get('update'),
            requestcategory = "Data Request",
            requesttype = request.json.get('requesttype'),
            subtypes = subtypes,
            team = request.json.get('team'),
            modules = request.json.get('modules'),
            relation =  request.json.get('relation'),
            nickname =  nickname,
            suspect =  request.json.get('suspect'),
            name =  request.json.get('name'),
            location =  request.json.get('location'),
            source = request.json.get('source'),
            reason = request.json.get('reason'),
            refno = request.json.get('refno'),
            refdate=str(request.json.get('refdate')),
            priority = request.json.get('priority'),
            token = tokens,
            pending = pending,
            approval = "-",
            others='',
            newnumber = newnumber,
            remarks = request.json.get('remarks'),
            date = request.json.get("date"),
            result = result,
            raise_time = request.json.get("raise_time")
        
    ).save()
    coll = db['tickets']
    user1= coll.find_one({'token':tokens})
    print(user1['subtypes'],'///////////')
    if any(('result' in res2 and "CO" in res2["result"]["MNP"]) for res2 in user1["newnumber"]):
        split_documents(tokens,'CDR')

    elif any('MNP' in res2 and "CO" in res2["MNP"] for res2 in user1["newnumber"]):
        split_documents(tokens,'Others')

    elif user1['subtypes'] == 'CGI':
        split_documents(tokens,'CGI')

    return jsonify({'message': 'Data Insert successful'})

@tickect_bp.route('/statussuccess', methods=["POST", "GET"])
def statussuccess():
    team = request.json.get("team")
    select = request.json.get("select")
    role = request.json.get("role")
    default = (Q(pending="Mail Under Process"))
    status_value = (Q(pending="Mail Under Process",team=team))
    successdata = []
    if len(select) > 0:
        for res in select:
            # #print(res)
            status_value1 = (Q(pending="Mail Under Process",id = ObjectId(res)))
            results = [ticket.to_mongo().to_dict() for ticket in Tickets.objects_safe(status_value1)]
            for i in results:
                i['_id'] = str(i['_id'])
                filtered_dict = {key: value for key, value in i.items() if key != 'result'}
                newnumber_dict = {'newnumber': [res for res in i['newnumber'] if 'mailer_hold' not in res]}
                filtered_dict.update(newnumber_dict)
                successdata.append(filtered_dict)
    else:
        results = [ticket.to_mongo().to_dict() for ticket in Tickets.objects_safe(default)]
        for i in results:
                i['_id'] = str(i['_id'])
                filtered_dict = {key: value for key, value in i.items() if key != 'result'}
                newnumber_dict = {'newnumber': [res for res in i['newnumber'] if 'mailer_hold' not in res]}
                filtered_dict.update(newnumber_dict)
                successdata.append(filtered_dict)

    return jsonify(successdata)
@tickect_bp.route('/statuspending', methods=["POST", "GET"])
def statuspending():
    team = request.json.get("team")
    email = request.json.get("email")
    status_value = (Q(useremail = email) & Q(team = team) & Q(pending = "Fetch_Pending_Data"))
    # status_value = "Fetch_Pending_Data"
    results = [tickets.to_mongo().to_dict() for tickets in Tickets.objects_safe(status_value)]
    pendingdata = []
    for i in results:
        i['_id'] = str(i['_id'])
        pendingdata.append(i)
    return jsonify(pendingdata)

@tickect_bp.route('/mytickets', methods=["POST", "GET"])
def mytickets():
    email = request.json.get("email")
    status_value = (Q(useremail =email))
    results = [tickets.to_mongo().to_dict() for tickets in Tickets.objects_safe(status_value)]
    Rejecteddata = []
    for i in results:
        if 'result' in i:
            del i['result']
            
    for i in results:
        i['_id'] = str(i['_id'])
        Rejecteddata.append(i)
    return jsonify(Rejecteddata)

# End of Export Endpoint---

@tickect_bp.route('/comments/<id>', methods=['POST'])
def comments(id):
    data1 = request.json['comments_input']
    commentor = request.json['commentor']
    
    # Assuming you have an established MongoDB connection
 
    collection = db['tickets']
    comment = data1 + ' comment by ' + commentor
    # Construct the query
    filter_query = {'_id': ObjectId(id)}
    update_query = {'$set': {'Comments': comment}}

    # Update the document in the collection
    collection.update_one(filter_query, update_query)

    return {'message': 'Comments added successfully'}

@tickect_bp.route('/submitedRecord', methods=['POST'])
def submitedRecord():
 
    ticket = db['tickets']
    user = db["users"]
    data = request.json
    id = data.get("id")
    email = data.get("email")
    officername = data.get("officername")
    pending = data.get("pending")
    newinspr = data.get("newinspr")
    approval = data.get("approval")
    designtion = data.get("designation")
    role = data.get("role")
    options = data.get("options")

    sender = user.find_one({"email":email})
    
    date_time = data.get('sendDateTime')
    #print(date_time)
    
    if not sender:
        return jsonify({"message":"Sender user not found"}), 400
    if designtion == "SP":
        Superior = user.find_one({"superior":sender["superior"]})
        if options == 'Self Approve':
            ticket.update_one({'_id': ObjectId(id)}, {'$set': {"pending" : pending,"approval" : approval, 'SP_Approval' : date_time}})
        else:
            ticket.update_one({'_id': ObjectId(id)}, {'$set': {"pending" : pending,"approval" : approval, 'SP_Reject' : date_time}})

            
    elif role == 'Analyst':
        if options == 'Self Approve':
            ticket.update_one({'_id': ObjectId(id)}, {'$set': {"pending" : pending, "approval" : approval,"others":newinspr, 'Analyst_Approval' : date_time}})
        else:
            ticket.update_one({'_id': ObjectId(id)}, {'$set': {"pending" : pending, "approval" : approval,"others":newinspr, 'Send_to_INS' : date_time}})

    elif designtion == "ADDL-SP/DSP":
        Superior = user.find_one({"officername":sender["superior"]})
        if options == 'Self Approve':
            ticket.update_one({'_id': ObjectId(id), "superior":Superior["officername"]}, {'$set': {"pending" : pending,"approval" : approval, 'DSP_Approval' : date_time}})
        elif options == 'Reject':
            ticket.update_one({'_id': ObjectId(id), "superior":Superior["officername"]}, {'$set': {"pending" : pending,"approval" : approval, 'DSP_Reject' : date_time}})
        else:
            ticket.update_one({'_id': ObjectId(id),"superior":Superior["officername"]}, {'$set': {"pending" : pending, "approval" : approval,'Send_to_SP' : date_time}})
            
    elif designtion == "INSPR":
        Superior = user.find_one({"officername":sender["superior"]})
        if options == 'Self Approve':
            ticket.update_one({'_id': ObjectId(id), "superior":Superior["officername"]}, {'$set': {"pending" : pending,"approval" : approval, 'INSPR_Approval' : date_time}})
        elif options == 'Reject':
            ticket.update_one({'_id': ObjectId(id), "superior":Superior["officername"]}, {'$set': {"pending" : pending,"approval" : approval, 'INS_Reject' : date_time}})
        else:
            ticket.update_one({'_id': ObjectId(id),"superior":Superior["officername"]}, {'$set': {"pending" : pending, "approval" : approval,'Send_to_DSP' : date_time}})

    else:
        Superior = user.find_one({"officername":sender["superior"]})
        if not Superior:
            return jsonify({"user_creation_error":"Superior not found/User Not Created!!!!!!"}), 400
        else:
            ticket.update_one({'_id': ObjectId(id),"superior":Superior["officername"]}, {'$set': {"pending" : pending, "approval" : approval,'Analyst_Approval' : date_time}})
    return jsonify({"message": "Ticket sent to superior successfully"}), 200
    

# End of Team Incharge Comments and record submition Endpoint------------------------------------------------------->

# Endpoint for Logout ---------------------------------------------------------------------------------------------->

@tickect_bp.route("/logout", methods=["POST", "GET"])
def logout():
    # if request.method == 'POST':
    #     data = request.get_json()
    #     email = data.get('email')
    #     print(f"Logging out user with email: {email}")
    #     collection = db['users']
    #     collection.update_one({'email': email}, {'$set': {'logged': '0'}})
    #     return jsonify({'message': 'Logged out successfully'}), 200

    #     # # Perform logout operations here
    #     # if "email" in session:
    #     #     session.pop("email", None)
    #     #     # Update the "login" field to "0" for the logged-out user
    #     #     return jsonify({'message': 'Logged out successfully'}), 200            
    #     # else:
    #     #     return jsonify({'message': 'failed to logout'}), 400
    # else:
    #     return jsonify({'message': 'Invalid request method'}), 405
   
    if "email" in session:
        session.pop("email", None)
        return jsonify({'message': 'failed to logout'})
    else:
        return jsonify({'message': 'Logged out successfully'})
    
# End of Logout Endpoint------------------------------------------------------------------------------------------->

# Endpoint for Update form----------------------------------------------------------------------------------------->

@tickect_bp.route('/updatefields', methods=['POST', 'GET'])
def updatefields():
 
    coll = db['tickets']
    id = request.json.get('id')
    msisdn = request.json.get('msisdn')
    value = request.json.get('value')
    field = request.json.get('field')
    edit = request.json.get("edit")
    #print("demooo",id,'Phone--------->',msisdn,'<-------PHONE',value,field,edit)
    filter_criteria = {
        '_id': ObjectId(id),
        'result.MSISDN': msisdn
    }
    document = coll.find_one({'_id': ObjectId(id),'result.MSISDN': msisdn})
    #print(document)
    if document:
        for item in document["result"]:
            if item['MSISDN'] == msisdn:
                item[field] = value
        coll.update_one({'_id': ObjectId(id),"result.MSISDN": msisdn}, {'$set': {f"result.$.{field}": value,"edit":edit}})
    return {'message': 'Field updated successfully'}

# End of Update form Endpoint------------------------------------------------------------------------------------------>

# Endpoint for Automation file Upload---------------------------------------------------------------------------------->
# Endpoint for Automation file Upload---------------------------------------------------------------------------------->
@tickect_bp.route('/uploadfiles', methods=['POST'])
def uploadfiles():
        coll = db['tickets']
        coll2 = db["users"]
        
        files = request.json.get("files")
        files = json.loads(files)
        my_key = ''
        if files:
            for key, value in files.items():
                #print(key, "sgdvg")
                first_key = key
                my_key = key
                user1= coll.find_one({'_id': ObjectId(first_key)})
                # #print(user1)
                final_output = {'demo1': value}  # Store the dictionary directly
                if "pending" in value[0]:
                    #print(value[0]["pending"])
                    coll.update_one({'_id': ObjectId(first_key)}, {'$set': {'pending': value[0]["pending"]}})
                else:
                    user2 = coll2.find_one({"email": user1["useremail"]})
                    #print(user2["designation"])
                    coll.update_one({'_id': ObjectId(first_key),"useremail":user2["email"]}, {'$set': {'result': final_output['demo1']}})
                #     # Update the entire 'result' field and the 'status' field
                    try:
                        for val in value:
                            demo = coll.find_one({"newnumber.Phno": val["MSISDN"]})
                            if demo:
                                query = {'_id': ObjectId(first_key), 'useremail': user2["email"], 'newnumber.Phno': val["MSISDN"]}
                                query1 = {'_id': ObjectId(first_key), 'useremail': user2["email"]}

                                update_data = {'$set': {'newnumber.$.result': val}}
                                coll.update_one(query, update_data)                    
                                if user2["role"] == "Logger":
                                    pending_status = 'Approval Pending'
                                elif user2["designation"] == "INSPR":
                                    pending_status = 'Approval Pending'
                                elif user2["designation"] == "ADDL-SP/DSP":
                                    pending_status = 'Approval Pending'
                                elif user2["designation"] == "SP":
                                    pending_status = 'Approval Pending'
                                else:
                                    pending_status = 'Approval Pending'
                                coll.update_one(query, {'$set': {'pending': pending_status}})
                    except Exception as e:
                        print(e)
                # try:
                user = coll.find({'_id': ObjectId(my_key)})
                for doc in user:
                    # #print(doc)
                    if doc['requesttype']=='CAF' or doc['requesttype']=='SDR':
                        # #print('Jakkas')
                        for res in doc['result']:
                            for res1 in doc['newnumber']:
                                mobile_number = res["MSISDN"]
                                activation_date = res["ActivationDate"]
                                # #print('wwwwww',activation_date,mobile_number)
                                mnp=res['MNP']
                                if activation_date != "":
                                    if 'CO' in mnp:
                                        bsnl = datetime.strptime(activation_date, "%d-%b-%y")
                                        bsnl_date=bsnl.strftime("%d/%m/%Y")
                                        coll.update_one(
                                                    {'_id': ObjectId(my_key), "result.MSISDN": res1["Phno"]},
                                                    {"$set": {"result.$.ActivationDate": bsnl_date}}
                                                )
                                    if 'VI' in mnp:
                                        # #print('xxxxxxxxxxx',activation_date)
                                        vi=datetime.strptime(activation_date, "%d-%b-%Y")
                                        vi_date=vi.strftime("%d/%m/%Y")
                                        coll.update_one(
                                                        {'_id': ObjectId(my_key), "result.MSISDN": res1["Phno"]},
                                                        {"$set": {"result.$.ActivationDate": vi_date}}
                                                    )
                                    if 'AT' in mnp:
                                        try:
                                            at=datetime.strptime(activation_date, "%a, %d %b %Y %H:%M:%S %Z")
                                            at_date=at.strftime("%d/%m/%Y")
                                            coll.update_one(
                                                        {'_id': ObjectId(my_key), "result.MSISDN": res1["Phno"]},
                                                        {"$set": {"result.$.ActivationDate": at_date}}
                                                    )
                                        except Exception as e:
                                            print(e)
                                            
                                    if 'JO' in mnp or 'Jio' in mnp or 'jio' in mnp or 'JIO' in mnp:
                                        try:    
                                            if len(activation_date)>10:
                                                jo=datetime.strptime(activation_date,"%Y-%m-%d %H:%M:%S")
                                                #jo=datetime.strptime(activation_date,"%Y-%m-%d %H:%M:%S")
                                                jo_date=jo.strftime("%d/%m/%Y")
                                                coll.update_one(
                                                    {'_id': ObjectId(my_key), "result.MSISDN": res1["Phno"]},
                                                    {"$set": {"result.$.ActivationDate": jo_date}})
                                            else:
                                                jo=datetime.strptime(activation_date,"%Y-%m-%d")
                                                # #print(jo)
                                                #jo=datetime.strptime(activation_date,"%Y-%m-%d %H:%M:%S")
                                                jo_date=jo.strftime("%d/%m/%Y")
                                                # #print(jo_date)
                                                coll.update_one(
                                                    {'_id': ObjectId(my_key), "result.MSISDN": res1["Phno"]},
                                                    {"$set": {"result.$.ActivationDate": jo_date}}
                                                )
                                        except Exception as e:
                                            print(e)
                                else:
                                    print('No Activation Date')
                                            
                    else:
                        for res in doc['result']:
                            for res1 in doc['newnumber']:
                                mobile_number = res["MSISDN"]
                                activation_date = res["ActivationDate"]
                                mnp=res['MNP']
                                from_date=res1['From_Date']
                                to_date=res1['To_Date']
                                # #print(res["MSISDN"],res1['Phno'],"==========================")
                                if activation_date == "":
                                    # #print(from_date,mobile_number,'//////////////////////////////////////////////////////////////////////////////////////')
                                    coll.update_one(
                                        {'_id': ObjectId(my_key), "result.MSISDN": res1["Phno"]},
                                        {"$set": {"result.$.From_Date": res1["From_Date"],"result.$.To_Date": res1['To_Date']}}
                                    )
                                else:
                                    if ((res["MSISDN"]==res1['Phno']) and ('CO' in mnp)):
                                        from_date1=res1['From_Date']
                                        from_date=datetime.strptime(from_date1,"%d/%m/%Y")
                                        bsnl = datetime.strptime(activation_date, "%d-%b-%y")

                                        bsnl_date1=bsnl.strftime("%d/%m/%Y")
                                        bsnl_date=datetime.strptime(bsnl_date1,'%d/%m/%Y')
                                        if from_date >= bsnl_date:
                                            coll.update_one(
                                                {'_id': ObjectId(my_key), "result.MSISDN": res1["Phno"]},
                                                {"$set": {"result.$.From_Date": from_date1,"result.$.To_Date": to_date}}
                                            )
                                        else:
                                            coll.update_one(
                                                {'_id': ObjectId(my_key), "result.MSISDN": res1["Phno"]},
                                                {"$set": {"result.$.From_Date": bsnl_date1,"result.$.To_Date": to_date}}
                                            )
                                    
                                    if (res["MSISDN"]==res1['Phno']) and ('VI' in mnp):
                                        try:
                                            from_date1=res1['From_Date']
                                            from_date=datetime.strptime(from_date1,"%d/%m/%Y")
                                            vi=datetime.strptime(activation_date, "%d-%b-%Y")
                                            vi_date1=vi.strftime("%d/%m/%Y")
                                            vi_date=datetime.strptime(vi_date1,'%d/%m/%Y')
                                            if from_date > vi_date :
                                                coll.update_one(
                                                    {'_id': ObjectId(my_key), "result.MSISDN": res1["Phno"]},
                                                    {"$set": {"result.$.From_Date": from_date1,"result.$.To_Date": to_date}}
                                                )
                                            else:
                                                coll.update_one(
                                                    {'_id': ObjectId(my_key), "result.MSISDN": res1["Phno"]},
                                                    {"$set": {"result.$.From_Date": vi_date1,"result.$.To_Date": to_date}}
                                                    ) 
                                        except Exception as e:
                                                print(e)                                         
                                    if (res["MSISDN"]==res1['Phno']) and ('AT' in mnp):
                                        from_date1=res1['From_Date']
                                        from_date=datetime.strptime(from_date1,"%d/%m/%Y")
                                        at=datetime.strptime(activation_date, "%a, %d %b %Y %H:%M:%S %Z")
                                        at_date1=at.strftime("%d/%m/%Y")
                                        at_date=datetime.strptime(at_date1,'%d/%m/%Y')

                                        if from_date >= at_date:
                                            coll.update_one(
                                                {'_id': ObjectId(my_key), "result.MSISDN": res1["Phno"]},
                                                {"$set": {"result.$.From_Date": from_date1,"result.$.To_Date": to_date}}
                                            )
                                        else:
                                            coll.update_one(
                                                {'_id': ObjectId(my_key), "result.MSISDN": res1["Phno"]},
                                                {"$set": {"result.$.From_Date": at_date1,"result.$.To_Date": to_date}}
                                            )

                                    if (res["MSISDN"]==res1['Phno']) and ('JO' in mnp or 'Jio' in mnp or 'jio' in mnp or 'JIO' in mnp) :
                                        try:
                                            from_date1=res1['From_Date']
                                            from_date=datetime.strptime(from_date1,"%d/%m/%Y")

                                            if len(activation_date)>10:
                                                jo=datetime.strptime(activation_date,"%Y-%m-%d %H:%M:%S")

                                                jo_date1=jo.strftime("%d/%m/%Y")
                                                jo_date=datetime.strptime(jo_date1,'%d/%m/%Y')
                            
                                                if from_date >= jo_date:
                                                    coll.update_one(
                                                        {'_id': ObjectId(my_key), "result.MSISDN": res1["Phno"]},
                                                        {"$set": {"result.$.From_Date": from_date1,"result.$.To_Date": to_date}}
                                                    )
                                                else:
                                                    coll.update_one(
                                                        {'_id': ObjectId(my_key), "result.MSISDN": res1["Phno"]},
                                                        {"$set": {"result.$.From_Date": jo_date1,"result.$.To_Date": to_date}}
                                                    )
                                            else:
                                                jo=datetime.strptime(activation_date,"%Y-%m-%d")
                                                jo_date1=jo.strftime("%d/%m/%Y")
                                                jo_date=datetime.strptime(jo_date1,'%d/%m/%Y')
                                                if from_date >= jo_date:
                                                    coll.update_one(
                                                        {'_id': ObjectId(my_key), "result.MSISDN": res1["Phno"]},
                                                        {"$set": {"result.$.From_Date": from_date1,"result.$.To_Date": to_date}}
                                                    )
                                                else:
                                                   coll.update_one(
                                                        {'_id': ObjectId(my_key), "result.MSISDN": res1["Phno"]},
                                                        {"$set": {"result.$.From_Date": jo_date1,"result.$.To_Date": to_date}}
                                                    )
                                        except Exception as e:
                                            print(e,"---------------")

                user = coll.find({'_id': ObjectId(my_key)})
                for doc in user:
                    if doc['requesttype']=='CAF' or doc['requesttype']=='SDR':
                        #print('Jakkas')
                        for res1 in doc['newnumber']:
                            mobile_number = res1["Phno"]
                            activation_date = res1['result']["ActivationDate"]
                            mnp=res1['result']['MNP']
                            if activation_date != "":
                                if 'CO' in mnp:
                                    bsnl = datetime.strptime(activation_date, "%d-%b-%y")
                                    bsnl_date=bsnl.strftime("%d/%m/%Y")
                                    coll.update_one(
                                        {'_id': ObjectId(my_key), "newnumber.Phno": mobile_number},
                                        {"$set": {"newnumber.$.From_Date": bsnl_date}}
                                    )
                                if 'VI' in mnp:
                                    vi=datetime.strptime(activation_date, "%d-%b-%Y")
                                    vi_date=vi.strftime("%d/%m/%Y")
                                    coll.update_one(
                                            {'_id': ObjectId(my_key), "newnumber.Phno": mobile_number},
                                            {"$set": {"newnumber.$.From_Date": vi_date}}
                                        )
                                
                                if 'AT' in mnp:
                                        at=datetime.strptime(activation_date, "%a, %d %b %Y %H:%M:%S %Z")
                                        at_date=at.strftime("%d/%m/%Y")
                                        coll.update_one(
                                            {'_id': ObjectId(my_key), "newnumber.Phno": mobile_number},
                                            {"$set": {"newnumber.$.From_Date": at_date}}
                                        )
                                        
                                if 'JO' in mnp or 'Jio' in mnp or 'jio' in mnp or 'JIO' in mnp:
                                        if len(activation_date)>10:
                                            jo=datetime.strptime(activation_date,"%Y-%m-%d %H:%M:%S")
                                            #jo=datetime.strptime(activation_date,"%Y-%m-%d %H:%M:%S")
                                            jo_date=jo.strftime("%d/%m/%Y")
                                            coll.update_one(
                                            {'_id': ObjectId(my_key), "newnumber.Phno": mobile_number},
                                            {"$set": {"newnumber.$.From_Date": jo_date}}
                                        )
                                        else:
                                            jo=datetime.strptime(activation_date,"%Y-%m-%d")
                                            #jo=datetime.strptime(activation_date,"%Y-%m-%d %H:%M:%S")
                                            jo_date=jo.strftime("%d/%m/%Y")
                                            coll.update_one(
                                            {'_id': ObjectId(my_key), "newnumber.Phno": mobile_number},
                                            {"$set": {"newnumber.$.From_Date": jo_date}}
                                        )
                    else:
                        for res1 in doc['newnumber']:
                            try:
                                mobile_number = res1["Phno"]
                                activation_date = res1['result']["ActivationDate"]
                                mnp=res1['result']['MNP']
                                # to_date=res1['To_Date']
                                if activation_date != "":
                                    if 'CO' in mnp:
                                        from_date1=res1['From_Date']
                                        from_date=datetime.strptime(from_date1,"%d/%m/%Y")
                                        #print("MNP : ",mnp)
                                        bsnl = datetime.strptime(activation_date, "%d-%b-%y")
                                        bsnl_date1=bsnl.strftime("%d/%m/%Y")
                                        bsnl_date=datetime.strptime(bsnl_date1,'%d/%m/%Y')
                                        if from_date >= bsnl_date:
                                            coll.update_one(
                                                {'_id': ObjectId(my_key), "newnumber.Phno": mobile_number},
                                                {"$set": {"newnumber.$.From_Date": from_date1}}
                                            )
                                        else:
                                            coll.update_one(
                                                {'_id': ObjectId(my_key), "newnumber.Phno": mobile_number},
                                                {"$set": {"newnumber.$.From_Date": bsnl_date1}}
                                            )
                                    if 'VI' in mnp:
                                        from_date1=res1['From_Date']
                                        from_date=datetime.strptime(from_date1,"%d/%m/%Y")
                                        vi=datetime.strptime(activation_date, "%d-%b-%Y")
                                        vi_date1=vi.strftime("%d/%m/%Y")
                                        vi_date = datetime.strptime(vi_date1,'%d/%m/%Y')
                                        if from_date >= vi_date :
                                            coll.update_one(
                                                {'_id': ObjectId(my_key), "newnumber.Phno": mobile_number},
                                                {"$set": {"newnumber.$.From_Date": from_date1}}
                                            )
                                        else:
                                            coll.update_one(
                                                {'_id': ObjectId(my_key), "newnumber.Phno": mobile_number},
                                                {"$set": {"newnumber.$.From_Date": vi_date1}}
                                            )
                                    if 'AT' in mnp:
                                        from_date1=res1['From_Date']
                                        from_date=datetime.strptime(from_date1,"%d/%m/%Y")
                                        at=datetime.strptime(activation_date, "%a, %d %b %Y %H:%M:%S %Z")
                                        at_date1=at.strftime("%d/%m/%Y")
                                        at_date = datetime.strptime(at_date1,'%d/%m/%Y')
                                        if from_date >= at_date:
                                            coll.update_one(
                                                {'_id': ObjectId(my_key), "newnumber.Phno": mobile_number},
                                                {"$set": {"newnumber.$.From_Date": from_date1}}
                                            )
                                        else:
                                            coll.update_one(
                                                {'_id': ObjectId(my_key), "newnumber.Phno": mobile_number},
                                                {"$set": {"newnumber.$.From_Date": at_date1}}
                                            )
                                    if 'JO' in mnp or 'Jio' in mnp or 'jio' in mnp or 'JIO' in mnp:
                                        from_date1=res1['From_Date']
                                        #print('iniffff',type(from_date1))
                                        from_date=datetime.strptime(from_date1,"%d/%m/%Y")
                                        if len(activation_date)>10:
                                            jo=datetime.strptime(activation_date,"%Y-%m-%d %H:%M:%S")
                                            jo_date1=jo.strftime("%d/%m/%Y")
                                            jo_date=datetime.strptime(jo_date1,'%d/%m/%Y')
                                            
                                            if from_date >= jo_date:
                                                
                                                coll.update_one(
                                                    {'_id': ObjectId(my_key), "newnumber.Phno": mobile_number},
                                                    {"$set": {"newnumber.$.From_Date": from_date1}}
                                                )
                                            else:
                                                
                                                coll.update_one(
                                                    {'_id': ObjectId(my_key), "newnumber.Phno": mobile_number},
                                                    {"$set": {"newnumber.$.From_Date": jo_date1}}
                                                )
                                        else:
                                            jo=datetime.strptime(activation_date,"%Y-%m-%d")
                                            jo_date1=jo.strftime("%d/%m/%Y")
                                            jo_date=datetime.strptime(jo_date1,'%d/%m/%Y')
                                            if from_date >= jo_date:
                                                coll.update_one(
                                                    {'_id': ObjectId(my_key), "newnumber.Phno": mobile_number},
                                                    {"$set": {"newnumber.$.From_Date": from_date1}}
                                                )
                                            else:
                                                coll.update_one(
                                                    {'_id': ObjectId(my_key), "newnumber.Phno": mobile_number},
                                                    {"$set": {"newnumber.$.From_Date": jo_date1}}
                                                )
                                else:
                                    coll.update_one(
                                        {'_id': ObjectId(my_key), "newnumber.Phno": mobile_number},
                                        {"$set": {"newnumber.$.From_Date": from_date1}}
                                    )
                            except Exception as e:
                                print(e,'/////sasasagsaghsjasasa')
                                    
                # except Exception as e:
                #     #print(e)
        else:
            print('WELCOME FILES UPLOAD Fn........')

        return jsonify({'message': 'Data updated successfully'})
# End of Automation file Upload Endpoint------------------------------------------------------------------------------------------>

# Endpoint for deleteRows--------------------------------------------------------------------------------------------------------->

@tickect_bp.route('/delete_row', methods=['POST'])
def delete_row():
    id = request.form['id']

    msisdn = request.form['msisdn']
    
    # Assuming your collection is named 'tickets'
    tickets_collection = db['tickets']
    
    # Convert id string to ObjectId
    obj_id = ObjectId(id)
    
    # Update the document using pymongo
    tickets_collection.update_one(
        {'_id': obj_id},
        {'$pull': {'result.MSISDN': msisdn}}
    )
    
    return jsonify({'message': 'Deleted successfully'})

# End of DeleteRows Endpoint----------------------------------------------------------------------------------------------------->

@tickect_bp.route('/delete_ticket', methods=['POST'])
def delete_ticket():
    id = request.json.get('id')
    coll = db['tickets']
    # #print(id,token)
    user = coll.delete_one({'_id':ObjectId(id)})
    # user = Tickets.objects_safe(id=id1, token = token).first()
    if user:
        return jsonify({'message': 'Tickect deleted successfully'})
    else:
        return jsonify({'message': 'Ticket not found'})
# Endpoint for Forgot Password--------------------------------------------------------------------------------------------------->
@tickect_bp.route('/newanalysis', methods=['POST'])
def newanalysis():
    now = datetime.now()
    day_with_millisec = now.strftime("%d%m%Y%H%M%S")+f"{now.microsecond // 1000:03d}"
    sts = "Approval Pending"
    coll2 = db['tickets']
    sum = 1001
    db_len = coll2.find_one({"status": {"$ne": "Converted"}}, sort=[('_id', -1)])
    last_token = db_len.get('token','').split("_")
    if last_token[0] != day_with_millisec:
        last_token=int(last_token[1])+1
        tokens = f"{day_with_millisec}_" + str(last_token)
    Tickets(
        useremail=request.json.get('useremail'),
        superior=request.json.get('superior'),
        officername=request.json.get('officername'),
        designation = request.json.get("designation"),
        role=request.json.get('role'),
        type=request.json.get('update'),
        requestcategory="Analysis Request",
        requesttypes=request.json.get('requesttypes'),
        requesttype=request.json.get('requesttype'),
        team=request.json.get('team'),
        modules=request.json.get('modules'),
        reason=request.json.get('reason'),
        refno=request.json.get('refno'),
        refdate=str(request.json.get('refdate')),
        priority=request.json.get('priority'),
        token=tokens,
        pending=sts,
        approval = "-",
        assign_Officer="",
        others='',
        newnumber = request.json.get('newnumber'),
        subtypes = request.json.get('subtype'),
        date = request.json.get("date"),
        remarks=request.json.get('remarks'),
        raise_time = request.json.get("raise_time")
    ).save()

    return jsonify({'message': 'Data Insert successful'})
@tickect_bp.route('/asign', methods=['POST'])
def asign():
    data = request.json
    ticket_id = data.get('id')
    assign_Officer = data.get('officer')
    pending = data.get("pending")
    approval = data.get("approval")
    
    collection = db['tickets']

    # Find the ticket in the database using the provided ticket ID
    ticket = collection.find_one({"_id": ObjectId(ticket_id)})

    if not ticket:
        return jsonify({"message": "Ticket not found"}), 404

    # Check if the ticket i of Analysis  Request type and is raised by AP or ISOT
    if ticket["requestcategory"] != "Analysis Request" or ticket["team"] not in ["AP", "ISOT","CAT",'ADMIN']:
        return jsonify({"message": "Ticket is not eligible for assignment to CAT team"}), 400

    # Update the status of the ticket
    if pending == 'ADDL-SP/DSP Approval Pending' or pending == 'SP Approval Pending' or pending == "--":
        ticket["assign_Officer"] = assign_Officer
        ticket["pending"] = pending
        ticket["approval"] = approval
    elif pending == 'Rejected by DSP' or pending == 'Rejected by SP':
        ticket["pending"] = pending
        ticket["approval"] = approval
        ticket["assign_Officer"] = ''
    else:
        ticket["assign_Officer"] = ""

    # Save the changes back to the database
    collection.update_one({"_id": ObjectId(ticket_id)}, {"$set": ticket})
    return jsonify({"message": "Ticket assigned to Analyst CAT team successfully"}), 200


@tickect_bp.route('/nickname_search', methods=['GET',
'POST'])
def nickname_search():
    phone_match=request.json.get('phone')
    coll=sus_db['cdat_suspect']
    sample={'phno':'',
    'nickname':''}
    if len(phone_match) == 10:
        data = coll.find({'phone':phone_match}).sort([('_id', -1)]).limit(1)
        for alldata in data:
            try:
                if phone_match == alldata['phone']:
                    sample["phno"] = alldata['phone']
                    sample['nickname'] = alldata['nickname']
            except:
                pass
    return jsonify(sample), 200

@tickect_bp.route('/get_tower_by_radius', methods=['POST', 'GET'])
def get_tower_in_polygon():
    coll = sus_db['cellidchart_jio']
    types = request.json.get("type")
    cgi = request.json.get("cgi")
    rad = request.json.get("rad")
    #print(types)
    # if types == "CGI" and cgi !="" and rad != "":
    #     sample={"CELL_ID":[],"AreaDescription":[],"Operator":[],"State":[],"RAD":""}
    #     #print(cgi,"CGI ID    <==================================")
    #     cgi_list =[]
    #     try:
    #         areafind = coll.find()
    #         for i in areafind:
    #             if cgi == i["celltowerid"]:
    #                 cgi_list.append(i)
    #         #print(cgi_list,"<==================================>")
    #         # if len(rad) >= 1:
    #         rad_in_meters =int(rad)*1000
    #         rad_in_meters=int(rad_in_meters)
    #         if cgi_list[0]['celltowerid']==cgi:
    #             towers_in_circle = list(coll.find({
    #             'location': {
    #                 '$near': {
    #                     '$geometry': {
    #                         'type': "Point" ,
    #                         'coordinates': [float(cgi_list[0]["long"]),float(cgi_list[0]["lat"])]
    #                     },
    #                     '$maxDistance': rad_in_meters   
    #                 }
    #             }
    #             },{"_id":0}))
    #             #print(towers_in_circle)
    #             for res in towers_in_circle:
    #                 sample['CELL_ID'].append(res['bts_id'])
    #                 sample['RAD']=rad
    #                 sample["Lat"] = (cgi_list[0]["lat"])
    #                 sample["Long"] = (cgi_list[0]["long"])
    #                 sample['AreaDescription'].append(res['areadescription'])
    #                 sample['Operator'].append(res['provider'])
    #                 sample['State'].append(res['state'])
    #             #print(sample)
    #             return jsonify(sample)

    #     except Exception as error:
    #         #print("errroorooroorroor: ",error)
        
    #     return jsonify(sample)
    
    if types == "Co-Ordinates":
        lat = request.json.get("lat")
        lat=float(lat)
        long = request.json.get("long")
        long=float(long)
        rad = request.json.get("rad")
        # #print(lat,long,rad)
        rad_in_meters =int(rad)*1000
        rad_in_meters=int(rad_in_meters)
        # #print(rad_in_meters)
        towers_in_circle = list(coll.find({
            'location': {
                '$near': {
                    '$geometry': {
                        'type': "Point" ,
                        'coordinates': [long,lat]
                    },
                    '$maxDistance': rad_in_meters   
                }
            }
            
        },{"_id":0}))
        sample={"CELL_ID":[],"AreaDescription":[],"Operator":[],"State":[],"RAD":""}
        for res in towers_in_circle:
            sample['CELL_ID'].append(res['bts_id'])
            sample['RAD']=rad
            sample['AreaDescription'].append(res['areadescription'])
            sample['Operator'].append(res['provider'])
            sample['State'].append(res['state'])
        # #print(sample)
        return jsonify(sample)
    
    return jsonify("NOT FOUND ANY VALUEs......")

@tickect_bp.route('/download-excel', methods=['POST', 'GET'])
def download_excel():
    # Your existing code to fetch data from MongoDB
    coll = db["tickets"]
    id = request.json.get("id")
    email = request.json.get("email")
    request_type = request.json.get("type1")
    request_type1 = request.json.get("type2")
    category = request.json.get("category")
    subtypes = request.json.get("sub")
    cel = coll.find({"_id":ObjectId(id),"useremail":email})
    if (category == 'Data Request'):
        if (request_type == "TOWER CDR" and subtypes == "Co-Ordinates") or (request_type == "TOWER GPRS" and subtypes == "Co-Ordinates") or (request_type == "TOWER IPDR" and subtypes == "Co-Ordinates"):
            concatenated_df = pd.DataFrame(columns=['cellid', 'latitude', 'longitude', 'area', 'operator', 'state'])
            for i in cel:
                for j in range(len(i['newnumber'])):
                    cellid = i["newnumber"][j]['Cell_id']
                    data = {
                        'cellid': cellid,
                        'latitude': i['newnumber'][j]['Latitude'],
                        'longitude': i['newnumber'][j]['Longitude'],
                        'area': i['newnumber'][j]['AreaDescription'],
                        'operator': i['newnumber'][j]['Operator'],
                        'state': i['newnumber'][j]['State']
                    }
                    df = pd.DataFrame(data, columns=['cellid', 'latitude', 'longitude', 'area', 'operator', 'state'])
                    concatenated_df = pd.concat([concatenated_df, df], ignore_index=True)
                    excel_data = io.BytesIO()
                    with pd.ExcelWriter(excel_data, engine='xlsxwriter') as writer:
                        concatenated_df.to_excel(writer, sheet_name='Sheet1', index=False)

                    excel_data.seek(0)
                    response = send_file(
                        excel_data,
                        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                        as_attachment=True,
                        download_name='/home/bruce/Downloads/output.xlsx'
                    )
            
        elif (request_type == "SDR"):
            concatenated_df = pd.DataFrame(columns=["Phno"])
            for i in cel:
                for j in range(len(i['newnumber'])):
                    Phno = i["newnumber"][j]['Phno']
                    data = {
                        'Phno': Phno
                    }
                    df = pd.DataFrame([data], columns=["Phno"])
                    concatenated_df = pd.concat([concatenated_df, df], ignore_index=True)
                    # #print(concatenated_df)
                    excel_data = io.BytesIO()
                    with pd.ExcelWriter(excel_data, engine='xlsxwriter') as writer:
                        concatenated_df.to_excel(writer, sheet_name='Sheet1', index=False)
                    excel_data.seek(0)
                    response = send_file(
                        excel_data,
                        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                        as_attachment=True,
                        download_name='/home/bruce/Downloads/output.xlsx'
                    )
        elif (request_type == "CAF") or (request_type == "CAF/CDR") :
            if request_type == "CAF":
                concatenated_df = pd.DataFrame(columns=["Phno","CAF"])
                for i in cel:
                    for j in range(len(i['newnumber'])):
                        Phno = i["newnumber"][j]['Phno']
                        data = {
                            'Phno': Phno,
                            'CAF': i['newnumber'][j]['CAF'],
                        }
                        df = pd.DataFrame([data], columns=["Phno","CAF"])
                        concatenated_df = pd.concat([concatenated_df, df], ignore_index=True)
                        # #print(concatenated_df)
                        excel_data = io.BytesIO()
                        with pd.ExcelWriter(excel_data, engine='xlsxwriter') as writer:
                            concatenated_df.to_excel(writer, sheet_name='Sheet1', index=False)

                        excel_data.seek(0)
                        response = send_file(
                            excel_data,
                            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                            as_attachment=True,
                            download_name='/home/bruce/Downloads/output.xlsx'
                        )
            else:
                concatenated_df = pd.DataFrame(columns=["Phno/CAF","From_Date","To_Date"])
                for i in cel:
                    for j in range(len(i['newnumber'])):
                        Phno = i["newnumber"][j]['Phno']
                        data = {
                            'Phno/CAF': Phno,
                            'From_Date': i['newnumber'][j]['From_Date'],
                            'To_Date': i['newnumber'][j]['To_Date']
                        }
                        df = pd.DataFrame([data], columns=["Phno/CAF","From_Date","To_Date"])
                        concatenated_df = pd.concat([concatenated_df, df], ignore_index=True)
                        # #print(concatenated_df)
                        excel_data = io.BytesIO()
                        with pd.ExcelWriter(excel_data, engine='xlsxwriter') as writer:
                            concatenated_df.to_excel(writer, sheet_name='Sheet1', index=False)

                        excel_data.seek(0)
                        response = send_file(
                            excel_data,
                            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                            as_attachment=True,
                            download_name='/home/bruce/Downloads/output.xlsx'
                        )

        elif request_type == "IMEI CDR" or (request_type == "GPRS" and subtypes == "IMEI") or (request_type == "IPDR" and subtypes == "IMEI"):
            if request_type == "IMEI CDR":
                concatenated_df = pd.DataFrame(columns=["IMEI","ISP","From_Date","To_Date"])
                for i in cel:
                    for j in range(len(i['newnumber'])):
                        IMEI = i["newnumber"][j]['IMEI']
                        data = {
                            'IMEI': IMEI,
                            'ISP' : i['newnumber'][j]['ISP'],
                            'From_Date': i['newnumber'][j]['From_Date'],
                            'To_Date': i['newnumber'][j]['To_Date']
                            
                        }
                        df = pd.DataFrame([data], columns=["IMEI","ISP","From_Date","To_Date"])
                        concatenated_df = pd.concat([concatenated_df, df], ignore_index=True)
                        # #print(concatenated_df)
                        excel_data = io.BytesIO()
                        with pd.ExcelWriter(excel_data, engine='xlsxwriter') as writer:
                            concatenated_df.to_excel(writer, sheet_name='Sheet1', index=False)

                        excel_data.seek(0)
                        response = send_file(
                            excel_data,
                            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                            as_attachment=True,
                            download_name='/home/bruce/Downloads/output.xlsx'
                        )
            else:
                concatenated_df = pd.DataFrame(columns=["IMEI","ISP","From_Date","From_Time","To_Date","To_Time"])
                for i in cel:
                    for j in range(len(i['newnumber'])):
                        IMEI = i["newnumber"][j]['IMEI']
                        data = {
                            'IMEI': IMEI,
                            'ISP': i['newnumber'][j]['ISP'],
                            'From_Date': i['newnumber'][j]['From_Date'],
                            'From_Time': i['newnumber'][j]['From_Time'],
                            'To_Date': i['newnumber'][j]['To_Date'],
                            'To_Time': i['newnumber'][j]['To_Time']
                            
                        }
                        df = pd.DataFrame([data], columns=["IMEI","ISP","From_Date","From_Time","To_Date","To_Time"])
                        concatenated_df = pd.concat([concatenated_df, df], ignore_index=True)
                        # #print(concatenated_df)
                        excel_data = io.BytesIO()
                        with pd.ExcelWriter(excel_data, engine='xlsxwriter') as writer:
                            concatenated_df.to_excel(writer, sheet_name='Sheet1', index=False)

                        excel_data.seek(0)
                        response = send_file(
                            excel_data,
                            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                            as_attachment=True,
                            download_name='/home/bruce/Downloads/output.xlsx'
                        )
                        

        elif (request_type == "IPDR") and (subtypes=="IPV6" or subtypes=="IPV4"):
            concatenated_df = pd.DataFrame(columns=["IP_Address","ISP","From_Date","From_Time","To_Date","To_Time"])
            for i in cel:
                for j in range(len(i['newnumber'])):
                    IP_Address = i["newnumber"][j]['IP_Address']
                    data = {
                        'IP_Address': IP_Address,
                        'ISP': i['newnumber'][j]['ISP'],
                        'From_Date': i['newnumber'][j]['From_Date'],
                        'From_Time': i['newnumber'][j]['From_Time'],
                        'To_Date': i['newnumber'][j]['To_Date'],
                        'To_Time': i['newnumber'][j]['To_Time']
                        
                    }
                    df = pd.DataFrame([data], columns=["IP_Address","ISP","From_Date","From_Time","To_Date","To_Time"])
                    concatenated_df = pd.concat([concatenated_df, df], ignore_index=True)
                    # #print(concatenated_df)
                    excel_data = io.BytesIO()
                    with pd.ExcelWriter(excel_data, engine='xlsxwriter') as writer:
                        concatenated_df.to_excel(writer, sheet_name='Sheet1', index=False)

                    excel_data.seek(0)
                    response = send_file(
                        excel_data,
                        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                        as_attachment=True,
                        download_name='/home/bruce/Downloads/output.xlsx'
                    )

        elif (request_type == "RH" and subtypes=="Retailer/Dealer Details") or (request_type == "POA" and subtypes=="Retailer/Dealer Details"):
            if request_type == "RH":
                concatenated_df = pd.DataFrame(columns=["RH_Dealer","RH_code","ISP","From_Date","To_Date"])
                for i in cel:
                    for j in range(len(i['newnumber'])):
                        RH_Dealer = i["newnumber"][j]['RH_Dealer']
                        data = {
                            'RH_Dealer': RH_Dealer,
                            'RH_code': i['newnumber'][j]['RH_code'],
                            'ISP': i['newnumber'][j]['ISP'],
                            'From_Date': i['newnumber'][j]['From_Date'],
                            'To_Date': i['newnumber'][j]['To_Date']                       
                        }
                        df = pd.DataFrame([data], columns=["RH_Dealer","RH_code","ISP","From_Date","To_Date"])
                        concatenated_df = pd.concat([concatenated_df, df], ignore_index=True)
                        # #print(concatenated_df)
                        excel_data = io.BytesIO()
                        with pd.ExcelWriter(excel_data, engine='xlsxwriter') as writer:
                            concatenated_df.to_excel(writer, sheet_name='Sheet1', index=False)
                        excel_data.seek(0)
                        response = send_file(
                            excel_data,
                            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                            as_attachment=True,
                            download_name='/home/bruce/Downloads/output.xlsx'
                        )
            else:
                concatenated_df = pd.DataFrame(columns=["POA_Dealer","POA_code","ISP","From_Date","To_Date"])
                for i in cel:
                    for j in range(len(i['newnumber'])):
                        RH_Dealer = i["newnumber"][j]['POA_Dealer']
                        data = {
                            'POA_Dealer': RH_Dealer,
                            'POA_code': i['newnumber'][j]['POA_code'],
                            'ISP': i['newnumber'][j]['ISP'],
                            'From_Date': i['newnumber'][j]['From_Date'],
                            'To_Date': i['newnumber'][j]['To_Date']                       
                        }
                        df = pd.DataFrame([data], columns=["POA_Dealer","POA_code","ISP","From_Date","To_Date"])
                        concatenated_df = pd.concat([concatenated_df, df], ignore_index=True)
                        # #print(concatenated_df)
                        excel_data = io.BytesIO()
                        with pd.ExcelWriter(excel_data, engine='xlsxwriter') as writer:
                            concatenated_df.to_excel(writer, sheet_name='Sheet1', index=False)
                        excel_data.seek(0)
                        response = send_file(
                            excel_data,
                            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                            as_attachment=True,
                            download_name='/home/bruce/Downloads/output.xlsx'
                        )

        elif (request_type == "TOWER CDR" and subtypes == "CGI") or (request_type == "TOWER GPRS" and subtypes == "CGI") or (request_type == "TOWER IPDR" and subtypes == "CGI"):
            # concatenated_df = pd.DataFrame(columns=[,'ISP', "From_Date",'To_Date'])
            concatenated_df = pd.DataFrame(columns=['Cellid', "From_Date","To_Date"])
            for i in cel:
                for j in range(len(i['newnumber'])):
                    cellid = i["newnumber"][j]['CGI']
                    data = {
                        'Cellid': cellid,
                        'From_Date': i['newnumber'][j]['From_Date'],
                        'To_Date': i['newnumber'][j]['To_Date']  
                    }
                    df = pd.DataFrame([data], columns=['Cellid', "From_Date","To_Date"])
                    concatenated_df = pd.concat([concatenated_df, df], ignore_index=True)
                    excel_data = io.BytesIO()
                    with pd.ExcelWriter(excel_data, engine='xlsxwriter') as writer:
                        concatenated_df.to_excel(writer, sheet_name='Sheet1', index=False)

                    excel_data.seek(0)
                    response = send_file(
                        excel_data,
                        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                        as_attachment=True,
                        download_name='/home/bruce/Downloads/output.xlsx'
                    )
        else:
            # #print("Heloooooooooooooooooooooooooooo")
            concatenated_df = pd.DataFrame(columns=["Phno","Nickname","From_Date","To_Date"])
            for i in cel:
                for j in range(len(i['newnumber'])):
                    Phno = i["newnumber"][j]['Phno']
                    data = {
                        'Phno': Phno,
                        'Nickname': i['newnumber'][j]['Nickname'],
                        'From_Date': i['newnumber'][j]['From_Date'],
                        'To_Date': i['newnumber'][j]['To_Date']
                    }
                    df = pd.DataFrame([data], columns=["Phno","Nickname","From_Date","To_Date"])
                    concatenated_df = pd.concat([concatenated_df, df], ignore_index=True)
                    # #print(concatenated_df)
                    excel_data = io.BytesIO()
                    with pd.ExcelWriter(excel_data, engine='xlsxwriter') as writer:
                        concatenated_df.to_excel(writer, sheet_name='Sheet1', index=False)
                    excel_data.seek(0)
                    response = send_file(
                        excel_data,
                        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                        as_attachment=True,
                        download_name='/home/bruce/Downloads/output.xlsx'
                    )
    else:
        if ("TOWER DATA" in request_type1 and subtypes == "Co-Ordinates"):
            concatenated_df = pd.DataFrame(columns=['cellid', 'latitude', 'longitude', 'area', 'operator', 'state'])
            for i in cel:
                for j in range(len(i['newnumber'])):
                    cellid = i["newnumber"][j]['Cell_id']
                    data = {
                        'cellid': cellid,
                        'latitude': i['newnumber'][j]['Latitude'],
                        'longitude': i['newnumber'][j]['Longitude'],
                        'area': i['newnumber'][j]['AreaDescription'],
                        'operator': i['newnumber'][j]['Operator'],
                        'state': i['newnumber'][j]['State']
                    }
                    df = pd.DataFrame(data, columns=['cellid', 'latitude', 'longitude', 'area', 'operator', 'state'])
                    concatenated_df = pd.concat([concatenated_df, df], ignore_index=True)
                    excel_data = io.BytesIO()
                    with pd.ExcelWriter(excel_data, engine='xlsxwriter') as writer:
                        concatenated_df.to_excel(writer, sheet_name='Sheet1', index=False)

                    excel_data.seek(0)
                    response = send_file(
                        excel_data,
                        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                        as_attachment=True,
                        download_name='/home/bruce/Downloads/output.xlsx'
                    )
            
        elif (("SDR" in request_type1) or ("CAF" in request_type1) or ("CDR" in request_type1)):
            concatenated_df = pd.DataFrame(columns=["Phno","ISP","From_Date","To_Date"])
            for i in cel:
                for j in range(len(i['newnumber'])):
                    Phno = i["newnumber"][j]['Phno']
                    data = {
                        'Phno': Phno,
                        'ISP': i['newnumber'][j]['ISP'],
                        'From_Date': i['newnumber'][j]['From_Date'],
                        'To_Date': i['newnumber'][j]['To_Date']
                    }
                    df = pd.DataFrame([data], columns=["Phno","ISP","From_Date","To_Date"])
                    concatenated_df = pd.concat([concatenated_df, df], ignore_index=True)
                    # #print(concatenated_df)
                    excel_data = io.BytesIO()
                    with pd.ExcelWriter(excel_data, engine='xlsxwriter') as writer:
                        concatenated_df.to_excel(writer, sheet_name='Sheet1', index=False)
                    excel_data.seek(0)
                    response = send_file(
                        excel_data,
                        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                        as_attachment=True,
                        download_name='/home/bruce/Downloads/output.xlsx'
                    )
        
        elif (("IMEI" in request_type1) or ("IPDR" in request_type1 and subtypes=="IMEI") or ("GPRS" in request_type1 and subtypes=="IMEI")):
            concatenated_df = pd.DataFrame(columns=["IMEI","From_Date","To_Date"])
            for i in cel:
                for j in range(len(i['newnumber'])):
                    IMEI = i["newnumber"][j]['IMEI']
                    data = {
                        'IMEI': IMEI,
                        'From_Date': i['newnumber'][j]['From_Date'],
                        'To_Date': i['newnumber'][j]['To_Date']
                        
                    }
                    df = pd.DataFrame([data], columns=["IMEI","From_Date","To_Date"])
                    concatenated_df = pd.concat([concatenated_df, df], ignore_index=True)
                    # #print(concatenated_df)
                    excel_data = io.BytesIO()
                    with pd.ExcelWriter(excel_data, engine='xlsxwriter') as writer:
                        concatenated_df.to_excel(writer, sheet_name='Sheet1', index=False)

                    excel_data.seek(0)
                    response = send_file(
                        excel_data,
                        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                        as_attachment=True,
                        download_name='/home/bruce/Downloads/output.xlsx'
                    )                        

        elif ("IPDR" in request_type1) and (subtypes=="IPV6" or subtypes=="IPV4"):
            concatenated_df = pd.DataFrame(columns=["IP_Address","From_Date","To_Date"])
            for i in cel:
                for j in range(len(i['newnumber'])):
                    IP_Address = i["newnumber"][j]['IP_Address']
                    data = {
                        'IP_Address': IP_Address,
                        # 'ISP': i['newnumber'][j]['ISP'],
                        'From_Date': i['newnumber'][j]['From_Date'],
                        'To_Date': i['newnumber'][j]['To_Date']
                        
                    }
                    df = pd.DataFrame([data], columns=["IP_Address","From_Date","To_Date"])
                    concatenated_df = pd.concat([concatenated_df, df], ignore_index=True)
                    # #print(concatenated_df)
                    excel_data = io.BytesIO()
                    with pd.ExcelWriter(excel_data, engine='xlsxwriter') as writer:
                        concatenated_df.to_excel(writer, sheet_name='Sheet1', index=False)

                    excel_data.seek(0)
                    response = send_file(
                        excel_data,
                        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                        as_attachment=True,
                        download_name='/home/bruce/Downloads/output.xlsx'
                    )

        elif ("RH" in request_type1 and subtypes=="Retailer/Dealer Details") or ("POA" in request_type1 and subtypes=="Retailer/Dealer Details"):
            if request_type == "RH":
                concatenated_df = pd.DataFrame(columns=["RH_Dealer","RH_code","ISP","From_Date","To_Date"])
                for i in cel:
                    for j in range(len(i['newnumber'])):
                        RH_Dealer = i["newnumber"][j]['RH_Dealer']
                        data = {
                            'RH_Dealer': RH_Dealer,
                            'RH_code': i['newnumber'][j]['RH_code'],
                            'ISP': i['newnumber'][j]['ISP'],
                            'From_Date': i['newnumber'][j]['From_Date'],
                            'To_Date': i['newnumber'][j]['To_Date']                       
                        }
                        df = pd.DataFrame([data], columns=["IP_Address","ISP","From_Date","From_Time","To_Date","To_Time"])
                        concatenated_df = pd.concat([concatenated_df, df], ignore_index=True)
                        # #print(concatenated_df)
                        excel_data = io.BytesIO()
                        with pd.ExcelWriter(excel_data, engine='xlsxwriter') as writer:
                            concatenated_df.to_excel(writer, sheet_name='Sheet1', index=False)
                        excel_data.seek(0)
                        response = send_file(
                            excel_data,
                            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                            as_attachment=True,
                            download_name='/home/bruce/Downloads/output.xlsx'
                        )
            else:
                concatenated_df = pd.DataFrame(columns=["POA_Dealer","POA_code","ISP","From_Date","To_Date"])
                for i in cel:
                    for j in range(len(i['newnumber'])):
                        RH_Dealer = i["newnumber"][j]['POA_Dealer']
                        data = {
                            'POA_Dealer': RH_Dealer,
                            'POA_code': i['newnumber'][j]['POA_code'],
                            'ISP': i['newnumber'][j]['ISP'],
                            'From_Date': i['newnumber'][j]['From_Date'],
                            'To_Date': i['newnumber'][j]['To_Date']                       
                        }
                        df = pd.DataFrame([data], columns=["IP_Address","ISP","From_Date","From_Time","To_Date","To_Time"])
                        concatenated_df = pd.concat([concatenated_df, df], ignore_index=True)
                        # #print(concatenated_df)
                        excel_data = io.BytesIO()
                        with pd.ExcelWriter(excel_data, engine='xlsxwriter') as writer:
                            concatenated_df.to_excel(writer, sheet_name='Sheet1', index=False)
                        excel_data.seek(0)
                        response = send_file(
                            excel_data,
                            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                            as_attachment=True,
                            download_name='/home/bruce/Downloads/output.xlsx'
                        )

        elif ("TOWER DATA" in request_type1 and subtypes == "CGI"):
            # concatenated_df = pd.DataFrame(columns=[,'ISP', "From_Date",'To_Date'])
            concatenated_df = pd.DataFrame(columns=['Cellid', 'Latitude', 'Longitude', 'Area', 'Operator', 'State'])
            for i in cel:
                for j in range(len(i['newnumber'])):
                    cellid = i["newnumber"][j]['Cell_id']
                    data = {
                        'Cellid': cellid,
                        'Latitude': i['newnumber'][j]['Latitude'],
                        'Longitude': i['newnumber'][j]['Longitude'],
                        'Area': i['newnumber'][j]['AreaDescription'],
                        'Operator': i['newnumber'][j]['Operator'],
                        'State': i['newnumber'][j]['State']
                    }
                    df = pd.DataFrame(data, columns=['Cellid', 'Latitude', 'Longitude', 'Area', 'Operator', 'State'])
                    concatenated_df = pd.concat([concatenated_df, df], ignore_index=True)
                    excel_data = io.BytesIO()
                    with pd.ExcelWriter(excel_data, engine='xlsxwriter') as writer:
                        concatenated_df.to_excel(writer, sheet_name='Sheet1', index=False)

                    excel_data.seek(0)
                    response = send_file(
                        excel_data,
                        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                        as_attachment=True,
                        download_name='/home/bruce/Downloads/output.xlsx'
                    )
        else:
            # #print("Heloooooooooooooooooooooooooooo")
            concatenated_df = pd.DataFrame(columns=["Phno","ISP","From_Date","To_Date"])
            for i in cel:
                for j in range(len(i['newnumber'])):
                    Phno = i["newnumber"][j]['Phno']
                    data = {
                        'Phno': Phno,
                        'ISP': i['newnumber'][j]['ISP'],
                        'From_Date': i['newnumber'][j]['From_Date'],
                        'To_Date': i['newnumber'][j]['To_Date']
                    }
                    df = pd.DataFrame([data], columns=["Phno","ISP","From_Date","To_Date"])
                    concatenated_df = pd.concat([concatenated_df, df], ignore_index=True)
                    # #print(concatenated_df)
                    excel_data = io.BytesIO()
                    with pd.ExcelWriter(excel_data, engine='xlsxwriter') as writer:
                        concatenated_df.to_excel(writer, sheet_name='Sheet1', index=False)
                    excel_data.seek(0)
                    response = send_file(
                        excel_data,
                        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                        as_attachment=True,
                        download_name='/home/bruce/Downloads/output.xlsx'
                    )
    return response


@tickect_bp.route('/update_nickname', methods=['POST', 'GET'])
def update_nickname():
    coll=sus_db['cdat_suspect']
    # collec=sus_db['cdat_suspect']
    num=request.json.get('num')
    email=request.json.get('email')

    ls1 = []
    now = datetime.now()
    date = now.strftime("%d/%m/%Y")
    time = now.strftime("%H:%M:%S")
    if num:
        for i in num:
            data = coll.find_one({"phone":i['Phno']})
            if data:
                if data["phone"] == i['Phno']:
                    pass
            else:
                #print(i['Phno'],"new numbers")
                ls1.append(i['Phno'])
                coll.insert_one({
                    "phone":i['Phno'],
                    "nickname":i['Nickname'],
                    'useremail':email,
                    "isactive": None,
                      "last_name": None,
                      "first_name": None,
                      "address": None,
                      "city": None,
                      "state": None,
                      "country": None,
                      "pin": None,
                      "remark": None,
                      "checkflag": None,
                      "imeinumber": None,
                      "as_on_datetime": now,
                      "as_on_date": date,
                      "as_on_time": time,
                      "inc_officer": None,
                      "module_name": None,
                      "monit_status": None,
                      "category": None,
                      "organization": None,
                    "module_code": None
                    })
        return jsonify({"message" :  ls1})
    else:
        return jsonify("OK")


@tickect_bp.route("/Data_search", methods=["POST"])
def search():
    coll = db['tickets']
    request_data = request.json
    value = request_data.get("value")
    value = value.strip()
    email = request_data.get("email")
    type = request_data.get("type")

    keys = ['useremail', 'superior', 'date', 'officername', 'type', 'role', 'designation', 'others', 'requestcategory', 'requesttype', 'requesttypes', 'subtypes', 'team', 
    'modules', 'relation', 'nickname', 'suspect', 'name', 'location', 'source', 'reason', 'refno', 'refdate', 'date1', 'date2', 'time', 'totimepicker', 'remarks', 
    'token', 'result', 'pending', 'approval', 'Comments', 'Signature', 'assign_Officer', 'edit', 'newnumber', 'priority', 'status', 'status_count', 'raise_time', 'Analyst_Approval', 
    'Send_to_INS', 'SP_Approval', 'DSP_Approval', 'Send_to_SP', 'SP_Reject', 'DSP_Reject', 'INS_Reject', 'INSPR_Approval', 'Send_to_DSP', 'proforma_data']

    search_fields = ['Phno','Nickname','Latitude','Longitude','IMEI','ISP','CGI','IP_Address','Area',
                     'Tower_name','RH_Dealer','RH_code','POA_Dealer','POA_code','Cell_id','Rad','CAF','Provider','From_Date','From_Time','To_Date','To_Time']
    data = []
    ids = set()
    # #print(keys)
    projection = {
            '_id' : 1,
            'priority' : 1,
            'useremail':1,
            'date' : 1,
            'token' : 1,
            'approval' : 1,
            'pending' : 1,
            'requestcategory' : 1,
            'requesttype' : 1,
            'subtypes' : 1,
        }
    
    if type == 'My_tickets':
        for res in coll.find({"$or": [{"useremail": email}]},projection).sort([('_id', -1)]):
            for res2 in res.get('newnumber', []):
                for field in search_fields:
                    if value == res2.get(field) and res['_id'] not in ids:
                      data.append(res)
                      ids.add(res['_id'])

        for key in keys:
            for res in coll.find({"useremail": email,key: re.compile(value)},projection).sort([('_id', -1)]):
                if res['_id'] not in ids:
                    data.append(res)
                    ids.add(res['_id'])
    pagination_info = {
                'total_pages': 1,
                'current_page': 1,
                'total_tickets': len(data)
            }
    # #print(data)
    return jsonify({'mytickets': data,'pagination': pagination_info})

@tickect_bp.route("/Data_search2", methods=["POST"])
def search2():
    coll = db['tickets']
    request_data = request.json
    value = request_data.get("value")
    value = value.strip()
    type = request.json.get('type')
    email = request.json.get('email')
    officer = request.json.get('officer')
    designation = request_data.get("designation")
    role = request_data.get("role")
    team = request_data.get("team")
    modules = request_data.get("modules")
    #print(value)
    keys = ['useremail', 'superior', 'date', 'officername', 'type', 'role', 'designation', 'others', 'requestcategory', 'requesttype', 'requesttypes', 'subtypes', 'team', 
    'modules', 'relation', 'nickname', 'suspect', 'name', 'location', 'source', 'reason', 'refno', 'refdate', 'date1', 'date2', 'time', 'totimepicker', 'remarks', 
    'token', 'result', 'pending', 'approval', 'Comments', 'Signature', 'assign_Officer', 'edit', 'newnumber', 'priority', 'status', 'status_count', 'raise_time', 'Analyst_Approval', 
    'Send_to_INS', 'SP_Approval', 'DSP_Approval', 'Send_to_SP', 'SP_Reject', 'DSP_Reject', 'INS_Reject', 'INSPR_Approval', 'Send_to_DSP', 'proforma_data']

    search_fields = ['Phno','Nickname','Latitude','Longitude','IMEI','ISP','CGI','IP_Address','Area',
                     'Tower_name','RH_Dealer','RH_code','POA_Dealer','POA_code','Cell_id','Rad','CAF','Provider','From_Date','From_Time','To_Date','To_Time']
    data = []
    # #print(keys)
    projection = {
            '_id' : 1,
            'priority' : 1,
            'useremail':1,
            'date' : 1,
            'token' : 1,
            'approval' : 1,
            'pending' : 1,
            'requestcategory' : 1,
            'requesttype' : 1,
            'requesttypes' : 1,
            'subtypes' : 1,
            'status_count': 1,
            'result': 1,
            'newnumber' : 1,
            'officername' : 1,
            'edit' : 1,


        }
    
    res_len = 0
    newno_len = 0
    total_len = 0
    id = set()
    if type == 'Others_Tickets':
        if team == 'ADMIN':
            for res in coll.find({'useremail': {'$nin' : [email]}},projection).sort([('_id', -1)]):
                for res2 in res.get('newnumber', []):
                    for field in search_fields:
                        if value == res2.get(field):
                            
                            if res['pending'] == 'Mail Under Process' and (res['subtypes'] == 'Phone' or res['requesttype'] == 'CDR'):
                                res_len = len(res['result'])
                                total_len = res_len - int(res['status_count'])
                                
                            elif res['pending'] == 'Ticket_Closed' and (res['subtypes'] == 'Phone' or res['requesttype'] == 'CDR'):
                                res_len = len(res['result'])
                                total_len = res_len
                                
                            elif res['pending'] == 'Mail Under Process' and (res['subtypes'] == 'IMEI' or res['requesttype'] == 'IMEI CDR'):
                                newno_len = len(res['newnumber'])
                                total_len = newno_len - int(res['status_count'])
                                
                            elif res['pending'] == 'Ticket_Closed' and (res['subtypes'] == 'IMEI' or res['requesttype'] == 'IMEI CDR'):
                                newno_len = len(res['newnumber'])
                                total_len = newno_len
                                
                            if res['_id'] not in id:
                                data.append({
                                    '_id': res['_id'],
                                    'priority': res['priority'],
                                    'useremail':res['useremail'],
                                    'date': res['date'],
                                    'token': res['token'],
                                    'approval': res['approval'],
                                    'pending': res['pending'],
                                    'requestcategory': res['requestcategory'],
                                    'requesttype': res['requesttype'],
                                    'subtypes': res['subtypes'],
                                    'status_count': res['status_count'],
                                    'total_len': total_len,
                                    "res_len": res_len,
                                    "newno_len": newno_len,
                                    'officername': res['officername'],
                                    'edit': res['edit'],
                                })
                            id.add(res['_id'])
            
            for key in keys:
                for res in coll.find({"$or": [{'useremail': {'$nin' : [email]}}],key: re.compile(value)},projection).sort([('_id', -1)]):
                    if res['pending'] == 'Mail Under Process' and (res['subtypes'] == 'Phone' or res['requesttype'] == 'CDR'):
                        res_len = len(res['result'])
                        total_len = res_len - int(res['status_count'])
                        
                    elif res['pending'] == 'Ticket_Closed' and (res['subtypes'] == 'Phone' or res['requesttype'] == 'CDR'):
                        res_len = len(res['result'])
                        total_len = res_len
                        
                    elif res['pending'] == 'Mail Under Process' and (res['subtypes'] == 'IMEI' or res['requesttype'] == 'IMEI CDR'):
                        newno_len = len(res['newnumber'])
                        total_len = newno_len - int(res['status_count'])
                        
                    elif res['pending'] == 'Ticket_Closed' and (res['subtypes'] == 'IMEI' or res['requesttype'] == 'IMEI CDR'):
                        newno_len = len(res['newnumber'])
                        total_len = newno_len
                        
                    if res['_id'] not in id:
                        data.append({
                            '_id': res['_id'],
                            'priority': res['priority'],
                            'useremail':res['useremail'],
                            'date': res['date'],
                            'token': res['token'],
                            'approval': res['approval'],
                            'pending': res['pending'],
                            'requestcategory': res['requestcategory'],
                            'requesttype': res['requesttype'],
                            'subtypes': res['subtypes'],
                            'status_count': res['status_count'],
                            'total_len': total_len,
                            "res_len": res_len,
                            "newno_len": newno_len,
                            'officername': res['officername'],
                            'edit': res['edit'],
                        })
                    id.add(res['_id'])
        elif designation == 'SP' and team != 'ADMIN':
            if team == 'CAT':
                query = {'$and' : [
                    {'useremail': {'$nin': [email]}},
                    {'team': team},
                    {"$or": [
                        {"pending": "ADDL-SP/DSP Approval Pending", "team": team},
                        {"$and": [{"pending": {"$regex": officer}}, {"team": team}]},
                        {"pending": "Mail Under Process", "token": {"$regex": "_CO"}},
                        {"approval": "Approved by CAT DSP", "token": {"$regex": "_CO"}},
                        {"pending": "--", "requestcategory": "Analysis Request"},
                        {"approval": {"$regex": "^Assigned"}, "requestcategory": "Analysis Request"},
                        {"pending": "SP Approval Pending", "requestcategory": "Analysis Request"},
                        {"pending": "SP Approval Pending", "team": team},
                        {"pending": "Mail Under Process", "team": team},
                        {"pending": "Ticket_Closed", "team": team},
                        {"pending": {"$regex": "^Rejected"}, "team": team}
                    ]}]}
            else:
                query = {'$and' : [
                    {'useremail': {'$nin': [email]}},
                    {'team': team},
                    {"$or": [
                        {"pending": "SP Approval Pending", "team": team,},
                        {"pending": "SP Approval Pending", "team": team,},
                        {"pending": "SP Approval Pending", "team": team},
                        {"approval": "Approved by SP", "team": team,},
                        {"approval": "Approved by SP", "team": team,},
                        {"token": {"$regex": "_CO"}, "team": team},
                        {"pending": {"$regex": officer}, "team": team},
                        {"pending": "Assign to CAT Ins", "requestcategory": "Analysis Request", "team": team,},
                        {"pending": "SP Approval Pending", "requestcategory": "Analysis Request", "team": team,},
                        {"pending": "Assign to CAT Ins", "requestcategory": "Analysis Request", "team": team,},
                        {"pending": "Assign to CAT Ins", "requestcategory": "Analysis Request", "team": team,},
                        {"pending": "SP Approval Pending", "requestcategory": "Analysis Request", "team": team,},
                        {"pending": "SP Approval Pending", "requestcategory": "Analysis Request", "team": team,},
                        {"approval": {"$regex": "^Assigned"}, "requestcategory": "Analysis Request", "team": team},
                        {"pending": "Mail Under Process", "team": team},
                        {"pending": {"$regex": "^Rejected"}, "team": team}
                    ]}]}
            for res in coll.find(query,projection):
                for res2 in res.get('newnumber', []):
                    for field in search_fields:
                        if value == res2.get(field):
                            if res['pending'] == 'Mail Under Process' and (res['subtypes'] == 'Phone' or res['requesttype'] == 'CDR'):
                                res_len = len(res['result'])
                                total_len = res_len - int(res['status_count'])
                                
                            elif res['pending'] == 'Ticket_Closed' and (res['subtypes'] == 'Phone' or res['requesttype'] == 'CDR'):
                                res_len = len(res['result'])
                                total_len = res_len
                                
                            elif res['pending'] == 'Mail Under Process' and (res['subtypes'] == 'IMEI' or res['requesttype'] == 'IMEI CDR'):
                                newno_len = len(res['newnumber'])
                                total_len = newno_len - int(res['status_count'])
                                
                            elif res['pending'] == 'Ticket_Closed' and (res['subtypes'] == 'IMEI' or res['requesttype'] == 'IMEI CDR'):
                                newno_len = len(res['newnumber'])
                                total_len = newno_len
                                
                            if res['_id'] not in id:
                                data.append({
                                    '_id': res['_id'],
                                    'priority': res['priority'],
                                    'useremail':res['useremail'],
                                    'date': res['date'],
                                    'token': res['token'],
                                    'approval': res['approval'],
                                    'pending': res['pending'],
                                    'requestcategory': res['requestcategory'],
                                    'requesttype': res['requesttype'],
                                    'subtypes': res['subtypes'],
                                    'status_count': res['status_count'],
                                    'total_len': total_len,
                                    "res_len": res_len,
                                    "newno_len": newno_len,
                                    'officername': res['officername'],
                                    'edit': res['edit'],
                                })
                            id.add(res['_id'])
            
            for key in keys:
                if team == 'CAT':
                    query = {'$and' : [
                        {'useremail': {'$nin': [email]}},
                        {key: re.compile(value)},
                        {'team': team},
                        {"$or": [
                            {"pending": "ADDL-SP/DSP Approval Pending", "team": team},
                            {"$and": [{"pending": {"$regex": officer}}, {"team": team}]},
                            {"pending": "Mail Under Process", "token": {"$regex": "_CO"}},
                            {"approval": "Approved by CAT DSP", "token": {"$regex": "_CO"}},
                            {"pending": "--", "requestcategory": "Analysis Request"},
                            {"approval": {"$regex": "^Assigned"}, "requestcategory": "Analysis Request"},
                            {"pending": "SP Approval Pending", "requestcategory": "Analysis Request"},
                            {"pending": "SP Approval Pending", "team": team},
                            {"pending": "Mail Under Process", "team": team},
                            {"pending": "Ticket_Closed", "team": team},
                            {"pending": {"$regex": "^Rejected"}, "team": team}
                        ]}]}
                else:
                    query = {'$and' : [
                        {'useremail': {'$nin': [email]}},
                        {key: re.compile(value)},
                        {'team': team},
                        {"$or": [
                            {"pending": "SP Approval Pending", "team": team,},
                            {"pending": "SP Approval Pending", "team": team,},
                            {"pending": "SP Approval Pending", "team": team},
                            {"approval": "Approved by SP", "team": team,},
                            {"approval": "Approved by SP", "team": team,},
                            {"token": {"$regex": "_CO"}, "team": team},
                            {"pending": {"$regex": officer}, "team": team},
                            {"pending": "Assign to CAT Ins", "requestcategory": "Analysis Request", "team": team,},
                            {"pending": "SP Approval Pending", "requestcategory": "Analysis Request", "team": team,},
                            {"pending": "Assign to CAT Ins", "requestcategory": "Analysis Request", "team": team,},
                            {"pending": "Assign to CAT Ins", "requestcategory": "Analysis Request", "team": team,},
                            {"pending": "SP Approval Pending", "requestcategory": "Analysis Request", "team": team,},
                            {"pending": "SP Approval Pending", "requestcategory": "Analysis Request", "team": team,},
                            {"approval": {"$regex": "^Assigned"}, "requestcategory": "Analysis Request", "team": team},
                            {"pending": "Mail Under Process", "team": team},
                            {"pending": {"$regex": "^Rejected"}, "team": team}
                        ]}]}
                for res in coll.find(query,projection).sort([('_id', -1)]):
                    if res['pending'] == 'Mail Under Process' and (res['subtypes'] == 'Phone' or res['requesttype'] == 'CDR'):
                        res_len = len(res['result'])
                        total_len = res_len - int(res['status_count'])
                        
                    elif res['pending'] == 'Ticket_Closed' and (res['subtypes'] == 'Phone' or res['requesttype'] == 'CDR'):
                        res_len = len(res['result'])
                        total_len = res_len
                        
                    elif res['pending'] == 'Mail Under Process' and (res['subtypes'] == 'IMEI' or res['requesttype'] == 'IMEI CDR'):
                        newno_len = len(res['newnumber'])
                        total_len = newno_len - int(res['status_count'])
                        
                    elif res['pending'] == 'Ticket_Closed' and (res['subtypes'] == 'IMEI' or res['requesttype'] == 'IMEI CDR'):
                        newno_len = len(res['newnumber'])
                        total_len = newno_len
                        
                    
                    if res['_id'] not in id:
                        data.append({
                            '_id': res['_id'],
                            'priority': res['priority'],
                            'useremail':res['useremail'],
                            'date': res['date'],
                            'token': res['token'],
                            'approval': res['approval'],
                            'pending': res['pending'],
                            'requestcategory': res['requestcategory'],
                            'requesttype': res['requesttype'],
                            'subtypes': res['subtypes'],
                            'status_count': res['status_count'],
                            'total_len': total_len,
                            "res_len": res_len,
                            "newno_len": newno_len,
                            'officername': res['officername'],
                            'edit': res['edit'],
                        })
                    id.add(res['_id'])
        elif designation == 'ADDL-SP/DSP':
            if team == "CAT":
                query = {'$and' : [
                    {'useremail': {'$nin': [email]}},
                    {"$or": [
                        {"pending": "Mail Under Process", "token": {"$regex": "_CO"}, "designation": {"$ne": "SP"}},
                        {"$and": [{"pending": {"$regex": officer}}, {"team": team}]},
                        {"pending": "Ticket_Closed", "team": team, "designation": {"$ne": "SP"}},
                        {"approval": "Approved by CAT Ins", "token": {"$regex": "_CO"}, "designation": {"$ne": "SP"}},
                        {"pending": "ADDL-SP/DSP Approval Pending", "requestcategory": "Analysis Request"},
                        {"$and": [{"pending": "--"}, {"requestcategory": "Analysis Request"}, {"designation": {"$ne": "SP"}}]},
                        {"$and": [{"pending": "ADDL-SP/DSP Approval Pending"}, {"team": team}]},
                        {"approval": {"$regex": "^Assigned"}, "requestcategory": "Analysis Request", "designation": {"$ne": "SP"}},
                        {"approval": {"$regex": "^Assigned"}, "requestcategory": "Analysis Request", "designation": {"$ne": "SP"}},
                        {"pending": "SP Approval Pending", "team": team, "designation": {"$ne": "ADDL-SP/DSP"}},
                        {"pending": "Mail Under Process", "team": team, "designation": {"$ne": "SP"}},
                        {"pending": {"$regex": "^Rejected"}, "team": team}
                        ]}
                ]
                    }
            else:
                query = {'$and' : [
                    {'useremail': {'$nin': [email]}},
                    {"$or": [
                        {"$and": [{"pending": {"$regex": officer}}, {"team": team}]},
                        {"$and": [{"token": {"$regex": "_CO"}}, {"team": team}, {"others": {"$in": [officer]}}]},
                        {"pending": "ADDL-SP/DSP Approval Pending", "team": team, "requestcategory": "Data Request", "superior": officer},
                        {"pending": "ADDL-SP/DSP Approval Pending", "team": team, "requestcategory": "Data Request", "others": officer},
                        {"pending": "SP Approval Pending", "team": team, "others": {"$in": [officer]}},
                        {"pending": "SP Approval Pending", "team": team, "superior": officer},
                        {"pending": "SP Approval Pending", "team": team, "others": officer},
                        {"pending": {"$regex": "Rejected"}, "team": team, "superior": officer},
                        {"pending": {"$regex": "Rejected"}, "team": team, "others": officer},
                        {"pending": "ADDL-SP/DSP Approval Pending", "requestcategory": "Analysis Request", "team": team, "others": {"$in": [officer]}},
                        {"pending": "ADDL-SP/DSP Approval Pending", "requestcategory": "Analysis Request", "team": team, "superior": officer},
                        {"pending": "ADDL-SP/DSP Approval Pending", "requestcategory": "Analysis Request", "team": team, "others": officer},
                        {"pending": "Assign to CAT Ins", "requestcategory": "Analysis Request", "team": team, "designation": {"$nin": ["SP"]}, "others": {"$in": [officer]}},
                        {"pending": "Assign to CAT Ins", "requestcategory": "Analysis Request", "team": team, "designation": {"$nin": ["SP"]}, "superior": officer},
                        {"pending": "Assign to CAT Ins", "requestcategory": "Analysis Request", "team": team, "designation": {"$nin": ["SP"]}, "others": officer},
                        {"pending": "Mail Under Process", "team": team, "designation": {"$ne": "SP"}, "others": officer},
                        {"pending": "Mail Under Process", "team": team, "designation": {"$ne": "SP"}, "superior": officer},
                        {"approval": {"$regex": "Assigned"}, "requestcategory": "Analysis Request", "team": team, "designation": {"$nin": ["SP"]}, "others": {"$in": [officer]}},
                        {"approval": {"$regex": "Assigned"}, "requestcategory": "Analysis Request", "team": team, "designation": {"$nin": ["SP"]}, "superior": officer},
                        {"approval": {"$regex": "Assigned"}, "requestcategory": "Analysis Request", "team": team, "designation": {"$nin": ["SP"]}, "others": officer}
                        ]}
                ]}
            for res in coll.find(query,projection):
                for res2 in res.get('newnumber', []):
                    for field in search_fields:
                        if value == res2.get(field):
                            if res['pending'] == 'Mail Under Process' and (res['subtypes'] == 'Phone' or res['requesttype'] == 'CDR'):
                                res_len = len(res['result'])
                                total_len = res_len - int(res['status_count'])
                                
                            elif res['pending'] == 'Ticket_Closed' and (res['subtypes'] == 'Phone' or res['requesttype'] == 'CDR'):
                                res_len = len(res['result'])
                                total_len = res_len
                                
                            elif res['pending'] == 'Mail Under Process' and (res['subtypes'] == 'IMEI' or res['requesttype'] == 'IMEI CDR'):
                                newno_len = len(res['newnumber'])
                                total_len = newno_len - int(res['status_count'])
                                
                            elif res['pending'] == 'Ticket_Closed' and (res['subtypes'] == 'IMEI' or res['requesttype'] == 'IMEI CDR'):
                                newno_len = len(res['newnumber'])
                                total_len = newno_len
                                
                            if res['_id'] not in id:
                                data.append({
                                    '_id': res['_id'],
                                    'priority': res['priority'],
                                    'useremail':res['useremail'],
                                    'date': res['date'],
                                    'token': res['token'],
                                    'approval': res['approval'],
                                    'pending': res['pending'],
                                    'requestcategory': res['requestcategory'],
                                    'requesttype': res['requesttype'],
                                    'subtypes': res['subtypes'],
                                    'status_count': res['status_count'],
                                    'total_len': total_len,
                                    "res_len": res_len,
                                    "newno_len": newno_len,
                                    'officername': res['officername'],
                                    'edit': res['edit'],
                                })
                            id.add(res['_id'])

            
            for key in keys:
                if team == "CAT":
                    query = {'$and' : [
                        {'useremail': {'$nin': [email]}},
                        {key: re.compile(value)},
                        {"$or": [
                            {"pending": "Mail Under Process", "token": {"$regex": "_CO"}, "designation": {"$ne": "SP"}},
                            {"$and": [{"pending": {"$regex": officer}}, {"team": team}]},
                            {"pending": "Ticket_Closed", "team": team, "designation": {"$ne": "SP"}},
                            {"approval": "Approved by CAT Ins", "token": {"$regex": "_CO"}, "designation": {"$ne": "SP"}},
                            {"pending": "ADDL-SP/DSP Approval Pending", "requestcategory": "Analysis Request"},
                            {"$and": [{"pending": "--"}, {"requestcategory": "Analysis Request"}, {"designation": {"$ne": "SP"}}]},
                            {"$and": [{"pending": "ADDL-SP/DSP Approval Pending"}, {"team": team}]},
                            {"approval": {"$regex": "^Assigned"}, "requestcategory": "Analysis Request", "designation": {"$ne": "SP"}},
                            {"approval": {"$regex": "^Assigned"}, "requestcategory": "Analysis Request", "designation": {"$ne": "SP"}},
                            {"pending": "SP Approval Pending", "team": team, "designation": {"$ne": "ADDL-SP/DSP"}},
                            {"pending": "Mail Under Process", "team": team, "designation": {"$ne": "SP"}},
                            {"pending": {"$regex": "^Rejected"}, "team": team}
                            ]}
                    ]
                        }
                else:
                    query = {'$and' : [
                        {'useremail': {'$nin': [email]}},
                        {key: re.compile(value)},
                        {"$or": [
                            {"$and": [{"pending": {"$regex": officer}}, {"team": team}]},
                            {"$and": [{"token": {"$regex": "_CO"}}, {"team": team}, {"others": {"$in": [officer]}}]},
                            {"pending": "ADDL-SP/DSP Approval Pending", "team": team, "requestcategory": "Data Request", "superior": officer},
                            {"pending": "ADDL-SP/DSP Approval Pending", "team": team, "requestcategory": "Data Request", "others": officer},
                            {"pending": "SP Approval Pending", "team": team, "others": {"$in": [officer]}},
                            {"pending": "SP Approval Pending", "team": team, "superior": officer},
                            {"pending": "SP Approval Pending", "team": team, "others": officer},
                            {"pending": {"$regex": "Rejected"}, "team": team, "superior": officer},
                            {"pending": {"$regex": "Rejected"}, "team": team, "others": officer},
                            {"pending": "ADDL-SP/DSP Approval Pending", "requestcategory": "Analysis Request", "team": team, "others": {"$in": [officer]}},
                            {"pending": "ADDL-SP/DSP Approval Pending", "requestcategory": "Analysis Request", "team": team, "superior": officer},
                            {"pending": "ADDL-SP/DSP Approval Pending", "requestcategory": "Analysis Request", "team": team, "others": officer},
                            {"pending": "Assign to CAT Ins", "requestcategory": "Analysis Request", "team": team, "designation": {"$nin": ["SP"]}, "others": {"$in": [officer]}},
                            {"pending": "Assign to CAT Ins", "requestcategory": "Analysis Request", "team": team, "designation": {"$nin": ["SP"]}, "superior": officer},
                            {"pending": "Assign to CAT Ins", "requestcategory": "Analysis Request", "team": team, "designation": {"$nin": ["SP"]}, "others": officer},
                            {"pending": "Mail Under Process", "team": team, "designation": {"$ne": "SP"}, "others": officer},
                            {"pending": "Mail Under Process", "team": team, "designation": {"$ne": "SP"}, "superior": officer},
                            {"approval": {"$regex": "Assigned"}, "requestcategory": "Analysis Request", "team": team, "designation": {"$nin": ["SP"]}, "others": {"$in": [officer]}},
                            {"approval": {"$regex": "Assigned"}, "requestcategory": "Analysis Request", "team": team, "designation": {"$nin": ["SP"]}, "superior": officer},
                            {"approval": {"$regex": "Assigned"}, "requestcategory": "Analysis Request", "team": team, "designation": {"$nin": ["SP"]}, "others": officer}
                            ]}
                    ]}
                for res in coll.find(query,projection).sort([('_id', -1)]):
                    if res['pending'] == 'Mail Under Process' and (res['subtypes'] == 'Phone' or res['requesttype'] == 'CDR'):
                        res_len = len(res['result'])
                        total_len = res_len - int(res['status_count'])
                        
                    elif res['pending'] == 'Ticket_Closed' and (res['subtypes'] == 'Phone' or res['requesttype'] == 'CDR'):
                        res_len = len(res['result'])
                        total_len = res_len
                        
                    elif res['pending'] == 'Mail Under Process' and (res['subtypes'] == 'IMEI' or res['requesttype'] == 'IMEI CDR'):
                        newno_len = len(res['newnumber'])
                        total_len = newno_len - int(res['status_count'])
                        
                    elif res['pending'] == 'Ticket_Closed' and (res['subtypes'] == 'IMEI' or res['requesttype'] == 'IMEI CDR'):
                        newno_len = len(res['newnumber'])
                        total_len = newno_len
                        
                    if res['_id'] not in id:
                        data.append({
                            '_id': res['_id'],
                            'priority': res['priority'],
                            'useremail':res['useremail'],
                            'date': res['date'],
                            'token': res['token'],
                            'approval': res['approval'],
                            'pending': res['pending'],
                            'requestcategory': res['requestcategory'],
                            'requesttype': res['requesttype'],
                            'subtypes': res['subtypes'],
                            'status_count': res['status_count'],
                            'total_len': total_len,
                            "res_len": res_len,
                            "newno_len": newno_len,
                            'officername': res['officername'],
                            'edit': res['edit'],
                        })
                    id.add(res['_id'])
        elif designation == 'INSPR' and role != 'Analyst':
            if team == 'CAT':
                query = {'$and' : [
                {'useremail': {'$nin': [email]}},
                {"$or": [
                    {"pending": "Ins Approval Pending", "team": team, 'modules':modules},
                    {"pending": "ADDL-SP/DSP Approval Pending", "team": team, 'modules':modules,"designation": {"$nin": ["SP", "DSP", "ADDL-SP/DSP"]}},
                    {"pending": "ADDL-SP/DSP Approval Pending", 'modules':modules,"token": {"$regex": "_CO"}, "designation": {"$nin": ["SP", "DSP", "ADDL-SP/DSP"]}},
                    {"pending": {"$regex": officer}, "team": team,'modules':modules},
                    {"pending": "Ticket_Closed", "team": team,'modules':modules, "designation": {"$nin": ["SP", "ADDL-SP/DSP"]}},
                    {"pending": "Mail Under Process", "token": {"$regex": "_CO"},'modules':modules, "designation": {"$nin": ["SP", "DSP", "ADDL-SP/DSP"]}},
                    {"pending": "Annexe Send to CAT INS", "token": {"$regex": "_CO"},'modules':modules, "designation": {"$nin": ["SP", "DSP", "ADDL-SP/DSP"]}},
                    {"pending": "Assign to CAT Ins", "requestcategory": "Analysis Request",'modules':modules, "designation": {"$nin": ["SP", "DSP", "ADDL-SP/DSP"]}},
                    {"team": {"$ne": "CAT"}, "approval": {"$regex": "^Assigned"},'modules':modules, "requestcategory": "Analysis Request", "designation": {"$nin": ["SP", "DSP", "ADDL-SP/DSP"]}},
                    {"approval": {"$regex": "^Assigned"},'modules':modules, "requestcategory": "Analysis Request", "team": team, "designation": {"$nin": ["SP", "DSP", "ADDL-SP/DSP"]}},
                    {"pending": "Mail Under Process",'modules':modules, "team": team, "designation": {"$nin": ["SP", "DSP", "ADDL-SP/DSP"]}},
                    {"pending": {"$regex": "^Rejected"}, "team": team, 'modules':modules,"designation": {"$nin": ["SP", "DSP", "ADDL-SP/DSP"]}}
                ]
                }
                ]}
            else:
                query = {'$and' : [
                {'useremail': {'$nin': [email]}},{
                    "$or": [
                        # DATA REQ
                        {"pending": "Ins Approval Pending", "team": team,'modules':modules, "others": {"$in": [officer, '']},"designation": {"$nin": ["SP", "DSP", "ADDL-SP/DSP"]}},
                        {"pending": "ADDL-SP/DSP Approval Pending",'modules':modules, "team": team, "others": {"$in": [officer, '']},"designation": {"$nin": ["SP", "DSP", "ADDL-SP/DSP"]}},
                        {"pending": {"$regex": officer}, "team": team,'modules':modules,"designation": {"$nin": ["SP", "DSP", "ADDL-SP/DSP"]}},
                        {"pending": "Annexe Send to CAT INS", "team": team, 'modules':modules,"others": {"$in": [officer, '']},"designation": {"$nin": ["SP", "DSP", "ADDL-SP/DSP"]}},
                        {"pending": {"$regex": "^Rejected"}, "team": team, 'modules':modules,"others": {"$in": [officer, '']},"designation": {"$nin": ["SP", "DSP", "ADDL-SP/DSP"]}},
                        {"pending": "Mail Under Process", "team": team,'modules':modules, "others": {"$in": [officer, '']},"designation": {"$nin": ["SP", "DSP", "ADDL-SP/DSP"]}},
                        {"pending": "SP Approval Pending", "team": team, 'modules':modules,"others": {"$in": [officer, '']},"designation": {"$nin": ["SP", "DSP", "ADDL-SP/DSP"]}},
                        # ANALYSIS REQ
                        {"pending": "Ins Approval Pending", 'modules':modules,"requestcategory": "Analysis Request", "team": team,"designation": {"$nin": ["SP", "DSP", "ADDL-SP/DSP"]}},
                        {"approval": {"$regex": "^Assigned"}, 'modules':modules,"requestcategory": "Analysis Request", "team": team,"designation": {"$nin": ["SP", "DSP", "ADDL-SP/DSP"]}},
                        {"pending": "ADDL-SP/DSP Approval Pending", 'modules':modules,"requestcategory": "Analysis Request", "team": team,"designation": {"$nin": ["SP", "DSP", "ADDL-SP/DSP"]}},
                        {"pending": "SP Approval Pending", 'modules':modules,"requestcategory": "Analysis Request", "team": team,"designation": {"$nin": ["SP", "DSP", "ADDL-SP/DSP"]}},
                        {"pending": "Assign to CAT Ins", 'modules':modules,"requestcategory": "Analysis Request", "team": team,"designation": {"$nin": ["SP", "DSP", "ADDL-SP/DSP"]}}
                    ]
                }]}
            for res in coll.find(query,projection):
                for res2 in res.get('newnumber', []):
                    for field in search_fields:
                        if value == res2.get(field):
                            if res['pending'] == 'Mail Under Process' and (res['subtypes'] == 'Phone' or res['requesttype'] == 'CDR'):
                                res_len = len(res['result'])
                                total_len = res_len - int(res['status_count'])
                                
                            elif res['pending'] == 'Ticket_Closed' and (res['subtypes'] == 'Phone' or res['requesttype'] == 'CDR'):
                                res_len = len(res['result'])
                                total_len = res_len
                                
                            elif res['pending'] == 'Mail Under Process' and (res['subtypes'] == 'IMEI' or res['requesttype'] == 'IMEI CDR'):
                                newno_len = len(res['newnumber'])
                                total_len = newno_len - int(res['status_count'])
                                
                            elif res['pending'] == 'Ticket_Closed' and (res['subtypes'] == 'IMEI' or res['requesttype'] == 'IMEI CDR'):
                                newno_len = len(res['newnumber'])
                                total_len = newno_len
                                
                            if res['_id'] not in id:
                                data.append({
                                    '_id': res['_id'],
                                    'priority': res['priority'],
                                    'useremail':res['useremail'],
                                    'date': res['date'],
                                    'token': res['token'],
                                    'approval': res['approval'],
                                    'pending': res['pending'],
                                    'requestcategory': res['requestcategory'],
                                    'requesttype': res['requesttype'],
                                    'subtypes': res['subtypes'],
                                    'status_count': res['status_count'],
                                    'total_len': total_len,
                                    "res_len": res_len,
                                    "newno_len": newno_len,
                                    'officername': res['officername'],
                                    'edit': res['edit'],
                                })
                            id.add(res['_id'])
            for key in keys:
                if team == 'CAT':
                    query = {'$and' : [
                    {'useremail': {'$nin': [email]},key: re.compile(value)},
                    {"$or": [
                        {"pending": "Ins Approval Pending", "team": team, 'modules':modules},
                        {"pending": "ADDL-SP/DSP Approval Pending", "team": team, 'modules':modules,"designation": {"$nin": ["SP", "DSP", "ADDL-SP/DSP"]}},
                        {"pending": "ADDL-SP/DSP Approval Pending", 'modules':modules,"token": {"$regex": "_CO"}, "designation": {"$nin": ["SP", "DSP", "ADDL-SP/DSP"]}},
                        {"pending": {"$regex": officer}, "team": team,'modules':modules},
                        {"pending": "Ticket_Closed", "team": team,'modules':modules, "designation": {"$nin": ["SP", "ADDL-SP/DSP"]}},
                        {"pending": "Mail Under Process", "token": {"$regex": "_CO"},'modules':modules, "designation": {"$nin": ["SP", "DSP", "ADDL-SP/DSP"]}},
                        {"pending": "Annexe Send to CAT INS", "token": {"$regex": "_CO"},'modules':modules, "designation": {"$nin": ["SP", "DSP", "ADDL-SP/DSP"]}},
                        {"pending": "Assign to CAT Ins", "requestcategory": "Analysis Request",'modules':modules, "designation": {"$nin": ["SP", "DSP", "ADDL-SP/DSP"]}},
                        {"team": {"$ne": "CAT"}, "approval": {"$regex": "^Assigned"},'modules':modules, "requestcategory": "Analysis Request", "designation": {"$nin": ["SP", "DSP", "ADDL-SP/DSP"]}},
                        {"approval": {"$regex": "^Assigned"},'modules':modules, "requestcategory": "Analysis Request", "team": team, "designation": {"$nin": ["SP", "DSP", "ADDL-SP/DSP"]}},
                        {"pending": "Mail Under Process",'modules':modules, "team": team, "designation": {"$nin": ["SP", "DSP", "ADDL-SP/DSP"]}},
                        {"pending": {"$regex": "^Rejected"}, "team": team, 'modules':modules,"designation": {"$nin": ["SP", "DSP", "ADDL-SP/DSP"]}}
                    ]
                }
                ]}
                else:
                    query = {'$and' : [
                    {'useremail': {'$nin': [email]},key: re.compile(value)},{
                        "$or": [
                            # DATA REQ
                            {"pending": "Ins Approval Pending", "team": team,'modules':modules, "others": {"$in": [officer, '']},"designation": {"$nin": ["SP", "DSP", "ADDL-SP/DSP"]}},
                            {"pending": "ADDL-SP/DSP Approval Pending",'modules':modules, "team": team, "others": {"$in": [officer, '']},"designation": {"$nin": ["SP", "DSP", "ADDL-SP/DSP"]}},
                            {"pending": {"$regex": officer}, "team": team,'modules':modules,"designation": {"$nin": ["SP", "DSP", "ADDL-SP/DSP"]}},
                            {"pending": "Annexe Send to CAT INS", "team": team, 'modules':modules,"others": {"$in": [officer, '']},"designation": {"$nin": ["SP", "DSP", "ADDL-SP/DSP"]}},
                            {"pending": {"$regex": "^Rejected"}, "team": team, 'modules':modules,"others": {"$in": [officer, '']},"designation": {"$nin": ["SP", "DSP", "ADDL-SP/DSP"]}},
                            {"pending": "Mail Under Process", "team": team,'modules':modules, "others": {"$in": [officer, '']},"designation": {"$nin": ["SP", "DSP", "ADDL-SP/DSP"]}},
                            {"pending": "SP Approval Pending", "team": team, 'modules':modules,"others": {"$in": [officer, '']},"designation": {"$nin": ["SP", "DSP", "ADDL-SP/DSP"]}},
                            # ANALYSIS REQ
                            {"pending": "Ins Approval Pending", 'modules':modules,"requestcategory": "Analysis Request", "team": team,"designation": {"$nin": ["SP", "DSP", "ADDL-SP/DSP"]}},
                            {"approval": {"$regex": "^Assigned"}, 'modules':modules,"requestcategory": "Analysis Request", "team": team,"designation": {"$nin": ["SP", "DSP", "ADDL-SP/DSP"]}},
                            {"pending": "ADDL-SP/DSP Approval Pending", 'modules':modules,"requestcategory": "Analysis Request", "team": team,"designation": {"$nin": ["SP", "DSP", "ADDL-SP/DSP"]}},
                            {"pending": "SP Approval Pending", 'modules':modules,"requestcategory": "Analysis Request", "team": team,"designation": {"$nin": ["SP", "DSP", "ADDL-SP/DSP"]}},
                            {"pending": "Assign to CAT Ins", 'modules':modules,"requestcategory": "Analysis Request", "team": team,"designation": {"$nin": ["SP", "DSP", "ADDL-SP/DSP"]}}
                        ]
                    }]}
                for res in coll.find(query,projection).sort([('_id', -1)]):
                    if res['pending'] == 'Mail Under Process' and (res['subtypes'] == 'Phone' or res['requesttype'] == 'CDR'):
                        res_len = len(res['result'])
                        total_len = res_len - int(res['status_count'])
                        
                    elif res['pending'] == 'Ticket_Closed' and (res['subtypes'] == 'Phone' or res['requesttype'] == 'CDR'):
                        res_len = len(res['result'])
                        total_len = res_len
                        
                    elif res['pending'] == 'Mail Under Process' and (res['subtypes'] == 'IMEI' or res['requesttype'] == 'IMEI CDR'):
                        newno_len = len(res['newnumber'])
                        total_len = newno_len - int(res['status_count'])
                        
                    elif res['pending'] == 'Ticket_Closed' and (res['subtypes'] == 'IMEI' or res['requesttype'] == 'IMEI CDR'):
                        newno_len = len(res['newnumber'])
                        total_len = newno_len
                        
                    if res['_id'] not in id:
                        data.append({
                            '_id': res['_id'],
                            'priority': res['priority'],
                            'useremail':res['useremail'],
                            'date': res['date'],
                            'token': res['token'],
                            'approval': res['approval'],
                            'pending': res['pending'],
                            'requestcategory': res['requestcategory'],
                            'requesttype': res['requesttype'],
                            'subtypes': res['subtypes'],
                            'status_count': res['status_count'],
                            'total_len': total_len,
                            "res_len": res_len,
                            "newno_len": newno_len,
                            'officername': res['officername'],
                            'edit': res['edit'],
                        })
                    id.add(res['_id'])
        elif role == 'Mailer':
            query = {"$or": [
                    {"requestcategory": "Data Request", 'pending' : {'$in': ['Mail Under Process','Ticket_Closed']}},
                ]
            }
            for res in coll.find(query,projection).sort([('_id', -1)]):
                for res2 in res.get('newnumber', []):
                    for field in search_fields:
                        if value == res2.get(field):
                            if res['pending'] == 'Mail Under Process' and (res['subtypes'] == 'Phone' or res['requesttype'] == 'CDR'):
                                res_len = len(res['result'])
                                total_len = res_len - int(res['status_count'])
                                
                            elif res['pending'] == 'Mail Under Process' and (res['subtypes'] == 'IMEI' or res['requesttype'] == 'IMEI CDR'):
                                newno_len = len(res['newnumber'])
                                total_len = newno_len - int(res['status_count'])
                                
                            if res['_id'] not in id:
                                data.append({
                                    '_id': res['_id'],
                                    'priority': res['priority'],
                                    'useremail':res['useremail'],
                                    'date': res['date'],
                                    'token': res['token'],
                                    'approval': res['approval'],
                                    'pending': res['pending'],
                                    'requestcategory': res['requestcategory'],
                                    'requesttype': res['requesttype'],
                                    'subtypes': res['subtypes'],
                                    'status_count': res['status_count'],
                                    'total_len': total_len,
                                    "res_len": res_len,
                                    "newno_len": newno_len,
                                    'officername': res['officername'],
                                    'edit': res['edit'],
                                })
                            id.add(res['_id'])
            for key in keys:
                query = {"$or": [
                    {"requestcategory": "Data Request", 'pending' : {'$in': ['Mail Under Process','Ticket_Closed']},key: re.compile(value)},
                ]
            }
                for res in coll.find(query,projection).sort([('_id', -1)]):
                    if res['pending'] == 'Mail Under Process' and (res['subtypes'] == 'Phone' or res['requesttype'] == 'CDR'):
                        res_len = len(res['result'])
                        total_len = res_len - int(res['status_count'])
            
                    elif res['pending'] == 'Mail Under Process' and (res['subtypes'] == 'IMEI' or res['requesttype'] == 'IMEI CDR'):
                        newno_len = len(res['newnumber'])
                        total_len = newno_len - int(res['status_count'])
                        
                    if res['_id'] not in id:
                        data.append({
                            '_id': res['_id'],
                            'priority': res['priority'],
                            'useremail':res['useremail'],
                            'date': res['date'],
                            'token': res['token'],
                            'approval': res['approval'],
                            'pending': res['pending'],
                            'requestcategory': res['requestcategory'],
                            'requesttype': res['requesttype'],
                            'subtypes': res['subtypes'],
                            'status_count': res['status_count'],
                            'total_len': total_len,
                            "res_len": res_len,
                            "newno_len": newno_len,
                            'officername': res['officername'],
                            'edit': res['edit'],
                        })
                    id.add(res['_id'])
        elif role == 'Analyst':
            query = {"$and": [
                    {'assign_Officer': {'$in':[officer]}}
                ]
            }
            for res in coll.find(query,projection):
                for res2 in res.get('newnumber', []):
                    for field in search_fields:
                        if value == res2.get(field):
                            if res['pending'] == 'Mail Under Process' and (res['subtypes'] == 'Phone' or res['requesttype'] == 'CDR'):
                                res_len = len(res['result'])
                                total_len = res_len - int(res['status_count'])
                                
                            elif res['pending'] == 'Ticket_Closed' and (res['subtypes'] == 'Phone' or res['requesttype'] == 'CDR'):
                                res_len = len(res['result'])
                                total_len = res_len
                                
                            elif res['pending'] == 'Mail Under Process' and (res['subtypes'] == 'IMEI' or res['requesttype'] == 'IMEI CDR'):
                                newno_len = len(res['newnumber'])
                                total_len = newno_len - int(res['status_count'])
                                
                            elif res['pending'] == 'Ticket_Closed' and (res['subtypes'] == 'IMEI' or res['requesttype'] == 'IMEI CDR'):
                                newno_len = len(res['newnumber'])
                                total_len = newno_len
                                
                            if res['_id'] not in id:
                                data.append({
                                    '_id': res['_id'],
                                    'priority': res['priority'],
                                    'useremail':res['useremail'],
                                    'date': res['date'],
                                    'token': res['token'],
                                    'approval': res['approval'],
                                    'pending': res['pending'],
                                    'requestcategory': res['requestcategory'],
                                    'requesttype': res['requesttype'],
                                    'subtypes': res['subtypes'],
                                    'status_count': res['status_count'],
                                    'total_len': total_len,
                                    "res_len": res_len,
                                    "newno_len": newno_len,
                                    'officername': res['officername'],
                                    'edit': res['edit'],
                                })
                            id.add(res['_id'])
            for key in keys:
                query = {"$and": [
                    {'assign_Officer': {'$in':[officer]},key: re.compile(value)}
                ]
            }
                for res in coll.find(query,projection).sort([('_id', -1)]):                       
                    if res['_id'] not in id:
                        data.append({
                            '_id': res['_id'],
                            'priority': res['priority'],
                            'useremail':res['useremail'],
                            'date': res['date'],
                            'token': res['token'],
                            'approval': res['approval'],
                            'pending': res['pending'],
                            'requestcategory': res['requestcategory'],
                            'requesttypes': res['requesttypes'],
                            'subtypes': res['subtypes'],
                            'officername': res['officername'],
                            'edit': res['edit'],
                        })
                    id.add(res['_id'])
        
        
    pagination_info = {
                'total_pages1': 1,
                'current_page1': 1,
                'total_tickets1': len(data)
            }
    return jsonify({'mytickets': data,'pagination': pagination_info})


@tickect_bp.route('/imei_convert', methods=['POST'])
def imei_convert():
    imei = request.json.get('imei')
    req = request.json.get('req')
    sub = request.json.get('sub')
    req2 = request.json.get('req2')
    sample = {'IMEI':''}
    if (req == "GPRS" and sub == "IMEI") or (req == "IPDR" and sub == "IMEI") or (req == 'IMEI CDR') or ('IMEI' in req2) or (sub == "IMEI") :
        if (len(imei) == 15 or len(imei) == 14):
            #print(req2,'IFFFFFFFFFFFFFFFF')
            imei = imei[:14] + '0'
            sample["IMEI"] = imei
            #print(sample)
        else:
            #print(req2,'ELSEEEEEEEEEEEEEEEEEEEEE')
            sample["IMEI"] = imei
    return jsonify(sample)

@tickect_bp.route('/raise_new_data', methods=['POST'])
def raise_new_data():
    id = request.json.get('id')
    #print(id)
    coll = db['tickets']
    data = []
    res = coll.find_one({'_id': ObjectId(id)})
    data.append(res)
    return jsonify(data)



@tickect_bp.route('/check_email', methods=['POST'])
def check_email():
    email = request.json.get('email')
    coll = db['users']
    check = coll.find_one({'email':email})
    
    if check:
        if check['email'] == email:
            #print('message1')
            return jsonify({'message1':'User Email Already Taken'})
    else:
        #print('message2')
        return jsonify({'message2':'This User Email Available for you!!!!!'})
    return jsonify('OKKK WORKING!!!!!!!!!!!!!!!')

@tickect_bp.route('/report2', methods=['POST','GET'])
@flask_cors.cross_origin(origin=cors_allowed_ip, supports_credentials=True)
@token_required
def report2(current_user):
    print("in reports2=================")
    team = request.json.get('team')
    module = request.json.get('module')
    designation = request.json.get('designation')
    role = request.json.get('role')
    email = request.json.get('email')
    print(team,'TEAM',designation,'DESIGN',role,'ROLE',email,'EMAILLLL',module, 'modules')
    coll = db['tickets']
    data = {}
    main_data = []
    if designation == 'SP' and team != 'ADMIN':
        data['Total_raise'] = len([res for res in coll.find({'team':team,"requestcategory":"Data Request"}, {'requestcategory': 1, 'team': 1})])
        data['Total_analysis_raise'] = len([res for res in coll.find({'team':team,"requestcategory":"Analysis Request"}, {'requestcategory': 1, 'team': 1})])
        # APPROVAL
        data['Total_approval'] = len([res for res in coll.find({'team':team,"requestcategory":"Data Request",'pending': 'Mail Under Process'}, {'requestcategory': 1, 'team': 1, 'pending': 1})])
        data['Total_analysis_approval'] = len([res for res in coll.find({'team':team,"requestcategory":"Analysis Request",'$or': [{'pending': '--'},{'pending': re.compile(r"Assign")}]}, {'requestcategory': 1, 'team': 1, 'pending': 1})])
        data['Total_close'] = len([res for res in coll.find({'team':team,"requestcategory":"Data Request",'pending': 'Ticket_Closed'}, {'requestcategory': 1, 'team': 1, 'pending': 1})])
        data['Total_under_process'] = len([res for res in coll.find({'team':team,"requestcategory":"Data Request",'$or': [{'result.status': 'Data_Received'}, {'result.status': 'Data Not Received'}, {'newnumber.status': 'Data_Received'}, {'newnumber.status': 'Data Not Received'}],'pending': 'Mail Under Process'}, {'requestcategory': 1, 'team': 1, 'result.status': 1, 'newnumber.status': 1, 'pending': 1})])
        # PENDING
        data['Total_pending'] = len([res for res in coll.find({'team':team,"requestcategory":"Data Request",'pending': re.compile(r"Pending")}, {'requestcategory': 1, 'team': 1, 'pending': 1})])
        data['Total_analysis_pending'] = len([res for res in coll.find({'team':team,"requestcategory":"Analysis Request",'pending': re.compile(r"Pending")}, {'requestcategory': 1, 'team': 1, 'pending': 1})])

    
    elif designation == 'ADDL-SP/DSP':
        data['Total_raise'] = len([res for res in coll.find({"requestcategory":"Data Request","team":team,"modules":module,'designation': {'$nin': ['SP','DIG','IG']}},{'requestcategory':1, 'team':1,'modules':1,'designation':1})])
        data['Total_analysis_raise'] = len([res for res in coll.find({"requestcategory":"Analysis Request","team":team,"modules":module,'designation': {'$nin': ['SP','DIG','IG']}},{'requestcategory':1, 'team':1,'modules':1,'designation':1})])
        # APPROVAAAAAAAAAAAAAAAAAAAL
        data['Total_approval'] = len([res for res in coll.find({"requestcategory":"Data Request",'pending': 'Mail Under Process',"team":team,"modules":module,'designation': {'$nin': ['SP','DIG','IG']}},{'requestcategory':1, 'team':1,'modules':1,'designation':1,'pending': 1})])

        data['Total_analysis_approval'] = len([res for res in coll.find({"requestcategory":"Analysis Request","team":team,"modules":module,'designation': {'$nin': ['SP','DIG','IG']},'$or': [{'pending': '--'},{'pending': re.compile(r"Assign")}]},{'requestcategory':1, 'team':1,'modules':1,'designation':1,'pending': 1})])

        data['Total_close'] = len([res for res in coll.find({"requestcategory":"Data Request",'pending': 'Ticket_Closed',"team":team,"modules":module,'designation': {'$nin': ['SP','DIG','IG']}},{'requestcategory':1, 'team':1,'modules':1,'designation':1,'pending': 1})])

        data['Total_under_process'] = len([res for res in coll.find({"requestcategory":"Data Request","team":team,"modules":module,'designation': {'$nin': ['SP','DIG','IG']},
                                                                     '$or': [{'result.status': 'Data_Received'}, {'result.status': 'Data Not Received'}, 
                                                                             {'newnumber.status': 'Data_Received'}, {'newnumber.status': 'Data Not Received'}],'pending': 'Mail Under Process'},{'requestcategory':1, 'team':1,'modules':1,'designation':1,'pending': 1,'result.status':1,'newnumber.status':1})])
        # PENDIIIIIIIIIIIIIIIIIIIIIIIIIIIIIING
        data['Total_pending'] = len([res for res in coll.find({"requestcategory":"Data Request",'pending': re.compile(r"Pending"),"team":team,"modules":module,'designation': {'$nin': ['SP','DIG','IG']}},{'requestcategory':1, 'team':1,'modules':1,'designation':1,'pending': 1})])
        data['Total_analysis_pending'] = len([res for res in coll.find({"requestcategory":"Analysis Request",'pending': re.compile(r"Pending"),"team":team,"modules":module,'designation': {'$nin': ['SP','DIG','IG']}},{'requestcategory':1, 'team':1,'modules':1,'designation':1,'pending': 1})])
    
    elif designation == 'INSPR' and role != 'Analyst':
        data['Total_raise'] = len([res for res in coll.find({"requestcategory":"Data Request","team":team,"modules":module,'designation': {'$nin': ['SP','DIG','IG','ADDL-SP/DSP']}}, {'requestcategory': 1, 'team': 1, 'modules': 1, 'designation': 1})])
        data['Total_analysis_raise'] = len([res for res in coll.find({"requestcategory":"Analysis Request","team":team,"modules":module,'designation': {'$nin': ['SP','DIG','IG','ADDL-SP/DSP']}}, {'requestcategory': 1, 'team': 1, 'modules': 1, 'designation': 1})])
        # APPROVAL
        data['Total_approval'] = len([res for res in coll.find({"requestcategory":"Data Request",'pending': 'Mail Under Process',"team":team,"modules":module,'designation': {'$nin': ['SP','DIG','IG','ADDL-SP/DSP']}}, {'requestcategory': 1, 'team': 1, 'modules': 1, 'designation': 1, 'pending': 1})])
        data['Total_analysis_approval'] = len([res for res in coll.find({"requestcategory":"Analysis Request","team":team,"modules":module,'designation': {'$nin': ['SP','DIG','IG','ADDL-SP/DSP']},'$or': [{'pending': '--'},{'pending': re.compile(r"Assign")}]}, {'requestcategory': 1, 'team': 1, 'modules': 1, 'designation': 1, 'pending': 1})])
        data['Total_close'] = len([res for res in coll.find({"requestcategory":"Data Request",'pending': 'Ticket_Closed' ,"team":team,"modules":module,'designation': {'$nin': ['SP','DIG','IG','ADDL-SP/DSP']}}, {'requestcategory': 1, 'team': 1, 'modules': 1, 'designation': 1, 'pending': 1})])
        data['Total_under_process'] = len([res for res in coll.find({"requestcategory":"Data Request","team":team,"modules":module,'designation': {'$nin': ['SP','DIG','IG','ADDL-SP/DSP']},'$or': [{'result.status': 'Data_Received'}, {'result.status': 'Data Not Received'}, {'newnumber.status': 'Data_Received'}, {'newnumber.status': 'Data Not Received'}],'pending': 'Mail Under Process'}, {'requestcategory': 1, 'team': 1, 'modules': 1, 'designation': 1, 'pending': 1, 'result.status': 1, 'newnumber.status': 1})])
        # PENDING
        data['Total_pending'] = len([res for res in coll.find({"requestcategory":"Data Request",'pending': re.compile(r"Pending"),"team":team,"modules":module,'designation': {'$nin': ['SP','DIG','IG','ADDL-SP/DSP']}}, {'requestcategory': 1, 'team': 1, 'modules': 1, 'designation': 1, 'pending': 1})])
        data['Total_analysis_pending'] = len([res for res in coll.find({"requestcategory":"Analysis Request",'pending': re.compile(r"Pending"),"team":team,"modules":module,'designation': {'$nin': ['SP','DIG','IG','ADDL-SP/DSP']}}, {'requestcategory': 1, 'team': 1, 'modules': 1, 'designation': 1, 'pending': 1})])


    elif team == 'ADMIN':
        # RAISEEEEEEEEEEEEEEEEEEEEE
        data['Total_raise'] = len([res for res in coll.find({"requestcategory":"Data Request"},{'requestcategory': 1})])
        data['Total_analysis_raise'] = len([res for res in coll.find({"requestcategory":"Analysis Request"},{'requestcategory': 1})])
        # APPROVAAAAAAAAAAAAAAAAAAAL
        data['Total_approval'] = len([res for res in coll.find({"requestcategory":"Data Request",'pending': 'Mail Under Process'},{'requestcategory': 1,'pending':1})])
        data['Total_close'] = len([res for res in coll.find({"requestcategory":"Data Request",'pending': 'Ticket_Closed'},{'requestcategory': 1,'pending':1})])
        data['Total_under_process'] = len([res for res in coll.find({"requestcategory":"Data Request",
                                                                     '$or': [{'result.status': 'Data_Received'}, {'result.status': 'Data Not Received'}, 
                                                                    {'newnumber.status': 'Data_Received'}, {'newnumber.status': 'Data Not Received'}],'pending': 'Mail Under Process'},{'requestcategory': 1,'pending': 1,'newnumber.status': 1,'result.status': 1})])
        data['Total_analysis_approval'] = len([res for res in coll.find({"requestcategory":"Analysis Request",'$or': [{'pending': '--'},{'pending': re.compile(r"Assign")}]},{'requestcategory': 1,'pending':1})])
        # PENDIIIIIIIIIIIIIIIIIIIIIIIIIIIIIING
        data['Total_pending'] = len([res for res in coll.find({"requestcategory":"Data Request",'pending': re.compile(r"Pending")},{'requestcategory': 1,'pending':1})])
        data['Total_analysis_pending'] = len([res for res in coll.find({"requestcategory":"Analysis Request",'pending': re.compile(r"Pending")},{'requestcategory': 1,'pending':1})])
    
    else:
        # RAISE
        data['Total_raise'] = len([res for res in coll.find({'useremail':email, "requestcategory":"Data Request"}, {'requestcategory': 1, 'useremail': 1})])
        data['Total_analysis_raise'] = len([res for res in coll.find({'useremail':email, "requestcategory":"Analysis Request"}, {'requestcategory': 1, 'useremail': 1})])
        # APPROVAL
        data['Total_approval'] = len([res for res in coll.find({'useremail':email, "requestcategory":"Data Request", 'pending': 'Mail Under Process'}, {'requestcategory': 1, 'useremail': 1, 'pending': 1})])
        data['Total_analysis_approval'] = len([res for res in coll.find({'useremail':email, "requestcategory":"Analysis Request", '$or': [{'pending': '--'},{'pending': re.compile(r"Assign")}]}, {'requestcategory': 1, 'useremail': 1, 'pending': 1})])
        data['Total_close'] = len([res for res in coll.find({'useremail':email, "requestcategory":"Data Request", 'pending': 'Ticket_Closed'}, {'requestcategory': 1, 'useremail': 1, 'pending': 1})])
        data['Total_under_process'] = len([res for res in coll.find({'useremail':email, "requestcategory":"Data Request", '$or': [{'result.status': 'Data_Received'}, {'result.status': 'Data Not Received'}, {'newnumber.status': 'Data_Received'}, {'newnumber.status': 'Data Not Received'}],'pending': 'Mail Under Process'}, {'requestcategory': 1, 'useremail': 1, 'result.status': 1, 'newnumber.status': 1, 'pending': 1})])
        # PENDING
        data['Total_pending'] = len([res for res in coll.find({'useremail':email, "requestcategory":"Data Request", 'pending': re.compile(r"Pending")}, {'requestcategory': 1, 'useremail': 1, 'pending': 1})])
        data['Total_analysis_pending'] = len([res for res in coll.find({'useremail':email, "requestcategory":"Analysis Request", 'pending': re.compile(r"Pending")}, {'requestcategory': 1, 'useremail': 1, 'pending': 1})])


    return jsonify(data)

@tickect_bp.route('/report', methods=['POST'])
def report():
    team = request.json.get('team')
    otherteam = request.json.get('teams')
    module = request.json.get('modules')
    types = request.json.get('types')
    category = request.json.get('category')
    #print(category)
    status = request.json.get('status')
    period = request.json.get('period')
    ISP = request.json.get('ISP')
    officer = request.json.get('officer')
    date_range = request.json.get('date_range')
    designation = request.json.get('designation')
    role = request.json.get('role')
    email = request.json.get('email')
    now = datetime.now()
    day = now.strftime("%d/%m/%Y")
    coll = db['tickets']
    coll2 = db['users']
    users = coll2.find_one({'email':email})
    officername = users['officername']

    main_data = []
    results = {'reqtypes' : ''}
    reqtypes = ['SDR','CAF','CDR','IMEI CDR','GPRS','IPDR','RH','POA','TOWER CDR','TOWER GPRS','TOWER IPDR']
    analysistypes = ["CAF","SDR","CDR","IMEI","GPRS","IPDR","RH","POA","TOWER DATA"]
    # if category in ['Data Request']:
    #     if team == 'ADMIN':
    #         req = set([ticket['requesttype'] for ticket in coll.find({'requestcategory': 'Data Request'}) if ticket['requesttype'] != ''])
    #     elif designation == 'SP' and team != 'ADMIN':
    #         req = set([ticket['requesttype'] for ticket in coll.find({'requestcategory': 'Data Request', 'team':team}) if ticket['requesttype'] != ''])
    #     elif designation == 'ADDL-SP/DSP':
    #         req = set([ticket['requesttype'] for ticket in coll.find({'requestcategory': 'Data Request', 'team':team,'modules':module,'designation': {'$nin': ['SP','DIG','IG']}}) if ticket['requesttype'] != ''])
    #     elif designation == 'INSPR' and role != 'Analyst':
    #         req = set([ticket['requesttype'] for ticket in coll.find({'requestcategory': 'Data Request', 'team':team,'modules':module,'designation': {'$nin': ['SP','DIG','IG','ADDL-SP/DSP']}}) if ticket['requesttype'] != ''])
    #     else:
    #         req = set([ticket['requesttype'] for ticket in coll.find({'requestcategory': 'Data Request','useremail':email}) if ticket['requesttype'] != ''])
    # else:
    #     if team == 'ADMIN':
    #         req = set([ticket['requesttypes'][0] for ticket in coll.find({'requestcategory': 'Analysis Request'}) if ticket['requesttypes'] != []])
    #     elif designation == 'SP' and team != 'ADMIN':
    #         req = set([ticket['requesttypes'][0] for ticket in coll.find({'requestcategory': 'Analysis Request', 'team':team}) if ticket['requesttypes'] != []])
    #     elif designation == 'ADDL-SP/DSP':
    #         req = set([ticket['requesttypes'][0] for ticket in coll.find({'requestcategory': 'Analysis Request', 'team':team,'modules':module,'designation': {'$nin': ['SP','DIG','IG']}}) if ticket['requesttypes'] != []])
    #     elif designation == 'INSPR' and role != 'Analyst':
    #         req = set([ticket['requesttypes'][0] for ticket in coll.find({'requestcategory': 'Analysis Request', 'team':team,'modules':module,'designation': {'$nin': ['SP','DIG','IG','ADDL-SP/DSP']}}) if ticket['requesttypes'] != []])
    #     else:
    #         req = set([ticket['requesttypes'][0] for ticket in coll.find({'requestcategory': 'Analysis Request','useremail':email}) if ticket['requesttypess'] != []])
    # reqtypes = list(req)
    # print(req)
    if category in ['Data Request']:
        if types in reqtypes:
            results['reqtypes'] = [types]
        else:
            results['reqtypes'] = reqtypes
    else:
        # changed reqtypes to analysistypes for if and else
        if types in analysistypes:
            results['reqtypes'] = [types]
        else:
            results['reqtypes'] = analysistypes
    if team == 'ADMIN':
        for req in results['reqtypes']:
            if category in ['Data Request', '', 'All']:
                if officer not in ['','All'] and otherteam in ['','All']:
                    query1 = {'requesttype': req,'requestcategory': 'Data Request','officername':officer}
                    project1 = {'requesttype': 1, 'date': 1,'requestcategory': 1,'officername':1, '_id': 0}

                    providers = {"requesttype": req, 'requestcategory': 'Data Request',"result.ProviderName": {"$exists": True},'officername':officer}
                    project_providers = {'requesttype': 1, 'date': 1,'requestcategory': 1,'officername':1, '_id': 0,"result.ProviderName": 1,'officername':1}

                    query2 = {'requesttype': req,'requestcategory': 'Data Request','pending': 'Mail Under Process','officername':officer}
                    project2 = {'requesttype': 1, 'date': 1,'requestcategory': 1,'_id': 0, 'pending': 1,'officername':1}

                    query3 = {'requesttype': req, 'requestcategory': 'Data Request', 'pending': re.compile(r"Pending"), 'officername': officer}
                    project3 = {'requesttype': 1, 'date': 1,'requestcategory': 1,'_id': 0, 'pending': 1,'officername':1}

                    query4 = {'requesttype': req, 'requestcategory': 'Data Request', 'pending': 'Mail Under Process', 'newnumber.status': {'$in': ['Data_Received', 'Data Not Received']}, 'officername': officer}
                    project4 = {'requesttype': 1, 'date': 1,'requestcategory': 1, '_id': 0, 'newnumber.status': 1, 'pending': 1,'officername':1}

                    query5 = {'requesttype': req, 'requestcategory': 'Data Request', 'pending': 'Ticket_Closed', 'officername': officer}
                    project5 = {'requesttype': 1, 'date': 1,'requestcategory': 1, '_id': 0,'pending': 1,'officername':1}

                elif otherteam not in ['','All'] and officer in ['','All']:
                    query1 = {'requesttype': req,'requestcategory': 'Data Request','team':otherteam}
                    project1 = {'requesttype': 1, 'date': 1,'requestcategory': 1,'team':1, '_id': 0}

                    providers = {"requesttype": req, 'requestcategory': 'Data Request',"result.ProviderName": {"$exists": True},'team':otherteam}
                    project_providers = {'requesttype': 1, 'date': 1,'requestcategory': 1, '_id': 0,"result.ProviderName": 1,'team':1}

                    query2 = {'requesttype': req,'requestcategory': 'Data Request','pending': 'Mail Under Process','team':otherteam}
                    project2 = {'requesttype': 1, 'date': 1,'requestcategory': 1, '_id': 0, 'pending': 1,'team':1}

                    query3 = {'requesttype': req, 'requestcategory': 'Data Request', 'pending': re.compile(r"Pending"), 'team': otherteam}
                    project3 = {'requesttype': 1, 'date': 1,'requestcategory': 1, '_id': 0, 'pending': 1,'team':1}

                    query4 = {'requesttype': req, 'requestcategory': 'Data Request', 'pending': 'Mail Under Process', 'newnumber.status': {'$in': ['Data_Received', 'Data Not Received']}, 'team': otherteam}
                    project4 = {'requesttype': 1, 'date': 1,'requestcategory': 1, '_id': 0, 'pending': 1,'team':1,'newnumber.status': 1}

                    query5 = {'requesttype': req, 'requestcategory': 'Data Request', 'pending': 'Ticket_Closed', 'team': otherteam}
                    project5 = {'requesttype': 1, 'date': 1,'requestcategory': 1, '_id': 0, 'pending': 1,'team':1}

                elif officer not in ['','All'] and otherteam not in ['','All']:
                    query1 = {'requesttype': req,'requestcategory': 'Data Request','officername':officer,'team':otherteam}
                    project1 = {'requesttype': 1, 'date': 1,'requestcategory': 1,'team':1, '_id': 0,'officername': 1,}

                    query2 = {'requesttype': req,'requestcategory': 'Data Request','pending': 'Mail Under Process','team':otherteam,'officername':officer}
                    project2 = {'requesttype': 1, 'date': 1,'requestcategory': 1,'team':1, '_id': 0,'officername': 1,'pending': 1}

                    query3 = {'requesttype': req, 'requestcategory': 'Data Request', 'pending': re.compile(r"Pending"), 'team': otherteam, 'officername': officer}
                    project3 = {'requesttype': 1, 'date': 1,'requestcategory': 1,'team':1, '_id': 0,'officername': 1,'pending': 1}

                    query4 = {'requesttype': req, 'requestcategory': 'Data Request', 'pending': 'Mail Under Process', 'newnumber.status': {'$in': ['Data_Received', 'Data Not Received']}, 'team': otherteam, 'officername': officer}
                    project4 = {'requesttype': 1, 'date': 1,'requestcategory': 1,'team':1, '_id': 0,'officername': 1,'pending': 1, 'newnumber.status': 1}

                    query5 = {'requesttype': req, 'requestcategory': 'Data Request', 'pending': 'Ticket_Closed', 'team': otherteam, 'officername': officer}
                    project5 = {'requesttype': 1, 'date': 1,'requestcategory': 1,'team':1, '_id': 0,'officername': 1,'pending': 1}

                    providers = {"requesttype": req, 'requestcategory': 'Data Request',"result.ProviderName": {"$exists": True},'team': otherteam, 'officername': officer}
                    project_providers = {'requesttype': 1, 'date': 1,'requestcategory': 1,'team':1, '_id': 0,'officername': 1,"result.ProviderName": 1}

                else:
                    query1 = {'requesttype': req,'requestcategory': 'Data Request'}
                    project1 = {'requesttype': 1, 'date': 1,'requestcategory': 1, '_id': 0,}
                    query2 = {'requesttype': req,'requestcategory': 'Data Request','pending': 'Mail Under Process'}
                    project2 = {'requesttype': 1, 'date': 1,'requestcategory': 1,'pending':1, '_id': 0,}
                    query3 = {'requesttype': req, 'requestcategory': 'Data Request', 'pending': re.compile(r"Pending")}
                    project3 = {'requesttype': 1, 'date': 1,'requestcategory': 1,'pending':1, '_id': 0,}
                    query4 = {'requesttype': req, 'requestcategory': 'Data Request', 'pending': 'Mail Under Process', 'newnumber.status': {'$in': ['Data_Received', 'Data Not Received']}}
                    project4 = {'requesttype': 1, 'date': 1,'requestcategory': 1,'pending':1, '_id': 0,'newnumber.status': 1}

                    query5 = {'requesttype': req, 'requestcategory': 'Data Request', 'pending': 'Ticket_Closed'}
                    project5 = {'requesttype': 1, 'date': 1,'requestcategory': 1,'pending':1, '_id': 0,}

                    providers = {"requesttype": req,'requestcategory': 'Data Request', "result.ProviderName": {"$exists": True}}
                    project_providers = {'requesttype': 1, 'date': 1,'requestcategory': 1,'_id': 0,"result.ProviderName": 1}

                if period == 'All':
                    results[req] = {
                        'Raise': len([ticket for ticket in coll.find(query1,project1)]),
                        'jo': len([i['ProviderName'] for res in coll.find(providers, project_providers) for i in res['result'] if i.get('ProviderName') == 'Jio' and res.get('requesttype') == req]),
                        'at': len([i['ProviderName'] for res in coll.find(providers, project_providers) for i in res['result'] if i.get('ProviderName') == 'Airtel' and res.get('requesttype') == req]),
                        'vi': len([i['ProviderName'] for res in coll.find(providers, project_providers) for i in res['result'] if i.get('ProviderName') == 'Vodafone' and res.get('requesttype') == req]),
                        'co': len([i['ProviderName'] for res in coll.find(providers, project_providers) for i in res['result'] if i.get('ProviderName') == 'Cellone' and res.get('requesttype') == req]),
                        'Approval': len([ticket for ticket in coll.find(query2,project2)]),
                        'Pending':  len([ticket for ticket in coll.find(query3,project3)]),
                        'Under_Process': len([ticket for ticket in coll.find(query4,project4)]),
                        'Ticket_closed': len([ticket for ticket in coll.find(query5,project5)]),
                        'Total_Pending' : len([ticket for ticket in coll.find(query3,project3)]) + len([ticket for ticket in coll.find(query4,project4)])
                    }
                elif period == 'Today':
                    results[req] = {
                        'Raise': len([ticket for ticket in coll.find(query1,project1) if ticket.get('date')== day]),
                        'jo': len([i['ProviderName'] for res in coll.find(providers, project_providers) for i in res['result'] if i.get('ProviderName') == 'Jio' and res.get('requesttype') == req and res.get('date')== day]),
                        'at': len([i['ProviderName'] for res in coll.find(providers, project_providers) for i in res['result'] if i.get('ProviderName') == 'Airtel' and res.get('requesttype') == req and res.get('date')== day]),
                        'vi': len([i['ProviderName'] for res in coll.find(providers, project_providers) for i in res['result'] if i.get('ProviderName') == 'Vodafone' and res.get('requesttype') == req and res.get('date')== day]),
                        'co': len([i['ProviderName'] for res in coll.find(providers, project_providers) for i in res['result'] if i.get('ProviderName') == 'Cellone' and res.get('requesttype') == req and res.get('date')== day]),
                        'Approval': len([ticket for ticket in coll.find(query2,project2) if ticket.get('date')== day]),
                        'Pending': len([ticket for ticket in coll.find(query3,project3) if ticket.get('date')== day]),
                        'Under_Process': len([ticket for ticket in coll.find(query4,project4) if ticket.get('date')== day]),
                        'Ticket_closed': len([ticket for ticket in coll.find(query5,project5) if ticket.get('date')== day]),
                        'Total_Pending' : len([ticket for ticket in coll.find(query3,project3) if ticket.get('date')== day]) + len([ticket for ticket in coll.find(query4,project4) if ticket.get('date')== day])
                    }
                elif date_range =='Date Range':
                    start_date, end_date = map(str.strip, period.split(' to '))
                    start_date1 = datetime.strptime(start_date, '%d/%m/%Y')
                    end_date1 = datetime.strptime(end_date, '%d/%m/%Y')
                    samples = [res for res in coll.find({"requestcategory": category},{'date':1,'_id':0})]
                    date_objects = [datetime.strptime(date['date'], '%d/%m/%Y') for date in samples]
                    result_dates = [date.strftime('%d/%m/%Y') for date in date_objects if start_date1 <= date <= end_date1]
                    results[req] = {
                        'Raise': len([ticket for ticket in coll.find(query1,project1) if ticket.get('date') in result_dates]),
                        'jo': len([i['ProviderName'] for res in coll.find(providers, project_providers) for i in res['result'] if i.get('ProviderName') == 'Jio' and res.get('requesttype') == req and res.get('date') in result_dates]),
                        'at': len([i['ProviderName'] for res in coll.find(providers, project_providers) for i in res['result'] if i.get('ProviderName') == 'Airtel' and res.get('requesttype') == req and res.get('date') in result_dates]),
                        'vi': len([i['ProviderName'] for res in coll.find(providers, project_providers) for i in res['result'] if i.get('ProviderName') == 'Vodafone' and res.get('requesttype') == req and res.get('date') in result_dates]),
                        'co': len([i['ProviderName'] for res in coll.find(providers, project_providers) for i in res['result'] if i.get('ProviderName') == 'Cellone' and res.get('requesttype') == req and res.get('date') in result_dates]),
                        'Approval': len([ticket for ticket in coll.find(query2, project2) if ticket.get('date') in result_dates]),
                        'Pending': len([ticket for ticket in coll.find(query3, project3) if ticket.get('date') in result_dates]),
                        'Under_Process': len([ticket for ticket in coll.find(query4,project4) if ticket.get('date') in result_dates]),
                        'Ticket_closed': len([ticket for ticket in coll.find(query5,project5) if ticket.get('date') in result_dates]),
                        'Total_Pending' : len([ticket for ticket in coll.find(query3,project3) if ticket.get('date') in result_dates]) + len([ticket for ticket in coll.find(query4,project4) if ticket.get('date') in result_dates])
                    }
            else:
                if officer not in ['','All'] and otherteam in ['','All']:
                    query1 = {'requesttypes':{'$in': [re.compile(req)]},'requestcategory': 'Analysis Request','officername':officer}
                    project1 = {'requesttypes': 1, 'date': 1,'requestcategory': 1,'officername':1, '_id': 0}
                    providers = {'requesttypes':{'$in': [re.compile(req)]},'requestcategory': 'Analysis Request',"newnumber.ISP": {"$exists": True},'officername':officer}
                    project_providers = {'requesttypes': 1, 'date': 1,'requestcategory': 1,'officername':1, '_id': 0,"newnumber.ISP": 1}

                    query2 = {'requesttypes':{'$in': [re.compile(req)]},'requestcategory': 'Analysis Request','pending': {'$in':['--']},'officername':officer}
                    project2 = {'requesttypes': 1, 'date': 1,'requestcategory': 1,'officername':1, '_id': 0,'pending': 1}

                    query3 = {'requesttypes':{'$in': [re.compile(req)]}, 'requestcategory': 'Analysis Request','pending': {'$in':['Assign to CAT Ins',re.compile(r"Pending")]}, 'officername': officer}
                    project3 = {'requesttypes': 1, 'date': 1,'requestcategory': 1,'officername':1, '_id': 0,'pending': 1}

                    query4 = {'requesttypes':{'$in': [re.compile(req)]}, 'requestcategory': 'Analysis Request','pending': {'$in':[re.compile(r"Assigned")]}, 'officername': officer}
                    project4 = {'requesttypes': 1, 'date': 1,'requestcategory': 1,'officername':1, '_id': 0,'pending': 1}

                    query5 = {'requesttypes':{'$in': [re.compile(req)]}, 'requestcategory': 'Analysis Request','pending': {'$in':[re.compile(r"Closed")]}, 'officername': officer}
                    project5 = {'requesttypes': 1, 'date': 1,'requestcategory': 1,'officername':1, '_id': 0,'pending': 1}
                    
                    
                elif otherteam not in ['','All'] and officer in ['','All']:
                    query1 = {'requesttypes':{'$in': [re.compile(req)]},'requestcategory': 'Analysis Request','team':otherteam}
                    project1 = {'requesttypes': 1, 'date': 1,'requestcategory': 1,'team':1, '_id': 0}

                    providers = {'requesttypes':{'$in': [re.compile(req)]},'requestcategory': 'Analysis Request',"newnumber.ISP": {"$exists": True},'team':otherteam}
                    project_providers = {'requesttypes': 1, 'date': 1,'requestcategory': 1,'team':1, '_id': 0,"newnumber.ISP": 1}

                    query2 = {'requesttypes':{'$in': [re.compile(req)]},'requestcategory': 'Analysis Request','pending': {'$in':['--']},'team':otherteam}
                    project2 = {'requesttypes': 1, 'date': 1,'requestcategory': 1,'team':1, '_id': 0,"pending": 1}

                    query3 = {'requesttypes':{'$in': [re.compile(req)]}, 'requestcategory': 'Analysis Request', 'pending': {'$in':['Assign to CAT Ins',re.compile(r"Pending")]}, 'team': otherteam}
                    project3 = {'requesttypes': 1, 'date': 1,'requestcategory': 1,'team':1, '_id': 0,"pending": 1}
                
                    query4 = {'requesttypes':{'$in': [re.compile(req)]}, 'requestcategory': 'Analysis Request', 'approval': {'$in':[re.compile(r"Assigned")]}, 'team': otherteam}
                    project4 = {'requesttypes': 1, 'date': 1,'requestcategory': 1,'team':1, '_id': 0,"approval": 1}

                    query5 = {'requesttypes':{'$in': [re.compile(req)]}, 'requestcategory': 'Analysis Request', 'pending': {'$in':[re.compile(r"Closed")]}, 'team': otherteam}
                    project5 = {'requesttypes': 1, 'date': 1,'requestcategory': 1,'team':1, '_id': 0,"pending": 1}

                    
                elif officer not in ['','All'] and otherteam not in ['','All']:
                    query1 = {'requesttypes':{'$in': [re.compile(req)]},'requestcategory': 'Analysis Request','officername':officer,'team':otherteam}
                    project1 = {'requesttypes': 1, 'date': 1,'requestcategory': 1,'team':1, '_id': 0,'officername' : 1}

                    query2 = {'requesttypes':{'$in': [re.compile(req)]},'requestcategory': 'Analysis Request','pending': {'$in':['--']},'team':otherteam,'officername':officer}
                    project2 = {'requesttypes': 1, 'date': 1,'requestcategory': 1,'team':1, '_id': 0,'officername' : 1,'pending' : 1}

                    query3 = {'requesttypes':{'$in': [re.compile(req)]},'requestcategory': 'Analysis Request', 'pending': {'$in':['Assign to CAT Ins',re.compile(r"Pending")]}, 'team': otherteam, 'officername': officer}
                    project3 = {'requesttypes': 1, 'date': 1,'requestcategory': 1,'team':1, '_id': 0,'officername' : 1,'pending' : 1}

                    query4 = {'requesttypes':{'$in': [re.compile(req)]},'requestcategory': 'Analysis Request', 'approval': {'$in':[re.compile(r"Assigned")]}, 'team': otherteam, 'officername': officer}
                    project4 = {'requesttypes': 1, 'date': 1,'requestcategory': 1,'team':1, '_id': 0,'officername' : 1,'approval' : 1}

                    query5 = {'requesttypes':{'$in': [re.compile(req)]},'requestcategory': 'Analysis Request', 'pending': {'$in':[re.compile(r"Closed")]}, 'team': otherteam, 'officername': officer}
                    project5 = {'requesttypes': 1, 'date': 1,'requestcategory': 1,'team':1, '_id': 0,'officername' : 1,'pending' : 1}

                    providers = {'requesttypes':{'$in': [re.compile(req)]},'requestcategory': 'Analysis Request',"newnumber.ISP": {"$exists": True},'team': otherteam, 'officername': officer}
                    project_providers = {'requesttypes': 1, 'date': 1,'requestcategory': 1,'team':1, '_id': 0,'officername' : 1,'pending' : 1,"newnumber.ISP": 1}

                    
                else:
                    query1 = {'requesttypes':{'$in': [re.compile(req)]},'requestcategory': 'Analysis Request'}
                    project1 = {'requesttypes': 1, 'date': 1,'requestcategory': 1,'_id': 0}

                    query2 = {'requesttypes':{'$in': [re.compile(req)]},'requestcategory': 'Analysis Request','pending': {'$in':['--']},}
                    project2 = {'requesttypes': 1, 'date': 1,'requestcategory': 1,'_id': 0,'pending': 1}

                    query3 = {'requesttypes':{'$in': [re.compile(req)]},'requestcategory': 'Analysis Request', 'pending': {'$in':['Assign to CAT Ins',re.compile(r"Pending")]}}
                    project3 = {'requesttypes': 1, 'date': 1,'requestcategory': 1,'_id': 0,'pending': 1}

                    query4 = {'requesttypes':{'$in': [re.compile(req)]},'requestcategory': 'Analysis Request', 'approval': {'$in':[re.compile(r"Assigned")]}}
                    project4 = {'requesttypes': 1, 'date': 1,'requestcategory': 1,'_id': 0,'approval': 1}

                    query5 = {'requesttypes':{'$in': [re.compile(req)]},'requestcategory': 'Analysis Request', 'pending': {'$in':[re.compile(r"Closed")]}}
                    project5 = {'requesttypes': 1, 'date': 1,'requestcategory': 1,'_id': 0,'pending': 1}

                    providers = {'requesttypes':{'$in': [re.compile(req)]},'requestcategory': 'Analysis Request', "newnumber.ISP": {"$exists": True}}
                    project_providers = {'requesttypes': 1, 'date': 1,'requestcategory': 1,'_id': 0,"newnumber.ISP": 1}

                if period == 'All':
                    results[req] = {
                    'Raise': len([ticket for ticket in coll.find(query1,project1)]),
                    'Approval': len([ticket for ticket in coll.find(query2,project2)]),
                    'Pending': len([ticket for ticket in coll.find(query3,project3)]),
                    'Assigned': len([ticket for ticket in coll.find(query4,project4)]),
                    'Ticket_closed': len([ticket for ticket in coll.find(query5,project5)]),
                    }
                elif period =='Today':
                    results[req] = {
                    'Raise':    len([ticket for ticket in coll.find(query1,project1) if ticket.get('date')== day]),
                    'Approval': len([ticket for ticket in coll.find(query2,project2) if ticket.get('date')== day]),
                    'Pending':  len([ticket for ticket in coll.find(query3,project3) if ticket.get('date')== day]),
                    'Assigned': len([ticket for ticket in coll.find(query4,project4)]),
                    'Ticket_closed': len([ticket for ticket in coll.find(query5,project5)]),
                    }
                elif date_range == 'Date Range':
                    start_date, end_date = map(str.strip, period.split(' to '))
                    start_date1 = datetime.strptime(start_date, '%d/%m/%Y')
                    end_date1 = datetime.strptime(end_date, '%d/%m/%Y')
                    samples = [res for res in coll.find({"requestcategory": category},{'date': 1,'_id': 0})]
                    date_objects = [datetime.strptime(date['date'], '%d/%m/%Y') for date in samples]
                    result_dates = [date.strftime('%d/%m/%Y') for date in date_objects if start_date1 <= date <= end_date1]
                    results[req] = {
                    'Raise': len([ticket for ticket in coll.find(query1,project1) if ticket.get('date') in result_dates]),
                    'Approval': len([ticket for ticket in coll.find(query2,project2) if ticket.get('date') in result_dates]),
                    'Pending':  len([ticket for ticket in coll.find(query3,project3) if ticket.get('date') in result_dates]),
                    'Assigned': len([ticket for ticket in coll.find(query4,project4)]),
                    'Ticket_closed': len([ticket for ticket in coll.find(query5,project5)]),
                    }

    elif designation == 'SP' and team != 'ADMIN':
        query1_projection = {'_id':1,'date':1}
        providers_projection = {'requesttype': 1,'result.ProviderName': 1,'_id':0,'date':1}
        query2_projection = {'pending': 1,'_id':0,'date':1}
        query3_projection = {'pending': 1,'_id':0,'date':1}
        query4_projection = {'pending': 1,'newnumber.status': 1,'_id': 0,'date':1}
        query5_projection = {'pending': 1, '_id':0,'date':1}
        additional_providers_projection = {'requesttypes': 1,'newnumber.ISP': 1, '_id':0,'date':1,'pending' : 1,'approval':1}
        for req in results['reqtypes']:
            if category in ['Data Request', '', 'All']:
                if officer not in ['','All'] and otherteam in ['','All']:
                    query1 = {'requesttype': req,'requestcategory': 'Data Request','officername':officer,"team":team}
                    providers = {"requesttype": req,'requestcategory': 'Data Request', "result.ProviderName": {"$exists": True},'officername':officer,"team":team}
                    query2 = {'requesttype': req,'requestcategory': 'Data Request','pending': 'Mail Under Process','officername':officer,"team":team}
                    query3 = {'requesttype': req, 'requestcategory': 'Data Request', 'pending': re.compile(r"Pending"), 'officername': officer,"team":team}
                    query4 = {'requesttype': req, 'requestcategory': 'Data Request', 'pending': 'Mail Under Process', "team":team, 'newnumber.status': {'$in': ['Data_Received', 'Data Not Received']}, 'officername': officer}
                    query5 = {'requesttype': req, 'requestcategory': 'Data Request', 'pending': 'Ticket_Closed', 'officername': officer,"team":team}
                else:
                    query1 = {'requesttype': req,'requestcategory': 'Data Request',"team":team}
                    query2 = {'requesttype': req,'requestcategory': 'Data Request','pending': 'Mail Under Process',"team":team}
                    query3 = {'requesttype': req, 'requestcategory': 'Data Request', 'pending': re.compile(r"Pending"),"team":team}
                    query4 = {'requesttype': req, 'requestcategory': 'Data Request', 'pending': 'Mail Under Process', "team":team, 'newnumber.status': {'$in': ['Data_Received', 'Data Not Received']}}
                    query5 = {'requesttype': req, 'requestcategory': 'Data Request', 'pending': 'Ticket_Closed',"team":team}
                    providers = {"requesttype": req,'requestcategory': 'Data Request', "result.ProviderName": {"$exists": True},"team":team}

                if period == 'All':
                    results[req] = {
                        'Raise': len([ticket for ticket in coll.find(query1,query1_projection)]),
                        'jo': len([i['ProviderName'] for res in coll.find(providers, providers_projection) for i in res['result'] if i.get('ProviderName') == 'Jio' and res.get('requesttype') == req]),
                        'at': len([i['ProviderName'] for res in coll.find(providers, providers_projection) for i in res['result'] if i.get('ProviderName') == 'Airtel' and res.get('requesttype') == req]),
                        'vi': len([i['ProviderName'] for res in coll.find(providers, providers_projection) for i in res['result'] if i.get('ProviderName') == 'Vodafone' and res.get('requesttype') == req]),
                        'co': len([i['ProviderName'] for res in coll.find(providers, providers_projection) for i in res['result'] if i.get('ProviderName') == 'Cellone' and res.get('requesttype') == req]),
                        'Approval': len([ticket for ticket in coll.find(query2,query2_projection)]),
                        'Pending': len([ticket for ticket in coll.find(query3,query3_projection)]),
                        'Under_Process': len([ticket for ticket in coll.find(query4,query4_projection)]),
                        'Ticket_closed': len([ticket for ticket in coll.find(query5,query5_projection)]),
                        'Total_Pending' : len([ticket for ticket in coll.find(query3)]) + len([ticket for ticket in coll.find(query4)])

                    }
                elif period == 'Today':
                    results[req] = {
                        'Raise': len([ticket for ticket in coll.find(query1,query1_projection) if ticket.get('date')== day]),
                        'jo':    len([i['ProviderName'] for res in coll.find(providers,providers_projection) for i in res['result'] if i.get('ProviderName') == 'Jio' and res.get('requesttype') == req and res.get('date')== day]),
                        'at':    len([i['ProviderName'] for res in coll.find(providers,providers_projection) for i in res['result'] if i.get('ProviderName') == 'Airtel' and res.get('requesttype') == req and res.get('date')== day]),
                        'vi':    len([i['ProviderName'] for res in coll.find(providers,providers_projection) for i in res['result'] if i.get('ProviderName') == 'Vodafone' and res.get('requesttype') == req and res.get('date')== day]),
                        'co':    len([i['ProviderName'] for res in coll.find(providers,providers_projection) for i in res['result'] if i.get('ProviderName') == 'Cellone' and res.get('requesttype') == req and res.get('date')== day]),
                        'Approval':      len([ticket for ticket in coll.find(query2,query2_projection) if ticket.get('date')== day]),
                        'Pending':       len([ticket for ticket in coll.find(query3,query3_projection) if ticket.get('date')== day]),
                        'Under_Process': len([ticket for ticket in coll.find(query4,query4_projection) if ticket.get('date')== day]),
                        'Ticket_closed': len([ticket for ticket in coll.find(query5,query5_projection) if ticket.get('date')== day]),
                        'Total_Pending' : len([ticket for ticket in coll.find(query3) if ticket.get('date')== day]) + len([ticket for ticket in coll.find(query4) if ticket.get('date')== day])
                    }
                elif date_range =='Date Range':
                    start_date, end_date = map(str.strip, period.split(' to '))
                    start_date1 = datetime.strptime(start_date, '%d/%m/%Y')
                    end_date1 = datetime.strptime(end_date, '%d/%m/%Y')
                    samples = [res for res in coll.find({"requestcategory": category,"team":team},{'date': 1,"_id": 0})]
                    date_objects = [datetime.strptime(date['date'], '%d/%m/%Y') for date in samples]
                    result_dates = [date.strftime('%d/%m/%Y') for date in date_objects if start_date1 <= date <= end_date1]
                    results[req] = {
                        'Raise': len([ticket for ticket in coll.find(query1,query1_projection) if ticket.get('date') in result_dates]),
                        'jo':    len([i['ProviderName'] for res in coll.find(providers, providers_projection) for i in res['result'] if i.get('ProviderName') == 'Jio' and res.get('requesttype') == req and res.get('date') in result_dates]),
                        'at':    len([i['ProviderName'] for res in coll.find(providers, providers_projection) for i in res['result'] if i.get('ProviderName') == 'Airtel' and res.get('requesttype') == req and res.get('date') in result_dates]),
                        'vi':    len([i['ProviderName'] for res in coll.find(providers, providers_projection) for i in res['result'] if i.get('ProviderName') == 'Vodafone' and res.get('requesttype') == req and res.get('date') in result_dates]),
                        'co':    len([i['ProviderName'] for res in coll.find(providers, providers_projection) for i in res['result'] if i.get('ProviderName') == 'Cellone' and res.get('requesttype') == req and res.get('date') in result_dates]),
                        'Approval':      len([ticket for ticket in coll.find(query2, query2_projection) if ticket.get('date') in result_dates]),
                        'Pending':       len([ticket for ticket in coll.find(query3, query3_projection) if ticket.get('date') in result_dates]),
                        'Under_Process': len([ticket for ticket in coll.find(query4, query4_projection) if ticket.get('date') in result_dates]),
                        'Ticket_closed': len([ticket for ticket in coll.find(query5, query5_projection) if ticket.get('date') in result_dates]),
                        'Total_Pending' : len([ticket for ticket in coll.find(query3) if ticket.get('date') in result_dates]) + len([ticket for ticket in coll.find(query4) if ticket.get('date') in result_dates])
                    }
            else:
                if officer not in ['','All']:
                    query1 = {'requesttypes':{'$in': [re.compile(req)]},'requestcategory': 'Analysis Request','officername':officer,"team":team}
                    providers = {'requesttypes':{'$in': [re.compile(req)]},'requestcategory': 'Analysis Request',"newnumber.ISP": {"$exists": True},'officername':officer,"team":team}
                    query2 = {'requesttypes':{'$in': [re.compile(req)]},'requestcategory': 'Analysis Request','pending': {'$in':['--']},'officername':officer,"team":team}
                    query3 = {'requesttypes':{'$in': [re.compile(req)]}, 'requestcategory': 'Analysis Request','pending': {'$in':['Assign to CAT Ins',re.compile(r"Pending")]},"team":team, 'officername': officer}  
                    query4 = {'requesttypes':{'$in': [re.compile(req)]}, 'requestcategory': 'Analysis Request','approval': {'$in':[re.compile(r"Assigned")]},"team":team, 'officername': officer}
                    query5 = {'requesttypes':{'$in': [re.compile(req)]}, 'requestcategory': 'Analysis Request','pending': {'$in':[re.compile(r"Closed")]},"team":team, 'officername': officer}                  

                else:
                    query1 = {'requesttypes':{'$in': [re.compile(req)]},'requestcategory': 'Analysis Request',"team":team}
                    query2 = {'requesttypes':{'$in': [re.compile(req)]},'requestcategory': 'Analysis Request',"team":team,'pending': {'$in':['--']},}
                    query3 = {'requesttypes':{'$in': [re.compile(req)]},'requestcategory': 'Analysis Request',"team":team, 'pending': {'$in':['Assign to CAT Ins',re.compile(r"Pending")]}}
                    query4 = {'requesttypes':{'$in': [re.compile(req)]},'requestcategory': 'Analysis Request',"team":team, 'approval': {'$in':[re.compile(r"Assigned")]}}
                    query5 = {'requesttypes':{'$in': [re.compile(req)]},'requestcategory': 'Analysis Request',"team":team, 'pending': {'$in':[re.compile(r"Closed")]}}
                    
                    providers = {'requesttypes':{'$in': [re.compile(req)]},'requestcategory': 'Analysis Request',"newnumber.ISP": {"$exists": True},"team":team}
                if period == 'All':
                    results[req] = {
                    'Raise':    len([ticket for ticket in coll.find(query1,additional_providers_projection)]),
                    'Approval': len([ticket for ticket in coll.find(query2,additional_providers_projection)]),
                    'Pending':  len([ticket for ticket in coll.find(query3,additional_providers_projection)]),
                    'Assigned':  len([ticket for ticket in coll.find(query4,additional_providers_projection)]),
                    'Ticket_closed':  len([ticket for ticket in coll.find(query5,additional_providers_projection)]),
                    
                    }
                elif period =='Today':
                    results[req] = {
                    'Raise':    len([ticket for ticket in coll.find(query1,additional_providers_projection) if ticket.get('date')== day]),
                    'Approval': len([ticket for ticket in coll.find(query2,additional_providers_projection) if ticket.get('date')== day]),
                    'Pending':  len([ticket for ticket in coll.find(query3,additional_providers_projection) if ticket.get('date')== day]),
                    'Assigned':  len([ticket for ticket in coll.find(query4,additional_providers_projection)]),
                    'Ticket_closed':  len([ticket for ticket in coll.find(query5,additional_providers_projection)]),
                    }
                elif date_range == 'Date Range':
                    start_date, end_date = map(str.strip, period.split(' to '))
                    start_date1 = datetime.strptime(start_date, '%d/%m/%Y')
                    end_date1 = datetime.strptime(end_date, '%d/%m/%Y')
                    samples = [res for res in coll.find({"requestcategory": category,"team":team},{'date' : 1, "_id" : 0})]
                    date_objects = [datetime.strptime(date['date'], '%d/%m/%Y') for date in samples]
                    result_dates = [date.strftime('%d/%m/%Y') for date in date_objects if start_date1 <= date <= end_date1]
                    results[req] = {
                    'Raise':    len([ticket for ticket in coll.find(query1,additional_providers_projection) if ticket.get('date') in result_dates]),
                    'Approval': len([ticket for ticket in coll.find(query2,additional_providers_projection) if ticket.get('date') in result_dates]),
                    'Pending':  len([ticket for ticket in coll.find(query3,additional_providers_projection) if ticket.get('date') in result_dates]),
                    'Assigned':  len([ticket for ticket in coll.find(query4,additional_providers_projection)]),
                    'Ticket_closed':  len([ticket for ticket in coll.find(query5,additional_providers_projection)]),
                    }
                    
    elif designation == 'ADDL-SP/DSP':
        for req in results['reqtypes']:
            if category in ['Data Request', '', 'All']:
                if officer not in ['','All']:
                    query1 = {'requesttype': req,'requestcategory': 'Data Request','officername':officer,"team":team,"modules":module,'designation': {'$nin': ['SP','DIG','IG']}}
                    query1_projection = {'requesttype': 1,'_id': 0, 'date':1}

                    providers = {"requesttype": req, "requestcategory":"Data Request","result.ProviderName": {"$exists": True},'officername':officer,"team":team,"modules":module,'designation': {'$nin': ['SP','DIG','IG']}}
                    providers_projection = {'requesttype': 1,'result.ProviderName': 1,'_id': 0,'date':1}

                    query2 = {'requesttype': req,'requestcategory': 'Data Request','pending': 'Mail Under Process','officername':officer,"team":team,"modules":module,'designation': {'$nin': ['SP','DIG','IG']}}
                    query2_projection = {'pending': 1,  'date': 1 ,'_id': 0}

                    query3 = {'requesttype': req, 'requestcategory': 'Data Request', 'pending': re.compile(r"Pending"), 'officername': officer,"team":team,"modules":module,'designation': {'$nin': ['SP','DIG','IG']}}
                    query3_projection = {'pending': 1,  'date': 1 ,'_id': 0}

                    query4 = {'requesttype': req, 'requestcategory': 'Data Request', 'pending': 'Mail Under Process', "team":team, "modules":module,'designation': {'$nin': ['SP','DIG','IG']}, 'newnumber.status': {'$in': ['Data_Received', 'Data Not Received']}, 'officername': officer}
                    query4_projection = {'pending': 1,'newnumber.status': 1, 'date': 1 ,'_id': 0}

                    query5 = {'requesttype': req, 'requestcategory': 'Data Request', 'pending': 'Ticket_Closed', 'officername': officer,"team":team,"modules":module,'designation': {'$nin': ['SP','DIG','IG']}}
                    query5_projection = {'pending': 1,  'date': 1 ,'_id': 0}
                else:
                    query1 = {'requesttype': req,'requestcategory': 'Data Request',"team":team, "modules":module,'designation': {'$nin': ['SP','DIG','IG']}}
                    query1_projection = {'requesttype': 1,'date': 1, '_id': 0}

                    query2 = {'requesttype': req,'requestcategory': 'Data Request','pending': 'Mail Under Process',"team":team,"modules":module,'designation': {'$nin': ['SP','DIG','IG']}}
                    query2_projection = {'pending': 1,'date': 1, '_id': 0}

                    query3 = {'requesttype': req, 'requestcategory': 'Data Request', 'pending': re.compile(r"Pending"),"team":team,"modules":module,'designation': {'$nin': ['SP','DIG','IG']}}
                    query3_projection = {'pending': 1,'date': 1 ,'_id': 0}

                    query4 = {'requesttype': req, 'requestcategory': 'Data Request', 'pending': 'Mail Under Process', "team":team, 'newnumber.status': {'$in': ['Data_Received', 'Data Not Received']},"modules":module,'designation': {'$nin': ['SP','DIG','IG']}}
                    query4_projection = { 'pending': 1, 'newnumber.status': 1,'date': 1 ,'_id': 0}

                    query5 = {'requesttype': req, 'requestcategory': 'Data Request', 'pending': 'Ticket_Closed',"team":team,"modules":module,'designation': {'$nin': ['SP','DIG','IG']}}
                    query5_projection = {'pending': 1,'date': 1, '_id': 0}
                    providers = {"requesttype": req, "requestcategory":"Data Request","result.ProviderName": {"$exists": True},"team":team,"modules":module,'designation': {'$nin': ['SP', 'DIG', 'IG']}}
                    providers_projection = {'requesttype': 1, 'result.ProviderName': 1,'date': 1 ,'_id': 0}

                if period == 'All':
                    results[req] = {
                        'Raise': len([ticket for ticket in coll.find(query1,query1_projection)]),
                        'jo':    sum(len(res['result']) for res in coll.find({'result.ProviderName': 'Jio', 'requesttype': req, 'useremail': {'$ne': email}, 'designation': {'$nin': ['SP']}, 'team': team, 'modules': module},{'result.ProviderName':1,'_id':0})),
                        'at':    sum(len(res['result']) for res in coll.find({'result.ProviderName': 'Airtel', 'requesttype': req, 'useremail': {'$ne': email}, 'designation': {'$nin': ['SP']}, 'team': team, 'modules': module},{'result.ProviderName':1,'_id':0})),
                        'vi':    sum(len(res['result']) for res in coll.find({'result.ProviderName': 'Vodafone', 'requesttype': req, 'useremail': {'$ne': email}, 'designation': {'$nin': ['SP']}, 'team': team, 'modules': module},{'result.ProviderName':1,'_id':0})),
                        'co':    sum(len(res['result']) for res in coll.find({'result.ProviderName': 'Cellone', 'requesttype': req, 'useremail': {'$ne': email}, 'designation': {'$nin': ['SP']}, 'team': team, 'modules': module},{'result.ProviderName':1,'_id':0})),
                        'Approval':      len([ticket for ticket in coll.find(query2, query2_projection)]),
                        'Pending':       len([ticket for ticket in coll.find(query3, query3_projection)]),
                        'Under_Process': len([ticket for ticket in coll.find(query4, query4_projection)]),
                        'Ticket_closed': len([ticket for ticket in coll.find(query5, query5_projection)]),
                        'Total_Pending' : len([ticket for ticket in coll.find(query3)]) + len([ticket for ticket in coll.find(query4)])
                    }
                elif period == 'Today':
                    results[req] = {
                        'Raise': len([ticket for ticket in coll.find(query1,query1_projection) if ticket.get('date')== day]),
                        'jo':    sum(len(res['result']) for res in coll.find({'result.ProviderName': 'Jio', 'requesttype': req, 'useremail': {'$ne': email}, 'designation': {'$nin': ['SP']}, 'team': team, 'modules': module,'date': day},{'result.ProviderName':1,'_id':0})),
                        'at':    sum(len(res['result']) for res in coll.find({'result.ProviderName': 'Airtel', 'requesttype': req, 'useremail': {'$ne': email}, 'designation': {'$nin': ['SP']}, 'team': team, 'modules': module,'date': day},{'result.ProviderName':1,'_id':0})),
                        'vi':    sum(len(res['result']) for res in coll.find({'result.ProviderName': 'Vodafone', 'requesttype': req, 'useremail': {'$ne': email}, 'designation': {'$nin': ['SP']}, 'team': team, 'modules': module,'date': day},{'result.ProviderName':1,'_id':0})),
                        'co':    sum(len(res['result']) for res in coll.find({'result.ProviderName': 'Cellone', 'requesttype': req, 'useremail': {'$ne': email}, 'designation': {'$nin': ['SP']}, 'team': team, 'modules': module,'date': day},{'result.ProviderName':1,'_id':0})),
                        'Approval':      len([ticket for ticket in coll.find(query2, query2_projection) if ticket.get('date')== day]),
                        'Pending':       len([ticket for ticket in coll.find(query3, query3_projection) if ticket.get('date')== day]),
                        'Under_Process': len([ticket for ticket in coll.find(query4, query4_projection) if ticket.get('date')== day]),
                        'Ticket_closed': len([ticket for ticket in coll.find(query5, query5_projection) if ticket.get('date')== day]),
                        'Total_Pending' : len([ticket for ticket in coll.find(query3) if ticket.get('date')== day]) + len([ticket for ticket in coll.find(query4) if ticket.get('date')== day])
                    }
                elif date_range =='Date Range':
                        start_date, end_date = map(str.strip, period.split(' to '))
                        start_date1 = datetime.strptime(start_date, '%d/%m/%Y')
                        end_date1 = datetime.strptime(end_date, '%d/%m/%Y')
                        samples = [res for res in coll.find({"requestcategory": category,"team":team,"modules":module,'designation': {'$nin': ['SP','DIG','IG']}},{'date': 1, '_id': 0})]
                        date_objects = [datetime.strptime(date['date'], '%d/%m/%Y') for date in samples]
                        result_dates = [date.strftime('%d/%m/%Y') for date in date_objects if start_date1 <= date <= end_date1]
                        results[req] = {
                            'Raise': len([ticket for ticket in coll.find(query1,query1_projection) if ticket.get('date') in result_dates]),
                            'jo': sum(len(res['result']) for res in coll.find({'result.ProviderName': 'Jio', 'requesttype': req, 'useremail': {'$ne': email}, 'designation': {'$nin': ['SP']}, 'team': team, 'modules': module,'date': {'$in': result_dates}},{'result.ProviderName':1,'_id':0})),
                            'at': sum(len(res['result']) for res in coll.find({'result.ProviderName': 'Airtel', 'requesttype': req, 'useremail': {'$ne': email}, 'designation': {'$nin': ['SP']}, 'team': team, 'modules': module,'date': {'$in': result_dates}},{'result.ProviderName':1,'_id':0})),
                            'vi': sum(len(res['result']) for res in coll.find({'result.ProviderName': 'Vodafone', 'requesttype': req, 'useremail': {'$ne': email}, 'designation': {'$nin': ['SP']}, 'team': team, 'modules': module,'date': {'$in': result_dates}},{'result.ProviderName':1,'_id':0})),
                            'co': sum(len(res['result']) for res in coll.find({'result.ProviderName': 'Cellone', 'requesttype': req, 'useremail': {'$ne': email}, 'designation': {'$nin': ['SP']}, 'team': team, 'modules': module,'date': {'$in': result_dates}},{'result.ProviderName':1,'_id':0})),
                            'Approval': len([ticket for ticket in coll.find(query2,query2_projection) if ticket.get('date') in result_dates]),
                            'Pending': len([ticket for ticket in coll.find(query3, query3_projection) if ticket.get('date') in result_dates]),
                            'Under_Process': len([ticket for ticket in coll.find(query4, query4_projection) if ticket.get('date') in result_dates]),
                            'Ticket_closed': len([ticket for ticket in coll.find(query5, query5_projection) if ticket.get('date') in result_dates]),
                            'Total_Pending' : len([ticket for ticket in coll.find(query3) if ticket.get('date') in result_dates]) + len([ticket for ticket in coll.find(query4) if ticket.get('date') in result_dates])
                        }
            else:
                # additional_providers_projection = {'requesttypes': 1, 'requestcategory': 1, 'newnumber.ISP': 1,  'date': 1 , "designation" : 1,'_id': 0}
                if officer not in ['','All']:
                    query1 = {'requesttypes':{'$in': [re.compile(req)]},'requestcategory': 'Analysis Request','officername':officer,"team":team,"modules":module,'designation': {'$nin': ['SP','DIG','IG']}}
                    query1_projection = {'requesttypes': 1, 'date': 1 ,'_id': 0}

                    query2 = {'requesttypes':{'$in': [re.compile(req)]},'requestcategory': 'Analysis Request','pending': {'$in':['--']},'officername':officer,"team":team,"modules":module,'designation': {'$nin': ['SP','DIG','IG']}}
                    query2_projection = {'pending': 1,'date': 1 ,'_id': 0}
                    query3 = {'requesttypes':{'$in': [re.compile(req)]}, 'requestcategory': 'Analysis Request','pending': {'$in':['Assign to CAT Ins',re.compile(r"Pending")]},"team":team, 'officername': officer,"modules":module,'designation': {'$nin': ['SP','DIG','IG']}}                    
                    query3_projection = {'pending': 1,'date': 1 ,'_id': 0}

                    query4 = {'requesttypes':{'$in': [re.compile(req)]}, 'requestcategory': 'Analysis Request','approval': {'$in':[re.compile(r"Assigned")]},"team":team, 'officername': officer,"modules":module,'designation': {'$nin': ['SP','DIG','IG']}}                    
                    query4_projection = {'approval': 1,'date': 1 ,'_id': 0}

                    query5 = {'requesttypes':{'$in': [re.compile(req)]}, 'requestcategory': 'Analysis Request','pending': {'$in':[re.compile(r"Closed")]},"team":team, 'officername': officer,"modules":module,'designation': {'$nin': ['SP','DIG','IG']}}                    
                    query5_projection = {'pending': 1,'date': 1 ,'_id': 0}
                
                else:
                    query1 = {'requesttypes':{'$in': [re.compile(req)]},'requestcategory': 'Analysis Request',"team":team,"modules":module,'designation': {'$nin': ['SP','DIG','IG']}}
                    query1_projection = {'requesttypes': 1,'date': 1 ,'_id': 0}
                    query2 = {'requesttypes':{'$in': [re.compile(req)]},'requestcategory': 'Analysis Request',"team":team,'pending': {'$in':['--']},"modules":module,'designation': {'$nin': ['SP','DIG','IG']}}
                    query2_projection = {'pending': 1, 'date': 1 ,'_id': 0}
                    query3 = {'requesttypes':{'$in': [re.compile(req)]},'requestcategory': 'Analysis Request',"team":team, 'pending': {'$in':['Assign to CAT Ins',re.compile(r"Pending")]},"modules":module,'designation': {'$nin': ['SP','DIG','IG']}}
                    query3_projection = {'pending': 1,'date': 1 ,'_id': 0}

                    query4 = {'requesttypes':{'$in': [re.compile(req)]}, 'requestcategory': 'Analysis Request','approval': {'$in':[re.compile(r"Assigned")]},"team":team,"modules":module,'designation': {'$nin': ['SP','DIG','IG']}}                    
                    query4_projection = {'approval': 1,'date': 1 ,'_id': 0}

                    query5 = {'requesttypes':{'$in': [re.compile(req)]}, 'requestcategory': 'Analysis Request','pending': {'$in':[re.compile(r"Closed")]},"team":team, "modules":module,'designation': {'$nin': ['SP','DIG','IG']}}                    
                    query5_projection = {'pending': 1,'date': 1 ,'_id': 0}
                    
                if period == 'All':
                    results[req] = {
                        'Raise': len([ticket for ticket in coll.find(query1,query1_projection)]),
                        'Approval': len([ticket for ticket in coll.find(query2,query2_projection)]),
                        'Pending': len([ticket for ticket in coll.find(query3,query3_projection)]),
                        'Assigned':  len([ticket for ticket in coll.find(query4)]),
                        'Ticket_closed':  len([ticket for ticket in coll.find(query5)]),
                    }
                elif period =='Today':
                    results[req] = {
                        'Raise': len([ticket for ticket in coll.find(query1,additional_providers_projection) if ticket.get('date')== day]),
                        'Approval': len([ticket for ticket in coll.find(query2,additional_providers_projection) if ticket.get('date')== day]),
                        'Pending': len([ticket for ticket in coll.find(query3,additional_providers_projection) if ticket.get('date')== day]),
                        'Assigned':  len([ticket for ticket in coll.find(query4)]),
                        'Ticket_closed':  len([ticket for ticket in coll.find(query5)]),
                    }
                elif date_range == 'Date Range':
                    start_date, end_date = map(str.strip, period.split(' to '))
                    start_date1 = datetime.strptime(start_date, '%d/%m/%Y')
                    end_date1 = datetime.strptime(end_date, '%d/%m/%Y')
                    samples = [res for res in coll.find({"requestcategory": category,"team":team,"modules":module,'designation': {'$nin': ['SP','DIG','IG']}},{'date': 1, '_id': 0})]
                    date_objects = [datetime.strptime(date['date'], '%d/%m/%Y') for date in samples]
                    result_dates = [date.strftime('%d/%m/%Y') for date in date_objects if start_date1 <= date <= end_date1]
                    results[req] = {
                        'Raise': len([ticket for ticket in coll.find(query1,additional_providers_projection) if ticket.get('date') in result_dates]),
                        'Approval': len([ticket for ticket in coll.find(query2,additional_providers_projection) if ticket.get('date') in result_dates]),
                        'Pending': len([ticket for ticket in coll.find(query3,additional_providers_projection) if ticket.get('date') in result_dates]),
                        'Assigned':  len([ticket for ticket in coll.find(query4)]),
                        'Ticket_closed':  len([ticket for ticket in coll.find(query5)]),
                    }

    elif designation == 'INSPR' and role != 'Analyst':
        query1_projection = {'date': 1 ,'_id': 1}
        providers_projection = {'result.ProviderName': 1,  'date': 1 ,'_id': 1}
        query2_projection = {'pending': 1,  'date': 1 ,'_id': 1}
        query3_projection = {'pending': 1,  'date': 1 ,'_id': 1}
        query4_projection = {'pending': 1, 'newnumber.status': 1,'date': 1 ,'_id': 1}
        query5_projection = {'pending': 1,  'date': 1 ,'_id': 1}
        additional_providers_projection = {'newnumber.ISP': 1,  'date': 1 , "pending" : 1,'_id': 1}
        for req in results['reqtypes']:
            if category in ['Data Request', '', 'All']:
                if officer not in ['','All'] and otherteam in ['','All']:
                    query1 = {'requesttype': req,'requestcategory': 'Data Request','officername':officer,"team":team,"modules":module,'designation': {'$nin': ['ADDL-SP/DSP','SP','DIG','IG']}}
                    providers = {"requesttype": req, "requestcategory":"Data Request","result.ProviderName": {"$exists": True},'officername':officer,"team":team,"modules":module,'designation': {'$nin': ['ADDL-SP/DSP','SP','DIG','IG']}}
                    query2 = {'requesttype': req,'requestcategory': 'Data Request','pending': 'Mail Under Process','officername':officer,"team":team,"modules":module,'designation': {'$nin': ['ADDL-SP/DSP','SP','DIG','IG']}}
                    query3 = {'requesttype': req, 'requestcategory': 'Data Request', 'pending': re.compile(r"Pending"), 'officername': officer,"team":team,"modules":module,'designation': {'$nin': ['ADDL-SP/DSP','SP','DIG','IG']}}
                    query4 = {'requesttype': req, 'requestcategory': 'Data Request', 'pending': 'Mail Under Process', "team":team, "modules":module,'designation': {'$nin': ['ADDL-SP/DSP','SP','DIG','IG']}, 'newnumber.status': {'$in': ['Data_Received', 'Data Not Received']}, 'officername': officer}
                    query5 = {'requesttype': req, 'requestcategory': 'Data Request', 'pending': 'Ticket_Closed', 'officername': officer,"team":team,"modules":module,'designation': {'$nin': ['ADDL-SP/DSP','SP','DIG','IG']}}
                else:
                    query1 = {'requesttype': req,'requestcategory': 'Data Request',"team":team,"modules":module,'designation': {'$nin': ['ADDL-SP/DSP','SP','DIG','IG']}}
                    query2 = {'requesttype': req,'requestcategory': 'Data Request','pending': 'Mail Under Process',"team":team,"modules":module,'designation': {'$nin': ['ADDL-SP/DSP','SP','DIG','IG']}}
                    query3 = {'requesttype': req, 'requestcategory': 'Data Request', 'pending': re.compile(r"Pending"),"team":team,"modules":module,'designation': {'$nin': ['ADDL-SP/DSP','SP','DIG','IG']}}
                    query4 = {'requesttype': req, 'requestcategory': 'Data Request', 'pending': 'Mail Under Process', "team":team, 'newnumber.status': {'$in': ['Data_Received', 'Data Not Received']},"modules":module,'designation': {'$nin': ['ADDL-SP/DSP','SP','DIG','IG']}}
                    query5 = {'requesttype': req, 'requestcategory': 'Data Request', 'pending': 'Ticket_Closed',"team":team,"modules":module,'designation': {'$nin': ['ADDL-SP/DSP','SP','DIG','IG']}}
                    providers = {"requesttype": req, "requestcategory":"Data Request","result.ProviderName": {"$exists": True},"team":team,"modules":module,'designation': {'$nin': ['ADDL-SP/DSP','SP','DIG','IG']}}

                if period == 'All':
                    results[req] = {
                        'Raise': len([ticket for ticket in coll.find(query1,query1_projection)]),
                        'jo': len([i['ProviderName'] for res in coll.find(providers) for i in res['result'] if i.get('ProviderName') == 'Jio' and res.get('requesttype') == req]),
                        'at': len([i['ProviderName'] for res in coll.find(providers) for i in res['result'] if i.get('ProviderName') == 'Airtel' and res.get('requesttype') == req]),
                        'vi': len([i['ProviderName'] for res in coll.find(providers) for i in res['result'] if i.get('ProviderName') == 'Vodafone' and res.get('requesttype') == req]),
                        'co': len([i['ProviderName'] for res in coll.find(providers) for i in res['result'] if i.get('ProviderName') == 'Cellone' and res.get('requesttype') == req]),
                        'Approval': len([ticket for ticket in coll.find(query2,query2_projection)]),
                        'Pending': len([ticket for ticket in coll.find(query3, query3_projection)]),
                        'Under_Process': len([ticket for ticket in coll.find(query4, query4_projection)]),
                        'Ticket_closed': len([ticket for ticket in coll.find(query5, query5_projection)]),
                        'Total_Pending' : len([ticket for ticket in coll.find(query3)]) + len([ticket for ticket in coll.find(query4)])

                    }
                elif period == 'Today':
                    results[req] = {
                        'Raise': len([ticket for ticket in coll.find(query1,query1_projection) if ticket.get('date')== day]),
                        'jo': len([i['ProviderName'] for res in coll.find(providers, providers_projection) for i in res['result'] if i.get('ProviderName') == 'Jio' and res.get('requesttype') == req and res.get('date')== day]),
                        'at': len([i['ProviderName'] for res in coll.find(providers, providers_projection) for i in res['result'] if i.get('ProviderName') == 'Airtel' and res.get('requesttype') == req and res.get('date')== day]),
                        'vi': len([i['ProviderName'] for res in coll.find(providers, providers_projection) for i in res['result'] if i.get('ProviderName') == 'Vodafone' and res.get('requesttype') == req and res.get('date')== day]),
                        'co': len([i['ProviderName'] for res in coll.find(providers, providers_projection) for i in res['result'] if i.get('ProviderName') == 'Cellone' and res.get('requesttype') == req and res.get('date')== day]),
                        'Approval': len([ticket for ticket in coll.find(query2,query2_projection) if ticket.get('date')== day]),
                        'Pending': len([ticket for ticket in coll.find(query3, query3_projection) if ticket.get('date')== day]),
                        'Under_Process': len([ticket for ticket in coll.find(query4, query4_projection) if ticket.get('date')== day]),
                        'Ticket_closed': len([ticket for ticket in coll.find(query5, query5_projection) if ticket.get('date')== day]),
                        'Total_Pending' : len([ticket for ticket in coll.find(query3) if ticket.get('date')== day]) + len([ticket for ticket in coll.find(query4) if ticket.get('date')== day])

                    }
                elif date_range == 'Date Range':
                    start_date, end_date = map(str.strip, period.split(' to '))
                    start_date1 = datetime.strptime(start_date, '%d/%m/%Y')
                    end_date1 = datetime.strptime(end_date, '%d/%m/%Y')
                    samples = [res for res in coll.find({"requestcategory": category,"team":team,"modules":module,'designation': {'$nin': ['ADDL-SP/DSP','SP','DIG','IG']}},{'date' : 1, '_id':0})]
                    date_objects = [datetime.strptime(date['date'], '%d/%m/%Y') for date in samples]
                    result_dates = [date.strftime('%d/%m/%Y') for date in date_objects if start_date1 <= date <= end_date1]
                    results[req] = {
                        'Raise': len([ticket for ticket in coll.find(query1,query1_projection) if ticket.get('date') in result_dates]),
                        'jo': len([i['ProviderName'] for res in coll.find(providers, providers_projection) for i in res['result'] if i.get('ProviderName') == 'Jio' and res.get('requesttype') == req and res.get('date') in result_dates]),
                        'at': len([i['ProviderName'] for res in coll.find(providers, providers_projection) for i in res['result'] if i.get('ProviderName') == 'Airtel' and res.get('requesttype') == req and res.get('date') in result_dates]),
                        'vi': len([i['ProviderName'] for res in coll.find(providers, providers_projection) for i in res['result'] if i.get('ProviderName') == 'Vodafone' and res.get('requesttype') == req and res.get('date') in result_dates]),
                        'co': len([i['ProviderName'] for res in coll.find(providers, providers_projection) for i in res['result'] if i.get('ProviderName') == 'Cellone' and res.get('requesttype') == req and res.get('date') in result_dates]),
                        'Approval': len([ticket for ticket in coll.find(query2, query2_projection) if ticket.get('date') in result_dates]),
                        'Pending': len([ticket for ticket in coll.find(query3, query3_projection) if ticket.get('date') in result_dates]),
                        'Under_Process': len([ticket for ticket in coll.find(query4, query4_projection) if ticket.get('date') in result_dates]),
                        'Ticket_closed': len([ticket for ticket in coll.find(query5, query5_projection) if ticket.get('date') in result_dates]),
                        'Total_Pending' : len([ticket for ticket in coll.find(query3) if ticket.get('date') in result_dates]) + len([ticket for ticket in coll.find(query4) if ticket.get('date') in result_dates])
                    }
            else:
                if officer not in ['','All']:
                    query1 = {'requesttypes':{'$in': [re.compile(req)]},'requestcategory':'Analysis Request','officername':officer,"team":team,"modules":module,'designation': {'$nin': ['ADDL-SP/DSP','SP','DIG','IG']}}
                    query2 = {'requesttypes':{'$in': [re.compile(req)]},'requestcategory':'Analysis Request','pending': {'$in':['--']},'officername':officer,"team":team,"modules":module,'designation': {'$nin': ['ADDL-SP/DSP','SP','DIG','IG']}}
                    query3 = {'requesttypes':{'$in': [re.compile(req)]}, 'requestcategory':'Analysis Request','pending': {'$in':['Assign to CAT Ins',re.compile(r"Pending")]},"team":team, 'officername': officer,"modules":module,'designation': {'$nin': ['ADDL-SP/DSP','SP','DIG','IG']}}    
                    query4 = {'requesttypes':{'$in': [re.compile(req)]}, 'requestcategory':'Analysis Request','approval': {'$in':[re.compile(r"Assigned")]},"team":team, 'officername': officer,"modules":module,'designation': {'$nin': ['ADDL-SP/DSP','SP','DIG','IG']}}                    
                    query5 = {'requesttypes':{'$in': [re.compile(req)]}, 'requestcategory':'Analysis Request','pending': {'$in':[re.compile(r"Closed")]},"team":team, 'officername': officer,"modules":module,'designation': {'$nin': ['ADDL-SP/DSP','SP','DIG','IG']}}                    

                else:
                    query1 = {'requesttypes':{'$in': [re.compile(req)]},'requestcategory':'Analysis Request',"team":team,"modules":module,'designation': {'$nin': ['ADDL-SP/DSP','SP','DIG','IG']}}
                    query2 = {'requesttypes':{'$in': [re.compile(req)]},'requestcategory':'Analysis Request',"team":team,'pending': {'$in':['--']},"modules":module,'designation': {'$nin': ['ADDL-SP/DSP', 'SP', 'DIG', 'IG']}}
                    query3 = {'requesttypes':{'$in': [re.compile(req)]},'requestcategory':'Analysis Request',"team":team, 'pending': {'$in':['Assign to CAT Ins',re.compile(r"Pending")]},"modules":module,'designation': {'$nin': ['ADDL-SP/DSP','SP','DIG','IG']}}
                    query4 = {'requesttypes':{'$in': [re.compile(req)]}, 'requestcategory':'Analysis Request','approval': {'$in':[re.compile(r"Assigned")]},"team":team, 'officername': officer,"modules":module,'designation': {'$nin': ['ADDL-SP/DSP','SP','DIG','IG']}}                    
                    query5 = {'requesttypes':{'$in': [re.compile(req)]}, 'requestcategory':'Analysis Request','pending': {'$in':[re.compile(r"Closed")]},"team":team, 'officername': officer,"modules":module,'designation': {'$nin': ['ADDL-SP/DSP','SP','DIG','IG']}}                    

                if period == 'All':
                    results[req] = {
                        'Raise': len([ticket for ticket in coll.find(query1,additional_providers_projection)]),
                        'Approval': len([ticket for ticket in coll.find(query2,additional_providers_projection)]),
                        'Pending': len([ticket for ticket in coll.find(query3,additional_providers_projection)]),
                        'Assigned':  len([ticket for ticket in coll.find(query4)]),
                        'Ticket_closed':  len([ticket for ticket in coll.find(query5)]),
                    }
                elif period =='Today':
                    results[req] = {
                        'Raise': len([ticket for ticket in coll.find(query1,additional_providers_projection) if ticket.get('date')== day]),
                        'Approval': len([ticket for ticket in coll.find(query2,additional_providers_projection) if ticket.get('date')== day]),
                        'Pending': len([ticket for ticket in coll.find(query3,additional_providers_projection) if ticket.get('date')== day]),
                        'Assigned':  len([ticket for ticket in coll.find(query4)]),
                        'Ticket_closed':  len([ticket for ticket in coll.find(query5)]),
                    }
                elif date_range == 'Date Range':
                    start_date, end_date = map(str.strip, period.split(' to '))
                    start_date1 = datetime.strptime(start_date, '%d/%m/%Y')
                    end_date1 = datetime.strptime(end_date, '%d/%m/%Y')
                    samples = [res for res in coll.find({"requestcategory": category,"team":team,"modules":module,'designation': {'$nin': ['ADDL-SP/DSP','SP','DIG','IG']}},{'date': 1, "_id" : 0})]
                    date_objects = [datetime.strptime(date['date'], '%d/%m/%Y') for date in samples]
                    result_dates = [date.strftime('%d/%m/%Y') for date in date_objects if start_date1 <= date <= end_date1]
                    results[req] = {
                        'Raise': len([ticket for ticket in coll.find(query1,additional_providers_projection) if ticket.get('date') in result_dates]),
                        'Approval': len([ticket for ticket in coll.find(query2,additional_providers_projection) if ticket.get('date') in result_dates]),
                        'Pending': len([ticket for ticket in coll.find(query3,additional_providers_projection) if ticket.get('date') in result_dates]),
                        'Assigned':  len([ticket for ticket in coll.find(query4)]),
                        'Ticket_closed':  len([ticket for ticket in coll.find(query5)]),
                    }
    else:

        query1_projection = {'requesttype': 1,"date": 1,'_id': 0}
        providers_projection = {'requesttype': 1,'result.ProviderName': 1, "date": 1,'_id': 0}
        query2_projection = {'pending': 1,"date": 1,'_id': 0}
        query3_projection = {'pending': 1,"date": 1,'_id': 0}
        query4_projection = {'pending': 1,"date": 1,'_id': 0}
        query5_projection = {'pending': 1,"date": 1,'_id': 0}
        additional_providers_projection = {'requesttypes': 1,'newnumber.ISP': 1, 'pending' : 1,"date": 1,'_id': 0}
        for req in results['reqtypes']:
            if category in ['Data Request', '', 'All']:
                query1 = {'requesttype': req,'requestcategory': 'Data Request','useremail':email}
                query2 = {'requesttype': req,'requestcategory': 'Data Request','pending': 'Mail Under Process','useremail':email}
                query3 = {'requesttype': req, 'requestcategory': 'Data Request', 'pending': re.compile(r"Pending"),'useremail':email}
                query4 = {'requesttype': req, 'requestcategory': 'Data Request', 'pending': 'Mail Under Process', 'newnumber.status': {'$in': ['Data_Received', 'Data Not Received']},'useremail':email}
                query5 = {'requesttype': req, 'requestcategory': 'Data Request', 'pending': 'Ticket_Closed','useremail':email}
                providers = {"requesttype": req, "requestcategory":"Data Request","result.ProviderName": {"$exists": True},'useremail':email}
                if period == 'All':
                    results[req] = {
                        'Raise': len([ticket for ticket in coll.find(query1,query1_projection)]),
                        'jo': len([i['ProviderName'] for res in coll.find(providers, providers_projection) for i in res['result'] if i.get('ProviderName') == 'Jio' and res.get('requesttype') == req]),
                        'at': len([i['ProviderName'] for res in coll.find(providers, providers_projection) for i in res['result'] if i.get('ProviderName') == 'Airtel' and res.get('requesttype') == req]),
                        'vi': len([i['ProviderName'] for res in coll.find(providers, providers_projection) for i in res['result'] if i.get('ProviderName') == 'Vodafone' and res.get('requesttype') == req]),
                        'co': len([i['ProviderName'] for res in coll.find(providers, providers_projection) for i in res['result'] if i.get('ProviderName') == 'Cellone' and res.get('requesttype') == req]),
                        'Approval': len([ticket for ticket in coll.find(query2, query2_projection)]),
                        'Pending': len([ticket for ticket in coll.find(query3, query3_projection)]),
                        'Under_Process': len([ticket for ticket in coll.find(query4, query4_projection)]),
                        'Ticket_closed': len([ticket for ticket in coll.find(query5, query5_projection)]),
                        'Total_Pending' : len([ticket for ticket in coll.find(query3)]) + len([ticket for ticket in coll.find(query4)])
                    }
                elif period == 'Today':
                    results[req] = {
                        'Raise': len([ticket for ticket in coll.find(query1,query1_projection) if ticket.get('date') == day]),
                        'jo': len([i['ProviderName'] for res in coll.find(providers, providers_projection) for i in res['result'] if i.get('ProviderName') == 'Jio' and res.get('requesttype') == req and res.get('date') == day]),
                        'at': len([i['ProviderName'] for res in coll.find(providers, providers_projection) for i in res['result'] if i.get('ProviderName') == 'Airtel' and res.get('requesttype') == req and res.get('date') == day]),
                        'vi': len([i['ProviderName'] for res in coll.find(providers, providers_projection) for i in res['result'] if i.get('ProviderName') == 'Vodafone' and res.get('requesttype') == req and res.get('date') == day]),
                        'co': len([i['ProviderName'] for res in coll.find(providers, providers_projection) for i in res['result'] if i.get('ProviderName') == 'Cellone' and res.get('requesttype') == req and res.get('date') == day]),
                        'Approval': len([ticket for ticket in coll.find(query2, query2_projection) if ticket.get('date') == day]),
                        'Pending': len([ticket for ticket in coll.find(query3, query3_projection) if ticket.get('date') == day]),
                        'Under_Process': len([ticket for ticket in coll.find(query4, query4_projection) if ticket.get('date') == day]),
                        'Ticket_closed': len([ticket for ticket in coll.find(query5, query5_projection) if ticket.get('date') == day]),
                        'Total_Pending' : len([ticket for ticket in coll.find(query3) if ticket.get('date')== day]) + len([ticket for ticket in coll.find(query4) if ticket.get('date')== day])
                    }
                elif date_range == 'Date Range':
                    start_date, end_date = map(str.strip, period.split(' to '))
                    start_date1 = datetime.strptime(start_date, '%d/%m/%Y')
                    end_date1 = datetime.strptime(end_date, '%d/%m/%Y')
                    samples = [res for res in coll.find({"requestcategory": category, 'useremail': email}, {'date': 1, "_id": 0})]
                    date_objects = [datetime.strptime(date['date'], '%d/%m/%Y') for date in samples]
                    result_dates = [date.strftime('%d/%m/%Y') for date in date_objects if start_date1 <= date <= end_date1]
                    results[req] = {
                        'Raise': len([ticket for ticket in coll.find(query1, query1_projection) if ticket.get('date') in result_dates]),
                        'jo': len([i['ProviderName'] for res in coll.find(providers, providers_projection) for i in res['result'] if i.get('ProviderName') == 'Jio' and res.get('requesttype') == req and res.get('date') in result_dates]),
                        'at': len([i['ProviderName'] for res in coll.find(providers, providers_projection) for i in res['result'] if i.get('ProviderName') == 'Airtel' and res.get('requesttype') == req and res.get('date') in result_dates]),
                        'vi': len([i['ProviderName'] for res in coll.find(providers, providers_projection) for i in res['result'] if i.get('ProviderName') == 'Vodafone' and res.get('requesttype') == req and res.get('date') in result_dates]),
                        'co': len([i['ProviderName'] for res in coll.find(providers, providers_projection) for i in res['result'] if i.get('ProviderName') == 'Cellone' and res.get('requesttype') == req and res.get('date') in result_dates]),
                        'Approval': len([ticket for ticket in coll.find(query2, query2_projection) if ticket.get('date') in result_dates]),
                        'Pending': len([ticket for ticket in coll.find(query3, query3_projection) if ticket.get('date') in result_dates]),
                        'Under_Process': len([ticket for ticket in coll.find(query4, query4_projection) if ticket.get('date') in result_dates]),
                        'Ticket_closed': len([ticket for ticket in coll.find(query5, query5_projection) if ticket.get('date') in result_dates]),
                        'Total_Pending' : len([ticket for ticket in coll.find(query3) if ticket.get('date') in result_dates]) + len([ticket for ticket in coll.find(query4) if ticket.get('date') in result_dates])
                    }

            else:
                query1 = {'requesttypes':{'$in': [re.compile(req)]},'requestcategory': 'Analysis Request','useremail':email}
                query2 = {'requesttypes':{'$in': [re.compile(req)]},'requestcategory': 'Analysis Request',"team":team,'pending': {'$in':['--']},"modules":module,'designation': {'$nin': ['ADDL-SP/DSP','SP','DIG','IG']}}
                query3 = {'requesttypes':{'$in': [re.compile(req)]},'requestcategory': 'Analysis Request',"team":team, 'pending': {'$in':['Assign to CAT Ins',re.compile(r"Pending")]},"modules":module,'designation': {'$nin': ['ADDL-SP/DSP','SP','DIG','IG']}}
                query4 = {'requesttypes':{'$in': [re.compile(req)]},'requestcategory': 'Analysis Request',"team":team, 'assign_Officer': officername,"modules":module,'designation': {'$nin': ['ADDL-SP/DSP','SP','DIG','IG']}}
                query5 = {'requesttypes':{'$in': [re.compile(req)]},'requestcategory': 'Analysis Request',"team":team, 'assign_Officer': officername,'pending': {'$in':[re.compile(r"Closed")]},"modules":module,'designation': {'$nin': ['ADDL-SP/DSP','SP','DIG','IG']}}

                if period == 'All':
                    results[req] = {
                        'Raise': len([ticket for ticket in coll.find(query1,additional_providers_projection)]),
                        'Approval': len([ticket for ticket in coll.find(query2,additional_providers_projection)]),
                        'Pending':  len([ticket for ticket in coll.find(query3,additional_providers_projection)]),
                        'Assigned':  len([ticket for ticket in coll.find(query4)]),
                        'Ticket_closed':  len([ticket for ticket in coll.find(query5)]),
                    }
                elif period =='Today':
                    results[req] = {
                        'Raise':    len([ticket for ticket in coll.find(query1,additional_providers_projection) if ticket.get('date')== day]),
                        'Approval': len([ticket for ticket in coll.find(query2,additional_providers_projection) if ticket.get('date')== day]),
                        'Pending':  len([ticket for ticket in coll.find(query3,additional_providers_projection) if ticket.get('date')== day]),
                        'Assigned':  len([ticket for ticket in coll.find(query4)]),
                        'Ticket_closed':  len([ticket for ticket in coll.find(query5)]),
                    }
                elif date_range == 'Date Range':
                    start_date, end_date = map(str.strip, period.split(' to '))
                    start_date1 = datetime.strptime(start_date, '%d/%m/%Y')
                    end_date1 = datetime.strptime(end_date, '%d/%m/%Y')
                    samples = [res for res in coll.find({"requestcategory": category,'useremail':email},{'date': 1, "_id": 0})]
                    date_objects = [datetime.strptime(date['date'], '%d/%m/%Y') for date in samples]
                    result_dates = [date.strftime('%d/%m/%Y') for date in date_objects if start_date1 <= date <= end_date1]
                    results[req] = {
                        'Raise':    len([ticket for ticket in coll.find(query1,additional_providers_projection) if ticket.get('date') in result_dates]),
                        'Approval': len([ticket for ticket in coll.find(query2,additional_providers_projection) if ticket.get('date') in result_dates]),
                        'Pending':  len([ticket for ticket in coll.find(query3,additional_providers_projection) if ticket.get('date') in result_dates]),
                        'Assigned':  len([ticket for ticket in coll.find(query4)]),
                        'Ticket_closed':  len([ticket for ticket in coll.find(query5)]),
                    }

                   
    main_data.append(results)
    return jsonify(main_data)

@tickect_bp.route('/fetch_ref', methods=['POST', 'GET'])
def fetch_ref():
    email = request.json.get('email')
    team = request.json.get("team")
    officername = request.json.get("name")
    coll = db['tickets']
    data1 = []
    data = coll.find({'useremail':email,'team':team})
    for res in data:
        data1.append(res)
    return jsonify(data1)

@tickect_bp.route('/user_data_upload', methods=['POST', 'GET'])
def user_data_upload():
    coll = db['users']
    # coll2 = db['users']
    data = request.json.get('data')
    for res in data:
        if coll.find_one({'email':res['email']}):
            print('FOUND!!!!!')
        else:
            #print('NOT FOUND!!!!!!')
            coll.insert_one({
            'name': res['name'],
            'email': res['email'],
            'password1': res['password1'],
            'password2': res['password2'],
            'mobilenumber': res['mobilenumber'],
            'designation': res['designation'],
            'officername': res['officername'],
            'team': res['team'],
            'modules': res['modules'],
            'role': res['role'],
            'superior': res['superior'],
            'Application':res['Application'],
            'status': "Active",
            'availability': "Active"
            })
    return jsonify({'messege':'Data Uploaded!!!!!!!!!!'})

@tickect_bp.route('/ref_no', methods=['POST', 'GET'])
def ref_no():
    coll = db['tickets']
    req = request.json.get('request')
    team = request.json.get('team')
    modules = request.json.get('modules')
    new_refno = ''
    # Step 1: Find the maximum refno based on team and modules
    if req == "Analysis Request":
        filter_criteria = {'requestcategory': req, 'team': team, 'modules': modules}
        max_refno_entry = coll.find_one(filter_criteria, sort=[("refno", -1)])
        max_refno = max_refno_entry['refno'] if max_refno_entry else -1

        # Step 2: Increment the refno for the new entry
        new_refno = int(max_refno) + 1 if int(max_refno) >= 1 else 1
        #print(new_refno)

    return jsonify({'new_ref': str(new_refno)})

# @tickect_bp.route('/emergency_alert', methods=['POST', 'GET'])
# def emergency_alert():
#     coll2 = db['dummy_emergency_counter']
#     alerts = []
#     data = request.json.get('data')
#     for res in data:
#         # #print(res['priority'])
#         if res['priority'] == "Emergency" and res['pending'] == 'Mail Under Process' and res['requestcategory'] == 'Data Request':
#             alerts.append(res)
            
#     if coll2.find_one({'emergency_count':str(len(alerts))}):
#         pass
#     else:
#         coll2.insert_one({'emergency_count':str(len(alerts))})
#         #print('data inserted.........')
#         return jsonify({'count':str(len(alerts))})
        
#     return jsonify('NO NEW UPDATE....')

@tickect_bp.route('/data_filters', methods=['POST', 'GET'])
def data_filters():
    coll = db['tickets']
    team = request.json.get('team')
    otherteam = request.json.get('teams')
    otherofficer = request.json.get('officer')

    page = request.json.get('page')
    limit = request.json.get('limit')
    category = request.json.get('category')
    designation = request.json.get('designation')
    role = request.json.get('role')
    module = request.json.get('module')
    email = request.json.get('email')
    types = request.json.get('types')
    status = request.json.get('status')
    period = request.json.get('period')
    date_range = request.json.get('date_range')
    search = []
    now = datetime.now()
    day = now.strftime("%d/%m/%Y")
    yesterday = now - timedelta(days=1)
    yesterday_date = yesterday.strftime("%d/%m/%Y")
    # Get the date period for the last 3 months from today
    three_months_ago = now - timedelta(days=90)
    last_three_months_start_date = three_months_ago.strftime("%d/%m/%Y")
    last_three_months_end_date = now.strftime("%d/%m/%Y")            

    if team == 'ADMIN':
        query = {'requestcategory': category,}
    elif designation =='SP' and team != 'ADMIN':
        query = {'requestcategory': category,'team' : team}
    elif designation == 'ADDL-SP/DSP':
        query = {'requestcategory': category,'team' : team, 'modules': module, 'designation' : {'$nin' : ['SP','DIG','IG']}}
    elif designation == 'INSPR' and role != 'Analyst':
        query = {'requestcategory': category,'team' : team, 'modules': module, 'designation' : {'$nin' : ['SP','DIG','IG','ADDL-SP/DSP']}}
    else:
        query = {'requestcategory': category,'useremail':email}

    projection = {
            '_id' : 1,
            'priority' : 1,
            'date' : 1,
            'token' : 1,
            'team' : 1,
            'officername' : 1,
            'approval' : 1,
            'pending' : 1,
            'requestcategory' : 1,
            'requesttype' : 1,
            'requesttypes' : 1,
            'subtypes' : 1,
            'status_count': 1,
            'result': 1,
            'newnumber': 1,            
        }
    
    if page is None or limit is None:
        page = 1
        limit = 6
    skip = (page - 1) * limit
    if category == 'Data Request':
        if period == 'All':
            if (types not in ['All'] and status not in ['Total_raise','Total_approval','Total_pending','Total_close','Total_under_process']):

                if otherofficer not in ['','All'] and otherteam in ['','All']:
                    query.update({'requesttype': types, 'officername': otherofficer})
                elif otherofficer in ['','All'] and otherteam not in ['','All']:
                    query.update({'requesttype': types, 'team': otherteam})
                elif otherofficer not in ['','All'] and otherteam not in ['','All']:
                    query.update({'requesttype': types, 'team': otherteam,'officername': otherofficer})
                else:
                    query.update({'requesttype': types})

            elif (types not in ['All'] and status == 'Total_approval') or (types == 'All' and status == 'Total_approval'):

                if (types not in ['All'] and status == 'Total_approval'):
                    if otherofficer not in ['','All'] and otherteam in ['','All']:
                        query.update({'requesttype': types, 'officername': otherofficer, 'pending': {'$in': ["Mail Under Process"]}})
                    elif otherofficer in ['','All'] and otherteam not in ['','All']:
                        query.update({'requesttype': types, 'team': otherteam, 'pending': {'$in': ["Mail Under Process"]}})
                    elif otherofficer not in ['','All'] and otherteam not in ['','All']:
                        query.update({'requesttype': types, 'team': otherteam,'officername': otherofficer,'pending': {'$in': ["Mail Under Process"]}})
                    else:
                        query.update({"requesttype":types,'pending': {'$in': ["Mail Under Process"]}})

                else:
                    if otherofficer not in ['','All'] and otherteam in ['','All']:
                        query.update({'officername': otherofficer, 'pending': {'$in': ["Mail Under Process"]}})
                    elif otherofficer in ['','All'] and otherteam not in ['','All']:
                        query.update({'team': otherteam, 'pending': {'$in': ["Mail Under Process"]}})
                    elif otherofficer not in ['','All'] and otherteam not in ['','All']:
                        query.update({'team': otherteam,'officername': otherofficer,'pending': {'$in': ["Mail Under Process"]}})
                    else:
                        query.update({'pending': {'$in': ["Mail Under Process"]}})

            elif (types not in ['All'] and status == 'Total_pending') or (types == 'All' and status == 'Total_pending'):
                if (types not in ['All'] and status == 'Total_pending'):
                    if otherofficer not in ['','All'] and otherteam in ['','All']:
                        query.update({'requesttype': types, 'officername': otherofficer, 'pending': re.compile(r"Pending")})

                    elif otherofficer in ['','All'] and otherteam not in ['','All']:
                        query.update({'requesttype': types, 'team': otherteam, 'pending': re.compile(r"Pending")})

                    elif otherofficer not in ['','All'] and otherteam not in ['','All']:
                        query.update({'requesttype': types, 'team': otherteam,'officername': otherofficer,'pending': re.compile(r"Pending")})
                    else:
                        query.update({"requesttype":types,'pending': re.compile(r"Pending")})
                else:
                    if otherofficer not in ['','All'] and otherteam in ['','All']:
                        query.update({'officername': otherofficer, 'pending': re.compile(r"Pending")})

                    elif otherofficer in ['','All'] and otherteam not in ['','All']:
                        query.update({'team': otherteam, 'pending': re.compile(r"Pending")})
                    elif otherofficer not in ['','All'] and otherteam not in ['','All']:
                        query.update({'team': otherteam,'officername': otherofficer,'pending': re.compile(r"Pending")})
                    else:
                        query.update({'pending': re.compile(r"Pending")})

            elif (types not in ['All'] and status == 'Total_close') or (types == 'All' and status == 'Total_close'):
                if (types not in ['All'] and status == 'Total_close'):
                    if otherofficer not in ['','All'] and otherteam in ['','All']:
                        query.update({'requesttype': types, 'officername': otherofficer, 'pending': re.compile(r"Closed")})

                    elif otherofficer in ['','All'] and otherteam not in ['','All']:
                        query.update({'requesttype': types, 'team': otherteam, 'pending': re.compile(r"Closed")})

                    elif otherofficer not in ['','All'] and otherteam not in ['','All']:
                        query.update({'requesttype': types, 'team': otherteam,'officername': otherofficer,'pending': re.compile(r"Closed")})
                    else:
                        query.update({"requesttype":types,'pending': re.compile(r"Closed")})
                else:
                    if otherofficer not in ['','All'] and otherteam in ['','All']:
                        query.update({'officername': otherofficer, 'pending': re.compile(r"Closed")})

                    elif otherofficer in ['','All'] and otherteam not in ['','All']:
                        query.update({'team': otherteam, 'pending': re.compile(r"Closed")})

                    elif otherofficer not in ['','All'] and otherteam not in ['','All']:
                        query.update({'team': otherteam,'officername': otherofficer,'pending': re.compile(r"Closed")})
                    else:
                        query.update({'pending': re.compile(r"Closed")})

            elif (types not in ['All'] and status == 'Total_under_process') or (types == 'All' and status == 'Total_under_process'):
                if (types not in ['All'] and status == 'Total_under_process'):
                    if otherofficer not in ['','All'] and otherteam in ['','All']:
                        query.update({'requesttype': types, 'officername': otherofficer, 'pending': re.compile(r"Mail"),
                                        '$or': [{'result.status': 'Data_Received'}, {'result.status': 'Data Not Received'}, {'newnumber.status': 'Data_Received'}, {'newnumber.status': 'Data Not Received'}]})

                    elif otherofficer in ['','All'] and otherteam not in ['','All']:
                        query.update({'requesttype': types, 'team': otherteam, 'pending': re.compile(r"Mail"),
                                        '$or': [{'result.status': 'Data_Received'}, {'result.status': 'Data Not Received'}, {'newnumber.status': 'Data_Received'}, {'newnumber.status': 'Data Not Received'}]})
                        
                    elif otherofficer not in ['','All'] and otherteam not in ['','All']:
                        query.update({'requesttype': types,'team': otherteam,'officername': otherofficer,'pending': re.compile(r"Mail"),
                                      '$or': [{'result.status': 'Data_Received'}, {'result.status': 'Data Not Received'}, {'newnumber.status': 'Data_Received'}, {'newnumber.status': 'Data Not Received'}]})
                    else:
                        query.update({"requesttype":types,'pending': re.compile(r"Mail"),
                                '$or': [{'result.status': 'Data_Received'}, {'result.status': 'Data Not Received'}, {'newnumber.status': 'Data_Received'}, {'newnumber.status': 'Data Not Received'}]})
                else:
                    if otherofficer not in ['','All'] and otherteam in ['','All']:
                        query.update({'officername': otherofficer, 'pending': re.compile(r"Mail"),
                                        '$or': [{'result.status': 'Data_Received'}, {'result.status': 'Data Not Received'}, {'newnumber.status': 'Data_Received'}, {'newnumber.status': 'Data Not Received'}]})

                    elif otherofficer in ['','All'] and otherteam not in ['','All']:
                        query.update({'team': otherteam, 'pending': re.compile(r"Mail"),
                                        '$or': [{'result.status': 'Data_Received'}, {'result.status': 'Data Not Received'}, {'newnumber.status': 'Data_Received'}, {'newnumber.status': 'Data Not Received'}]})
                        
                    elif otherofficer not in ['','All'] and otherteam not in ['','All']:
                        query.update({'team': otherteam,'officername': otherofficer,'pending': re.compile(r"Mail"),
                                      '$or': [{'result.status': 'Data_Received'}, {'result.status': 'Data Not Received'}, {'newnumber.status': 'Data_Received'}, {'newnumber.status': 'Data Not Received'}]})
                    else:
                        query.update({'pending': re.compile(r"Mail"),'$or': [{'result.status': 'Data_Received'}, {'result.status': 'Data Not Received'}, {'newnumber.status': 'Data_Received'}, {'newnumber.status': 'Data Not Received'}]})
            
            else:
                if otherofficer not in ['','All'] and otherteam in ['','All']:
                    query.update({'officername': otherofficer})

                elif otherofficer in ['','All'] and otherteam not in ['','All']:
                    query.update({'team': otherteam})
                elif otherofficer not in ['','All'] and otherteam not in ['','All']:
                    query.update({'team': otherteam,'officername': otherofficer})
                    

        elif period == 'Today':
            if (types not in ['All'] and status not in ['Total_raise','Total_approval','Total_pending','Total_close','Total_under_process']):
                if otherofficer not in ['','All'] and otherteam in ['','All']:
                    query.update({'requesttype': types, 'officername': otherofficer,'date': day})
                elif otherofficer in ['','All'] and otherteam not in ['','All']:
                    query.update({'requesttype': types, 'team': otherteam,'date': day})
                else:
                    query.update({'requesttype': types,'date': day})

            elif (types not in ['All'] and status == 'Total_approval') or (types == 'All' and status == 'Total_approval'):
                if (types not in ['All'] and status == 'Total_approval'):
                    if otherofficer not in ['','All'] and otherteam in ['','All']:
                        query.update({'requesttype': types,'date': day, 'officername': otherofficer, 'pending': {'$in': ["Mail Under Process"]}})
                    elif otherofficer in ['','All'] and otherteam not in ['','All']:
                        query.update({'requesttype': types,'date': day, 'team': otherteam, 'pending': {'$in': ["Mail Under Process"]}})
                    else:
                        query.update({"requesttype":types,'date': day, 'pending': {'$in': ["Mail Under Process"]}})
                else:
                    if otherofficer not in ['','All'] and otherteam in ['','All']:
                        query.update({'officername': otherofficer,'date': day, 'pending': {'$in': ["Mail Under Process"]}})
                    elif otherofficer in ['','All'] and otherteam not in ['','All']:
                        query.update({'team': otherteam,'date': day, 'pending': {'$in': ["Mail Under Process"]}})
                    else:
                        query.update({'pending': {'$in': ["Mail Under Process"]},'date': day})
            elif (types not in ['All'] and status == 'Total_pending') or (types == 'All' and status == 'Total_pending'):
                if (types not in ['All'] and status == 'Total_pending'):
                    if otherofficer not in ['','All'] and otherteam in ['','All']:
                        query.update({'requesttype': types,'date': day, 'officername': otherofficer, 'pending': re.compile(r"Pending")})
                    elif otherofficer in ['','All'] and otherteam not in ['','All']:
                        query.update({'requesttype': types,'date': day, 'team': otherteam, 'pending': re.compile(r"Pending")})
                    else:
                        query.update({"requesttype":types,'date': day, 'pending': re.compile(r"Pending")})
                else:
                    if otherofficer not in ['','All'] and otherteam in ['','All']:
                        query.update({'officername': otherofficer,'date': day, 'pending': re.compile(r"Pending")})
                    elif otherofficer in ['','All'] and otherteam not in ['','All']:
                        query.update({'team': otherteam,'date': day, 'pending': re.compile(r"Pending")})
                    else:
                        query.update({'pending': re.compile(r"Pending"),'date': day})
            elif (types not in ['All'] and status == 'Total_close') or (types == 'All' and status == 'Total_close'):
                if (types not in ['All'] and status == 'Total_close'):
                    if otherofficer not in ['','All'] and otherteam in ['','All']:
                        query.update({'requesttype': types,'date': day, 'officername': otherofficer, 'pending': re.compile(r"Closed")})
                    elif otherofficer in ['','All'] and otherteam not in ['','All']:
                        query.update({'requesttype': types,'date': day, 'team': otherteam, 'pending': re.compile(r"Closed")})
                    else:
                        query.update({"requesttype":types,'date': day,'pending': re.compile(r"Closed")})
                else:
                    if otherofficer not in ['','All'] and otherteam in ['','All']:
                        query.update({'officername': otherofficer, 'date': day,'pending': re.compile(r"Closed")})
                    elif otherofficer in ['','All'] and otherteam not in ['','All']:
                        query.update({'team': otherteam, 'date': day,'pending': re.compile(r"Closed")})
                    else:
                        query.update({'pending': re.compile(r"Closed"),'date': day})
            elif (types not in ['All'] and status == 'Total_under_process') or (types == 'All' and status == 'Total_under_process'):
                if (types not in ['All'] and status == 'Total_under_process'):
                    if otherofficer not in ['','All'] and otherteam in ['','All']:
                        query.update({'requesttype': types, 'date': day,'officername': otherofficer, 'pending': re.compile(r"Mail"),
                                        '$or': [{'result.status': 'Data_Received'}, {'result.status': 'Data Not Received'}, {'newnumber.status': 'Data_Received'}, {'newnumber.status': 'Data Not Received'}]})
                    elif otherofficer in ['','All'] and otherteam not in ['','All']:
                        query.update({'requesttype': types,'date': day, 'team': otherteam, 'pending': re.compile(r"Mail"),
                                        '$or': [{'result.status': 'Data_Received'}, {'result.status': 'Data Not Received'}, {'newnumber.status': 'Data_Received'}, {'newnumber.status': 'Data Not Received'}]})
                    else:
                        query.update({"requesttype":types,'date': day,'pending': re.compile(r"Mail"),
                                '$or': [{'result.status': 'Data_Received'}, {'result.status': 'Data Not Received'}, {'newnumber.status': 'Data_Received'}, {'newnumber.status': 'Data Not Received'}]})
                else:
                    if otherofficer not in ['','All'] and otherteam in ['','All']:
                        query.update({'officername': otherofficer,'date': day, 'pending': re.compile(r"Mail"),
                                        '$or': [{'result.status': 'Data_Received'}, {'result.status': 'Data Not Received'}, {'newnumber.status': 'Data_Received'}, {'newnumber.status': 'Data Not Received'}]})

                    elif otherofficer in ['','All'] and otherteam not in ['','All']:
                        query.update({'team': otherteam,'date': day, 'pending': re.compile(r"Mail"),
                                        '$or': [{'result.status': 'Data_Received'}, {'result.status': 'Data Not Received'}, {'newnumber.status': 'Data_Received'}, {'newnumber.status': 'Data Not Received'}]})
                    else:
                        query.update({'pending': re.compile(r"Mail"),'date': day,'$or': [{'result.status': 'Data_Received'}, {'result.status': 'Data Not Received'}, {'newnumber.status': 'Data_Received'}, {'newnumber.status': 'Data Not Received'}]})
            
            else:
                if otherofficer not in ['','All'] and otherteam in ['','All']:
                    query.update({'officername': otherofficer,'date': day,})

                elif otherofficer in ['','All'] and otherteam not in ['','All']:
                    query.update({'team': otherteam,'date': day,})
                else:
                    query.update({'date': day,})


        elif date_range == 'Date Range':
            start_date, end_date = map(str.strip, period.split(' to '))
            start_date1 = datetime.strptime(start_date, '%d/%m/%Y')
            end_date1 = datetime.strptime(end_date, '%d/%m/%Y')
            samples = [res for res in coll.find(query,{'date': 1,'_id': 0})]
            date_objects = [datetime.strptime(date['date'], '%d/%m/%Y') for date in samples]
            result_dates = [date.strftime('%d/%m/%Y') for date in date_objects if start_date1 <= date <= end_date1]
            
            if (types not in ['All'] and status not in ['Total_raise','Total_approval','Total_pending','Total_close','Total_under_process']):
                if otherofficer not in ['','All'] and otherteam in ['','All']:
                    query.update({'requesttype': types, 'officername': otherofficer,'date':  {'$in':result_dates}})
                elif otherofficer in ['','All'] and otherteam not in ['','All']:
                    query.update({'requesttype': types, 'team': otherteam,'date':  {'$in':result_dates}})
                else:
                    query.update({'requesttype': types,'date': {'$in':result_dates}})

            elif (types not in ['All'] and status == 'Total_approval') or (types == 'All' and status == 'Total_approval'):
                if (types not in ['All'] and status == 'Total_approval'):
                    if otherofficer not in ['','All'] and otherteam in ['','All']:
                        query.update({'requesttype': types,'date':  {'$in':result_dates}, 'officername': otherofficer, 'pending': {'$in': ["Mail Under Process"]}})
                    elif otherofficer in ['','All'] and otherteam not in ['','All']:
                        query.update({'requesttype': types,'date':  {'$in':result_dates}, 'team': otherteam, 'pending': {'$in': ["Mail Under Process"]}})
                    else:
                        query.update({"requesttype":types,'date':  {'$in':result_dates}, 'pending': {'$in': ["Mail Under Process"]}})
                else:
                    if otherofficer not in ['','All'] and otherteam in ['','All']:
                        query.update({'officername': otherofficer,'date':  {'$in':result_dates}, 'pending': {'$in': ["Mail Under Process"]}})
                    elif otherofficer in ['','All'] and otherteam not in ['','All']:
                        query.update({'team': otherteam,'date':  {'$in':result_dates}, 'pending': {'$in': ["Mail Under Process"]}})
                    else:
                        query.update({'pending': {'$in': ["Mail Under Process"]},'date':  {'$in':result_dates}})
            elif (types not in ['All'] and status == 'Total_pending') or (types == 'All' and status == 'Total_pending'):
                if (types not in ['All'] and status == 'Total_pending'):
                    if otherofficer not in ['','All'] and otherteam in ['','All']:
                        query.update({'requesttype': types,'date':  {'$in':result_dates}, 'officername': otherofficer, 'pending': re.compile(r"Pending")})
                    elif otherofficer in ['','All'] and otherteam not in ['','All']:
                        query.update({'requesttype': types,'date':  {'$in':result_dates}, 'team': otherteam, 'pending': re.compile(r"Pending")})
                    else:
                        query.update({"requesttype":types,'date':  {'$in':result_dates}, 'pending': re.compile(r"Pending")})
                else:
                    if otherofficer not in ['','All'] and otherteam in ['','All']:
                        query.update({'officername': otherofficer,'date':  {'$in':result_dates}, 'pending': re.compile(r"Pending")})
                    elif otherofficer in ['','All'] and otherteam not in ['','All']:
                        query.update({'team': otherteam,'date':  {'$in':result_dates}, 'pending': re.compile(r"Pending")})
                    else:
                        query.update({'pending': re.compile(r"Pending"),'date':  {'$in':result_dates}})
            elif (types not in ['All'] and status == 'Total_close') or (types == 'All' and status == 'Total_close'):
                if (types not in ['All'] and status == 'Total_close'):
                    if otherofficer not in ['','All'] and otherteam in ['','All']:
                        query.update({'requesttype': types,'date':  {'$in':result_dates}, 'officername': otherofficer, 'pending': re.compile(r"Closed")})
                    elif otherofficer in ['','All'] and otherteam not in ['','All']:
                        query.update({'requesttype': types,'date':  {'$in':result_dates}, 'team': otherteam, 'pending': re.compile(r"Closed")})
                    else:
                        query.update({"requesttype":types,'date':  {'$in':result_dates},'pending': re.compile(r"Closed")})
                else:
                    if otherofficer not in ['','All'] and otherteam in ['','All']:
                        query.update({'officername': otherofficer, 'date':  {'$in':result_dates},'pending': re.compile(r"Closed")})
                    elif otherofficer in ['','All'] and otherteam not in ['','All']:
                        query.update({'team': otherteam, 'date':  {'$in':result_dates},'pending': re.compile(r"Closed")})
                    else:
                        query.update({'pending': re.compile(r"Closed"),'date':  {'$in':result_dates}})
            elif (types not in ['All'] and status == 'Total_under_process') or (types == 'All' and status == 'Total_under_process'):
                if (types not in ['All'] and status == 'Total_under_process'):
                    if otherofficer not in ['','All'] and otherteam in ['','All']:
                        query.update({'requesttype': types, 'date':  {'$in':result_dates},'officername': otherofficer, 'pending': re.compile(r"Mail"),
                                        '$or': [{'result.status': 'Data_Received'}, {'result.status': 'Data Not Received'}, {'newnumber.status': 'Data_Received'}, {'newnumber.status': 'Data Not Received'}]})
                    elif otherofficer in ['','All'] and otherteam not in ['','All']:
                        query.update({'requesttype': types,'date':  {'$in':result_dates}, 'team': otherteam, 'pending': re.compile(r"Mail"),
                                        '$or': [{'result.status': 'Data_Received'}, {'result.status': 'Data Not Received'}, {'newnumber.status': 'Data_Received'}, {'newnumber.status': 'Data Not Received'}]})
                    else:
                        query.update({"requesttype":types,'date':  {'$in':result_dates},'pending': re.compile(r"Mail"),
                                '$or': [{'result.status': 'Data_Received'}, {'result.status': 'Data Not Received'}, {'newnumber.status': 'Data_Received'}, {'newnumber.status': 'Data Not Received'}]})
                else:
                    if otherofficer not in ['','All'] and otherteam in ['','All']:
                        query.update({'officername': otherofficer,'date':  {'$in':result_dates}, 'pending': re.compile(r"Mail"),
                                        '$or': [{'result.status': 'Data_Received'}, {'result.status': 'Data Not Received'}, {'newnumber.status': 'Data_Received'}, {'newnumber.status': 'Data Not Received'}]})

                    elif otherofficer in ['','All'] and otherteam not in ['','All']:
                        query.update({'team': otherteam,'date':  {'$in':result_dates}, 'pending': re.compile(r"Mail"),
                                        '$or': [{'result.status': 'Data_Received'}, {'result.status': 'Data Not Received'}, {'newnumber.status': 'Data_Received'}, {'newnumber.status': 'Data Not Received'}]})
                    else:
                        query.update({'pending': re.compile(r"Mail"),'date':  {'$in':result_dates},'$or': [{'result.status': 'Data_Received'}, {'result.status': 'Data Not Received'}, {'newnumber.status': 'Data_Received'}, {'newnumber.status': 'Data Not Received'}]})
            
            else:
                if otherofficer not in ['','All'] and otherteam in ['','All']:
                    query.update({'officername': otherofficer,'date':  {'$in':result_dates},})

                elif otherofficer in ['','All'] and otherteam not in ['','All']:
                    query.update({'team': otherteam,'date':  {'$in':result_dates},})
                else:
                    query.update({'date':  {'$in':result_dates},})


    if category == 'Analysis Request':
        if period == 'All':
            if (types not in ['All'] and status not in ['Total_raise','Total_approval','Total_pending','Total_close','Total_under_process']):

                if otherofficer not in ['','All'] and otherteam in ['','All']:
                    query.update({"requesttypes": {'$in': [types]}, 'officername': otherofficer})
                elif otherofficer in ['','All'] and otherteam not in ['','All']:
                    query.update({"requesttypes": {'$in': [types]}, 'team': otherteam})
                else:
                    query.update({"requesttypes": {'$in': [types]}})

            elif (types not in ['All'] and status == 'Total_approval') or (types == 'All' and status == 'Total_approval'):

                if (types not in ['All'] and status == 'Total_approval'):
                    if otherofficer not in ['','All'] and otherteam in ['','All']:
                        query.update({"requesttypes": {'$in': [types]}, 'officername': otherofficer, '$or': [{'pending': '--'},{'pending': re.compile(r"Assign")}]})
                    elif otherofficer in ['','All'] and otherteam not in ['','All']:
                        query.update({"requesttypes": {'$in': [types]}, 'team': otherteam, '$or': [{'pending': '--'},{'pending': re.compile(r"Assign")}]})
                    else:
                        query.update({"requesttypes": {'$in': [types]}, '$or': [{'pending': '--'},{'pending': re.compile(r"Assign")}]})

                else:
                    if otherofficer not in ['','All'] and otherteam in ['','All']:
                        query.update({'officername': otherofficer, '$or': [{'pending': '--'},{'pending': re.compile(r"Assign")}]})
                    elif otherofficer in ['','All'] and otherteam not in ['','All']:
                        query.update({'team': otherteam, '$or': [{'pending': '--'},{'pending': re.compile(r"Assign")}]})
                    else:
                        query.update({'$or': [{'pending': '--'},{'pending': re.compile(r"Assign")}]})

            elif (types not in ['All'] and status == 'Total_pending') or (types == 'All' and status == 'Total_pending'):
                if (types not in ['All'] and status == 'Total_pending'):
                    if otherofficer not in ['','All'] and otherteam in ['','All']:
                        query.update({"requesttypes": {'$in': [types]}, 'officername': otherofficer, 'pending': re.compile(r"Pending")})

                    elif otherofficer in ['','All'] and otherteam not in ['','All']:
                        query.update({"requesttypes": {'$in': [types]}, 'team': otherteam, 'pending': re.compile(r"Pending")})
                    else:
                        query.update({"requesttypes": {'$in': [types]},'pending': re.compile(r"Pending")})
                else:
                    if otherofficer not in ['','All'] and otherteam in ['','All']:
                        query.update({'officername': otherofficer, 'pending': re.compile(r"Pending")})

                    elif otherofficer in ['','All'] and otherteam not in ['','All']:
                        query.update({'team': otherteam, 'pending': re.compile(r"Pending")})
                    else:
                        query.update({'pending': re.compile(r"Pending")})
            else:
                if otherofficer not in ['','All'] and otherteam in ['','All']:
                    query.update({'officername': otherofficer})

                elif otherofficer in ['','All'] and otherteam not in ['','All']:
                    query.update({'team': otherteam})
        elif period == 'Today':
            if (types not in ['All'] and status not in ['Total_raise','Total_approval','Total_pending','Total_close','Total_under_process']):
                if otherofficer not in ['','All'] and otherteam in ['','All']:
                    query.update({"requesttypes": {'$in': [types]}, 'officername': otherofficer,'date': day})
                elif otherofficer in ['','All'] and otherteam not in ['','All']:
                    query.update({"requesttypes": {'$in': [types]}, 'team': otherteam,'date': day})
                else:
                    query.update({"requesttypes": {'$in': [types]},'date': day})

            elif (types not in ['All'] and status == 'Total_approval') or (types == 'All' and status == 'Total_approval'):
                if (types not in ['All'] and status == 'Total_approval'):
                    if otherofficer not in ['','All'] and otherteam in ['','All']:
                        query.update({"requesttypes": {'$in': [types]},'date': day, 'officername': otherofficer, '$or': [{'pending': '--'},{'pending': re.compile(r"Assign")}]})
                    elif otherofficer in ['','All'] and otherteam not in ['','All']:
                        query.update({"requesttypes": {'$in': [types]},'date': day, 'team': otherteam, '$or': [{'pending': '--'},{'pending': re.compile(r"Assign")}]})
                    else:
                        query.update({"requesttypes": {'$in': [types]},'date': day, '$or': [{'pending': '--'},{'pending': re.compile(r"Assign")}]})
                else:
                    if otherofficer not in ['','All'] and otherteam in ['','All']:
                        query.update({'officername': otherofficer,'date': day, '$or': [{'pending': '--'},{'pending': re.compile(r"Assign")}]})
                    elif otherofficer in ['','All'] and otherteam not in ['','All']:
                        query.update({'team': otherteam,'date': day, '$or': [{'pending': '--'},{'pending': re.compile(r"Assign")}]})
                    else:
                        query.update({'$or': [{'pending': '--'},{'pending': re.compile(r"Assign")}],'date': day})
            elif (types not in ['All'] and status == 'Total_pending') or (types == 'All' and status == 'Total_pending'):
                if (types not in ['All'] and status == 'Total_pending'):
                    if otherofficer not in ['','All'] and otherteam in ['','All']:
                        query.update({"requesttypes": {'$in': [types]},'date': day, 'officername': otherofficer, 'pending': re.compile(r"Pending")})
                    elif otherofficer in ['','All'] and otherteam not in ['','All']:
                        query.update({"requesttypes": {'$in': [types]},'date': day, 'team': otherteam, 'pending': re.compile(r"Pending")})
                    else:
                        query.update({"requesttypes": {'$in': [types]},'date': day, 'pending': re.compile(r"Pending")})
                else:
                    if otherofficer not in ['','All'] and otherteam in ['','All']:
                        query.update({'officername': otherofficer,'date': day, 'pending': re.compile(r"Pending")})
                    elif otherofficer in ['','All'] and otherteam not in ['','All']:
                        query.update({'team': otherteam,'date': day, 'pending': re.compile(r"Pending")})
                    else:
                        query.update({'pending': re.compile(r"Pending"),'date': day})
            
            else:
                if otherofficer not in ['','All'] and otherteam in ['','All']:
                    query.update({'officername': otherofficer,'date': day,})

                elif otherofficer in ['','All'] and otherteam not in ['','All']:
                    query.update({'team': otherteam,'date': day,})
                else:
                    query.update({'date': day,})


        elif date_range == 'Date Range':
            start_date, end_date = map(str.strip, period.split(' to '))
            start_date1 = datetime.strptime(start_date, '%d/%m/%Y')
            end_date1 = datetime.strptime(end_date, '%d/%m/%Y')
            samples = [res for res in coll.find(query,{'date': 1,'_id': 0})]
            date_objects = [datetime.strptime(date['date'], '%d/%m/%Y') for date in samples]
            result_dates = [date.strftime('%d/%m/%Y') for date in date_objects if start_date1 <= date <= end_date1]
            
            if (types not in ['All'] and status not in ['Total_raise','Total_approval','Total_pending','Total_close','Total_under_process']):
                if otherofficer not in ['','All'] and otherteam in ['','All']:
                    query.update({'requesttype': types, 'officername': otherofficer,'date':  {'$in':result_dates}})
                elif otherofficer in ['','All'] and otherteam not in ['','All']:
                    query.update({'requesttype': types, 'team': otherteam,'date':  {'$in':result_dates}})
                else:
                    query.update({'requesttype': types,'date': {'$in':result_dates}})

            elif (types not in ['All'] and status == 'Total_approval') or (types == 'All' and status == 'Total_approval'):
                if (types not in ['All'] and status == 'Total_approval'):
                    if otherofficer not in ['','All'] and otherteam in ['','All']:
                        query.update({'requesttype': types,'date':  {'$in':result_dates}, 'officername': otherofficer, '$or': [{'pending': '--'},{'pending': re.compile(r"Assign")}]})
                    elif otherofficer in ['','All'] and otherteam not in ['','All']:
                        query.update({'requesttype': types,'date':  {'$in':result_dates}, 'team': otherteam, '$or': [{'pending': '--'},{'pending': re.compile(r"Assign")}]})
                    else:
                        query.update({"requesttype":types,'date':  {'$in':result_dates}, '$or': [{'pending': '--'},{'pending': re.compile(r"Assign")}]})
                else:
                    if otherofficer not in ['','All'] and otherteam in ['','All']:
                        query.update({'officername': otherofficer,'date':  {'$in':result_dates}, '$or': [{'pending': '--'},{'pending': re.compile(r"Assign")}]})
                    elif otherofficer in ['','All'] and otherteam not in ['','All']:
                        query.update({'team': otherteam,'date':  {'$in':result_dates}, '$or': [{'pending': '--'},{'pending': re.compile(r"Assign")}]})
                    else:
                        query.update({'$or': [{'pending': '--'},{'pending': re.compile(r"Assign")}],'date':  {'$in':result_dates}})
            elif (types not in ['All'] and status == 'Total_pending') or (types == 'All' and status == 'Total_pending'):
                if (types not in ['All'] and status == 'Total_pending'):
                    if otherofficer not in ['','All'] and otherteam in ['','All']:
                        query.update({'requesttype': types,'date':  {'$in':result_dates}, 'officername': otherofficer, 'pending': re.compile(r"Pending")})
                    elif otherofficer in ['','All'] and otherteam not in ['','All']:
                        query.update({'requesttype': types,'date':  {'$in':result_dates}, 'team': otherteam, 'pending': re.compile(r"Pending")})
                    else:
                        query.update({"requesttype":types,'date':  {'$in':result_dates}, 'pending': re.compile(r"Pending")})
                else:
                    if otherofficer not in ['','All'] and otherteam in ['','All']:
                        query.update({'officername': otherofficer,'date':  {'$in':result_dates}, 'pending': re.compile(r"Pending")})
                    elif otherofficer in ['','All'] and otherteam not in ['','All']:
                        query.update({'team': otherteam,'date':  {'$in':result_dates}, 'pending': re.compile(r"Pending")})
                    else:
                        query.update({'pending': re.compile(r"Pending"),'date':  {'$in':result_dates}})
            else:
                if otherofficer not in ['','All'] and otherteam in ['','All']:
                    query.update({'officername': otherofficer,'date':  {'$in':result_dates},})

                elif otherofficer in ['','All'] and otherteam not in ['','All']:
                    query.update({'team': otherteam,'date':  {'$in':result_dates},})
                else:
                    query.update({'date':  {'$in':result_dates},})


    My_tickets = coll.find(query, projection).sort([('_id', -1)]).skip(skip).limit(limit)
    total_tickets = coll.count_documents(query)  
    filtered_tickets = []
    res_len = 0
    newno_len = 0
    total_len = 0
    for res in My_tickets:
        if res['pending'] == 'Mail Under Process' and (res['subtypes'] == 'Phone' or res['requesttype'] == 'CDR'):
            res_len = res['status_count']
            total_len = len(res['result'])
            
        elif res['pending'] == 'Mail Under Process' and (res['subtypes'] == 'IMEI' or res['requesttype'] == 'IMEI CDR'):
            newno_len = res['status_count']
            total_len = len(res['newnumber'])            

        filtered_tickets.append({
            'priority': res['priority'],
            'date': res['date'],
            '_id': res['_id'],
            'token': res['token'],
            'team': res['team'],
            'officername': res['officername'],
            'approval': res['approval'],
            'pending': res['pending'],
            'requestcategory': res['requestcategory'],
            'requesttype': res['requesttype'],
            'requesttypes': res['requesttypes'],
            'subtypes': res['subtypes'],
            'status_count': res['status_count'],
            'total_len': total_len,
            "res_len": res_len,
            "newno_len": newno_len
        })
    # Calculate pagination information
    total_pages = math.ceil(total_tickets / limit) # limit
    pagination_info = {
        'total_pages': int(total_pages),
        'current_page': page,
        'total_tickets': total_tickets
    }
    #print(pagination_info)
    # Send filtered_tickets and pagination_info to frontend
    return jsonify({'mytickets': filtered_tickets, 'pagination': pagination_info})

# @tickect_bp.route('/newform_converter', methods=['POST', 'GET'])
# def newform_converter():
#     newnumber = request.json.get('newnumber')
#     requesttype = request.json.get('requesttype')
#     subtypes = request.json.get('subtype')
#     if subtypes == "Phone" or requesttype == "CDR" or requesttype == "CAF" or requesttype == "SDR" or requesttype == "CAF/CDR" or requesttype == "ISD":
#         nickname =  request.json.get('nickname')
#         pending = "Approval Pending"
#     else:
#         nickname = ''
#         pending = "Approval Pending"
#     requesttype = request.json.get('requesttype')
#     demo = []
#     if (requesttype == "GPRS" and subtypes == "IMEI") or (requesttype == "IPDR" and subtypes == "IMEI") or (request == 'IMEI CDR'):
#         for phoneNumber in newnumber:
#             if "IMEI" in phoneNumber and (len(phoneNumber["IMEI"]) == 15 or len(phoneNumber["IMEI"]) == 14):
#                 phoneNumber["IMEI"] = phoneNumber["IMEI"][:14] + '0'
#             demo.append(phoneNumber)
#         newnumber=demo


#     result = []
#     try:
#         for res in newnumber:
#             #print(len(newnumber))
#             if res['result']['SDR'] != '':
#                 activation_date = res['result']['ActivationDate']
#                 res['result']['FetchedNickName'] = res['Nickname']
#                 from_date = res['From_Date']
#                 to_date = res['To_Date']
#                 mnp = res['result']['MNP']
#                 #print(activation_date)
#                 if activation_date == '' or activation_date == 'None' or activation_date != '':
#                     res['result']['From_Date'] = from_date
#                     res['result']['To_Date'] = to_date
#                 else:
#                     if 'JO' in mnp or 'Jio' in mnp or 'jio' in mnp or 'JIO' in mnp:
#                         #print(activation_date,'/////////////////////////////////////')
#                         from_date1=res['From_Date']
#                         #print('iniffff',type(from_date1))
#                         from_date=datetime.strptime(from_date1,"%d/%m/%Y")
#                         if len(activation_date)>10:
#                             jo=datetime.strptime(activation_date,"%Y-%m-%d %H:%M:%S")
#                             jo_date1=jo.strftime("%d/%m/%Y")
#                             jo_date=datetime.strptime(jo_date1,'%d/%m/%Y')
                            
#                             if from_date >= jo_date:
#                                 res['result']['From_Date'] = from_date1
#                                 res['result']['To_Date'] = to_date
#                             else:
#                                 res['result']['From_Date'] = jo_date1
#                                 res['result']['To_Date'] = to_date                        
#                         else:
#                             jo=datetime.strptime(activation_date,"%Y-%m-%d")
#                             jo_date1=jo.strftime("%d/%m/%Y")
#                             jo_date=datetime.strptime(jo_date1,'%d/%m/%Y')
#                             if from_date >= jo_date:
#                                 res['result']['From_Date'] = from_date1
#                                 res['result']['To_Date'] = to_date
#                             else:
#                                 res['result']['From_Date'] = jo_date1
#                                 res['result']['To_Date'] = to_date

#                     elif 'CO' in mnp:
#                         from_date1=res['From_Date']
#                         from_date=datetime.strptime(from_date1,"%d/%m/%Y")
#                         bsnl = datetime.strptime(activation_date, "%d-%b-%y")
#                         bsnl_date1=bsnl.strftime("%d/%m/%Y")
#                         bsnl_date=datetime.strptime(bsnl_date1,'%d/%m/%Y')
#                         if from_date >= bsnl_date:
#                             res['result']['From_Date'] = from_date1
#                             res['result']['To_Date'] = to_date
#                         else:
#                             res['result']['From_Date'] = bsnl_date1
#                             res['result']['To_Date'] = to_date

#                     elif 'VI' in mnp:
#                         from_date1=res['From_Date']
#                         from_date=datetime.strptime(from_date1,"%d/%m/%Y")
#                         vi=datetime.strptime(activation_date, "%d-%b-%Y")
#                         vi_date1=vi.strftime("%d/%m/%Y")
#                         vi_date = datetime.strptime(vi_date1,'%d/%m/%Y')
#                         if from_date >= vi_date :
#                             res['result']['From_Date'] = from_date1
#                             res['result']['To_Date'] = to_date
#                         else:
#                             res['result']['From_Date'] = vi_date1
#                             res['result']['To_Date'] = to_date

#                     elif 'AT' in mnp:
#                         from_date1=res['From_Date']
#                         from_date=datetime.strptime(from_date1,"%d/%m/%Y")
#                         at=datetime.strptime(activation_date, "%a, %d %b %Y %H:%M:%S %Z")
#                         at_date1=at.strftime("%d/%m/%Y")
#                         at_date = datetime.strptime(at_date1,'%d/%m/%Y')
#                         if from_date >= at_date:
#                             res['result']['From_Date'] = from_date1
#                             res['result']['To_Date'] = to_date
#                         else:
#                             res['result']['From_Date'] = at_date1
#                             res['result']['To_Date'] = to_date

#                 pending = "Approval Pending"
#                 result.append(res['result'])

#     except Exception as e:
#         print(e)

#     Tickets(
#             useremail =request.json.get('useremail'),
#             superior =request.json.get('superior'),
#             officername =request.json.get('officername'),
#             role =request.json.get('role'),
#             designation = request.json.get("designation"),
#             type = request.json.get('update'),
#             requestcategory = "Data Request",
#             requesttype = request.json.get('requesttype'),
#             subtypes = subtypes,
#             team = request.json.get('team'),
#             modules = request.json.get('modules'),
#             relation =  request.json.get('relation'),
#             nickname =  nickname,
#             suspect =  request.json.get('suspect'),
#             name =  request.json.get('name'),
#             location =  request.json.get('location'),
#             source = request.json.get('source'),
#             reason = request.json.get('reason'),
#             refno = request.json.get('refno'),
#             refdate=str(request.json.get('refdate')),
#             priority = request.json.get('priority'),
#             token = request.json.get('token'),
#             status = request.json.get('status'),
#             pending = pending,
#             approval = "-",
#             others='',
#             newnumber = newnumber,
#             result = result,
#             remarks = request.json.get('remarks'),
#             date = request.json.get("date")        
#             ).save()
#     return jsonify({'message': 'Data Insert successful'})
@tickect_bp.route('/newform_converter', methods=['POST', 'GET'])
def newform_converter():
    coll = db['tickets']
    newnumber = request.json.get('newnumber')
    id = request.json.get('id')    

    requesttype = request.json.get('requesttype')
    subtypes = request.json.get('subtype')
    if subtypes == "Phone" or requesttype == "CDR" or requesttype == "CAF" or requesttype == "SDR" or requesttype == "CAF/CDR" or requesttype == "ISD":
        nickname =  request.json.get('nickname')
        pending = "Approval Pending"
    else:
        nickname = ''
        pending = "Approval Pending"
    requesttype = request.json.get('requesttype')
    demo = []
    if (requesttype == "GPRS" and subtypes == "IMEI") or (requesttype == "IPDR" and subtypes == "IMEI") or (request == 'IMEI CDR'):
        for phoneNumber in newnumber:
            if "IMEI" in phoneNumber and (len(phoneNumber["IMEI"]) == 15 or len(phoneNumber["IMEI"]) == 14):
                phoneNumber["IMEI"] = phoneNumber["IMEI"][:14] + '0'
            demo.append(phoneNumber)
        newnumber=demo


    result = []
    try:
        for res in newnumber:
            #print(len(newnumber))
            if res['result']['SDR'] != '':
                activation_date = res['result']['ActivationDate']
                res['result']['FetchedNickName'] = res['Nickname']
                from_date = res['From_Date']
                to_date = res['To_Date']
                mnp = res['result']['MNP']
                #print(activation_date)
                if activation_date == '' or activation_date == 'None' or activation_date != '':
                    res['result']['From_Date'] = from_date
                    res['result']['To_Date'] = to_date
                else:
                    if 'JO' in mnp or 'Jio' in mnp or 'jio' in mnp or 'JIO' in mnp:
                        #print(activation_date,'/////////////////////////////////////')
                        from_date1=res['From_Date']
                        #print('iniffff',type(from_date1))
                        from_date=datetime.strptime(from_date1,"%d/%m/%Y")
                        if len(activation_date)>10:
                            jo=datetime.strptime(activation_date,"%Y-%m-%d %H:%M:%S")
                            jo_date1=jo.strftime("%d/%m/%Y")
                            jo_date=datetime.strptime(jo_date1,'%d/%m/%Y')
                            
                            if from_date >= jo_date:
                                res['result']['From_Date'] = from_date1
                                res['result']['To_Date'] = to_date
                            else:
                                res['result']['From_Date'] = jo_date1
                                res['result']['To_Date'] = to_date                        
                        else:
                            jo=datetime.strptime(activation_date,"%Y-%m-%d")
                            jo_date1=jo.strftime("%d/%m/%Y")
                            jo_date=datetime.strptime(jo_date1,'%d/%m/%Y')
                            if from_date >= jo_date:
                                res['result']['From_Date'] = from_date1
                                res['result']['To_Date'] = to_date
                            else:
                                res['result']['From_Date'] = jo_date1
                                res['result']['To_Date'] = to_date

                    elif 'CO' in mnp:
                        from_date1=res['From_Date']
                        from_date=datetime.strptime(from_date1,"%d/%m/%Y")
                        bsnl = datetime.strptime(activation_date, "%d-%b-%y")
                        bsnl_date1=bsnl.strftime("%d/%m/%Y")
                        bsnl_date=datetime.strptime(bsnl_date1,'%d/%m/%Y')
                        if from_date >= bsnl_date:
                            res['result']['From_Date'] = from_date1
                            res['result']['To_Date'] = to_date
                        else:
                            res['result']['From_Date'] = bsnl_date1
                            res['result']['To_Date'] = to_date

                    elif 'VI' in mnp:
                        from_date1=res['From_Date']
                        from_date=datetime.strptime(from_date1,"%d/%m/%Y")
                        vi=datetime.strptime(activation_date, "%d-%b-%Y")
                        vi_date1=vi.strftime("%d/%m/%Y")
                        vi_date = datetime.strptime(vi_date1,'%d/%m/%Y')
                        if from_date >= vi_date :
                            res['result']['From_Date'] = from_date1
                            res['result']['To_Date'] = to_date
                        else:
                            res['result']['From_Date'] = vi_date1
                            res['result']['To_Date'] = to_date

                    elif 'AT' in mnp:
                        from_date1=res['From_Date']
                        from_date=datetime.strptime(from_date1,"%d/%m/%Y")
                        at=datetime.strptime(activation_date, "%a, %d %b %Y %H:%M:%S %Z")
                        at_date1=at.strftime("%d/%m/%Y")
                        at_date = datetime.strptime(at_date1,'%d/%m/%Y')
                        if from_date >= at_date:
                            res['result']['From_Date'] = from_date1
                            res['result']['To_Date'] = to_date
                        else:
                            res['result']['From_Date'] = at_date1
                            res['result']['To_Date'] = to_date

                pending = "Approval Pending"
                result.append(res['result'])

    except Exception as e:
        print(e)

    coll.update_one({'_id': ObjectId(id)},{'$set' : {'status':'Converted'}})
    Tickets(
            useremail =request.json.get('useremail'),
            superior =request.json.get('superior'),
            officername =request.json.get('officername'),
            role =request.json.get('role'),
            designation = request.json.get("designation"),
            type = request.json.get('update'),
            requestcategory = "Data Request",
            requesttype = request.json.get('requesttype'),
            subtypes = subtypes,
            team = request.json.get('team'),
            modules = request.json.get('modules'),
            relation =  request.json.get('relation'),
            nickname =  nickname,
            suspect =  request.json.get('suspect'),
            name =  request.json.get('name'),
            location =  request.json.get('location'),
            source = request.json.get('source'),
            reason = request.json.get('reason'),
            refno = request.json.get('refno'),
            refdate=str(request.json.get('refdate')),
            priority = request.json.get('priority'),
            token = request.json.get('token'),
            status = 'Converted',
            pending = pending,
            approval = "-",
            others='',
            newnumber = newnumber,
            result = result,
            remarks = request.json.get('remarks'),
            raise_time = request.json.get('raise_time'),

            date = request.json.get("date")        
            ).save()
    
    tokens = request.json.get('token')
    user1= coll.find_one({'token':tokens,'requestcategory': 'Data Request'})
    if any(('result' in res2 and "CO" in res2["result"]["MNP"]) for res2 in user1["newnumber"]):
        split_documents(tokens,'CDR')

    elif any('MNP' in res2 and "CO" in res2["MNP"] for res2 in user1["newnumber"]):
        split_documents(tokens,'Others')

    elif user1['subtypes'] == 'CGI':
        split_documents(tokens,'CGI')
    return jsonify({'message': 'Data Insert successful'})

def allowed_file(filename):
    ALLOWED_EXTENSIONS = {'xlsx','xls'}
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

@tickect_bp.route('/upload', methods=['POST'])
def handle_file_upload():
    if 'file' not in request.files:
        return jsonify({'error': 'No file provided'}), 400
    file = request.files['file']
    reqtype = request.form['type']
    category = request.form['category']

    reqtype2 = request.form['type2']
    subtype = request.form['subtype']
    time = request.form['time']
    team = request.form['team']
    modules = request.form['modules']
    officername = request.form['officername']
    #print(category,reqtype,reqtype2,subtype)
    if file.filename == '':
        return jsonify({'error': 'No selected file'}), 400
    # Check if the file has an allowed extension
    if not allowed_file(file.filename):
        return jsonify({'error': 'Invalid file type'}), 400
    # Create the 'uploaded_excel_files' directory if it doesn't exist
    upload_path = f'{os.getcwd()}/app/ticketing/uploaded_excel_files/'
    if not os.path.exists(upload_path):
        os.makedirs(upload_path)
    # Save the file with a unique name in the 'uploaded_excel_files' folder

    filename_prefix = f"{officername}_{reqtype}_{time}"
    filename = secure_filename(f"{filename_prefix}_{file.filename}")
    filepath = os.path.join(upload_path, filename)
    serial_number = 1
    while os.path.exists(filepath):
        base_filename, extension = os.path.splitext(file.filename)
        serial_filename = f"{filename_prefix}_{base_filename}_{serial_number}{extension}"
        filepath = os.path.join(upload_path, serial_filename)
        serial_number += 1
    file.save(filepath)
    data = [{"newnumber":[],'others':[],'columns' : []}]
    # Perform file processing
    workbook = openpyxl.load_workbook(filepath)
    sheet = workbook.active
    columns = []
    for col in range(1, sheet.max_column + 1):
        column_name1 = sheet.cell(row=1, column=col).value or f'Column_{col}'
        columns.append(column_name1)
    data[0]['columns'] = columns
    provider = ''
    # LSA_ID = ''
    for row in sheet.iter_rows(min_row=2, values_only=True):
        row_data = {}
        for col, value in enumerate(row, start=1):
            column_name = sheet.cell(row=1, column=col).value or f'Column_{col}'
            # Check if the column is 'From_Date' or 'To_Date' and validate the date format
            if column_name in ['From_Date', 'To_Date'] and (reqtype != 'SDR' and reqtype != 'CAF'):
                try:
                    date_value = date_parser.parse(str(value))
                    date_value = date_value.strftime("%d/%m/%Y")
                except ParserError:
                    return jsonify({'error': f'Invalid date format in column {column_name}'}), 400
                row_data[column_name] = date_value
            else:
                row_data[column_name] = str(value).strip()

        if category == 'Data Request':
            if reqtype == 'CDR' or subtype == 'Phone':
                if row_data['MNP'] == 'None' and row_data['SDR'] == 'None':
                    return jsonify({'error': f'Please Fill MNP and SDR'}), 400
                else:
                    mnp = row_data['MNP']
                    if '-' in mnp:
                        split_values = mnp.split('-')
                        desired_value = split_values[1]
                        LSA_ID = desired_value
                        if 'JO' in mnp or 'Jio' in mnp or 'jio' in mnp or 'JIO' in mnp or 'Jo' in mnp:
                            provider = 'Jio'
                        elif 'CO' in mnp or 'co' in mnp or 'Co' in mnp:
                            provider = 'Cellone'
                        elif 'VI' in mnp or 'Vi' in mnp or 'vi' in mnp:
                            provider = 'Vodafone'
                        elif 'AT' in mnp or 'at' in mnp or 'At' in mnp:
                            provider = 'Airtel'
                    else:
                        return jsonify({'error': f'Invalid MNP Format: {mnp}, It should be JO-AP/AT-AP'}), 400
                
            if reqtype == 'CDR' or subtype == 'Phone':
                print('yayyyyyyyyyyyyyyy')
                data[0]['newnumber'].append({
                'Phno': row_data['Phone/IMEI/IP/CGI/Dealer'],
                'From_Date': row_data['From_Date'],
                'To_Date': row_data['To_Date'],
                'From_Time': row_data['From_Time'],
                'To_Time': row_data['To_Time'],
                'result' : {
                            'ActivationDate': row_data['ActivationDate'],
                            'Billing_City': "",
                            'Circle': "",
                            'Comments': "",
                            'Current_Owner_ID': "",
                            'E_sim': "",
                            'FetchedNickName': '',
                            'LSA_ID': LSA_ID,
                            'Local_Address': "",
                            'MNP': row_data['MNP'],
                            'MSISDN': row_data['Phone/IMEI/IP/CGI/Dealer'],
                            'Name': "",
                            'Permanent_Address': "",
                            'Portability': "",
                            'ProviderName': provider,
                            'SDR': row_data['SDR'],
                            'Truecaller': row_data['Truecaller'],
                            'Type': "",
                            'Unit': "",
                            'searchid': "",
                            'Remarks' : row_data['remarks']
                        }
                    })
                data[0]['others'].append({
                    'module': row_data['module'],
                    'nickname': row_data['nickname'],
                    'source': row_data['source'],
                    'priority': row_data['priority']
                })

            elif reqtype == 'IMEI CDR' or subtype == 'IMEI':
                if row_data['ISP'] == 'None':
                    return jsonify({'error': f'Please Fill ISP'}), 400
                else:
                    data[0]['newnumber'].append({
                    'IMEI': row_data['Phone/IMEI/IP/CGI/Dealer'],
                    'ISP': row_data['ISP'],
                    'From_Date': row_data['From_Date'],
                    'To_Date': row_data['To_Date'],
                    'From_Time': row_data['From_Time'],
                    'To_Time': row_data['To_Time']
                    })

            elif reqtype == 'CAF':
                if row_data['MNP'] == 'None':
                    return jsonify({'error': f'Please Fill MNP'}), 400
                mnp = row_data['MNP']
                # if '-' in mnp:
                #     split_values = mnp.split('-')
                #     desired_value = split_values[1]
                #     LSA_ID = desired_value
                if 'JO' in mnp or 'Jio' in mnp or 'jio' in mnp or 'JIO' in mnp:
                    provider = 'Jio'
                elif 'CO' in mnp or 'co' in mnp or 'Co' in mnp:
                    provider = 'Cellone'
                elif 'VI' in mnp or 'Vi' in mnp or 'vi' in mnp:
                    provider = 'Vodafone'
                elif 'AT' in mnp or 'at' in mnp or 'At' in mnp:
                    provider = 'Airtel'
                else:
                    return jsonify({'error': f'Invalid MNP Format: {mnp}, It should be JO-AP/AT-AP'}), 400
                
                data[0]['newnumber'].append({
                'Phno': row_data['Phone/IMEI/IP/CGI/Dealer'],
                'CAF': row_data['CAF'],
                'MNP' : row_data['MNP'],
                'result' : {
                            'ActivationDate': row_data['ActivationDate'],
                            'Billing_City': "",
                            'Circle': "",
                            'Comments': "",
                            'Current_Owner_ID': "",
                            'E_sim': "",
                            'FetchedNickName': '',
                            'LSA_ID': row_data['MNP'].split('-')[1],
                            'Local_Address': "",
                            'MNP': row_data['MNP'],
                            'MSISDN': row_data['Phone/IMEI/IP/CGI/Dealer'],
                            'Name': "",
                            'Permanent_Address': "",
                            'Portability': "",
                            'ProviderName': provider,
                            'SDR': '',
                            'Truecaller': row_data['Truecaller'],
                            'Type': "",
                            'Unit': "",
                            'searchid': "",
                            'Remarks' : row_data['remarks']
                        }
                })

            elif reqtype == 'SDR':
                if row_data['MNP'] == 'None':
                    return jsonify({'error': f'Please Fill MNP'}), 400
                else:
                    mnp = row_data['MNP']
                    if '-' in mnp:
                        split_values = mnp.split('-')
                        desired_value = split_values[1]
                        LSA_ID = desired_value
                        if 'JO' in mnp or 'Jio' in mnp or 'jio' in mnp or 'JIO' in mnp:
                            provider = 'Jio'
                        elif 'CO' in mnp or 'co' in mnp or 'Co' in mnp:
                            provider = 'Cellone'
                        elif 'VI' in mnp or 'Vi' in mnp or 'vi' in mnp:
                            provider = 'Vodafone'
                        elif 'AT' in mnp or 'at' in mnp or 'At' in mnp:
                            provider = 'Airtel'
                    else:
                        return jsonify({'error': f'Invalid MNP Format: {mnp}, It should be JO-AP/AT-AP'}), 400
                    
                    data[0]['newnumber'].append({
                    'Phno': row_data['Phone/IMEI/IP/CGI/Dealer'],
                    'MNP' : row_data['MNP'],
                    'result' : {
                                'ActivationDate': row_data['ActivationDate'],
                                'Billing_City': "",
                                'Circle': "",
                                'Comments': "",
                                'Current_Owner_ID': "",
                                'E_sim': "",
                                'FetchedNickName': '',
                                'LSA_ID': LSA_ID,
                                'Local_Address': "",
                                'MNP': row_data['MNP'],
                                'MSISDN': row_data['Phone/IMEI/IP/CGI/Dealer'],
                                'Name': "",
                                'Permanent_Address': "",
                                'Portability': "",
                                'ProviderName': provider,
                                'SDR': '',
                                'Truecaller': row_data['Truecaller'],
                                'Type': "",
                                'Unit': "",
                                'searchid': "",
                                'Remarks' : row_data['remarks']
                            }
                    })

            elif reqtype == 'IPDR' and (subtype == 'IPV6' or subtype == 'IPV4'):
                data[0]['newnumber'].append({
                'IP_Address': row_data['Phone/IMEI/IP/CGI/Dealer'],
                'ISP': row_data['ISP'],
                'From_Date': row_data['From_Date'],
                'To_Date': row_data['To_Date'],
                'From_Time': row_data['From_Time'],
                'To_Time': row_data['To_Time']
                })

            elif 'TOWER' in reqtype and subtype == 'CGI':
                data[0]['newnumber'].append({
                'CGI': row_data['Phone/IMEI/IP/CGI/Dealer'],
                'ISP': row_data['ISP'],
                'From_Date': row_data['From_Date'],
                'To_Date': row_data['To_Date'],
                'From_Time': row_data['From_Time'],
                'To_Time': row_data['To_Time'],
                'result' : {
                                'ActivationDate': row_data['ActivationDate'],
                                'Billing_City': "",
                                'Circle': "",
                                'Comments': "",
                                'Current_Owner_ID': "",
                                'E_sim': "",
                                'FetchedNickName': '',
                                'LSA_ID': row_data['ISP'].split('-')[1],
                                'Local_Address': "",
                                'MNP': row_data['ISP'],
                                'MSISDN': row_data['Phone/IMEI/IP/CGI/Dealer'],
                                'Name': "",
                                'Permanent_Address': "",
                                'Portability': "",
                                'ProviderName': "",
                                'SDR': '',
                                'Truecaller': row_data['Truecaller'],
                                'Type': "",
                                'Unit': "",
                                'searchid': "",
                                'Remarks' : row_data['remarks']
                            }
                })

            elif (reqtype == "RH" or reqtype == "POA") and subtype == 'Retailer/Dealer Details':
                if reqtype == "RH" and subtype == 'Retailer/Dealer Details':
                    data[0]['newnumber'].append({
                    'RH_Dealer': row_data['Phone/IMEI/IP/CGI/Dealer'],
                    'RH_code': row_data['Dealer Code'],
                    'ISP': row_data['ISP'],
                    'From_Date': row_data['From_Date'],
                    'To_Date': row_data['To_Date'],
                    'From_Time': row_data['From_Time'],
                    'To_Time': row_data['To_Time']
                    })
                elif reqtype == "POA" and subtype == 'Retailer/Dealer Details':
                    data[0]['newnumber'].append({
                    'POA_Dealer': row_data['Phone/IMEI/IP/CGI/Dealer'],
                    'POA_code': row_data['Dealer Code'],
                    'ISP': row_data['ISP'],
                    'From_Date': row_data['From_Date'],
                    'To_Date': row_data['To_Date'],
                    'From_Time': row_data['From_Time'],
                    'To_Time': row_data['To_Time']
                    })
        elif category == 'Analysis Request':
            if 'CDR' in reqtype2 or subtype == 'Phone':
                if row_data['MNP'] == 'None' and row_data['SDR'] == 'None':
                    return jsonify({'error': f'Please Fill MNP and SDR'}), 400
                mnp = row_data['MNP']
                if '-' in mnp:
                    split_values = mnp.split('-')
                    desired_value = split_values[1]
                    LSA_ID = desired_value
                    #print(LSA_ID)
                    if 'JO' in mnp or 'Jio' in mnp or 'jio' in mnp or 'JIO' in mnp:
                        provider = 'Jio'
                    elif 'CO' in mnp or 'co' in mnp or 'Co' in mnp:
                        provider = 'Cellone'
                    elif 'VI' in mnp or 'Vi' in mnp or 'vi' in mnp:
                        provider = 'Vodafone'
                    elif 'AT' in mnp or 'at' in mnp or 'At' in mnp:
                        provider = 'Airtel'
                else:
                    return jsonify({'error': f'Invalid MNP Format: {mnp}, It should be JO-AP/AT-AP'}), 400
                
                data[0]['newnumber'].append({
                    'Phno': row_data['Phone/IMEI/IP/CGI/Dealer'],
                    'ISP': row_data['MNP'],
                    'From_Date': row_data['From_Date'],
                    'To_Date': row_data['To_Date'],
                    'From_Time': row_data['From_Time'],
                    'To_Time': row_data['To_Time'],
                    'result' : {
                            'ActivationDate': row_data['ActivationDate'],
                            'Billing_City': "",
                            'Circle': "",
                            'Comments': "",
                            'Current_Owner_ID': "",
                            'E_sim': "",
                            'FetchedNickName': '',
                            'LSA_ID': LSA_ID,
                            'Local_Address': "",
                            'MNP': row_data['MNP'],
                            'MSISDN': row_data['Phone/IMEI/IP/CGI/Dealer'],
                            'Name': "",
                            'Permanent_Address': "",
                            'Portability': "",
                            'ProviderName': provider,
                            'SDR': row_data['SDR'],
                            'Truecaller': row_data['Truecaller'],
                            'Type': "",
                            'Unit': "",
                            'searchid': "",
                            'Remarks' : row_data['remarks']
                        }
                })
            elif 'IMEI' in reqtype2 or subtype == 'IMEI':
                data[0]['newnumber'].append({
                'IMEI': row_data['Phone/IMEI/IP/CGI/Dealer'],
                'ISP': row_data['ISP'],
                'From_Date': row_data['From_Date'],
                'To_Date': row_data['To_Date'],
                'From_Time': row_data['From_Time'],
                'To_Time': row_data['To_Time']
                })
            elif 'IPDR' in reqtype2 and (subtype == 'IPV6' or subtype == 'IPV4'):
                data[0]['newnumber'].append({
                'IP_Address': row_data['Phone/IMEI/IP/CGI/Dealer'],
                'From_Date': row_data['From_Date'],
                'To_Date': row_data['To_Date'],
                'From_Time': row_data['From_Time'],
                'To_Time': row_data['To_Time']
                })
            elif 'TOWER DATA' in reqtype2 and subtype == 'CGI':
                data[0]['newnumber'].append({
                'CGI': row_data['Phone/IMEI/IP/CGI/Dealer'],
                'From_Date': row_data['From_Date'],
                'To_Date': row_data['To_Date'],
                'From_Time': row_data['From_Time'],
                'To_Time': row_data['To_Time']
                })
            elif 'TOWER DATA' == reqtype2 and subtype == 'Co-Ordinates':
                #print('Im here')
                data[0]['newnumber'].append({
                'Latitude': row_data['lat'],
                'Longitude': row_data['long'],
                'From_Date': row_data['From_Date'],
                'To_Date': row_data['To_Date'],
                'From_Time': row_data['From_Time'],
                'To_Time': row_data['To_Time']
                })
            elif (reqtype2 == "RH" or reqtype2 == "POA") and subtype == 'Retailer/Dealer Details':
                if reqtype2 == "RH" and subtype == 'Retailer/Dealer Details':
                    data[0]['newnumber'].append({
                    'RH_Dealer': row_data['Phone/IMEI/IP/CGI/Dealer'],
                    'RH_code': row_data['Dealer Code'],
                    'ISP': row_data['ISP'],
                    'From_Date': row_data['From_Date'],
                    'To_Date': row_data['To_Date'],
                    'From_Time': row_data['From_Time'],
                    'To_Time': row_data['To_Time']
                    })
                elif reqtype2 == "POA" and subtype == 'Retailer/Dealer Details':
                    data[0]['newnumber'].append({
                    'POA_Dealer': row_data['Phone/IMEI/IP/CGI/Dealer'],
                    'POA_code': row_data['Dealer Code'],
                    'ISP': row_data['ISP'],
                    'From_Date': row_data['From_Date'],
                    'To_Date': row_data['To_Date'],
                    'From_Time': row_data['From_Time'],
                    'To_Time': row_data['To_Time']
                    })
    return jsonify(data)

@app.route('/status_update', methods=['POST'])
def status_update():
    coll = db['tickets']
    id = request.json.get('id')
    number = request.json.get('number')
    status = request.json.get('status')
    remark = request.json.get('remark')
    status1 = 'Hold by Mailer'
    result = coll.find_one({'_id': ObjectId(id)})
    if status == 'Hold':
        for res in result['newnumber']:
            for no in number:
                if res['Phno'] in number:
                    coll.update_one(
                        {'_id': ObjectId(id), "newnumber.Phno": no},
                        {"$set": {"newnumber.$.mailer_hold": status1,"newnumber.$.mailer_remarks" : remark}}
                    )
                    coll.update_one(
                        {'_id': ObjectId(id), "result.MSISDN": no},
                        {"$set": {"result.$.mailer_hold": status1,"result.$.mailer_remarks" : remark}}
                    )
    return jsonify('ALL OKKK') 

@tickect_bp.route('/analysis_proform', methods=['POST'])
def analysis_proform():
    coll = db['tickets']
    id = request.json.get('id')
    result = coll.find_one({'_id': ObjectId(id)})
    return jsonify(result)

@tickect_bp.route('/analysis_proform_save', methods=['POST'])
def analysis_proform_save():
    coll = db['tickets']
    id = request.json.get('id')
    refno = request.json.get('refno')
    summary1 = request.json.get('summary1')
    summary2 = request.json.get('summary2')
    raised_by = request.json.get('raised_by')

    analysis_type = request.json.get('type')
    officer = request.json.get('officer')

    data = request.json.get('data')
    result = [
        {
            'refno':refno,
            'analysistype':analysis_type,
            'officername':officer,
            'data':data,
            'summary1' : summary1,
            'summary2' : summary2,
            'raised_by' : raised_by,
        }
    ]
    if data[0]['sdr'] != '' and data[0]['sdb'] != '' and analysis_type != '':
        coll.update_one(
            {'_id': ObjectId(id)},
            {"$set": {
                "proforma_data": result,
                'pending':"Report Generated"
                }}
        )
        return jsonify({'result':'Analysis Proforma Save Successfully!!!!!'})
    else:
        return jsonify('NO DATA!!!!!')

# Define the backend route for email search
@tickect_bp.route('/mail_search', methods=['POST'])
def mail_search():
    email = request.json.get('email')
    pass1 = request.json.get('pass1')

    coll= db['users']
    result = coll.find_one({'email': email, 'password1': pass1})
    if result:
        return jsonify({'result': result})
    else:
        return jsonify({"Error": "Wrong Email ID or Password"})

# Define the backend route for password update
@tickect_bp.route('/password_update', methods=['POST'])
def password_update():
    email = request.json.get('email')
    pass1 = request.json.get('pass1')
    pass2 = request.json.get('pass2')
    coll = db['users']
    result = coll.find_one({'email': email})
    if result and pass1 == pass2:
        coll.update_one(
            {'_id': ObjectId(result['_id']), 'email': email},
            {"$set": {'password1': pass1, 'password2': pass2}}
        )
        return jsonify("UPDATED")
    else:
        return jsonify({"Error": "Wrong Email ID or Password Mismatch"})
    

@tickect_bp.route('/download_for_mailer', methods=['POST', 'GET'])
def download_for_mailer():
    coll = db['tickets']
    ids = request.json.get('id')
    if len(ids) == 0:
        ids = [str(doc['_id']) for doc in coll.find({'pending' : 'Mail Under Process'}, {'_id': 1})]

    col = ['Token','Request Type','Sub-Type','Officer Name',"Phno",'Fetch_sdr','MNP','IMEI','CAF','CAF/CDR','SDR','IP_Address',"RH_Dealer",
           "RH_code","POA_Dealer","POA_code","Nickname",
           'ISP','cellid', 'latitude', 'longitude','area',
           'operator','state',"From_Date","From_Time","To_Date","To_Time",]
    
    total_execution_time = 0  # Initialize total execution time
    concatenated_dfs = []  # List to store DataFrames for each document
    
    for id in ids:
        excel = coll.find_one({'_id': ObjectId(id)}, {
        'token': 1,
        'newnumber.Phno': 1,
        'newnumber.Nickname': 1,
        'newnumber.From_Date': 1,
        'newnumber.To_Date': 1,
        'newnumber.IMEI':1,
        'result.SDR': 1,
        'result.MNP': 1,
        'requesttype': 1,
        'subtypes': 1,
        '_id': 0,
        'officername': 1
    })
        if excel:
            start_time = time.time()
            newnumbers = excel.get('newnumber', [])
            if excel['subtypes'] == 'Phone' or excel['requesttype'] == 'CDR':
                try:
                    data = [{
                        'Token': excel['token'],
                        'Request Type': excel['requesttype'],
                        'Sub-Type': excel['subtypes'],
                        'Officer Name': excel['officername'],
                        'Phno': nn.get('Phno', ''),
                        'Fetch_sdr': excel['result'][i]['SDR'] if excel.get('result') else '',
                        'MNP': excel['result'][i]['MNP'] if excel.get('result') else '',
                        'Nickname': nn.get('Nickname', ''),
                        'From_Date': nn.get('From_Date', ''),
                        'To_Date': nn.get('To_Date', ''),
                    } for i, nn in enumerate(newnumbers)]
                except:
                    pass
                
                concatenated_df = pd.DataFrame(data, columns=col)
                concatenated_dfs.append(concatenated_df)
            
            elif (excel['requesttype'] == 'IMEI CDR') or (excel['requesttype'] == "GPRS" and excel['subtypes'] == "IMEI") or (excel['requesttype'] == "IPDR" and excel['subtypes'] == "IMEI"):
                data = [{
                    'Token': excel['token'],
                    'Request Type': excel['requesttype'],
                    'Sub-Type': excel['subtypes'],
                    'Officer Name': excel['officername'],
                    'IMEI': nn.get('IMEI', ''),
                    'ISP': nn.get('ISP', ''),
                    'From_Date': nn.get('From_Date', ''),
                    'To_Date': nn.get('To_Date', ''),
                } for i, nn in enumerate(newnumbers)]
                # for item in data:
                #     print(item)

                concatenated_df = pd.DataFrame(data, columns=col)
                concatenated_dfs.append(concatenated_df)

            elif (excel['requesttype'] == 'IMEI CDR') or (excel['requesttype'] == "GPRS" and excel['subtypes'] == "IMEI") or (excel['requesttype'] == "IPDR" and excel['subtypes'] == "IMEI"):
                data = [{
                    'Token': excel['token'],
                    'Request Type': excel['requesttype'],
                    'Sub-Type': excel['subtypes'],
                    'Officer Name': excel['officername'],
                    'IMEI': nn.get('IMEI', ''),
                    'ISP': nn.get('ISP', ''),
                    'From_Date': nn.get('From_Date', ''),
                    'From_Time': nn.get('From_Time', ''),
                    'To_Date': nn.get('To_Date', ''),
                    'To_Time': nn.get('To_Time', ''),
                } for i, nn in enumerate(newnumbers)]
                concatenated_df = pd.DataFrame(data, columns=col)
                concatenated_dfs.append(concatenated_df)

            elif excel['requesttype'] == "SDR":
                data = [{
                    'Token': excel['token'],
                    'Request Type': excel['requesttype'],
                    'Sub-Type': excel['subtypes'],
                    'Officer Name': excel['officername'],
                    'Phno': nn.get('Phno', ''),
                    'MNP': nn.get('MNP', ''),
                } for i, nn in enumerate(newnumbers)]
                concatenated_df = pd.DataFrame(data, columns=col)
                concatenated_dfs.append(concatenated_df)

            elif (excel['requesttype'] == "CAF"):
                if excel['requesttype'] == "CAF":
                    data = [{
                        'Token': excel['token'],
                        'Request Type': excel['requesttype'],
                        'Sub-Type': excel['subtypes'],
                        'Officer Name': excel['officername'],
                        'Phno': nn.get('Phno', ''),
                        'CAF': nn.get('CAF', ''),
                    } for i, nn in enumerate(newnumbers)]
                concatenated_df = pd.DataFrame(data, columns=col)
                concatenated_dfs.append(concatenated_df)

            elif (excel['requesttype'] == "IPDR") and (excel['subtypes'] == "IPV6" or excel['subtypes'] == "IPV4"):
                data = [{
                    'Token': excel['token'],
                    'Request Type': excel['requesttype'],
                    'Sub-Type': excel['subtypes'],
                    'Officer Name': excel['officername'],
                    'IP_Address': nn.get('IP_Address', ''),
                    'ISP':       nn.get('ISP', ''),
                    'From_Date': nn.get('From_Date', ''),
                    'From_Time': nn.get('From_Time', ''),
                    'To_Date':   nn.get('To_Date', ''),
                    'To_Time':   nn.get('To_Time', ''),
                } for i, nn in enumerate(newnumbers)]
                concatenated_df = pd.DataFrame(data, columns=col)
                concatenated_dfs.append(concatenated_df)

            elif (excel['requesttype'] == "RECHARGE HISTORY (RH)" and excel['subtypes'] == "Retailer/Dealer Details") or (excel['requesttype'] == "POINT OF ACTIVATION (POA)" and excel['subtypes'] == "Retailer/Dealer Details"):
                if excel['requesttype'] == "RECHARGE HISTORY (RH)":
                    data = [{
                        'Token': excel['token'],
                        'Request Type': excel['requesttype'],
                        'Sub-Type': excel['subtypes'],
                        'Officer Name': excel['officername'],
                        'RH_Dealer': nn.get('RH_Dealer', ''),
                        'RH_code':   nn.get('RH_code', ''),
                        'ISP':       nn.get('ISP', ''),
                        'From_Date': nn.get('From_Date', ''),
                        'To_Date':   nn.get('To_Date', ''),
                    } for i, nn in enumerate(newnumbers)]
                else:
                    data = [{
                        'Token': excel['token'],
                        'Request Type': excel['requesttype'],
                        'Sub-Type': excel['subtypes'],
                        'Officer Name': excel['officername'],
                        'POA_Dealer':   nn.get('POA_Dealer', ''),
                        'POA_code':     nn.get('POA_code', ''),
                        'ISP':          nn.get('ISP', ''),
                        'From_Date':    nn.get('From_Date', ''),
                        'To_Date':      nn.get('To_Date', ''),
                    } for i, nn in enumerate(newnumbers)]
                concatenated_df = pd.DataFrame(data, columns=col)
                concatenated_dfs.append(concatenated_df)

            elif (excel['requesttype'] == "TOWER CDR" and excel['subtypes'] == "CGI") or (excel['requesttype'] == "TOWER GPRS" and excel['subtypes'] == "CGI") or (excel['requesttype'] == "TOWER IPDR" and excel['subtypes'] == "CGI"):
                    data = [{
                        'Token': excel['token'],
                        'Request Type': excel['requesttype'],
                        'Sub-Type': excel['subtypes'],
                        'Officer Name': excel['officername'],
                        'cellid': nn.get('CGI', ''),
                        'ISP': nn.get('ISP', ''),
                        'From_Date': nn.get('From_Date', ''),
                        'To_Date': nn.get('To_Date', ''),
                    } for i, nn in enumerate(newnumbers)]
                    concatenated_df = pd.DataFrame(data, columns=col)
                    concatenated_dfs.append(concatenated_df)

            end_time = time.time()  # Record end time
            execution_time_ms = (end_time - start_time) * 1000  # Convert to milliseconds
            total_execution_time += execution_time_ms  # Add to total execution time
            execution_time_ms_rounded = round(execution_time_ms, 3)  # Round to 3 decimal places
            print(f"Execution time for document {id}: {execution_time_ms_rounded:.3f} milliseconds")
    
    concatenated_df = pd.concat(concatenated_dfs, ignore_index=True)
    concatenated_df.dropna(axis=1, how='all', inplace=True)

    excel_data = io.BytesIO()
    with pd.ExcelWriter(excel_data, engine='xlsxwriter') as writer:
        concatenated_df.to_excel(writer, sheet_name='Sheet1', index=False)
    excel_data.seek(0)
    response = send_file(
        excel_data,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='output.xlsx'
    )

    print(f"Total execution time for all documents: {total_execution_time:.3f} milliseconds")
    return response

@tickect_bp.route('/note_create', methods=['POST', 'GET'])
def note_create():
    refno = request.json.get('refno')
    newno = request.json.get('newnumber')
    sum1 = request.json.get('summary1')
    sum2 = request.json.get('summary2')
    analysis_type = request.json.get('type')
    creator = request.json.get('user')
    officer = request.json.get('officer')
    data = request.json.get('data')
    coll2 = db['tickets']
    now = datetime.now()
    day_with_millisec = now.strftime("%d%m%Y%H%M%S")+f"{now.microsecond // 1000:03d}"
    sum = 1001
    db_len = coll2.count_documents({})
    sum = sum + int(db_len)
    #print(sum)
    tokens = f"{day_with_millisec}_" + str(sum)
    result = [
        {
            'refno':refno,
            'analysistype':analysis_type,
            'summary1' : sum1,
            'summary2' : sum2,
            'officername':officer,
            'data':data,
            "raised_by":creator,
        }
    ]
    Tickets(
            useremail =request.json.get('useremail'),
            superior =request.json.get('superior'),
            officername =request.json.get('officername'),
            role = request.json.get('role'),
            type = '',
            designation = request.json.get("designation"),
            requestcategory = "Note",
            requesttype = 'Analysis Note',
            subtypes = '',
            team = request.json.get('team'),
            modules = request.json.get('modules'),
            relation =  '',
            nickname =  '',
            suspect =  '',
            name =  '',
            location = '',
            source = '',
            reason = '',
            refno = '',
            refdate= '',
            priority = request.json.get('priority'),
            token = tokens,
            pending = 'Analysis Note Raised',
            approval = "-",
            others='',
            newnumber = newno,
            remarks = request.json.get('remarks'),
            status = request.json.get('status'),
            date = request.json.get("date"),
            raise_time = request.json.get("raise_time"),
            proforma_data= result,
        
    ).save()

    return jsonify('ALLL OK')

@tickect_bp.route('/sample_excel', methods=['POST', 'GET'])
def sample_excel():
    col = ['Phone/IMEI/IP/CGI/Dealer', 'MNP', 'ActivationDate', 'SDR', 'lat',
    'long',
    'CAF', 
           'From_Date', 'From_Time', 'To_Date', 'To_Time', 'Dealer Code',
           'ISP', 'Pan_India', 'priority', 'Truecaller', 
           'name', 'team', 'module', 'nickname', 'source', 'remarks']
    # Create an empty DataFrame with columns
    df = pd.DataFrame(columns=col)
    excel_data = io.BytesIO()
    # Write the DataFrame to an Excel file
    with pd.ExcelWriter(excel_data, engine='xlsxwriter') as writer:
        df.to_excel(writer, sheet_name='Sheet1', index=False)
    excel_data.seek(0)
    # Send the Excel file as a response for download
    response = send_file(
        excel_data,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='/home/bruce/Downloads/output.xlsx'
    )
    return response

@tickect_bp.route('/newnoresult', methods=['POST', 'GET'])
def newnoresult():
    coll = db['tickets']
    id = request.json.get('id')
    reqt = request.json.get('reqt')
    reqc = request.json.get('reqc')
    sub = request.json.get('sub')
    projection = {}
    #print(reqt,sub)
    if reqc == 'Data Request':
        if reqt == 'CDR' or sub == 'Phone':
            projection = {
            'newnumber.Phno': 1,
            'newnumber.Nickname': 1,
            'newnumber.From_Date': 1,
            'newnumber.To_Date': 1,
            'newnumber.From_Time': 1,
            'newnumber.To_Time': 1,
            'newnumber.status': 1,
            'raise_time' : 1,
            'Analyst_Approval' : 1,
            'Send_to_INS' : 1,
            'SP_Approval' : 1,
            'DSP_Approval' : 1,
            'Send_to_SP' : 1,
            'INSPR_Approval' : 1,
            'Send_to_DSP' : 1,
            'SP_Reject' : 1,
            'DSP_Reject' : 1,
            'INS_Reject' : 1,
            'officername':1,
            'others' : 1,
            'nickname':1,
            'reason':1,
            'remarks':1,
            'refno':1
            }
        elif reqt == 'SDR':
            projection = {
            'newnumber.Phno': 1,
            'newnumber.MNP': 1,
            'officername':1,
            'raise_time' : 1,
            'Analyst_Approval' : 1,
            'Send_to_INS' : 1,
            'SP_Approval' : 1,
            'DSP_Approval' : 1,
            'Send_to_SP' : 1,
            'SP_Reject' : 1,
            'DSP_Reject' : 1,
            'INS_Reject' : 1,
            'INSPR_Approval' : 1,
            'Send_to_DSP' : 1,
            'nickname':1,
            'others' : 1,
            'reason':1,
            'remarks':1,
            'refno':1
            }
        elif reqt == 'CAF':
            projection = {
            'newnumber.Phno': 1,
            'newnumber.CAF': 1,
            'newnumber.MNP': 1,
            'raise_time' : 1,
            'Analyst_Approval' : 1,
            'Send_to_INS' : 1,
            'SP_Approval' : 1,
            'DSP_Approval' : 1,
            'Send_to_SP' : 1,
            'SP_Reject' : 1,
            'DSP_Reject' : 1,
            'INS_Reject' : 1,
            'INSPR_Approval' : 1,
            'Send_to_DSP' : 1,
            'officername':1,
            'nickname':1,
            'others' : 1,
            'reason':1,
            'remarks':1,
            'refno':1
            }
        elif reqt == 'IPDR' or (sub == 'IPV6' or sub == 'IPV4'):
            projection = {
            'newnumber.IP_Address': 1,
            'newnumber.ISP': 1,
            'newnumber.From_Date': 1,
            'raise_time' : 1,
            'Analyst_Approval' : 1,
            'Send_to_INS' : 1,
            'SP_Approval' : 1,
            'DSP_Approval' : 1,
            'Send_to_SP' : 1,
            'INSPR_Approval' : 1,
            'Send_to_DSP' : 1,
            'SP_Reject' : 1,
            'DSP_Reject' : 1,
            'INS_Reject' : 1,
            'newnumber.To_Date': 1,
            'newnumber.From_Time': 1,
            'newnumber.To_Time': 1,
            'newnumber.status': 1,
            'officername':1,
            'nickname':1,
            'others' : 1,
            'reason':1,
            'remarks':1,
            'refno':1
            }
        elif reqt == 'IMEI CDR' or sub == 'IMEI':
            projection = {
            'newnumber.IMEI': 1,
            'newnumber.ISP': 1,
            'newnumber.From_Date': 1,
            'raise_time' : 1,
            'Analyst_Approval' : 1,
            'Send_to_INS' : 1,
            'SP_Approval' : 1,
            'DSP_Approval' : 1,
            'Send_to_SP' : 1,
            'INSPR_Approval' : 1,
            'Send_to_DSP' : 1,
            'newnumber.To_Date': 1,
            'SP_Reject' : 1,
            'DSP_Reject' : 1,
            'INS_Reject' : 1,
            'newnumber.status': 1,
            'newnumber.From_Time': 1,
            'newnumber.To_Time': 1,
            'officername':1,
            'nickname':1,
            'others' : 1,
            'reason':1,
            'remarks':1,
            'refno':1
            }
        elif reqt == 'RH' and sub == 'Retailer/Dealer Details':
            projection = {
            'newnumber.RH_Dealer': 1,
            'newnumber.RH_code': 1,
            'newnumber.ISP': 1,
            'raise_time' : 1,
            'Analyst_Approval' : 1,
            'Send_to_INS' : 1,
            'SP_Approval' : 1,
            'DSP_Approval' : 1,
            'Send_to_SP' : 1,
            'INSPR_Approval' : 1,
            'SP_Reject' : 1,
            'DSP_Reject' : 1,
            'INS_Reject' : 1,
            'Send_to_DSP' : 1,
            'newnumber.From_Date': 1,
            'newnumber.To_Date': 1,
            'officername':1,
            'nickname':1,
            'reason':1,
            'remarks':1,
            'others' : 1,
            'refno':1
            }
        elif reqt == 'POA' and sub == 'Retailer/Dealer Details':
            projection = {
            'officername':1,
            'nickname':1,
            'reason':1,
            'remarks':1,
            'refno':1,
            'others' : 1,
            'newnumber.POA_Dealer': 1,
            'raise_time' : 1,
            'Analyst_Approval' : 1,
            'Send_to_INS' : 1,
            'SP_Approval' : 1,
            'SP_Reject' : 1,
            'DSP_Reject' : 1,
            'INS_Reject' : 1,
            'DSP_Approval' : 1,
            'Send_to_SP' : 1,
            'INSPR_Approval' : 1,
            'Send_to_DSP' : 1,
            'newnumber.POA_code': 1,
            'newnumber.ISP': 1,
            'newnumber.From_Date': 1,
            'newnumber.To_Date': 1,
            }
        elif sub == 'CGI':
            projection = {
            'officername':1,
            'nickname':1,
            'reason':1,
            'remarks':1,
            'others' : 1,
            'refno':1,
            'newnumber.CGI': 1,
            'raise_time' : 1,
            'Analyst_Approval' : 1,
            'Send_to_INS' : 1,
            'SP_Reject' : 1,
            'DSP_Reject' : 1,
            'INS_Reject' : 1,
            'SP_Approval' : 1,
            'DSP_Approval' : 1,
            'Send_to_SP' : 1,
            'INSPR_Approval' : 1,
            'Send_to_DSP' : 1,
            'newnumber.status': 1,
            'newnumber.ISP': 1,
            'newnumber.From_Date': 1,
            'newnumber.To_Date': 1,
            }
    else:
        if 'CDR' in reqt or sub == 'Phone':
            projection = {
            'newnumber.Phno': 1,
            'newnumber.ISP': 1,
            'newnumber.From_Date': 1,
            'newnumber.To_Date': 1,
            'newnumber.From_Time': 1,
            'newnumber.To_Time': 1,
            'officername':1,
            'others' : 1,
            'nickname':1,
            'raise_time' : 1,
            'Analyst_Approval' : 1,
            'SP_Reject' : 1,
            'DSP_Reject' : 1,
            'INS_Reject' : 1,
            'Send_to_INS' : 1,
            'SP_Approval' : 1,
            'DSP_Approval' : 1,
            'Send_to_SP' : 1,
            'INSPR_Approval' : 1,
            'Send_to_DSP' : 1,
            'team': 1,
            'Comments': 1,
            'requestcategory' : 1,
            'assign_Officer' : 1,
            'pending': 1,
            'approval' : 1,
            'reason':1,
            'remarks':1,
            'refno':1,
            }
        elif 'IMEI' in reqt or sub == 'IMEI':
            projection = {
            'newnumber.IMEI': 1,
            'newnumber.ISP': 1,
            'newnumber.From_Date': 1,
            'newnumber.To_Date': 1,
            'raise_time' : 1,
            'Analyst_Approval' : 1,
            'Send_to_INS' : 1,
            'SP_Approval' : 1,
            'DSP_Approval' : 1,
            'Send_to_SP' : 1,
            'INSPR_Approval' : 1,
            'Send_to_DSP' : 1,
            'newnumber.From_Time': 1,
            'SP_Reject' : 1,
            'DSP_Reject' : 1,
            'INS_Reject' : 1,
            'newnumber.To_Time': 1,
            'officername':1,
            'nickname':1,
            'team': 1,
            'others' : 1,
            'Comments': 1,
            'requestcategory' : 1,
            'assign_Officer' : 1,
            'pending': 1,
            'approval' : 1,
            'reason':1,
            'remarks':1,
            'refno':1,
            }
        elif 'RH' in reqt and sub == 'Retailer/Dealer Details':
            projection = {
            'newnumber.RH_Dealer': 1,
            'newnumber.RH_code': 1,
            'raise_time' : 1,
            'Analyst_Approval' : 1,
            'Send_to_INS' : 1,
            'SP_Approval' : 1,
            'SP_Reject' : 1,
            'DSP_Reject' : 1,
            'INS_Reject' : 1,
            'DSP_Approval' : 1,
            'Send_to_SP' : 1,
            'INSPR_Approval' : 1,
            'Send_to_DSP' : 1,
            'newnumber.ISP': 1,
            'newnumber.From_Date': 1,
            'newnumber.To_Date': 1,
            'officername':1,
            'nickname':1,
            'others' : 1,
            'team': 1,
            'Comments': 1,
            'requestcategory' : 1,
            'assign_Officer' : 1,
            'pending': 1,
            'approval' : 1,
            'reason':1,
            'remarks':1,
            'refno':1,
            }
        elif reqt == 'POA' and sub == 'Retailer/Dealer Details':
            projection = {
            'officername':1,
            'nickname':1,
            'others' : 1,
            'team': 1,
            'Comments': 1,
            'raise_time' : 1,
            'Analyst_Approval' : 1,
            'SP_Reject' : 1,
            'DSP_Reject' : 1,
            'INS_Reject' : 1,
            'Send_to_INS' : 1,
            'SP_Approval' : 1,
            'DSP_Approval' : 1,
            'Send_to_SP' : 1,
            'INSPR_Approval' : 1,
            'Send_to_DSP' : 1,
            'requestcategory' : 1,
            'assign_Officer' : 1,
            'pending': 1,
            'approval' : 1,
            'reason':1,
            'remarks':1,
            'refno':1,
            'newnumber.POA_Dealer': 1,
            'newnumber.POA_code': 1,
            'newnumber.ISP': 1,
            'newnumber.From_Date': 1,
            'newnumber.To_Date': 1,
            }
        elif sub == 'CGI':
            projection = {
            'officername':1,
            'nickname':1,
            'team': 1,
            'Comments': 1,
            'raise_time' : 1,
            'Analyst_Approval' : 1,
            'Send_to_INS' : 1,
            'SP_Approval' : 1,
            'DSP_Approval' : 1,
            'SP_Reject' : 1,
            'DSP_Reject' : 1,
            'INS_Reject' : 1,
            'Send_to_SP' : 1,
            'INSPR_Approval' : 1,
            'Send_to_DSP' : 1,
            'requestcategory' : 1,
            'assign_Officer' : 1,
            'pending': 1,
            'approval' : 1,
            'reason':1,
            'remarks':1,
            'refno':1,
            'others' : 1,
            'newnumber.CGI': 1,
            'newnumber.ISP': 1,
            'newnumber.From_Date': 1,
            'newnumber.To_Date': 1,
            }
    result = coll.find_one({'_id': ObjectId(id)}, projection)
    return jsonify(result)



@tickect_bp.route('/dashboardresult', methods=['POST', 'GET'])
def dashboardresult():
    coll = db['tickets']
    id = request.json.get('id')
    reqt = request.json.get('reqt')
    reqc = request.json.get('reqc')
    sub = request.json.get('sub')
    projection = {}
    if reqc == 'Data Request':
        if reqt == 'CDR' or sub == 'Phone':
            projection = {
            'newnumber.Phno': 1,
            'newnumber.Nickname': 1,
            'newnumber.From_Date': 1,
            'newnumber.To_Date': 1,
            'newnumber.From_Time': 1,
            'newnumber.To_Time': 1,
            'newnumber.status': 1,
            }
        elif reqt == 'SDR':
            projection = {
            'newnumber.Phno': 1,
            'newnumber.MNP': 1,
            }
        elif reqt == 'CAF':
            projection = {
            'newnumber.Phno': 1,
            'newnumber.CAF': 1,
            'newnumber.MNP': 1,
            
            }
        elif reqt == 'IPDR' or (sub == 'IPV6' or sub == 'IPV4'):
            projection = {
            'newnumber.IP_Address': 1,
            'newnumber.ISP': 1,
            'newnumber.From_Date': 1,
            'newnumber.To_Date': 1,
            'newnumber.From_Time': 1,
            'newnumber.To_Time': 1,
            'newnumber.status': 1,
           
            }
        elif reqt == 'IMEI CDR' or sub == 'IMEI':
            projection = {
            'newnumber.IMEI': 1,
            'newnumber.ISP': 1,
            'newnumber.From_Date': 1,
            'newnumber.To_Date': 1,
            'newnumber.status': 1,
            'newnumber.From_Time': 1,
            'newnumber.To_Time': 1,
            
            }
        elif reqt == 'RH' and sub == 'Retailer/Dealer Details':
            projection = {
            'newnumber.RH_Dealer': 1,
            'newnumber.RH_code': 1,
            'newnumber.ISP': 1,
            'newnumber.From_Date': 1,
            'newnumber.To_Date': 1,
            
            }
        elif reqt == 'POA' and sub == 'Retailer/Dealer Details':
            projection = {
           
            'newnumber.POA_Dealer': 1,
            'newnumber.POA_code': 1,
            'newnumber.ISP': 1,
            'newnumber.From_Date': 1,
            'newnumber.To_Date': 1,
            }
        elif sub == 'CGI':
            projection = {
           
            'newnumber.CGI': 1,
            'newnumber.status': 1,
            'newnumber.ISP': 1,
            'newnumber.From_Date': 1,
            'newnumber.To_Date': 1,
            }
    else:
        if 'CDR' in reqt or sub == 'Phone':
            projection = {
            'newnumber.Phno': 1,
            'newnumber.ISP': 1,
            'newnumber.From_Date': 1,
            'newnumber.To_Date': 1,
            'newnumber.From_Time': 1,
            'newnumber.To_Time': 1,
            
            }
        elif 'IMEI' in reqt or sub == 'IMEI':
            projection = {
            'newnumber.IMEI': 1,
            'newnumber.ISP': 1,
            'newnumber.From_Date': 1,
            'newnumber.To_Date': 1,
            'newnumber.From_Time': 1,
            'newnumber.To_Time': 1,
            
            }
        elif 'RH' in reqt and sub == 'Retailer/Dealer Details':
            projection = {
            'newnumber.RH_Dealer': 1,
            'newnumber.RH_code': 1,
            'newnumber.ISP': 1,
            'newnumber.From_Date': 1,
            'newnumber.To_Date': 1,
            
            }
        elif reqt == 'POA' and sub == 'Retailer/Dealer Details':
            projection = {
            
            'newnumber.POA_Dealer': 1,
            'newnumber.POA_code': 1,
            'newnumber.ISP': 1,
            'newnumber.From_Date': 1,
            'newnumber.To_Date': 1,
            }
        elif sub == 'CGI':
            projection = {
            
            'newnumber.CGI': 1,
            'newnumber.ISP': 1,
            'newnumber.From_Date': 1,
            'newnumber.To_Date': 1,
            }
    result = coll.find_one({'_id': ObjectId(id)}, projection)
    return jsonify(result)



@tickect_bp.route('/autoresult', methods=['POST', 'GET'])
def autoresult():
    coll = db['tickets']
    id = request.json.get('id')
    #print(id)
    projection = {
            "_id" : 1,
            "pending":1,
            'result.MSISDN': 1,
            'result.MNP': 1,
            'result.FetchedNickName': 1,
            'result.SDR': 1,
            'result.From_Date': 1,
            'result.To_Date': 1,
            'result.Truecaller': 1,
            'result.Remarks': 1,
            "result.mailer_remarks" : 1,
            'result.mailer_hold': 1,
            'result.status' : 1,
            }
    result = coll.find_one({'_id': ObjectId(id)}, projection)
    return jsonify(result)

@tickect_bp.route('/header_filter', methods=['POST', 'GET'])
def header_filter():
    coll = db['tickets']
    status = request.json.get('status')
    cat = request.json.get('cat')
    if cat == 'Data Request':
        category = ['Data Request']
    elif cat == 'Analysis Request':
        category = ['Analysis Request','Note']
    headertype = request.json.get('type')
    email = request.json.get('email')
    result = []
    #print(headertype)
    projection = {
            '_id' : 1,
            'priority' : 1,
            'useremail':1,
            'date' : 1,
            'token' : 1,
            'approval' : 1,
            'pending' : 1,
            'requestcategory' : 1,
            'requesttype' : 1,
            'subtypes' : 1,
            'requesttypes' : 1,
            'status_count': 1,
            'result': 1,
            'newnumber' : 1
        }
    query = {
        'useremail': email,
        'requestcategory': {'$in' :  category},
    }
    My_tickets = coll.find(query, projection).sort([('_id', -1)])
    res_len = 0
    newno_len = 0
    total_len = 0
    filtered_tickets = []
    for res in My_tickets:
        if res['pending'] == 'Mail Under Process' and (res['subtypes'] == 'Phone' or res['requesttype'] == 'CDR'):
            res_len = len(res['result'])
            total_len = res_len - int(res['status_count'])
            
        elif res['pending'] == 'Ticket_Closed' and (res['subtypes'] == 'Phone' or res['requesttype'] == 'CDR'):
            res_len = len(res['result'])
            total_len = res_len
            
        elif res['pending'] == 'Mail Under Process' and (res['subtypes'] == 'IMEI' or res['requesttype'] == 'IMEI CDR'):
            newno_len = len(res['newnumber'])
            total_len = newno_len - int(res['status_count'])
            
        elif res['pending'] == 'Ticket_Closed' and (res['subtypes'] == 'IMEI' or res['requesttype'] == 'IMEI CDR'):
            newno_len = len(res['newnumber'])
            total_len = newno_len
            


        if headertype == 'priority':
            if res['priority'] == status:
                filtered_tickets.append({
                    '_id': res['_id'],
                    'priority': res['priority'],
                    'useremail':res['useremail'],
                    'date': res['date'],
                    'token': res['token'],
                    'approval': res['approval'],
                    'pending': res['pending'],
                    'requestcategory': res['requestcategory'],
                    'requesttype': res['requesttype'],
                    'requesttypes': res['requesttypes'],
                    'subtypes': res['subtypes'],
                    'status_count': res['status_count'],
                    'total_len': total_len,
                    "res_len": res_len,
                    "newno_len": newno_len
                })
        elif headertype == 'approval':
            if res['approval'] == status:
                filtered_tickets.append({
                    '_id': res['_id'],
                    'priority': res['priority'],
                    'useremail':res['useremail'],
                    'date': res['date'],
                    'token': res['token'],
                    'approval': res['approval'],
                    'pending': res['pending'],
                    'requestcategory': res['requestcategory'],
                    'requesttype': res['requesttype'],
                    'requesttypes': res['requesttypes'],
                    'subtypes': res['subtypes'],
                    'status_count': res['status_count'],
                    'total_len': total_len,
                    "res_len": res_len,
                    "newno_len": newno_len
                })
        elif headertype == 'pending':
            if res['pending'] == status:
                filtered_tickets.append({
                    '_id': res['_id'],
                    'priority': res['priority'],
                    'useremail':res['useremail'],
                    'date': res['date'],
                    'token': res['token'],
                    'approval': res['approval'],
                    'pending': res['pending'],
                    'requestcategory': res['requestcategory'],
                    'requesttype': res['requesttype'],
                    'requesttypes': res['requesttypes'],
                    'subtypes': res['subtypes'],
                    'status_count': res['status_count'],
                    'total_len': total_len,
                    "res_len": res_len,
                    "newno_len": newno_len
                })
        elif headertype == 'category':
            if res['requestcategory'] == status:
                filtered_tickets.append({
                    '_id': res['_id'],
                    'priority': res['priority'],
                    'useremail':res['useremail'],
                    'date': res['date'],
                    'token': res['token'],
                    'approval': res['approval'],
                    'pending': res['pending'],
                    'requestcategory': res['requestcategory'],
                    'requesttype': res['requesttype'],
                    'requesttypes': res['requesttypes'],
                    'subtypes': res['subtypes'],
                    'status_count': res['status_count'],
                    'total_len': total_len,
                    "res_len": res_len,
                    "newno_len": newno_len
                })
        elif headertype == 'requesttype':
            if res['requesttype'] == status or status in res['requesttypes']:
                filtered_tickets.append({
                    '_id': res['_id'],
                    'priority': res['priority'],
                    'useremail':res['useremail'],
                    'date': res['date'],
                    'token': res['token'],
                    'approval': res['approval'],
                    'pending': res['pending'],
                    'requestcategory': res['requestcategory'],
                    'requesttype': res['requesttype'],
                    'requesttypes': res['requesttypes'],
                    'subtypes': res['subtypes'],
                    'status_count': res['status_count'],
                    'total_len': total_len,
                    "res_len": res_len,
                    "newno_len": newno_len
                })
    pagination_info = {
                'total_pages': 1,
                'current_page': 1,
                'total_tickets': len(filtered_tickets)
            }
    return jsonify({'mytickets': filtered_tickets,'pagination': pagination_info})



@tickect_bp.route('/header_filter2', methods=['POST', 'GET'])
def header_filter2():
    coll = db['tickets']
    status = request.json.get('status')
    cat = request.json.get('cat')
    if cat == 'Data Request':
        category = ['Data Request']
    elif cat == 'Analysis Request':
        category = ['Analysis Request','Note']
    headertype = request.json.get('type')
    designation = request.json.get('designation')
    role = request.json.get('role')
    email = request.json.get('email')
    team = request.json.get('team')
    modules = request.json.get('modules')
    officername = request.json.get("officer")

    result = []
    #print(headertype)
    projection = {
            '_id' : 1,
            'priority' : 1,
            'date' : 1,
            'useremail' : 1,
            'token' : 1,
            'approval' : 1,
            'officername' : 1,
            'pending' : 1,
            'requestcategory' : 1,
            'requesttype' : 1,
            'subtypes' : 1,
            'requesttypes' : 1,
            'edit':1,
            'status_count': 1,
            'result': 1,
            'newnumber' : 1
    }
    if team == 'ADMIN':
        query = {
                    'useremail': {'$nin': [email]},
                    'requestcategory': {'$in' :  category},
                }
    
    elif designation == 'SP' and team != 'ADMIN':
        if team == 'CAT':
                query = {'$and' : [
                    {'useremail': {'$nin': [email]}},
                    {'team': team},
                    {'requestcategory': {'$in' :  category}},
                    {"$or": [
                        {"pending": "ADDL-SP/DSP Approval Pending", "team": team},
                        {"$and": [{"pending": {"$regex": officername}}, {"team": team}]},
                        {"pending": "Mail Under Process", "token": {"$regex": "_CO"}},
                        {"approval": "Approved by CAT DSP", "token": {"$regex": "_CO"}},
                        {"pending": "--", "requestcategory": "Analysis Request"},
                        {"approval": {"$regex": "^Assigned"}, "requestcategory": "Analysis Request"},
                        {"pending": "SP Approval Pending", "requestcategory": "Analysis Request"},
                        {"pending": "SP Approval Pending", "team": team},
                        {"pending": "Mail Under Process", "team": team},
                        {"pending": "Ticket_Closed", "team": team},
                        {"pending": {"$regex": "^Rejected"}, "team": team}
                    ]}]}
        else:
            query = {'$and' : [
                {'useremail': {'$nin': [email]}},
                {'team': team},
                {'requestcategory': {'$in' :  category}},
                {"$or": [
                    {"pending": "SP Approval Pending", "team": team,},
                    {"pending": "SP Approval Pending", "team": team,},
                    {"pending": "SP Approval Pending", "team": team},
                    {"approval": "Approved by SP", "team": team,},
                    {"approval": "Approved by SP", "team": team,},
                    {"token": {"$regex": "_CO"}, "team": team},
                    {"pending": {"$regex": officername}, "team": team},
                    {"pending": "Assign to CAT Ins", "requestcategory": "Analysis Request", "team": team,},
                    {"pending": "SP Approval Pending", "requestcategory": "Analysis Request", "team": team,},
                    {"pending": "Assign to CAT Ins", "requestcategory": "Analysis Request", "team": team,},
                    {"pending": "Assign to CAT Ins", "requestcategory": "Analysis Request", "team": team,},
                    {"pending": "SP Approval Pending", "requestcategory": "Analysis Request", "team": team,},
                    {"pending": "SP Approval Pending", "requestcategory": "Analysis Request", "team": team,},
                    {"approval": {"$regex": "^Assigned"}, "requestcategory": "Analysis Request", "team": team},
                    {"pending": "Mail Under Process", "team": team},
                    {"pending": {"$regex": "^Rejected"}, "team": team}
                ]}]}
    elif designation == 'ADDL-SP/DSP':
        if team == "CAT":
                query = {'$and' : [
                    {'useremail': {'$nin': [email]}},
                    {'requestcategory': {'$in' :  category}},
                    {"$or": [
                        {"pending": "Mail Under Process", "token": {"$regex": "_CO"}, "designation": {"$ne": "SP"}},
                        {"$and": [{"pending": {"$regex": officername}}, {"team": team}]},
                        {"pending": "Ticket_Closed", "team": team, "designation": {"$ne": "SP"}},
                        {"approval": "Approved by CAT Ins", "token": {"$regex": "_CO"}, "designation": {"$ne": "SP"}},
                        {"pending": "ADDL-SP/DSP Approval Pending", "requestcategory": "Analysis Request"},
                        {"$and": [{"pending": "--"}, {"requestcategory": "Analysis Request"}, {"designation": {"$ne": "SP"}}]},
                        {"$and": [{"pending": "ADDL-SP/DSP Approval Pending"}, {"team": team}]},
                        {"approval": {"$regex": "^Assigned"}, "requestcategory": "Analysis Request", "designation": {"$ne": "SP"}},
                        {"approval": {"$regex": "^Assigned"}, "requestcategory": "Analysis Request", "designation": {"$ne": "SP"}},
                        {"pending": "SP Approval Pending", "team": team, "designation": {"$ne": "ADDL-SP/DSP"}},
                        {"pending": "Mail Under Process", "team": team, "designation": {"$ne": "SP"}},
                        {"pending": {"$regex": "^Rejected"}, "team": team}
                        ]}
                ]
                    }
        else:
            query = {'$and' : [
                {'useremail': {'$nin': [email]}},
                {'requestcategory': {'$in' :  category}},
                {"$or": [
                    {"$and": [{"pending": {"$regex": officername}}, {"team": team}]},
                    {"$and": [{"token": {"$regex": "_CO"}}, {"team": team}, {"others": {"$in": [officername]}}]},
                    {"pending": "ADDL-SP/DSP Approval Pending", "team": team, "requestcategory": "Data Request", "superior": officername},
                    {"pending": "ADDL-SP/DSP Approval Pending", "team": team, "requestcategory": "Data Request", "others": officername},
                    {"pending": "SP Approval Pending", "team": team, "others": {"$in": [officername]}},
                    {"pending": "SP Approval Pending", "team": team, "superior": officername},
                    {"pending": "SP Approval Pending", "team": team, "others": officername},
                    {"pending": {"$regex": "Rejected"}, "team": team, "superior": officername},
                    {"pending": {"$regex": "Rejected"}, "team": team, "others": officername},
                    {"pending": "ADDL-SP/DSP Approval Pending", "requestcategory": "Analysis Request", "team": team, "others": {"$in": [officername]}},
                    {"pending": "ADDL-SP/DSP Approval Pending", "requestcategory": "Analysis Request", "team": team, "superior": officername},
                    {"pending": "ADDL-SP/DSP Approval Pending", "requestcategory": "Analysis Request", "team": team, "others": officername},
                    {"pending": "Assign to CAT Ins", "requestcategory": "Analysis Request", "team": team, "designation": {"$nin": ["SP"]}, "others": {"$in": [officername]}},
                    {"pending": "Assign to CAT Ins", "requestcategory": "Analysis Request", "team": team, "designation": {"$nin": ["SP"]}, "superior": officername},
                    {"pending": "Assign to CAT Ins", "requestcategory": "Analysis Request", "team": team, "designation": {"$nin": ["SP"]}, "others": officername},
                    {"pending": "Mail Under Process", "team": team, "designation": {"$ne": "SP"}, "others": officername},
                    {"pending": "Mail Under Process", "team": team, "designation": {"$ne": "SP"}, "superior": officername},
                    {"approval": {"$regex": "Assigned"}, "requestcategory": "Analysis Request", "team": team, "designation": {"$nin": ["SP"]}, "others": {"$in": [officername]}},
                    {"approval": {"$regex": "Assigned"}, "requestcategory": "Analysis Request", "team": team, "designation": {"$nin": ["SP"]}, "superior": officername},
                    {"approval": {"$regex": "Assigned"}, "requestcategory": "Analysis Request", "team": team, "designation": {"$nin": ["SP"]}, "others": officername}
                    ]}
            ]}
    elif designation == "INSPR" and role != "Analyst":
        if team == 'CAT':
                query = {'$and' : [
                {'useremail': {'$nin': [email]}},
                {'requestcategory': {'$in' :  category}},
                {"$or": [
                        {"pending": "Ins Approval Pending", "team": team, 'modules':modules},
                        {"pending": "ADDL-SP/DSP Approval Pending", "team": team, 'modules':modules,"designation": {"$nin": ["SP", "DSP", "ADDL-SP/DSP"]}},
                        {"pending": "ADDL-SP/DSP Approval Pending", 'modules':modules,"token": {"$regex": "_CO"}, "designation": {"$nin": ["SP", "DSP", "ADDL-SP/DSP"]}},
                        {"pending": {"$regex": officername}, "team": team,'modules':modules},
                        {"pending": "Ticket_Closed", "team": team,'modules':modules, "designation": {"$nin": ["SP", "ADDL-SP/DSP"]}},
                        {"pending": "Mail Under Process", "token": {"$regex": "_CO"},'modules':modules, "designation": {"$nin": ["SP", "DSP", "ADDL-SP/DSP"]}},
                        {"pending": "Annexe Send to CAT INS", "token": {"$regex": "_CO"},'modules':modules, "designation": {"$nin": ["SP", "DSP", "ADDL-SP/DSP"]}},
                        {"pending": "Assign to CAT Ins"},
                        {"team": {"$ne": "CAT"}, "approval": {"$regex": "^Assigned"}},
                        {"approval": {"$regex": "^Assigned"}},
                        {"pending": "Mail Under Process",'modules':modules, "team": team, "designation": {"$nin": ["SP", "DSP", "ADDL-SP/DSP"]}},
                        {"pending": {"$regex": "^Rejected"}, "team": team, 'modules':modules,"designation": {"$nin": ["SP", "DSP", "ADDL-SP/DSP"]}}
                    ]
                }]}
        else:
            query = {'$and' : [
                {'useremail': {'$nin': [email]}},
                {'requestcategory': {'$in' :  category}},
                {"$or": [
                    # DATA REQ
                    {"pending": "Ins Approval Pending", "team": team,'modules':modules, "others": {"$in": [officername, '']},"designation": {"$nin": ["SP", "DSP", "ADDL-SP/DSP"]}},
                    {"pending": "ADDL-SP/DSP Approval Pending",'modules':modules, "team": team, "others": {"$in": [officername, '']},"designation": {"$nin": ["SP", "DSP", "ADDL-SP/DSP"]}},
                    {"pending": {"$regex": officername}, "team": team,'modules':modules,"designation": {"$nin": ["SP", "DSP", "ADDL-SP/DSP"]}},
                    {"pending": "Annexe Send to CAT INS", "team": team, 'modules':modules,"others": {"$in": [officername, '']},"designation": {"$nin": ["SP", "DSP", "ADDL-SP/DSP"]}},
                    {"pending": {"$regex": "^Rejected"}, "team": team, 'modules':modules,"others": {"$in": [officername, '']},"designation": {"$nin": ["SP", "DSP", "ADDL-SP/DSP"]}},
                    {"pending": "Mail Under Process", "team": team,'modules':modules, "others": {"$in": [officername, '']},"designation": {"$nin": ["SP", "DSP", "ADDL-SP/DSP"]}},
                    {"pending": "SP Approval Pending", "team": team, 'modules':modules,"others": {"$in": [officername, '']},"designation": {"$nin": ["SP", "DSP", "ADDL-SP/DSP"]}},
                    # ANALYSIS REQ
                    {"pending": "Ins Approval Pending", 'modules':modules,"requestcategory": "Analysis Request", "team": team,"designation": {"$nin": ["SP", "DSP", "ADDL-SP/DSP"]}},
                    {"approval": {"$regex": "^Assigned"}, 'modules':modules,"requestcategory": "Analysis Request", "team": team,"designation": {"$nin": ["SP", "DSP", "ADDL-SP/DSP"]}},
                    {"pending": "ADDL-SP/DSP Approval Pending", 'modules':modules,"requestcategory": "Analysis Request", "team": team,"designation": {"$nin": ["SP", "DSP", "ADDL-SP/DSP"]}},
                    {"pending": "SP Approval Pending", 'modules':modules,"requestcategory": "Analysis Request", "team": team,"designation": {"$nin": ["SP", "DSP", "ADDL-SP/DSP"]}},
                    {"pending": "Assign to CAT Ins", 'modules':modules,"requestcategory": "Analysis Request", "team": team,"designation": {"$nin": ["SP", "DSP", "ADDL-SP/DSP"]}}
                ]
            }]}
    elif role == "Analyst":
        query = {"$or": [
                    {'assign_Officer': {'$in':[officername]}}
                ]
            }
    elif role == "Mailer":
        query = {"$or": [
                    {"requestcategory": "Data Request", 'pending' : {'$in': ['Mail Under Process','Ticket_Closed']}},
                ]
            }
    My_tickets = coll.find(query, projection).sort([('_id', -1)])
    res_len = 0
    newno_len = 0
    total_len = 0
    filtered_tickets = []
    for res in My_tickets:
        if res['pending'] == 'Mail Under Process' and (res['subtypes'] == 'Phone' or res['requesttype'] == 'CDR'):
            res_len = len(res['result'])
            total_len = res_len - int(res['status_count'])

        elif res['pending'] == 'Mail Under Process' and (res['subtypes'] == 'IMEI' or res['requesttype'] == 'IMEI CDR'):
            newno_len = len(res['newnumber'])
            total_len = newno_len - int(res['status_count'])


        if headertype == 'priority':
            if res['priority'] == status:
                filtered_tickets.append({
                    '_id': res['_id'],
                    'priority': res['priority'],
                    'useremail':res['useremail'],
                    'date': res['date'],
                    'token': res['token'],
                    'approval': res['approval'],
                    'officername': res['officername'],
                    'edit': res['edit'],
                    'pending': res['pending'],
                    'requestcategory': res['requestcategory'],
                    'requesttype': res['requesttype'],
                    'requesttypes': res['requesttypes'],
                    'subtypes': res['subtypes'],
                    'status_count': res['status_count'],
                    'total_len': total_len,
                    "res_len": res_len,
                    "newno_len": newno_len
                })
        elif headertype == 'approval':
            if res['approval'] == status:
                filtered_tickets.append({
                    '_id': res['_id'],
                    'priority': res['priority'],
                    'useremail':res['useremail'],
                    'date': res['date'],
                    'token': res['token'],
                    'officername': res['officername'],
                    'edit': res['edit'],
                    'approval': res['approval'],
                    'pending': res['pending'],
                    'requestcategory': res['requestcategory'],
                    'requesttype': res['requesttype'],
                    'requesttypes': res['requesttypes'],
                    'subtypes': res['subtypes'],
                    'status_count': res['status_count'],
                    'total_len': total_len,
                    "res_len": res_len,
                    "newno_len": newno_len
                })
        elif headertype == 'pending':
            if res['pending'] == status:
                filtered_tickets.append({
                    '_id': res['_id'],
                    'priority': res['priority'],
                    'useremail':res['useremail'],
                    'date': res['date'],
                    'token': res['token'],
                    'approval': res['approval'],
                    'officername': res['officername'],
                    'edit': res['edit'],
                    'pending': res['pending'],
                    'requestcategory': res['requestcategory'],
                    'requesttype': res['requesttype'],
                    'requesttypes': res['requesttypes'],
                    'subtypes': res['subtypes'],
                    'status_count': res['status_count'],
                    'total_len': total_len,
                    "res_len": res_len,
                    "newno_len": newno_len
                })
        elif headertype == 'category':
            if res['requestcategory'] == status:
                filtered_tickets.append({
                    '_id': res['_id'],
                    'priority': res['priority'],
                    'useremail':res['useremail'],
                    'date': res['date'],
                    'token': res['token'],
                    'approval': res['approval'],
                    'pending': res['pending'],
                    'requestcategory': res['requestcategory'],
                    'requesttype': res['requesttype'],
                    'requesttypes': res['requesttypes'],
                    'subtypes': res['subtypes'],
                    'officername': res['officername'],
                    'edit': res['edit'],
                    'status_count': res['status_count'],
                    'total_len': total_len,
                    "res_len": res_len,
                    "newno_len": newno_len
                })
        elif headertype == 'requesttype':
            if res['requesttype'] == status or status in res['requesttypes']:
                filtered_tickets.append({
                    '_id': res['_id'],
                    'priority': res['priority'],
                    'useremail':res['useremail'],
                    'date': res['date'],
                    'token': res['token'],
                    'approval': res['approval'],
                    'pending': res['pending'],
                    'requestcategory': res['requestcategory'],
                    'requesttype': res['requesttype'],
                    'requesttypes': res['requesttypes'],
                    'subtypes': res['subtypes'],
                    'officername': res['officername'],
                    'edit': res['edit'],
                    'status_count': res['status_count'],
                    'total_len': total_len,
                    "res_len": res_len,
                    "newno_len": newno_len
                })
        elif headertype == 'officername':
            if res['officername'] == status:
                filtered_tickets.append({
                    '_id': res['_id'],
                    'priority': res['priority'],
                    'useremail':res['useremail'],
                    'date': res['date'],
                    'token': res['token'],
                    'approval': res['approval'],
                    'pending': res['pending'],
                    'officername': res['officername'],
                    'edit': res['edit'],
                    'requestcategory': res['requestcategory'],
                    'requesttype': res['requesttype'],
                    'requesttypes': res['requesttypes'],
                    'subtypes': res['subtypes'],
                    'status_count': res['status_count'],
                    'total_len': total_len,
                    "res_len": res_len,
                    "newno_len": newno_len
                })
        elif headertype == 'edit':
            if res['edit'] == status:
                filtered_tickets.append({
                    '_id': res['_id'],
                    'priority': res['priority'],
                    'useremail':res['useremail'],
                    'date': res['date'],
                    'token': res['token'],
                    'approval': res['approval'],
                    'pending': res['pending'],
                    'officername': res['officername'],
                    'edit': res['edit'],
                    'requestcategory': res['requestcategory'],
                    'requesttype': res['requesttype'],
                    'requesttypes': res['requesttypes'],
                    'subtypes': res['subtypes'],
                    'status_count': res['status_count'],
                    'total_len': total_len,
                    "res_len": res_len,
                    "newno_len": newno_len
                })
    


    pagination_info = {
                'total_pages1': 1,
                'current_page1': 1,
                'total_tickets1': len(filtered_tickets)
            }
    return jsonify({'mytickets': filtered_tickets,'pagination': pagination_info})

@tickect_bp.route('/date-filter', methods=['POST', 'GET'])
def date_filter():
    coll = db['tickets']
    email = request.json.get('email')
    cat = request.json.get('cat')
    if cat == 'Data Request':
        category = ['Data Request']
    elif cat == 'Analysis Request':
        category = ['Analysis Request','Note']
    fromdate = request.json.get('fromdate')
    todate = request.json.get('todate')
    #print(fromdate, todate)
    query = {}
    projection = {}
    result_dates = []
    if fromdate != 'Invalid Date' and todate != 'Invalid Date':
        # Convert string dates to datetime objects
        start_date1 = datetime.strptime(fromdate, '%d/%m/%Y')
        end_date1 = datetime.strptime(todate, '%d/%m/%Y')
        samples = [res for res in coll.find({'useremail': email,'requestcategory': {'$in' :  category}},{'date' : 1})]
        date_objects = [datetime.strptime(date['date'], '%d/%m/%Y') for date in samples]
        result_dates = [date.strftime('%d/%m/%Y') for date in date_objects if start_date1 <= date <= end_date1]
        projection = {
                '_id' : 1,
                'priority' : 1,
                'useremail':1,
                'date' : 1,
                'token' : 1,
                'approval' : 1,
                'pending' : 1,
                'requestcategory' : 1,
                'requesttype' : 1,
                'subtypes' : 1,
                'requesttypes' : 1,
                'status_count': 1,
                'result': 1,
                'newnumber' : 1
            }
        query = {'useremail': email,'requestcategory': {'$in' :  category}}
        My_tickets = coll.find(query, projection).sort([('_id', -1)])
        res_len = 0
        newno_len = 0
        total_len = 0
        filtered_tickets = []
        for res in My_tickets:
            if res['date'] in result_dates:
                if res['pending'] == 'Mail Under Process' and (res['subtypes'] == 'Phone' or res['requesttype'] == 'CDR'):
                    res_len = len(res['result'])
                    total_len = res_len - int(res['status_count'])
                    
                elif res['pending'] == 'Ticket_Closed' and (res['subtypes'] == 'Phone' or res['requesttype'] == 'CDR'):
                    res_len = len(res['result'])
                    total_len = res_len
                    
                elif res['pending'] == 'Mail Under Process' and (res['subtypes'] == 'IMEI' or res['requesttype'] == 'IMEI CDR'):
                    newno_len = len(res['newnumber'])
                    total_len = newno_len - int(res['status_count'])
                    
                elif res['pending'] == 'Ticket_Closed' and (res['subtypes'] == 'IMEI' or res['requesttype'] == 'IMEI CDR'):
                    newno_len = len(res['newnumber'])
                    total_len = newno_len
                    

                filtered_tickets.append({
                    '_id': res['_id'],
                    'priority': res['priority'],
                    'useremail':res['useremail'],
                    'date': res['date'],
                    'token': res['token'],
                    'approval': res['approval'],
                    'pending': res['pending'],
                    'requestcategory': res['requestcategory'],
                    'requesttype': res['requesttype'],
                    'requesttypes': res['requesttypes'],
                    'subtypes': res['subtypes'],
                    'status_count': res['status_count'],
                    'total_len': total_len,
                    "res_len": res_len,
                    "newno_len": newno_len
                })
        pagination_info = {
                    'total_pages': 1,
                    'current_page': 1,
                    'total_tickets': len(filtered_tickets)
                }
        return jsonify({'mytickets': filtered_tickets,'pagination': pagination_info})
    else:
        return jsonify("WAIT")

@tickect_bp.route('/date-filter2', methods=['POST', 'GET'])
def date_filter2():
    coll = db['tickets']
    designation = request.json.get('designation')
    role = request.json.get('role')
    email = request.json.get('email')
    cat = request.json.get('cat')
    if cat == 'Data Request':
        category = ['Data Request']
    elif cat == 'Analysis Request':
        category = ['Analysis Request','Note']
    team = request.json.get('team')
    modules = request.json.get('modules')
    officername = request.json.get("officer")    
    fromdate = request.json.get('fromdate')
    todate = request.json.get('todate')
    result_dates = []
    query = {
        'useremail': {'$nin': [email]},
        'requestcategory': {'$in' :  category},
    }
    projection = {}
    if fromdate != 'Invalid Date' and todate != 'Invalid Date':
        # Convert string dates to datetime objects
        start_date1 = datetime.strptime(fromdate, '%d/%m/%Y')
        end_date1 = datetime.strptime(todate, '%d/%m/%Y')
        samples = [res for res in coll.find(query,{'date' : 1})]
        date_objects = [datetime.strptime(date['date'], '%d/%m/%Y') for date in samples]
        result_dates = [date.strftime('%d/%m/%Y') for date in date_objects if start_date1 <= date <= end_date1]
        projection = {
            '_id' : 1,
            'priority' : 1,
            'date' : 1,
            'useremail' : 1,
            'token' : 1,
            'approval' : 1,
            'officername' : 1,
            'pending' : 1,
            'requestcategory' : 1,
            'requesttype' : 1,
            'subtypes' : 1,
            'requesttypes' : 1,
            'edit':1,
            'status_count': 1,
            'result': 1,
            'newnumber' : 1
    }
        if team == 'ADMIN':
            query = {
                    'useremail': {'$nin': [email]},
                    'requestcategory': {'$in' :  category},
                }
        
        elif designation == 'SP' and team != 'ADMIN':
            if team == 'CAT':
                    query = {'$and' : [
                        {'useremail': {'$nin': [email]}},
                        {'requestcategory': {'$in' :  category}},
                        {'team': team},
                        {"$or": [
                            {"pending": "ADDL-SP/DSP Approval Pending", "team": team},
                            {"$and": [{"pending": {"$regex": officername}}, {"team": team}]},
                            {"pending": "Mail Under Process", "token": {"$regex": "_CO"}},
                            {"approval": "Approved by CAT DSP", "token": {"$regex": "_CO"}},
                            {"pending": "--", "requestcategory": "Analysis Request"},
                            {"approval": {"$regex": "^Assigned"}, "requestcategory": "Analysis Request"},
                            {"pending": "SP Approval Pending", "requestcategory": "Analysis Request"},
                            {"pending": "SP Approval Pending", "team": team},
                            {"pending": "Mail Under Process", "team": team},
                            {"pending": "Ticket_Closed", "team": team},
                            {"pending": {"$regex": "^Rejected"}, "team": team}
                        ]}]}
            else:
                query = {'$and' : [
                    {'useremail': {'$nin': [email]}},
                    {'requestcategory': {'$in' :  category}},
                    {'team': team},
                    {"$or": [
                        {"pending": "SP Approval Pending", "team": team,},
                        {"pending": "SP Approval Pending", "team": team,},
                        {"pending": "SP Approval Pending", "team": team},
                        {"approval": "Approved by SP", "team": team,},
                        {"approval": "Approved by SP", "team": team,},
                        {"token": {"$regex": "_CO"}, "team": team},
                        {"pending": {"$regex": officername}, "team": team},
                        {"pending": "Assign to CAT Ins", "requestcategory": "Analysis Request", "team": team,},
                        {"pending": "SP Approval Pending", "requestcategory": "Analysis Request", "team": team,},
                        {"pending": "Assign to CAT Ins", "requestcategory": "Analysis Request", "team": team,},
                        {"pending": "Assign to CAT Ins", "requestcategory": "Analysis Request", "team": team,},
                        {"pending": "SP Approval Pending", "requestcategory": "Analysis Request", "team": team,},
                        {"pending": "SP Approval Pending", "requestcategory": "Analysis Request", "team": team,},
                        {"approval": {"$regex": "^Assigned"}, "requestcategory": "Analysis Request", "team": team},
                        {"pending": "Mail Under Process", "team": team},
                        {"pending": {"$regex": "^Rejected"}, "team": team}
                    ]}]}
        elif designation == 'ADDL-SP/DSP':
            if team == "CAT":
                    query = {'$and' : [
                        {'useremail': {'$nin': [email]}},
                        {'requestcategory': {'$in' :  category}},
                        {"$or": [
                            {"pending": "Mail Under Process", "token": {"$regex": "_CO"}, "designation": {"$ne": "SP"}},
                            {"$and": [{"pending": {"$regex": officername}}, {"team": team}]},
                            {"pending": "Ticket_Closed", "team": team, "designation": {"$ne": "SP"}},
                            {"approval": "Approved by CAT Ins", "token": {"$regex": "_CO"}, "designation": {"$ne": "SP"}},
                            {"pending": "ADDL-SP/DSP Approval Pending", "requestcategory": "Analysis Request"},
                            {"$and": [{"pending": "--"}, {"requestcategory": "Analysis Request"}, {"designation": {"$ne": "SP"}}]},
                            {"$and": [{"pending": "ADDL-SP/DSP Approval Pending"}, {"team": team}]},
                            {"approval": {"$regex": "^Assigned"}, "requestcategory": "Analysis Request", "designation": {"$ne": "SP"}},
                            {"approval": {"$regex": "^Assigned"}, "requestcategory": "Analysis Request", "designation": {"$ne": "SP"}},
                            {"pending": "SP Approval Pending", "team": team, "designation": {"$ne": "ADDL-SP/DSP"}},
                            {"pending": "Mail Under Process", "team": team, "designation": {"$ne": "SP"}},
                            {"pending": {"$regex": "^Rejected"}, "team": team}
                            ]}
                    ]
                        }
            else:
                query = {'$and' : [
                    {'useremail': {'$nin': [email]}},
                    {'requestcategory': {'$in' :  category}},
                    {"$or": [
                        {"$and": [{"pending": {"$regex": officername}}, {"team": team}]},
                        {"$and": [{"token": {"$regex": "_CO"}}, {"team": team}, {"others": {"$in": [officername]}}]},
                        {"pending": "ADDL-SP/DSP Approval Pending", "team": team, "requestcategory": "Data Request", "superior": officername},
                        {"pending": "ADDL-SP/DSP Approval Pending", "team": team, "requestcategory": "Data Request", "others": officername},
                        {"pending": "SP Approval Pending", "team": team, "others": {"$in": [officername]}},
                        {"pending": "SP Approval Pending", "team": team, "superior": officername},
                        {"pending": "SP Approval Pending", "team": team, "others": officername},
                        {"pending": {"$regex": "Rejected"}, "team": team, "superior": officername},
                        {"pending": {"$regex": "Rejected"}, "team": team, "others": officername},
                        {"pending": "ADDL-SP/DSP Approval Pending", "requestcategory": "Analysis Request", "team": team, "others": {"$in": [officername]}},
                        {"pending": "ADDL-SP/DSP Approval Pending", "requestcategory": "Analysis Request", "team": team, "superior": officername},
                        {"pending": "ADDL-SP/DSP Approval Pending", "requestcategory": "Analysis Request", "team": team, "others": officername},
                        {"pending": "Assign to CAT Ins", "requestcategory": "Analysis Request", "team": team, "designation": {"$nin": ["SP"]}, "others": {"$in": [officername]}},
                        {"pending": "Assign to CAT Ins", "requestcategory": "Analysis Request", "team": team, "designation": {"$nin": ["SP"]}, "superior": officername},
                        {"pending": "Assign to CAT Ins", "requestcategory": "Analysis Request", "team": team, "designation": {"$nin": ["SP"]}, "others": officername},
                        {"pending": "Mail Under Process", "team": team, "designation": {"$ne": "SP"}, "others": officername},
                        {"pending": "Mail Under Process", "team": team, "designation": {"$ne": "SP"}, "superior": officername},
                        {"approval": {"$regex": "Assigned"}, "requestcategory": "Analysis Request", "team": team, "designation": {"$nin": ["SP"]}, "others": {"$in": [officername]}},
                        {"approval": {"$regex": "Assigned"}, "requestcategory": "Analysis Request", "team": team, "designation": {"$nin": ["SP"]}, "superior": officername},
                        {"approval": {"$regex": "Assigned"}, "requestcategory": "Analysis Request", "team": team, "designation": {"$nin": ["SP"]}, "others": officername}
                        ]}
                ]}
        elif designation == "INSPR" and role != "Analyst":
            if team == 'CAT':
                    query = {'$and' : [
                    {'useremail': {'$nin': [email]}},
                    {'requestcategory': {'$in' :  category}},
                    {"$or": [
                            {"pending": "Ins Approval Pending", "team": team, 'modules':modules},
                            {"pending": "ADDL-SP/DSP Approval Pending", "team": team, 'modules':modules,"designation": {"$nin": ["SP", "DSP", "ADDL-SP/DSP"]}},
                            {"pending": "ADDL-SP/DSP Approval Pending", 'modules':modules,"token": {"$regex": "_CO"}, "designation": {"$nin": ["SP", "DSP", "ADDL-SP/DSP"]}},
                            {"pending": {"$regex": officername}, "team": team,'modules':modules},
                            {"pending": "Ticket_Closed", "team": team,'modules':modules, "designation": {"$nin": ["SP", "ADDL-SP/DSP"]}},
                            {"pending": "Mail Under Process", "token": {"$regex": "_CO"},'modules':modules, "designation": {"$nin": ["SP", "DSP", "ADDL-SP/DSP"]}},
                            {"pending": "Annexe Send to CAT INS", "token": {"$regex": "_CO"},'modules':modules, "designation": {"$nin": ["SP", "DSP", "ADDL-SP/DSP"]}},
                            {"pending": "Assign to CAT Ins", "designation": {"$nin": ["SP", "DSP", "ADDL-SP/DSP"]}},
                            {"team": {"$ne": "CAT"}, "approval": {"$regex": "^Assigned"}},
                            {"approval": {"$regex": "^Assigned"}},
                            {"pending": "Mail Under Process",'modules':modules, "team": team, "designation": {"$nin": ["SP", "DSP", "ADDL-SP/DSP"]}},
                            {"pending": {"$regex": "^Rejected"}, "team": team, 'modules':modules,"designation": {"$nin": ["SP", "DSP", "ADDL-SP/DSP"]}}
                        ]
                    }]}
            else:
                query = {'$and' : [
                    {'useremail': {'$nin': [email]}},
                    {'requestcategory': {'$in' :  category}},
                    {"$or": [
                        # DATA REQ
                        {"pending": "Ins Approval Pending", "team": team,'modules':modules, "others": {"$in": [officername, '']},"designation": {"$nin": ["SP", "DSP", "ADDL-SP/DSP"]}},
                        {"pending": "ADDL-SP/DSP Approval Pending",'modules':modules, "team": team, "others": {"$in": [officername, '']},"designation": {"$nin": ["SP", "DSP", "ADDL-SP/DSP"]}},
                        {"pending": {"$regex": officername}, "team": team,'modules':modules,"designation": {"$nin": ["SP", "DSP", "ADDL-SP/DSP"]}},
                        {"pending": "Annexe Send to CAT INS", "team": team, 'modules':modules,"others": {"$in": [officername, '']},"designation": {"$nin": ["SP", "DSP", "ADDL-SP/DSP"]}},
                        {"pending": {"$regex": "^Rejected"}, "team": team, 'modules':modules,"others": {"$in": [officername, '']},"designation": {"$nin": ["SP", "DSP", "ADDL-SP/DSP"]}},
                        {"pending": "Mail Under Process", "team": team,'modules':modules, "others": {"$in": [officername, '']},"designation": {"$nin": ["SP", "DSP", "ADDL-SP/DSP"]}},
                        {"pending": "SP Approval Pending", "team": team, 'modules':modules,"others": {"$in": [officername, '']},"designation": {"$nin": ["SP", "DSP", "ADDL-SP/DSP"]}},
                        # ANALYSIS REQ
                        {"pending": "Ins Approval Pending", 'modules':modules,"requestcategory": "Analysis Request", "team": team,"designation": {"$nin": ["SP", "DSP", "ADDL-SP/DSP"]}},
                        {"approval": {"$regex": "^Assigned"}, 'modules':modules,"requestcategory": "Analysis Request", "team": team,"designation": {"$nin": ["SP", "DSP", "ADDL-SP/DSP"]}},
                        {"pending": "ADDL-SP/DSP Approval Pending", 'modules':modules,"requestcategory": "Analysis Request", "team": team,"designation": {"$nin": ["SP", "DSP", "ADDL-SP/DSP"]}},
                        {"pending": "SP Approval Pending", 'modules':modules,"requestcategory": "Analysis Request", "team": team,"designation": {"$nin": ["SP", "DSP", "ADDL-SP/DSP"]}},
                        {"pending": "Assign to CAT Ins", 'modules':modules,"requestcategory": "Analysis Request", "team": team,"designation": {"$nin": ["SP", "DSP", "ADDL-SP/DSP"]}}
                    ]
                }]}
        elif role == "Analyst":
            query = {"$or": [
                        {'assign_Officer': {'$in':[officername]}}
                    ]
                }
        elif role == "Mailer":
            query = {"$or": [
                        {"requestcategory": "Data Request", 'pending' : {'$in': ['Mail Under Process',
                        'Ticket_Closed']}},
                    ]
                }

            # Use $gte (greater than or equal) and $lte (less than or equal) operators
        My_tickets = coll.find(query, projection).sort([('_id', -1)])
        res_len = 0
        newno_len = 0
        total_len = 0
        filtered_tickets = []
        for res in My_tickets:
            if res['date'] in result_dates and res['useremail'] != email:
                if res['pending'] == 'Mail Under Process' and (res['subtypes'] == 'Phone' or res['requesttype'] == 'CDR'):
                    res_len = len(res['result'])
                    total_len = res_len - int(res['status_count'])
                elif res['pending'] == 'Mail Under Process' and (res['subtypes'] == 'IMEI' or res['requesttype'] == 'IMEI CDR'):
                    newno_len = len(res['newnumber'])
                    total_len = newno_len - int(res['status_count'])
                    

                filtered_tickets.append({
                    '_id': res['_id'],
                    'priority': res['priority'],
                    'useremail':res['useremail'],
                    'date': res['date'],
                    'token': res['token'],
                    'approval': res['approval'],
                    'pending': res['pending'],
                    'officername': res['officername'],
                    'edit': res['edit'],
                    'requestcategory': res['requestcategory'],
                    'requesttype': res['requesttype'],
                    'requesttypes': res['requesttypes'],
                    'subtypes': res['subtypes'],
                    'status_count': res['status_count'],
                    'total_len': total_len,
                    "res_len": res_len,
                    "newno_len": newno_len
                })
        pagination_info = {
                    'total_pages1': 1,
                    'current_page1': 1,
                    'total_tickets1': len(filtered_tickets)
                }
        return jsonify({'mytickets': filtered_tickets,'pagination': pagination_info})
    else:
        return jsonify("WAIT")


@tickect_bp.route('/fetch_status', methods=['POST', 'GET'])
def fetch_status():
    coll = db['tickets']
    headertype = request.json.get('type')
    cat = request.json.get('cat')
    if cat == 'Data Request':
        category = ['Data Request']
    elif cat == 'Analysis Request':
        category = ['Analysis Request','Note']
    email = request.json.get('email')
    status = set()
    if headertype == 'approval':
        status = (set(res['approval'] for res in coll.find({'useremail': email,'requestcategory': {'$in' :  category}},{'approval' : 1,"_id":0})))
    elif headertype == 'pending':
        status = (set(res['pending'] for res in coll.find({'useremail': email,'requestcategory': {'$in' :  category}},{'pending' : 1,"_id":0})))
    elif headertype == 'category':
        status = (set(res['requestcategory'] for res in coll.find({'useremail': email,'requestcategory': {'$in' :  category}},{'requestcategory' : 1,"_id":0})))
    elif headertype == 'requesttype':
        list1 = set()
        list2 = set()
        # Iterate over the results
        for res in coll.find({'useremail': email,'requestcategory': {'$in' :  category}}, {'requesttype': 1, 'date': 1, 'requesttypes': 1, "_id": 0}):

            if cat == 'Data Request' and res['requesttype'] != '':
                list1.add(res['requesttype'])
            # Append 'requesttypes' value to the list if it exists
            elif cat == 'Analysis Request':
                if 'requesttypes' in res and res['requesttypes'] != []:
                    for req in res['requesttypes']:
                        list1.add(req)
        status = list(list1)
    #print(list(status))
    return jsonify(list(status))


@tickect_bp.route('/fetch_status2', methods=['POST', 'GET'])
def fetch_status2():
    coll = db['tickets']
    headertype = request.json.get('type')
    cat = request.json.get('cat')
    if cat == 'Data Request':
        category = ['Data Request']
    elif cat == 'Analysis Request':
        category = ['Analysis Request','Note']
    designation = request.json.get('designation')
    role = request.json.get('role')
    email = request.json.get('email')
    team = request.json.get('team')
    modules = request.json.get('modules')
    officername = request.json.get("officer")
    status = set()
    list1 = set()
    list2 = set()
    query = {}
    print(cat)
    if cat == 'Data Request':
        category = ['Data Request']
    elif cat == 'Analysis Request':
        category = ['Analysis Request','Note']
    projection = {
            'approval' : 1,
            'officername' : 1,
            'pending' : 1,
            'requestcategory' : 1,
            'requesttype' : 1,
            'requesttypes' : 1,
            'edit':1
    }
    if team == 'ADMIN':
        query = {
                'useremail': {'$nin': [email]},
                'requestcategory': {'$in' :  category}
            }
    
    elif designation == 'SP' and team != 'ADMIN':
        if team == 'CAT':
                query = {'$and' : [
                    {'useremail': {'$nin': [email]}},
                    {'team': team},
                    {'requestcategory': {'$in' :  category}},
                    {"$or": [
                        {"pending": "ADDL-SP/DSP Approval Pending", "team": team},
                        {"$and": [{"pending": {"$regex": officername}}, {"team": team}]},
                        {"pending": "Mail Under Process", "token": {"$regex": "_CO"}},
                        {"approval": "Approved by CAT DSP", "token": {"$regex": "_CO"}},
                        {"pending": "--", "requestcategory": "Analysis Request"},
                        {"approval": {"$regex": "^Assigned"}, "requestcategory": "Analysis Request"},
                        {"pending": "SP Approval Pending", "requestcategory": "Analysis Request"},
                        {"pending": "SP Approval Pending", "team": team},
                        {"pending": "Mail Under Process", "team": team},
                        {"pending": "Ticket_Closed", "team": team},
                        {"pending": {"$regex": "^Rejected"}, "team": team}
                    ]}]}
        else:
            query = {'$and' : [
                {'useremail': {'$nin': [email]}},
                {'team': team},
                {'requestcategory': {'$in' :  category}},
                {"$or": [
                    {"pending": "SP Approval Pending", "team": team,},
                    {"pending": "SP Approval Pending", "team": team,},
                    {"pending": "SP Approval Pending", "team": team},
                    {"approval": "Approved by SP", "team": team,},
                    {"approval": "Approved by SP", "team": team,},
                    {"token": {"$regex": "_CO"}, "team": team},
                    {"pending": {"$regex": officername}, "team": team},
                    {"pending": "Assign to CAT Ins", "requestcategory": "Analysis Request", "team": team,},
                    {"pending": "SP Approval Pending", "requestcategory": "Analysis Request", "team": team,},
                    {"pending": "Assign to CAT Ins", "requestcategory": "Analysis Request", "team": team,},
                    {"pending": "Assign to CAT Ins", "requestcategory": "Analysis Request", "team": team,},
                    {"pending": "SP Approval Pending", "requestcategory": "Analysis Request", "team": team,},
                    {"pending": "SP Approval Pending", "requestcategory": "Analysis Request", "team": team,},
                    {"approval": {"$regex": "^Assigned"}, "requestcategory": "Analysis Request", "team": team},
                    {"pending": "Mail Under Process", "team": team},
                    {"pending": {"$regex": "^Rejected"}, "team": team}
                ]}]}

    elif designation == 'ADDL-SP/DSP':
        if team == "CAT":
                query = {'$and' : [
                    {'useremail': {'$nin': [email]}},
                    {'requestcategory': {'$in' :  category}},
                    {"$or": [
                        {"pending": "Mail Under Process", "token": {"$regex": "_CO"}, "designation": {"$ne": "SP"}},
                        {"$and": [{"pending": {"$regex": officername}}, {"team": team}]},
                        {"pending": "Ticket_Closed", "team": team, "designation": {"$ne": "SP"}},
                        {"approval": "Approved by CAT Ins", "token": {"$regex": "_CO"}, "designation": {"$ne": "SP"}},
                        {"pending": "ADDL-SP/DSP Approval Pending", "requestcategory": "Analysis Request"},
                        {"$and": [{"pending": "--"}, {"requestcategory": "Analysis Request"}, {"designation": {"$ne": "SP"}}]},
                        {"$and": [{"pending": "ADDL-SP/DSP Approval Pending"}, {"team": team}]},
                        {"approval": {"$regex": "^Assigned"}, "requestcategory": "Analysis Request", "designation": {"$ne": "SP"}},
                        {"approval": {"$regex": "^Assigned"}, "requestcategory": "Analysis Request", "designation": {"$ne": "SP"}},
                        {"pending": "SP Approval Pending", "team": team, "designation": {"$ne": "ADDL-SP/DSP"}},
                        {"pending": "Mail Under Process", "team": team, "designation": {"$ne": "SP"}},
                        {"pending": {"$regex": "^Rejected"}, "team": team}
                        ]}
                ]
                    }
        else:
            query = {'$and' : [
                {'useremail': {'$nin': [email]}},
                {'requestcategory': {'$in' :  category}},
                {"$or": [
                    {"$and": [{"pending": {"$regex": officername}}, {"team": team}]},
                    {"$and": [{"token": {"$regex": "_CO"}}, {"team": team}, {"others": {"$in": [officername]}}]},
                    {"pending": "ADDL-SP/DSP Approval Pending", "team": team, "requestcategory": "Data Request", "superior": officername},
                    {"pending": "ADDL-SP/DSP Approval Pending", "team": team, "requestcategory": "Data Request", "others": officername},
                    {"pending": "SP Approval Pending", "team": team, "others": {"$in": [officername]}},
                    {"pending": "SP Approval Pending", "team": team, "superior": officername},
                    {"pending": "SP Approval Pending", "team": team, "others": officername},
                    {"pending": {"$regex": "Rejected"}, "team": team, "superior": officername},
                    {"pending": {"$regex": "Rejected"}, "team": team, "others": officername},
                    {"pending": "ADDL-SP/DSP Approval Pending", "requestcategory": "Analysis Request", "team": team, "others": {"$in": [officername]}},
                    {"pending": "ADDL-SP/DSP Approval Pending", "requestcategory": "Analysis Request", "team": team, "superior": officername},
                    {"pending": "ADDL-SP/DSP Approval Pending", "requestcategory": "Analysis Request", "team": team, "others": officername},
                    {"pending": "Assign to CAT Ins", "requestcategory": "Analysis Request", "team": team, "designation": {"$nin": ["SP"]}, "others": {"$in": [officername]}},
                    {"pending": "Assign to CAT Ins", "requestcategory": "Analysis Request", "team": team, "designation": {"$nin": ["SP"]}, "superior": officername},
                    {"pending": "Assign to CAT Ins", "requestcategory": "Analysis Request", "team": team, "designation": {"$nin": ["SP"]}, "others": officername},
                    {"pending": "Mail Under Process", "team": team, "designation": {"$ne": "SP"}, "others": officername},
                    {"pending": "Mail Under Process", "team": team, "designation": {"$ne": "SP"}, "superior": officername},
                    {"approval": {"$regex": "Assigned"}, "requestcategory": "Analysis Request", "team": team, "designation": {"$nin": ["SP"]}, "others": {"$in": [officername]}},
                    {"approval": {"$regex": "Assigned"}, "requestcategory": "Analysis Request", "team": team, "designation": {"$nin": ["SP"]}, "superior": officername},
                    {"approval": {"$regex": "Assigned"}, "requestcategory": "Analysis Request", "team": team, "designation": {"$nin": ["SP"]}, "others": officername}
                    ]}
            ]}
    elif designation == "INSPR" and role != "Analyst":
        if team == 'CAT':
                query = {'$and' : [
                {'useremail': {'$nin': [email]}},
                {'requestcategory': {'$in' :  category}},
                {"$or": [
                        {"pending": "Ins Approval Pending", "team": team, 'modules':modules},
                        {"pending": "ADDL-SP/DSP Approval Pending", "team": team, 'modules':modules,"designation": {"$nin": ["SP", "DSP", "ADDL-SP/DSP"]}},
                        {"pending": "ADDL-SP/DSP Approval Pending", 'modules':modules,"token": {"$regex": "_CO"}, "designation": {"$nin": ["SP", "DSP", "ADDL-SP/DSP"]}},
                        {"pending": {"$regex": officername}, "team": team,'modules':modules},
                        {"pending": "Ticket_Closed", "team": team,'modules':modules, "designation": {"$nin": ["SP", "ADDL-SP/DSP"]}},
                        {"pending": "Mail Under Process", "token": {"$regex": "_CO"},'modules':modules, "designation": {"$nin": ["SP", "DSP", "ADDL-SP/DSP"]}},
                        {"pending": "Annexe Send to CAT INS", "token": {"$regex": "_CO"},'modules':modules, "designation": {"$nin": ["SP", "DSP", "ADDL-SP/DSP"]}},
                        {"pending": "Assign to CAT Ins", "requestcategory": "Analysis Request",'modules':modules, "designation": {"$nin": ["SP", "DSP", "ADDL-SP/DSP"]}},
                        {"team": {"$ne": "CAT"}, "approval": {"$regex": "^Assigned"}},
                        {"approval": {"$regex": "^Assigned"}},
                        {"pending": "Mail Under Process",'modules':modules, "team": team, "designation": {"$nin": ["SP", "DSP", "ADDL-SP/DSP"]}},
                        {"pending": {"$regex": "^Rejected"}, "team": team, 'modules':modules,"designation": {"$nin": ["SP", "DSP", "ADDL-SP/DSP"]}}
                    ]
                }]}
        else:
            query = {'$and' : [
                {'useremail': {'$nin': [email]}},
                {'requestcategory': {'$in' :  category}},
                {"$or": [
                    # DATA REQ
                    {"pending": "Ins Approval Pending", "team": team,'modules':modules, "others": {"$in": [officername, '']},"designation": {"$nin": ["SP", "DSP", "ADDL-SP/DSP"]}},
                    {"pending": "ADDL-SP/DSP Approval Pending",'modules':modules, "team": team, "others": {"$in": [officername, '']},"designation": {"$nin": ["SP", "DSP", "ADDL-SP/DSP"]}},
                    {"pending": {"$regex": officername}, "team": team,'modules':modules,"designation": {"$nin": ["SP", "DSP", "ADDL-SP/DSP"]}},
                    {"pending": "Annexe Send to CAT INS", "team": team, 'modules':modules,"others": {"$in": [officername, '']},"designation": {"$nin": ["SP", "DSP", "ADDL-SP/DSP"]}},
                    {"pending": {"$regex": "^Rejected"}, "team": team, 'modules':modules,"others": {"$in": [officername, '']},"designation": {"$nin": ["SP", "DSP", "ADDL-SP/DSP"]}},
                    {"pending": "Mail Under Process", "team": team,'modules':modules, "others": {"$in": [officername, '']},"designation": {"$nin": ["SP", "DSP", "ADDL-SP/DSP"]}},
                    {"pending": "SP Approval Pending", "team": team, 'modules':modules,"others": {"$in": [officername, '']},"designation": {"$nin": ["SP", "DSP", "ADDL-SP/DSP"]}},
                    # ANALYSIS REQ
                    {"pending": "Ins Approval Pending", 'modules':modules,"requestcategory": "Analysis Request", "team": team,"designation": {"$nin": ["SP", "DSP", "ADDL-SP/DSP"]}},
                    {"approval": {"$regex": "^Assigned"}, 'modules':modules,"requestcategory": "Analysis Request", "team": team,"designation": {"$nin": ["SP", "DSP", "ADDL-SP/DSP"]}},
                    {"pending": "ADDL-SP/DSP Approval Pending", 'modules':modules,"requestcategory": "Analysis Request", "team": team,"designation": {"$nin": ["SP", "DSP", "ADDL-SP/DSP"]}},
                    {"pending": "SP Approval Pending", 'modules':modules,"requestcategory": "Analysis Request", "team": team,"designation": {"$nin": ["SP", "DSP", "ADDL-SP/DSP"]}},
                    {"pending": "Assign to CAT Ins", 'modules':modules,"requestcategory": "Analysis Request", "team": team,"designation": {"$nin": ["SP", "DSP", "ADDL-SP/DSP"]}}
                ]
            }]}
    elif role == "Analyst":
        query = {"$or": [
                    {'assign_Officer': {'$in':[officername]}}
                ]
            }
    elif role == "Mailer":
        query = {"$or": [
                    {"requestcategory": "Data Request", 'pending' : {'$in': ['Mail Under Process',
                    'Ticket_Closed']}},
                ]
            }
    My_tickets = coll.find(query, projection).sort([('_id', -1)])
    for res in My_tickets:
        if headertype == 'approval':
            status.add(res['approval'])

        elif headertype == 'pending':
            status.add(res['pending'])

        elif headertype == 'category':
            status.add(res['requestcategory'])

        elif headertype == 'officername':
            status.add(res['officername'])

        elif headertype == 'edit':
            status.add(res['edit'])

        elif headertype == 'requesttype':
            if cat == 'Data Request' and res['requesttype'] != '':
                list1.add(res['requesttype'])
            # Append 'requesttypes' value to the list if it exists
            elif cat == 'Analysis Request':
                if 'requesttypes' in res and res['requesttypes'] != []:
                    for req in res['requesttypes']:
                        list1.add(req)
            status = list(list1)
            # #print(status)
    return jsonify(list(status))

@tickect_bp.route('/fetch_ins', methods=['POST', 'GET'])
def fetch_ins():
    coll = db['users']
    types = request.json.get('type')
    team = request.json.get('team')
    data = []
    if types == 'ins':
        data = list(coll.find({"designation":"INSPR",'team':team},{'officername': 1,'_id': 0}))
    if types == 'dsp':
        data = list(coll.find({"designation":"ADDL-SP/DSP",'team':team},{'officername': 1,'_id': 0}))
    if types == 'sp':
        data = list(coll.find({"designation":"SP",'team':team},{'officername': 1,'_id': 0}))
    #print(data)
    return jsonify(data)

@tickect_bp.route('/fetch_officers', methods=['POST', 'GET'])
def fetch_officers():
    coll = db['tickets']
    team = request.json.get('team')
    designation = request.json.get('designation')
    role = request.json.get('role')
    email = request.json.get('email')
    modules = request.json.get('modules')
    officer = []
    officer_list = []
    main_data = set()
    if team == 'ADMIN':
        officer_list = coll.find({},{'officername': 1,'_id' : 0})
    elif designation == 'SP' and team != 'ADMIN':
        officer_list = coll.find({'team':team},{'officername': 1,'_id' : 0})
    elif designation == 'ADDL-SP/DSP':
        officer_list = coll.find({'team':team,'modules':modules,'designation': {'$nin': ['SP','DIG','IG']}},{'officername': 1,'_id' : 0})
    elif designation == 'INSPR' and role != 'Analyst':
        officer_list = coll.find({'team':team,'modules':modules,'designation': {'$nin': ['SP','DIG','IG','ADDL-SP/DSP']}},{'officername': 1,'_id' : 0})

    for res in officer_list:
        main_data.add(res['officername'])
    officer = list(main_data)
    return jsonify(officer)

@tickect_bp.route('/NickName_data', methods=['POST', 'GET'])
def NickName_data():
    email = request.json.get('email')
    nickname = sus_db['cdat_suspect']
    coll = db['users']
    main_data = []
    page = request.json.get('page')
    limit = request.json.get('limit')
    if page is None or limit is None:
        page = 1
        limit = 12
    skip = (page - 1) * limit
    user = coll.find_one({'email': email})
    if user:
        my_data= nickname.find({'action': {'$ne': 'deleted'}},{'phone': 1,'nickname' : 1,'as_on_date': 1,'useremail':1,'update_on_date':1}).sort([('_id', -1)]).skip(skip).limit(limit)
        total_data = nickname.count_documents({})  # Total count of tickets

    else:
        my_data= nickname.find({'action': {'$ne': 'deleted'}},{'phone': 1,'nickname' : 1,'as_on_date': 1,'useremail':1,'update_on_date':1}).sort([('_id', -1)]).skip(skip).limit(limit)
        total_data = nickname.count_documents({})
    for res in my_data:
        if user:
            if res['useremail'] != '':
                user2 = coll.find_one({'email': res['useremail']})
            else:
                user2 = {'team': '','officername': ''}

            if 'update_on_date' in res and res['update_on_date']:
                update = res['update_on_date']
            else:
                update = '-'
            main_data.append({
                '_id': res['_id'],
                'phone': res['phone'],
                'nickname': res['nickname'],
                'created_on': res['as_on_date'],
                'updated_on': update,
                'team': user2['team'],
                'officername': user2['officername'],
            })
        else:
            if res['useremail'] != '':
                user2 = coll.find_one({'email': res['useremail']})
            else:
                user2 = {'team': '','officername': ''}

            if 'update_on_date' in res and res['update_on_date']:
                update = res['update_on_date']
            else:
                update = '-'
            main_data.append({
                '_id': res['_id'],
                'phone': res['phone'],
                'nickname': res['nickname'],
                'created_on': res['as_on_date'],
                'updated_on': update,
                'team': user2['team'],
                'officername': user2['officername'],
            })
    total_pages = math.ceil(total_data / limit) # limit
    pagination_info = {
        'total_pages': int(total_pages),
        'current_page': page,
        'total_tickets': total_data
    }
    return jsonify({'my_data': main_data, 'pagination': pagination_info})

@tickect_bp.route('/nickname_search2', methods=['POST', 'GET'])
def nickname_search2():
    email = request.json.get('email')
    value = request.json.get('value')
    value = value.strip()
    nickanme = sus_db['cdat_suspect']
    main_data = []
    coll = db['users']
    user = coll.find_one({'email': email})
    if user:
        if len(value) == 10:
            data=nickanme.find({'phone': value},{'phone': 1,'nickname' : 1,'as_on_date': 1,'useremail':1}).sort([('_id', -1)])
        else:
            data=nickanme.find({'nickname': value},{'phone': 1,'nickname' : 1,'as_on_date': 1,'useremail':1}).sort([('_id', -1)])
    else:
        if len(value) == 10:
            data=nickanme.find({'phone': value},{'phone': 1,'nickname' : 1,'as_on_date': 1,'useremail':1}).sort([('_id', -1)])
        else:
            data=nickanme.find({'nickname': value},{'phone': 1,'nickname' : 1,'as_on_date': 1,'useremail':1}).sort([('_id', -1)])
    for res in data:
        if user:
            main_data.append({
                'phone': res['phone'],
                'nickname': res['nickname'],
                'created_on': res['as_on_date'],
                'team': user['team'],
                'officername': user['officername'],
            })
        else:
            if res['useremail'] != '':
                user2 = coll.find_one({'email': res['useremail']})
            else:
                user2 = {'team': '','officername': ''}
            main_data.append({
                'phone': res['phone'],
                'nickname': res['nickname'],
                'created_on': res['as_on_date'],
                'team': user2['team'],
                'officername': user2['officername'],
            })
    pagination_info = {
                'total_pages': 1,
                'current_page': 1,
                'total_tickets': len(main_data)
            }

    return jsonify({'my_data': main_data, 'pagination': pagination_info})

@tickect_bp.route('/nickname_upload', methods=['POST'])
def nickanme_upload():
    if 'file' not in request.files:
        return jsonify({'error': 'No file provided'}), 400
    file = request.files['file']
    time = request.form['time']

    email = request.form['email']
    #print(category,reqtype,reqtype2,subtype)
    if file.filename == '':
        return jsonify({'error': 'No selected file'}), 400
    # Check if the file has an allowed extension
    if not allowed_file(file.filename):
        return jsonify({'error': 'Invalid file type'}), 400
    # Create the 'uploaded_excel_files' directory if it doesn't exist
    upload_path = f'{os.getcwd()}/app/ticketing/uploaded_excel_files/'
    if not os.path.exists(upload_path):
        os.makedirs(upload_path)
    # Save the file with a unique name in the 'uploaded_excel_files' folder

    filename_prefix = f"Super_admin_{time}"
    print(file.filename)
    filename = secure_filename(f"{filename_prefix}_{file.filename}")
    filepath = os.path.join(upload_path, filename)
    serial_number = 1
    while os.path.exists(filepath):
        # If the file already exists, modify the filename
        base_filename, extension = os.path.splitext(file.filename)
        serial_filename = f"{filename_prefix}_{base_filename}_{serial_number}{extension}"
        filepath = os.path.join(upload_path, serial_filename)
        serial_number += 1

    file.save(filepath)
    data = []
    # Perform file processing
    workbook = openpyxl.load_workbook(filepath)
    sheet = workbook.active
    columns = []
    for col in range(1, sheet.max_column + 1):
        column_name1 = sheet.cell(row=1, column=col).value or f'Column_{col}'
        columns.append(column_name1)
    for row in sheet.iter_rows(min_row=2, values_only=True):
        row_data = {}
        for col, value in enumerate(row, start=1):
            column_name = sheet.cell(row=1, column=col).value or f'Column_{col}'
            row_data[column_name] = str(value).strip()
        print(row_data)
        # Phone	Nickname	Officername	Team	Created_date
        data.append({
            'phone': row_data["Phone"],
            'nickname' : row_data['Nickname'],
            'officername': row_data['Officername'],
            'team':row_data['Team'],
        })
    return jsonify(data)

@tickect_bp.route('/nickdata_update', methods=['POST'])
def nickdata_update():
    nickname = sus_db['cdat_suspect']
    coll = db['users']
    data = request.json.get('data')
    now = datetime.now()
    date = now.strftime("%d/%m/%Y")
    time = now.strftime("%H:%M:%S")
    for res in data:
        user = coll.find_one({'officername':res['officername']})
        if user:
            email = user['email']
        else:
            email = ''
        num = nickname.find_one({'useremail':email,'phone':res['phone']})
        nickname.insert_one({
                    "phone":res['phone'],
                    "nickname":res['nickname'],
                    'useremail':email,
                    "isactive": None,
                    "last_name": None,
                    "first_name": None,
                    "address": None,
                    "city": None,
                    "state": None,
                    "country": None,
                    "pin": None,
                    "remark": None,
                    "checkflag": None,
                    "imeinumber": None,
                    "as_on_datetime": num['as_on_datetime'],
                    "as_on_date": num['as_on_date'],
                    "as_on_time": num['as_on_time'],
                    'update_on_datetime': now,
                    "update_on_date": date,
                    "inc_officer": None,
                    "module_name": None,
                    "monit_status": None,
                    "category": None,
                    "organization": None,
                    "module_code": None
                    })
    return jsonify('')

@tickect_bp.route('/nickname_excel', methods=['POST', 'GET'])
def nickname_excel():
    # Phone	Nickname	Officername	Team	Created_date
    col = ['Phone', 'Nickname', 'Officername', 'Team']
    # Create an empty DataFrame with columns
    df = pd.DataFrame(columns=col)
    excel_data = io.BytesIO()
    # Write the DataFrame to an Excel file
    with pd.ExcelWriter(excel_data, engine='xlsxwriter') as writer:
        df.to_excel(writer, sheet_name='Sheet1', index=False)
    excel_data.seek(0)
    # Send the Excel file as a response for download
    response = send_file(
        excel_data,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='/home/bruce/Downloads/output.xlsx'
    )
    return response

@tickect_bp.route('/dashexcel', methods=['POST', 'GET'])
def dashexcel():
    coll = db['tickets']
    reqtype_list = ['SDR','CAF','CDR','IMEI CDR','GPRS','IPDR','RH',
                    'POA','TOWER CDR','TOWER GPRS','TOWER IPDR']
    cat = request.json.get('cat')
    data = request.json.get('data')
    tabledata = request.json.get('tabledata')
    reqtype = request.json.get('reqtype')
    status = request.json.get('status')
    daterange = request.json.get('daterange')
    sum = 0
    main_data = []
    concatenated_dfs = []
    total_recieved = []
    for total in tabledata:
        if total['requesttype'] in ['CDR' ,'IMEI CDR']:
            sum = sum + total['total_len']

    col = ['Request Type', 'Date Range', 'Filter Status', 'Raised','Closed','Approved','Pending','Data Receival Pending','Total Data Received']
    if reqtype != 'All':
        for req in reqtype_list:
            for res in data:
                if req in res:
                    main_data = [{
                        'Request Type' : req,
                        'Date Range' : daterange,
                        'Status' : status,
                        'Raised' : res[req]['Raise'],
                        'Closed' : res[req]['Ticket_closed'],
                        'Approved' : res[req]['Approval'],
                        'Pending' : res[req]['Pending'],
                        'Data Receival Pending' : res[req]['Under_Process'],
                        'Total Data Received' : sum,
                    }]
        concatenated_df = pd.DataFrame(main_data, columns=col)
        concatenated_dfs.append(concatenated_df)
    else:
        for req in reqtype_list:
            for res in data:
                    main_data = [{
                        'Request Type' : req,
                        'Date Range' : daterange,
                        'Status' : status,
                        'Raised' : res[req]['Raise'],
                        'Closed' : res[req]['Ticket_closed'],
                        'Approved' : res[req]['Approval'],
                        'Pending' : res[req]['Pending'],
                        'Data Receival Pending' : res[req]['Under_Process'],
                        'Total Data Received' : sum,
                    }]
            concatenated_df = pd.DataFrame(main_data, columns=col)
            concatenated_dfs.append(concatenated_df)

    concatenated_df = pd.concat(concatenated_dfs, ignore_index=True)
    concatenated_df.dropna(axis=1, how='all', inplace=True)

    excel_data = io.BytesIO()
    with pd.ExcelWriter(excel_data, engine='xlsxwriter') as writer:
        concatenated_df.to_excel(writer, sheet_name='Sheet1', index=False)
    excel_data.seek(0)
    response = send_file(
        excel_data,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='output.xlsx'
    )
    return response

# @tickect_bp.route('/split_CO', methods=['POST'])
# def split_documents(key):
#     print('entered in split func.........................................')
#     coll = db["tickets"]
#     newnumber = []
#     cellone = []
#     J_A_V = []
#     J_A_V_new = []
#     user1= coll.find_one({'token':key})
#     # ISP = user1["newnumber"]

#     # if (user1['requesttype'] =='TOWER CDR' or user1['requesttype'] =='TOWER GPRS' or user1['requesttype'] =='TOWER IPDR') :
#     #     for isp in ISP:
#     #         if "CO" in isp["ISP"]:
#     #             print(user1["_id"])

#     #     return('0')


#     # else:

#     if any(("CO" in res2["result"]["MNP"]) or ('MNP' in res2 and "CO" in res2["MNP"]) for res2 in user1["newnumber"]):
#         if len(user1["result"]) > 0:
#             for res in user1["result"]:
#                 if "CO" in res["MNP"]:
#                     cellone.append(res)
#                 else:
#                     J_A_V.append(res)
#         else:
#             for res2 in user1["newnumber"]:
#                 if ("CO" in res2["result"]["MNP"]) or ('MNP' in res2 and "CO" in res2["MNP"]):
#                     cellone.append(res2["result"])
#                 else:
#                     J_A_V.append(res2["result"])



#         for res2 in user1["newnumber"]:
#             if ("CO" in res2["result"]["MNP"]) or ('MNP' in res2 and "CO" in res2["MNP"]):
#                 newnumber.append(res2)
#             else:
#                 J_A_V_new.append(res2)

#         coll.insert_one({
#             'useremail' : user1['useremail'] ,
#             'superior' : user1['superior'] ,
#             'officername' : user1['officername'],
#             'role' : user1['role'],
#             'designation' : user1['designation'],
#             'type' : user1['type'],
#             'requestcategory' : user1['requestcategory'],
#             'requesttype' : user1['requesttype'],
#             'requesttypes' : user1['requesttypes'],
#             'subtypes' : user1['subtypes'],
#             'team' : user1['team'],
#             'modules' : user1['modules'],
#             'relation' : user1['relation'],
#             'nickname' : user1['nickname'],
#             'suspect' : user1['suspect'],
#             'name' : user1['name'],
#             'location' : user1['location'],
#             'source' : user1['source'],
#             'reason' : user1['reason'],
#             'refno' : user1['refno'],
#             'refdate' : user1['refdate'],
#             'priority' : user1['priority'],
#             'token' : user1['token'] + '_CO',
#             'pending' : user1['pending'],
#             'approval' : user1['approval'],
#             'others' : user1['others'],
#             'remarks' : user1['remarks'],
#             'date' : user1['date'],
#             'raise_time' : user1['raise_time'],
#             'status_count' : user1['status_count'],
#             'edit':user1['edit'],
#             'newnumber' : newnumber,
#             'result' : cellone,
#         })
#         coll.insert_one({
#             'useremail' : user1['useremail'] ,
#             'superior' : user1['superior'] ,
#             'officername' : user1['officername'],
#             'role' : user1['role'],
#             'designation' : user1['designation'],
#             'type' : user1['type'],
#             'requestcategory' : user1['requestcategory'],
#             'requesttype' : user1['requesttype'],
#             'requesttypes' : user1['requesttypes'],
#             'subtypes' : user1['subtypes'],
#             'team' : user1['team'],
#             'modules' : user1['modules'],
#             'relation' : user1['relation'],
#             'nickname' : user1['nickname'],
#             'suspect' : user1['suspect'],
#             'name' : user1['name'],
#             'location' : user1['location'],
#             'source' : user1['source'],
#             'reason' : user1['reason'],
#             'refno' : user1['refno'],
#             'refdate' : user1['refdate'],
#             'priority' : user1['priority'],
#             'token' : user1['token'] + '_Others',
#             'pending' : user1['pending'],
#             'approval' : user1['approval'],
#             'others' : user1['others'],
#             'remarks' : user1['remarks'],
#             'date' : user1['date'],
#             'raise_time' : user1['raise_time'],
#             'status_count' : user1['status_count'],
#             'edit':user1['edit'],
#             'newnumber' : J_A_V_new,
#             'result' : J_A_V,
#         })
#         newnumber.clear()
#         cellone.clear()
#         J_A_V.clear()
#         J_A_V_new.clear()
#         coll.delete_one({'token':key})
#     print('spliting done.......................................................')
#     return jsonify("Split Successfull........")

def split_documents(key,typess):
    print('entered in split func.........................................')
    coll = db["tickets"]
    newnumber = []
    cellone = []
    J_A_V = []
    J_A_V_new = []
    user1= coll.find_one({'token':key,'requestcategory': 'Data Request'})

    if typess != '':        
        print('im here...........................')
        if len(user1["result"]) > 0:
            for res in user1["result"]:
                if "CO" in res["MNP"]:
                    cellone.append(res)
                else:
                    J_A_V.append(res)
        else:
            for res2 in user1["newnumber"]:
                if 'result' in res2 or 'MNP' in res2:
                    if ('result' in res2 and "CO" in res2["result"]["MNP"]) or ('MNP' in res2 and "CO" in res2["MNP"]):
                        cellone.append(res2["result"])
                    else:
                        J_A_V.append(res2["result"])
                       
        for res2 in user1["newnumber"]:
            if ('result' in res2 and "CO" in res2["result"]["MNP"]) or ('MNP' in res2 and "CO" in res2["MNP"]):
                print('here alsooooooooooooooooooooooooooooooo')
                newnumber.append(res2)
            else:
                J_A_V_new.append(res2)

        coll.insert_one({
            'useremail' : user1['useremail'] ,
            'superior' : user1['superior'] ,
            'officername' : user1['officername'],
            'role' : user1['role'],
            'designation' : user1['designation'],
            'type' : user1['type'],
            'requestcategory' : user1['requestcategory'],
            'requesttype' : user1['requesttype'],
            'requesttypes' : user1['requesttypes'],
            'subtypes' : user1['subtypes'],
            'team' : user1['team'],
            'modules' : user1['modules'],
            'relation' : user1['relation'],
            'nickname' : user1['nickname'],
            'suspect' : user1['suspect'],
            'name' : user1['name'],
            'location' : user1['location'],
            'source' : user1['source'],
            'reason' : user1['reason'],
            'refno' : user1['refno'],
            'refdate' : user1['refdate'],
            'priority' : user1['priority'],
            'token' : user1['token'] + '_CO',
            'pending' : user1['pending'],
            'approval' : user1['approval'],
            'others' : user1['others'],
            'remarks' : user1['remarks'],
            'date' : user1['date'],
            'raise_time' : str(datetime.now()),
            'status_count' : user1['status_count'],
            'edit':user1['edit'],
            'newnumber' : newnumber,
            'result' : cellone,
        })
        coll.insert_one({
            'useremail' : user1['useremail'] ,
            'superior' : user1['superior'] ,
            'officername' : user1['officername'],
            'role' : user1['role'],
            'designation' : user1['designation'],
            'type' : user1['type'],
            'requestcategory' : user1['requestcategory'],
            'requesttype' : user1['requesttype'],
            'requesttypes' : user1['requesttypes'],
            'subtypes' : user1['subtypes'],
            'team' : user1['team'],
            'modules' : user1['modules'],
            'relation' : user1['relation'],
            'nickname' : user1['nickname'],
            'suspect' : user1['suspect'],
            'name' : user1['name'],
            'location' : user1['location'],
            'source' : user1['source'],
            'reason' : user1['reason'],
            'refno' : user1['refno'],
            'refdate' : user1['refdate'],
            'priority' : user1['priority'],
            'token' : user1['token'] + '_Others',
            'pending' : user1['pending'],
            'approval' : user1['approval'],
            'others' : user1['others'],
            'remarks' : user1['remarks'],
            'date' : user1['date'],
            'raise_time' : str(datetime.now()),
            'status_count' : user1['status_count'],
            'edit':user1['edit'],
            'newnumber' : J_A_V_new,
            'result' : J_A_V,
        })
        newnumber.clear()
        cellone.clear()
        J_A_V.clear()
        J_A_V_new.clear()
        coll.delete_one({'token':key,'requestcategory': 'Data Request'})
    print('spliting done.......................................................')
    return jsonify("Split Successfull........")
        

@tickect_bp.route('/upload-pdf', methods=['POST'])
def upload_pdf():
    id = request.json.get('id')
    id = id[0]
    st = request.json.get('st')
    url = request.json.get('url')
    coll = db['tickets']
    print('ID : ',id,'STATEs : ',st,'URL : ',url,'////////////////////////////////')
    main_url = []
    cdrs = coll.find_one({'_id': ObjectId(id)})
    if cdrs['requesttype'] == 'CDR':
       main_url = [url + '/ticketing/Annexure?user=' + id + '?st=' + 'All']
    for res in st:
        main_url.append(url + '/ticketing/Annexure?user=' + id + '?st=' + res)
        # return redirect(url + '/Annexure?user=' + id + '?st=' + res)
    print(main_url)
    return jsonify(main_url)
    



@tickect_bp.route('/view_annexure', methods=['POST'])
def view_annexure():
    data = []
    id = request.json.get("id")
    pattern = r'(\w+)\?st=(\w+)'
    match = re.search(pattern, id)
    annex_data={"data": [],'states':[],'Cell':[]}
    id = match.group(1)
    state = match.group(2)
    # print(state)
    if id != "undefined" and state != "All":
        print("helooooooooooooooooooooooooooooo")
        coll = db["tickets"]
        cell_one = coll.find_one({'_id': ObjectId(id)})
        annex_data["data"].append(cell_one)
        annex_data["data"].append({'state':state})
        if "_CO" in cell_one["token"]:
            for res in cell_one["newnumber"]:
              if res["result"]["LSA_ID"] == state:
                  annex_data['states'].append(res)
    else:
        print("byeeeeeeeeeeeeeeeeeeeee")
        coll = db["tickets"]
        cell_one = coll.find_one({'_id': ObjectId(id)})
        annex_data["data"].append(cell_one)
        annex_data["data"].append({'state':state})
        if "_CO" in cell_one["token"]:
            annex_data['states'] = cell_one["newnumber"]


        elif 'TOWER' in cell_one['requesttype']:
            for res in cell_one["newnumber"]:
                data.append({'CGI': res['CGI'],'From_Date':res['From_Date'],'To_Date':res['To_Date']})
            annex_data['states'] = data
        else:
             annex_data['states'] = cell_one["newnumber"]
    # print((annex_data["states"]))
    return jsonify(annex_data)
    
@tickect_bp.route('/unique_state', methods=['POST'])
def unique_state():
    annex_data={'states':[],"count" : []}
    id = request.json.get("id")
    print(id)
    coll = db["tickets"]
    uni = []
    cell_one = coll.find_one({'_id': ObjectId(id)})
    if "_CO" in cell_one["token"]:
        for res in cell_one["newnumber"]:
            uni.append(res["result"]["LSA_ID"])
            annex_data['states'].append(res["result"]["LSA_ID"])
    return jsonify(annex_data)

# its under working dont touch

# search_field for super_admin
@tickect_bp.route('/search_user', methods=['POST'])
def search_user():
    query = request.json.get('query')
    query=query.strip()
    # print(query)

    coll = db['users']
    users = coll.find({
        '$or': [
            {'officername': query},
            {'designation': query},
            {'email': query}
        ]
    })

    user_list = [user for user in users]
    if user_list:
        return jsonify({'users': user_list})
    else:
        return jsonify({'message': 'No users found for the provided query'})
    

@tickect_bp.route('/delete_case', methods=['POST'])
def delete_case():
    phone = request.json.get('phone')
    print(f"Phone No. to delete: {phone}")
    collect = sus_db['cdat_suspect']
    collect.update_one({"phone": phone}, {'$set': {'action': 'deleted'}})
    
    return jsonify({'message': f'Phone No. {phone} records deleted successfully'})

# @tickect_bp.route('/nickname_edit_case', methods=['POST'])
# def nickname_edit_case():
#     data = request.json  # Get the JSON data sent from the frontend
#     phone = data.get('phone')
#     nickname = data.get('nickname')
#     officername = data.get('officername')
#     team = data.get('team')
#     created_on = data.get('created_on')
#     updated_on = data.get('updated_on')

#     # Process the individual fields and update the database as needed
#     print(f"Phone: {phone}, Nickname: {nickname}, Officer Name: {officername}, Team: {team}, Created On: {created_on}, Updated On: {updated_on}")

#     return jsonify({'message': 'Records updated successfully'})




UPLOAD_FOLDER = 'uploads'
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

@app.route('/save-pdf', methods=['POST'])
def save_pdf():
    if 'pdfFile' not in request.files:
        return 'No PDF file received', 400

    pdf_files = request.files.getlist('pdfFile')
    my_file = request.form['pdfname']
    my_zip = request.form['ziptoken']

    if not pdf_files:
        return 'No PDF files selected', 400

    folder_path = os.path.join(app.config['UPLOAD_FOLDER'], secure_filename(my_zip))
    os.makedirs(folder_path, exist_ok=True)

    for idx, pdf_file in enumerate(pdf_files):
        if pdf_file.filename == '':
            continue
        filename = secure_filename(my_file)
        destination_path = os.path.join(folder_path, filename)
        pdf_file.save(destination_path)

    time.sleep(4)
    # Create a single zip file containing all PDF files
    zip_path = io.BytesIO()
    with zipfile.ZipFile(zip_path, 'w') as zip_file:
        for root, _, files in os.walk(folder_path):
            for file in files:
                file_path = os.path.join(root, file)
                zip_file.write(file_path, os.path.relpath(file_path, folder_path))

    # Move the pointer to the beginning of the BytesIO object
    zip_path.seek(0)

    # Send the zip file as a response
    return send_file(zip_path, mimetype='application/zip', as_attachment=True, download_name=f'{secure_filename(my_zip)}.zip')

@tickect_bp.route('/nickname_filter', methods=['POST'])
def nickname_filter():
    from_date = request.json.get('from_date')
    to_date = request.json.get('to_date')
    email = request.json.get('email')
    coll = db['users']
    user = coll.find_one({'email': email})
    nickname = sus_db['cdat_suspect']
    date_objects = []
    result_dates = []
    my_data = []
    total_data = []
    main_data = []
    if user:
        start_date1 = datetime.strptime(from_date, '%d/%m/%Y')
        end_date1 = datetime.strptime(to_date, '%d/%m/%Y')
        samples = [res for res in nickname.find({'useremail':email},{'as_on_date': 1,"_id": 0})]
        date_objects = [date_parser.parse(str(date['as_on_date'])).strftime("%d/%m/%Y") for date in samples]
        date_objects = [datetime.strptime(date, '%d/%m/%Y') for date in date_objects]
        result_dates = [date.strftime('%d/%m/%Y') for date in date_objects if start_date1 <= date <= end_date1]
        my_data = nickname.find({'useremail':email, 'as_on_date' : {'$in': result_dates}},{'phone': 1,'nickname' : 1,'as_on_date': 1,'useremail':1,'update_on_date':1}).sort([('_id', -1)])
        total_data = nickname.count_documents({'useremail':email, 'as_on_date' : {'$in': result_dates}})
    else:
        start_date1 = datetime.strptime(from_date, '%d/%m/%Y')
        end_date1 = datetime.strptime(to_date, '%d/%m/%Y')
        samples = [res for res in nickname.find({},{'as_on_date': 1,"_id": 0}) if res['as_on_date'] != None]
        date_objects = [date_parser.parse(str(date['as_on_date'])).strftime("%d/%m/%Y") for date in samples]
        date_objects = [datetime.strptime(date, '%d/%m/%Y') for date in date_objects]
        result_dates = [date.strftime('%d/%m/%Y') for date in date_objects if start_date1 <= date <= end_date1]
        my_data = nickname.find({'as_on_date' : {'$in': result_dates}},{'phone': 1,'nickname' : 1,'as_on_date': 1,'useremail':1,'update_on_date':1}).sort([('_id', -1)])
        total_data = nickname.count_documents({'as_on_date' : {'$in': result_dates}})

    for res in my_data:
        if user:
            if 'update_on_date' in res:
                update = res['update_on_date']
            else:
                update = '-'
            main_data.append({
                'phone': res['phone'],
                'nickname': res['nickname'],
                'created_on': res['as_on_date'],
                'updated_on': update,
                'team': user['team'],
                'officername': user['officername'],
            })
        else:
            if res['useremail'] != '':
                user2 = coll.find_one({'email': res['useremail']})
            else:
                user2 = {'team': '','officername': ''}

            if 'update_on_date' in res and res['update_on_date']:
                update = res['update_on_date']
            else:
                update = '-'
            main_data.append({
                'phone': res['phone'],
                'nickname': res['nickname'],
                'created_on': res['as_on_date'],
                'updated_on': update,
                'team': user2['team'],
                'officername': user2['officername'],
            })
    pagination_info = {
                'total_pages': 1,
                'current_page': 1,
                'total_tickets': total_data
            }
    return jsonify({'my_data': main_data, 'pagination': pagination_info})

@tickect_bp.route('/nickname_excel_download', methods=['POST'])
def nickname_excel_download():
    data = request.json.get('data')
    update = ''
    main_data = []
    concatenated_dfs = []
    col = ['Phone',"Nickname",'Officername','Team', 'Created On', "Updated On"]
    # {'created_on': '02/02/2024', 'nickname': 'CAT_RUSHIKSH_1234567890', 'officername': 'Mailer_CAT', 'phone': '9505516683', 'team': 'CAT', 'updated_on': '14/03/2024'},
    for res in data:
        if 'updated_on' not in res:
            update = ''
        else:
            update = res['updated_on']
        main_data = [{
                        'Phone' : res['phone'],
                        'Nickname' : res['nickname'],
                        'Officername' : res['officername'],
                        'Team' : res['team'],
                        'Created On' : res['created_on'],
                        'Updated On' : update,
                    }]
        concatenated_df = pd.DataFrame(main_data, columns=col)
        concatenated_dfs.append(concatenated_df)

    concatenated_df = pd.concat(concatenated_dfs, ignore_index=True)
    concatenated_df.dropna(axis=1, how='all', inplace=True)

    excel_data = io.BytesIO()
    with pd.ExcelWriter(excel_data, engine='xlsxwriter') as writer:
        concatenated_df.to_excel(writer, sheet_name='Sheet1', index=False)
    excel_data.seek(0)
    response = send_file(
        excel_data,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='output.xlsx'
    )
    return response

@tickect_bp.route('/nickname_save', methods=['POST', 'GET'])
def nickname_save():
    id = request.json.get('id')
    nickname = request.json.get('nickname')
    phone = request.json.get('phone')
    nick_db = sus_db['cdat_suspect']
    now = datetime.now()
    date = now.strftime("%d/%m/%Y")
    nick_db.update_one({'_id': ObjectId(id)},{'$set':{
        "phone": phone,
        "nickname":nickname,
        'update_on_datetime': now,
        "update_on_date": date,
    }})
    return jsonify({'messege':"Nickname Updated.....!"})